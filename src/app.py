# SPDX-License-Identifier: AGPL-3.0-only
# Copyright (c) 2025 R. A. Gardner

from __future__ import annotations

import os
import tkinter as tk
from contextlib import suppress
from sys import argv
from tkinter import filedialog, ttk

from openpyxl import load_workbook
from tksheet import (
    DotDict,
    alpha2idx,
)

from .classes import (
    Header,
    tk_trees_api,
)
from .constants import (
    BF,
    EF,
    EFB,
    ERR_ASK_FNT,
    STSF,
    TF,
    USER_OS,
    about_system,
    app_copyright,
    app_title,
    contact_email,
    ctrl_button,
    default_app_window_size,
    dropdown_font,
    menu_kwargs,
    std_font_size,
    top_left_icon,
    upone_dir,
    website1,
)
from .functions import (
    b32_x_dict,
    bytes_io_wb,
    center,
    csv_str_x_data,
    get_json_format,
    get_json_from_file,
    json_to_sheet,
    load_cfg,
    try_write_error_log,
    write_cfg,
    ws_x_data,
    ws_x_program_data_str,
)
from .toplevels import (
    Ask_Confirm_Quit,
    Error,
    First_Start_Popup,
    Help_Popup,
    License_Popup,  # noqa: F401
    Text_Popup,
)
from .tree_compare import Tree_Compare
from .tree_editor import Tree_Editor
from .widgets import (
    Column_Selection,
    Frame,
    Readonly_Entry,
    Status_Bar,
)

# ________________________ ALLOW USE OF API HERE ________________________
allow_api_use = True


class AppGUI(tk.Tk):
    def __init__(self, start_arg=None):
        tk.Tk.__init__(self)
        if start_arg is None:
            start_arg = []
        self.withdraw()
        # DEBUGGING
        # start_arg = ("","40k wo.xlsx")
        self.theme = "light_green"
        self.working = False
        self.save_menu_state = "save as"

        self.protocol("WM_DELETE_WINDOW", self.USER_HAS_CLOSED_WINDOW)
        if USER_OS == "darwin":
            self.createcommand("::tk::mac::Quit", self.USER_HAS_CLOSED_WINDOW)

        self.tk.call("wm", "iconphoto", self._w, tk.PhotoImage(format="gif", data=top_left_icon))
        self.title(app_title)
        style = ttk.Style()
        style.configure("TNotebook.Tab", font=("Calibri", std_font_size, "bold"))
        style.configure("Std.TButton", font=BF, borderwidth=0)
        style.configure("EF.Std.TButton", font=EF, borderwidth=0)
        style.configure("EFW.Std.TButton", font=EF, anchor="w", borderwidth=0)
        style.configure("TF.Std.TButton", font=TF, borderwidth=0)
        style.configure("STSF.Std.TButton", font=STSF, borderwidth=0)
        style.configure("EFB.Std.TButton", font=EFB, borderwidth=0)
        style.configure("ERR_ASK_FNT.Std.TButton", font=ERR_ASK_FNT, borderwidth=0)
        style.configure("x_button.Std.TButton", font=EF, anchor="e", borderwidth=0)
        style.configure("wx_button.Std.TButton", font=EF, anchor="w", borderwidth=0)
        self.option_add("*TCombobox*Listbox.font", dropdown_font)
        self.style = style

        self.menubar = tk.Menu(self, **menu_kwargs)
        self.EMPTYMENU = tk.Menu(self, **menu_kwargs)
        self.EMPTYMENU.add_radiobutton(**menu_kwargs)
        self.config(menu=self.menubar)
        self.file = tk.Menu(self.menubar, tearoff=0, **menu_kwargs)
        self.menubar.add_cascade(label="File", menu=self.file, **menu_kwargs)
        self.file.add_command(label="New", command=self.create_new_at_start, accelerator="Ctrl+N", **menu_kwargs)
        self.file.add_separator()
        self.file.add_command(label="Compare sheets", command=self.compare_at_start, **menu_kwargs)
        self.file.add_separator()
        self.file.add_command(label="Open", command=self.open_file_at_start, accelerator="Ctrl+O", **menu_kwargs)
        self.file.add_separator()
        self.file.add_command(label="Save", accelerator="Ctrl+S", state="disabled", **menu_kwargs)
        self.file.add_command(label="Save new version", state="disabled", **menu_kwargs)
        self.file.add_separator()
        self.file.add_command(label="Save as", **menu_kwargs)
        self.file.add_separator()
        self.file.add_command(label="Settings", state="disabled", **menu_kwargs)
        self.file.add_separator()
        self.file.add_command(label="Quit", command=self.USER_HAS_CLOSED_WINDOW, **menu_kwargs)

        self.wb = None
        self.frames = DotDict()
        self.open_dict = {}
        self.created_new = False

        self.USER_HAS_QUIT = False
        self.number_unsaved_changes = 0

        try:
            with open(upone_dir + "LICENSE.txt", "r") as fh:
                self.LICENSE = fh.read()
        except Exception as errormsg:
            Error(
                self,
                (
                    f"Error locating LICENSE file: '{errormsg}'.\n\n"
                    "LICENSE file must be in the same folder as '{app_title}.pyw'."
                ),
                theme=self.theme,
            )
            self.try_to_close_everything()
            return

        try:
            with open(upone_dir + "DOCUMENTATION.md", "r") as fh:
                self.DOCUMENTATION = fh.read()
        except Exception as errormsg:
            Error(
                self,
                (
                    f"Error locating DOCUMENTATION.md file: '{errormsg}'.\n\n"
                    "DOCUMENTATION.md file must be in the same folder as '{app_title}.pyw'."
                ),
                theme=self.theme,
            )
            self.DOCUMENTATION = "Could not locate documentation file."

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.frames["column_selection"] = Column_Selection(parent=self, C=self)
        self.frames["tree_compare"] = Tree_Compare(parent=self, C=self)
        self.frames["tree_edit"] = Tree_Editor(parent=self, C=self)
        self.current_frame = "tree_edit"
        self.frames["tree_edit"].grid(row=0, column=0, sticky="nsew")

        self.status_frame = Frame(self)
        self.status_frame.grid_rowconfigure(0, weight=1)
        self.status_frame.grid_columnconfigure(0, weight=1)
        self.status_frame.grid_columnconfigure(1, weight=1)
        self.status_frame.grid(row=1, column=0, sticky="ew")

        self.status_bar = Status_Bar(self.status_frame, text="")
        self.status_bar.grid(row=0, column=0, sticky="nsw", padx=(0, 5), pady=0)
        self.selection_info = Readonly_Entry(
            self.status_frame,
            font=BF,
            theme="light_green",
            use_status_fg=True,
            outline=0,
        )
        self.selection_info.grid(row=0, column=1, sticky="nse", padx=(5, 5), pady=0)

        self.file.entryconfig("Compare sheets", command=self.compare_at_start)
        self.file.entryconfig("Open", command=self.open_file_at_start)
        self.file.entryconfig("New", command=self.create_new_at_start)
        self.file.entryconfig("Save", state="disabled")
        self.file.entryconfig("Save new version", state="disabled")
        self.menubar_state("disabled")
        self.frames["tree_edit"].bind_or_unbind_save()

        try:
            d = load_cfg()
            if isinstance(d, str):
                self.default_configsettings()
                self.save_cfg(get_settings=False)
            self.set_settings(d)
        except Exception:
            self.default_configsettings()
            self.save_cfg(get_settings=False)
            self.set_settings(d)

        if len(start_arg) > 1:
            try:
                self.open_dict["filepath"] = os.path.normpath(start_arg[1])
                self.created_new = False
                self.load_from_file()
            except Exception:
                self.open_dict = {}
                self.create_new_at_start()
        else:
            self.create_new_at_start()

        self.bind("<Configure>", self.frames["tree_edit"].WINDOW_DIMENSIONS_CHANGED)
        self.deiconify()
        if self.configsettings["First GUI start"]:
            First_Start_Popup(
                self,
                text=f"""
Welcome to Tk-Trees!

This program is for management of tree/hierarchy data which is in table format.

It is licensed under GPL-3.0.
{app_copyright}
Contact: {contact_email}
Visit: {website1}

To get started once you have closed this popup, either:
- Right click in the empty space in the table.
- Or go to the File menu and select a desired option.
- Or go to the Help menu for documentation.
                """,
                theme=self.theme,
            )
        self.configsettings["First GUI start"] = False
        self.save_cfg()

    def USER_HAS_CLOSED_WINDOW(self, callback=None):
        self.USER_HAS_QUIT = True
        if self.working:
            return
        try:
            self.configsettings["Window state"] = self.state()
            if self.configsettings["Window state"] == "normal":
                self.configsettings["Window size"] = (self.winfo_width(), self.winfo_height())
            else:
                self.configsettings["Window size"] = default_app_window_size
            self.configsettings["Window coords"] = self.geometry().split("+")[1:]
        except Exception:
            pass
        self.check_window_size_settings()
        self.save_cfg()
        if self.number_unsaved_changes:
            confirm = Ask_Confirm_Quit(self, changes=self.number_unsaved_changes, theme=self.theme)
            if confirm.option == "save":
                if self.file.entrycget("Save new version", "state") == "normal":
                    success = self.frames["tree_edit"].save_(quitting=True)
                else:
                    success = self.frames["tree_edit"].save_as(quitting=True)
                if not success:
                    self.USER_HAS_QUIT = False
                    return
            elif confirm.option == "cancel":
                self.USER_HAS_QUIT = False
                return
        self.try_to_close_everything()

    def try_to_close_everything(self):
        self.USER_HAS_QUIT = True
        self.try_to_close_workbook()
        with suppress(Exception):
            self.quit()
        with suppress(Exception):
            self.destroy()

    def try_to_close_workbook(self):
        with suppress(Exception):
            self.wb.close()
        with suppress(Exception):
            self.wb = None

    def default_configsettings(self, event=None):
        self.configsettings = {
            "Save xlsx with program data": True,
            "Save xlsx with viewable changelog": False,
            "Save xlsx with flattened sheet": False,
            "Save xlsx with treeview": False,
            "Flatten include detail columns": True,
            "Flatten justify left": True,
            "Flatten reverse order": False,
            "Flatten add index": False,
            "Json output format": 1,
            "Save json with program data": True,
            "First GUI start": True,
            "Theme": self.theme,
            "Window state": "zoomed",
            "Window size": default_app_window_size,
            "Window coords": center(self, *default_app_window_size, get=True),
            # left, 50/50, adjustable, right
            "Editor display option": "left",
            "Treeview indent": self.frames["tree_edit"].tree.ops.treeview_indent,
            "Treeview levels": self.frames["tree_edit"].tv_lvls_bool,
            "Auto select sheet id": self.frames["tree_edit"].mirror_var,
            "Alternate color": self.frames["tree_edit"].tree.ops.alternate_color,
            "Auto resize row indexes": self.frames["tree_edit"].auto_resize_indexes,
            "Allow cell text overflow": self.frames["tree_edit"].tree.ops.allow_cell_overflow,
        }
        self.check_window_size_settings()

    def check_window_size_settings(self):
        try:
            screenw = self.winfo_screenwidth()
            screenh = self.winfo_screenheight()
            if self.configsettings["Window size"][0] > screenw:
                self.configsettings["Window size"] = (screenw - 10, self.configsettings["Window size"][1])
            if self.configsettings["Window size"][1] > screenh:
                self.configsettings["Window size"] = (self.configsettings["Window size"][0], screenh - 10)
        except Exception:
            pass

    def get_configsettings(self):
        self.configsettings = {
            "Save xlsx with program data": self.frames["tree_edit"].save_xlsx_with_program_data,
            "Save xlsx with viewable changelog": self.frames["tree_edit"].save_xlsx_with_changelog,
            "Save xlsx with flattened sheet": self.frames["tree_edit"].save_xlsx_with_flattened,
            "Save xlsx with treeview": self.frames["tree_edit"].save_xlsx_with_treeview,
            "Flatten include detail columns": self.frames["tree_edit"].xlsx_flattened_detail_columns,
            "Flatten justify left": self.frames["tree_edit"].xlsx_flattened_justify,
            "Flatten reverse order": self.frames["tree_edit"].xlsx_flattened_reverse_order,
            "Flatten add index": self.frames["tree_edit"].xlsx_flattened_add_index,
            "Json output format": self.frames["tree_edit"].json_format,
            "Save json with program data": self.frames["tree_edit"].save_json_with_program_data,
            "First GUI start": self.configsettings["First GUI start"],
            "Theme": f"{self.theme}",
            "Window state": self.configsettings["Window state"],
            "Window size": self.configsettings["Window size"],
            "Window coords": self.configsettings["Window coords"],
            "Editor display option": self.frames["tree_edit"].get_display_option(),
            "Treeview indent": self.frames["tree_edit"].tree.ops.treeview_indent,
            "Treeview levels": self.frames["tree_edit"].tv_lvls_bool,
            "Auto select sheet id": self.frames["tree_edit"].mirror_var,
            "Alternate color": self.frames["tree_edit"].tree.ops.alternate_color,
            "Auto resize row indexes": self.frames["tree_edit"].auto_resize_indexes,
            "Allow cell text overflow": self.frames["tree_edit"].tree.ops.allow_cell_overflow,
        }

    def save_cfg(self, event=None, get_settings=True):
        if get_settings:
            self.get_configsettings()
        write_cfg(self.configsettings)

    def set_settings(self, d: None | dict = None):
        if isinstance(d, dict):
            self.configsettings = d
        self.frames["tree_edit"].save_xlsx_with_program_data = self.configsettings["Save xlsx with program data"]
        self.frames["tree_edit"].save_xlsx_with_changelog = self.configsettings["Save xlsx with viewable changelog"]
        self.frames["tree_edit"].save_xlsx_with_flattened = self.configsettings["Save xlsx with flattened sheet"]
        self.frames["tree_edit"].save_xlsx_with_treeview = self.configsettings["Save xlsx with treeview"]
        self.frames["tree_edit"].xlsx_flattened_detail_columns = self.configsettings["Flatten include detail columns"]
        self.frames["tree_edit"].xlsx_flattened_justify = self.configsettings["Flatten justify left"]
        self.frames["tree_edit"].xlsx_flattened_reverse_order = self.configsettings["Flatten reverse order"]
        self.frames["tree_edit"].xlsx_flattened_add_index = self.configsettings["Flatten add index"]
        self.frames["tree_edit"].json_format = int(self.configsettings["Json output format"])
        if "Treeview indent" in self.configsettings:
            self.frames["tree_edit"].tree.ops.treeview_indent = self.configsettings["Treeview indent"]
        if "Alternate color" in self.configsettings:
            self.frames["tree_edit"].tree.ops.alternate_color = self.configsettings["Alternate color"]
        if "Treeview levels" in self.configsettings:
            self.frames["tree_edit"].tv_lvls_bool = self.configsettings["Treeview levels"]
        if "Auto select sheet id" in self.configsettings:
            self.frames["tree_edit"].mirror_var = self.configsettings["Auto select sheet id"]
        if "Auto resize row indexes" in self.configsettings:
            self.frames["tree_edit"].toggle_auto_resize_index(self.configsettings["Auto resize row indexes"])
        if "Allow cell text overflow" in self.configsettings:
            self.frames["tree_edit"].tree.ops.allow_cell_overflow = self.configsettings["Allow cell text overflow"]
            self.frames["tree_edit"].sheet.ops.allow_cell_overflow = self.configsettings["Allow cell text overflow"]
        self.theme = self.configsettings["Theme"]
        self.frames["tree_edit"].set_display_option(self.configsettings["Editor display option"])
        self.frames["tree_edit"].change_theme(self.theme, write=False)
        center(
            toplevel=self,
            desired_width=self.configsettings["Window size"][0],
            desired_height=self.configsettings["Window size"][1],
            x=int(self.configsettings["Window coords"][0]),
            y=int(self.configsettings["Window coords"][1]),
        )
        if self.configsettings["Window state"] not in ("zoomed", "normal"):
            self.configsettings["Window state"] = "zoomed"
        with suppress(Exception):
            self.state(self.configsettings["Window state"])

    def menubar_state(self, state="normal", start=False):
        if state == "disabled":
            for label in ("File", "Edit", "View", "Import", "Export", "Help"):
                self.menubar.entryconfig(label, state="disabled")
            self.unbind(f"<{ctrl_button}-O>")
            self.unbind(f"<{ctrl_button}-o>")
            self.unbind(f"<{ctrl_button}-N>")
            self.unbind(f"<{ctrl_button}-n>")
        else:
            self.bind(f"<{ctrl_button}-O>", self.open_file_at_start)
            self.bind(f"<{ctrl_button}-o>", self.open_file_at_start)
            self.bind(f"<{ctrl_button}-N>", self.create_new_at_start)
            self.bind(f"<{ctrl_button}-n>", self.create_new_at_start)
            self.menubar.entryconfig("File", state="normal")
            x = "disabled" if start else "normal"
            for label in ("Edit", "View", "Import", "Export"):
                self.menubar.entryconfig(label, state=x)
            self.menubar.entryconfig("Help", state="normal")

    def change_app_title(self, title=None, star=None):
        if title:
            self.title("".join((app_title, " - ", title)))
        elif star == "add":
            self.title("".join((app_title, " - ", os.path.basename(self.open_dict["filepath"]), "*")))
        elif star == "remove":
            self.title("".join((app_title, " - ", os.path.basename(self.open_dict["filepath"]))))
        else:
            self.title(app_title)
        self.update_idletasks()

    def load_from_file(self):
        self.status_bar.change_text("Loading...")
        self.frames["tree_edit"].sheet.MT.data = []
        self.change_app_title(title=os.path.basename(self.open_dict["filepath"]))
        if self.open_dict["filepath"].lower().endswith((".csv", ".tsv")):
            try:
                with open(self.open_dict["filepath"], "r") as fh:
                    temp_data = fh.read()
                self.frames["tree_edit"].sheet.MT.data = csv_str_x_data(temp_data)
            except Exception as error_msg:
                Error(self, f"Error: {error_msg}", theme=self.theme)
                self.create_new_at_start()
                return
            if not self.frames["tree_edit"].sheet.MT.data:
                Error(self, "File contains no data   ", theme=self.theme)
                self.create_new_at_start()
                return
            self.open_dict["sheet"] = "Sheet1"
            self.frames["column_selection"].populate(
                list(map(str, range(1, max(map(len, self.frames["tree_edit"].sheet.MT.data), default=0) + 1))),
                clear_dd=True,
            )

        elif self.open_dict["filepath"].lower().endswith(".json"):
            try:
                j = get_json_from_file(self.open_dict["filepath"])
            except Exception as error_msg:
                Error(self, f"Error: {error_msg}", theme=self.theme)
                self.create_new_at_start()
                return
            if "program_data" in j:
                try:
                    program_data = b32_x_dict(j["program_data"])
                    self.frames["tree_edit"].sheet.MT.data = program_data["records"]
                    self.open_dict["sheet"] = "Sheet1"
                except Exception as error_msg:
                    Error(self, f"Error: {error_msg}", theme=self.theme)
                    self.create_new_at_start()
                    return
                try:
                    self.frames["tree_edit"].populate(program_data=program_data)
                    self.frames["tree_edit"].show_warnings(self.open_dict["filepath"], self.open_dict["sheet"])
                except Exception as error_msg:
                    Error(self, f"Error opening program data: {error_msg}", theme=self.theme)
                    self.frames["tree_edit"].reset_tree()
                    self.json_go_to_column_selection(j)
                    return
            else:
                self.json_go_to_column_selection(j)

        elif self.open_dict["filepath"].lower().endswith((".xlsx", ".xlsm", ".xls")):
            try:
                in_mem = bytes_io_wb(self.open_dict["filepath"])
                self.wb = load_workbook(in_mem, read_only=True, data_only=True)
            except Exception as error_msg:
                Error(self, f"Error: {error_msg}", theme=self.theme)
                self.create_new_at_start()
                return
            if len(self.wb.sheetnames) < 1:
                Error(self, "File contains no data   ", theme=self.theme)
                self.create_new_at_start()
                return
            sheetnames = set(self.wb.sheetnames)
            if "program_data" in sheetnames:
                self.status_bar.change_text("Loading...")
                ws = self.wb["program_data"]
                ws.reset_dimensions()
                try:
                    d = b32_x_dict(ws_x_program_data_str(ws))
                    self.frames["tree_edit"].populate(program_data=d)
                    self.open_dict["sheet"] = d["sheetname"]
                    self.wb.close()
                    self.frames["tree_edit"].show_warnings(self.open_dict["filepath"], self.open_dict["sheet"])

                except Exception as error_msg:
                    self.wb.close()
                    self.frames["tree_edit"].sheet.MT.data = []
                    self.wb = load_workbook(in_mem, read_only=True, data_only=True)
                    self.frames["column_selection"].sheet_selector.updatesheets(self.wb.sheetnames)
                    self.frames["column_selection"].sheet_selector.cont()
                    self.show_frame("column_selection")
                    Error(self, f"Error opening program data: {error_msg}   ", theme=self.theme)

            else:
                self.frames["column_selection"].sheet_selector.updatesheets(self.wb.sheetnames)
                self.frames["column_selection"].sheet_selector.cont()
                self.show_frame("column_selection")
        else:
            Error(
                self,
                "Error: File must be one of these types - .xlsx, .xlsm, .xls, .csv, .tsv, .json",
                theme=self.theme,
            )
            self.create_new_at_start()

    def json_go_to_column_selection(self, d: dict) -> bool:
        try:
            json_format = get_json_format(d)
            if not json_format:
                Error(
                    self,
                    "Error opening file, could not find data of correct format   ",
                    theme=self.theme,
                )
                self.create_new_at_start()
                return False
            self.frames["tree_edit"].sheet.MT.data, self.frames["tree_edit"].row_len = json_to_sheet(
                d,
                format_=json_format[0],
                key=json_format[1],
                get_format=False,
                return_rowlen=True,
            )
            self.open_dict["sheet"] = "Sheet1"
            self.frames["column_selection"].populate(
                list(map(str, range(1, max(map(len, self.frames["tree_edit"].sheet.MT.data), default=0) + 1))),
                clear_dd=True,
            )
            return True
        except Exception as error_msg:
            Error(self, f"Error: {error_msg}   ", theme=self.theme)
            self.create_new_at_start()
            return False

    def wb_sheet_has_been_selected(self, selection):
        self.status_bar.change_text("Loading...")
        ws = self.wb[selection]
        ws.reset_dimensions()
        self.frames["tree_edit"].sheet.MT.data = ws_x_data(ws)
        if not self.frames["tree_edit"].sheet.MT.data:
            Error(self, "Sheet contains no data   ", theme=self.theme)
            self.frames["column_selection"].sheet_selector.updatesheets(self.wb.sheetnames)
            self.frames["column_selection"].sheet_selector.cont()
            self.show_frame("column_selection")
            return
        self.frames["tree_edit"].row_len = max(map(len, self.frames["tree_edit"].sheet.MT.data), default=0)
        self.open_dict["sheet"] = selection
        self.frames["column_selection"].populate(
            list(map(str, range(1, self.frames["tree_edit"].row_len + 1))),
            clear_dd=False,
        )

    def help_func(self):
        Help_Popup(self, self.DOCUMENTATION, theme=self.theme)

    def license_func(self):
        License_Popup(self, f"Tk-Trees {app_copyright}\n\n" + self.LICENSE, show_buttons=False, theme=self.theme)

    def about_func(self):
        Text_Popup(
            self,
            about_system,
            width_=600,
            height_=500,
            theme=self.theme,
            use_entry_bg=False,
            wrap="word",
            show_finder=False,
            heading=app_copyright,
        )

    def reset_data_change_app_title(self):
        self.frames["tree_edit"].sheet.MT.data = []
        self.change_app_title(title=None)
        if self.current_frame == "treecompare":
            self.frames["tree_compare"].reset()

    def compare_at_start(self):
        self.reset_data_change_app_title()
        self.frames["tree_compare"].populate()

    def create_new_at_start(self, event=None):
        if event is not None and self.current_frame == "tree_edit":
            self.frames["tree_edit"].create_new_from_within_treeframe()
            return
        self.reset_data_change_app_title()
        self.frames["tree_edit"].reset_tree(False)
        self.created_new = True
        self.open_dict["filepath"] = "New sheet"
        self.open_dict["sheet"] = "Sheet1"
        self.frames["tree_edit"].show_warnings("n/a - CREATED NEW", "n/a")
        self.frames["tree_edit"].headers = [Header("ID", "ID"), Header("DETAIL_1"), Header("PARENT_1", "Parent")]
        self.frames["tree_edit"].ic = 0
        self.frames["tree_edit"].pc = 2
        self.frames["tree_edit"].hiers = [2]
        self.frames["tree_edit"].row_len = 3
        self.change_app_title(title="New sheet")
        self.frames["tree_edit"].populate()
        self.bind(f"<{ctrl_button}-O>", self.open_file_at_start)
        self.bind(f"<{ctrl_button}-o>", self.open_file_at_start)
        self.bind(f"<{ctrl_button}-N>", self.create_new_at_start)
        self.bind(f"<{ctrl_button}-n>", self.create_new_at_start)

    def show_frame(self, name, start=True, msg=None):
        if self.current_frame != name:
            self.frames[self.current_frame].grid_forget()
            self.frames[name].grid(row=0, column=0, sticky="nsew")
            self.frames[name].tkraise()
            self.current_frame = name
        self.frames[name].enable_widgets()
        self.frames[name].focus_set()
        self.menubar_state("normal", start=start)
        self.status_bar.change_text("Program ready" if msg is None else msg)

    def disable_at_start(self):
        self.menubar_state("disabled")
        self.frames[self.current_frame].disable_widgets()

    def enable_at_start(self):
        self.menubar_state("normal", start=True)
        self.frames[self.current_frame].enable_widgets()

    def open_file_at_start(self, event=None):
        if event is not None and self.current_frame == "tree_edit":
            self.frames["tree_edit"].open_from_within_treeframe()
            return
        fp = filedialog.askopenfilename(parent=self, title="Select a file")
        if not fp:
            return
        self.disable_at_start()
        try:
            fp = os.path.normpath(fp)
        except Exception:
            Error(self, "Filepath invalid   ", theme=self.theme)
            self.enable_at_start()
            return
        if not fp.lower().endswith((".json", ".xlsx", ".xls", ".xlsm", ".csv", ".tsv")):
            Error(self, "Please select excel/csv/json   ", theme=self.theme)
            self.enable_at_start()
            return
        check = os.path.isfile(fp)
        if check:
            self.open_dict["filepath"] = fp
            self.reset_data_change_app_title()
            self.load_from_file()
        else:
            Error(self, "Filepath invalid   ", theme=self.theme)
            self.enable_at_start()


def run_app(startup_args):
    if len(startup_args) > 4 and allow_api_use:
        try:
            kwargs = DotDict()
            for ctr, arg in enumerate(startup_args):
                # 1, 2, 3, 4 are required
                # -id-<int> and -parent-<int> required for flatten operations
                if ctr == 1:
                    api_action = arg

                elif ctr == 2:
                    input_filepath = os.path.normpath(arg)

                elif ctr == 3:
                    output_filepath = os.path.normpath(arg)

                elif arg.startswith("-all-parent-columns-"):
                    all_parent_column_indexes = sorted(
                        int(c) if c.isdigit() else alpha2idx(c) for c in arg.split("-all-parent-columns-")[1].split(",")
                    )
                    for i in all_parent_column_indexes:
                        if i < 0:
                            raise ValueError(
                                f"Parent column index must be number of letter representing column, not '{i}'"
                            )

                # defaults to first sheet
                elif arg.startswith("-input-sheet-"):
                    kwargs["input_sheet"] = arg.split("-input-sheet-")[1]

                # defaults to input-sheet name
                elif arg.startswith("-output-sheet-"):
                    kwargs["output_sheet"] = arg.split("-output-sheet-")[1]

                # defaults to comma
                elif arg.startswith("-delim-"):
                    kwargs["csv_delimiter"] = arg.split("-delim-")[1]

                # -id- and -parent- required for flatten, not for unflatten
                elif arg.startswith("-id-"):
                    _arg = arg.split("-id-")[1]
                    if _arg.isdigit():
                        kwargs["flatten_id_column"] = int(_arg)
                    else:
                        kwargs["flatten_id_column"] = alpha2idx(_arg)

                elif arg.startswith("-parent-"):
                    _arg = arg.split("-parent-")[1]
                    if _arg.isdigit():
                        kwargs["flatten_parent_column"] = int(_arg)
                    else:
                        kwargs["flatten_parent_column"] = alpha2idx(_arg)

                # optional flags, e.g. -odjr
                elif arg.startswith("-"):
                    # flags
                    # o overwrite
                    # d detail_columns
                    # j justify_left
                    # r reverse
                    # i add index
                    flags = arg.split("-")[1]
                    for c in flags:
                        if c == "o":
                            kwargs["overwrite_file"] = True
                        elif c == "d":
                            kwargs["detail_columns"] = True
                        elif c == "j":
                            kwargs["justify_left"] = True
                        elif c == "r":
                            kwargs["reverse"] = True
                        elif c == "i":
                            kwargs["add_index"] = True
                        else:
                            break
        except Exception as error_msg:
            try_write_error_log(f"{error_msg}")

        tk_trees_api(
            api_action=api_action,
            input_filepath=input_filepath,
            output_filepath=output_filepath,
            all_parent_column_indexes=all_parent_column_indexes,
            **kwargs,
        )
    else:
        app = AppGUI(startup_args)
        app.mainloop()


if __name__ == "__main__":
    run_app(argv)
