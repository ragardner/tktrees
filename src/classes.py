# SPDX-License-Identifier: AGPL-3.0-only
# Copyright (c) 2025 R. A. Gardner

from __future__ import annotations

import contextlib
import csv
import io
from collections import defaultdict, deque
from collections.abc import Generator
from itertools import chain, islice, repeat
from operator import itemgetter
from typing import Literal

from openpyxl import load_workbook

from .functions import (
    csv_str_x_data,
    equalize_sublist_lens,
    get_json_format,
    get_json_from_file,
    json_to_sheet,
    shift_elements_to_end,
    shift_elements_to_start,
    to_csv,
    to_json,
    to_xlsx,
    try_write_error_log,
    ws_x_data,
)


class TreeBuilder:
    def check_cn(self, iid: str, h: int, nodes: dict[str, Node]) -> Generator[str]:
        stack = [iid]
        while stack:
            current = stack.pop()
            yield current
            stack.extend(reversed(nodes[current].cn[h]))

    def check_ps(self, iid: str, h: int, nodes: dict[str, Node]) -> Generator[str]:
        current = iid
        while True:
            yield current
            if not nodes[current].ps[h]:
                break
            current = nodes[current].ps[h]

    def build(
        self,
        input_sheet: list[list[str]],
        output_sheet: list[list[str]],
        row_len: int,
        ic: int,
        hiers: list[int],
        nodes: dict[str, Node],
        warnings: list[object] | None = None,
        rns: dict[str, int] | None = None,
        add_warnings: bool = True,
        skip_1st: bool = False,
        compare: bool = False,
        fix_associate: bool = False,
        strip: bool = True,
    ) -> (
        tuple[list[list[str]], dict[str, Node], list[object], dict[str, int]]
        | tuple[list[list[str]], dict[str, Node], list[object]]
        | tuple[list[list[str]], dict[str, Node]]
    ):
        if warnings is None:
            warnings = []
        if rns is None:
            rns = {}
        tally_of_ids = defaultdict(lambda: -1)
        qhsic = sorted(hiers.copy() + [ic])
        qhs = hiers
        for i, r in enumerate(islice(input_sheet, 0 if not skip_1st else 1, len(input_sheet))):
            rn = f"{i + 2}"
            if len(r) < row_len:
                r += list(repeat("", row_len - len(r)))
            if r[ic]:
                for c in qhsic:
                    e = r[c]
                    if strip:
                        if add_warnings:
                            if " " in e:
                                warnings.append(f" - Spaces in row #{rn} column #{c + 1}")
                            if "\n" in e:
                                warnings.append(f" - Newlines in row #{rn} column #{c + 1}")
                            if "\r" in e:
                                warnings.append(f" - Carriage returns in row #{rn} column #{c + 1}")
                            if "\t" in e:
                                warnings.append(f" - Tabs returns in row #{rn} column #{c + 1}")
                        r[c] = "".join(e.strip().split())
                ID = r[ic]
                ik = ID.lower()
                tally_of_ids[ik] += 1
                if tally_of_ids[ik] > 0:
                    if add_warnings:
                        warnings.append(f" - ID ({ID}) renamed due to repeat occurrence at row #{rn}")
                    orig = ID
                    x = 1
                    while ik in tally_of_ids:
                        ID = f"{orig}_DUPLICATED_{x}"
                        ik = ID.lower()
                        x += 1
                    tally_of_ids[ik] += 1
                    r[ic] = ID
                if ik not in nodes:
                    nodes[ik] = Node(ID, hiers)
                for h in qhs:
                    parent = r[h]
                    pk = parent.lower()
                    if ik == pk:
                        if add_warnings:
                            warnings.append(
                                f" - ID ({ID}) same as parent ({parent}). Set parent ({parent}) to none at row #{rn}"
                            )
                        r[h] = ""
                        parent = ""
                        pk = ""
                    elif pk:
                        for ck in chain(self.check_cn(ik, h, nodes), self.check_ps(ik, h, nodes)):
                            if pk == ck:
                                if add_warnings:
                                    warnings.append(
                                        (
                                            f" - Infinite loop of children avoided by setting "
                                            f"IDs ({ID}) parent ({parent}) to none at row #{rn}"
                                        )
                                    )
                                r[h] = ""
                                parent = ""
                                pk = ""
                                break
                    if pk:
                        if pk not in nodes:
                            nodes[pk] = Node(parent, hiers)
                        nodes[ik].ps[h] = pk
                        nodes[pk].cn[h].append(ik)
                    else:
                        nodes[ik].ps[h] = ""
                output_sheet.append(r)
            else:
                if add_warnings:
                    warnings.append(f" - Empty ID cell, row #{rn} excluded from sheet")
                continue
        if fix_associate:
            quick_hiers = hiers[1:]
            lh = len(hiers)
            rns = {r[ic].lower(): i for i, r in enumerate(output_sheet)}
            for iid, node in nodes.items():
                if all(p is None for p in node.ps.values()):
                    node.ps = {h: "" if node.cn[h] else None for h in hiers}
                    newrow = ["" for _ in range(row_len)]
                    newrow[ic] = node.name
                    output_sheet.append(newrow)
                    rns[iid] = len(output_sheet) - 1
                    if compare:
                        warnings.append(f" - ID ({node.name}) missing from ID column, new row added")
                tlly = 0
                for k, v in node.cn.items():
                    if not v and not node.ps[k]:
                        node.ps[k] = None
                        tlly += 1
                if tlly == lh:
                    node.ps[hiers[0]] = ""
                    for h in quick_hiers:
                        node.ps[h] = None

            return output_sheet, nodes, warnings, rns
        elif not fix_associate:
            if add_warnings:
                return output_sheet, nodes, warnings
            return output_sheet, nodes

    def gen_pc_base_ids(
        self,
        sheet: list[list[str]],
        nodes: dict[str, Node],
        ic: int,
        pc: int,
    ) -> Generator[str]:
        for row in sheet:
            if row[ic]:
                iid = row[ic].lower()
                node = nodes[iid]
                if node.ps[pc] is not None and not node.cn[pc]:
                    yield iid

    def build_flattened(
        self,
        input_sheet: list[list[str]],
        output_sheet: list[list[str]],
        nodes: dict[str, Node],
        headers: list[str],
        ic: int,
        pc: int,
        hiers: list[int],
        detail_columns: bool,
        justify_left: bool,
        reverse: bool,
        add_index: bool,
        empty_cells_to_none: bool = False,
    ) -> list[list[str]]:
        output_headers = []
        detail_columns = detail_columns and len(hiers) + 1 < len(headers)
        ic_plus_hiers = {ic} | set(hiers)
        detail_cols_idxs_names = {i: headers[i] for i in [i for i in range(len(headers)) if i not in ic_plus_hiers]}
        pc_name = headers[pc]
        self.n_lvls = 1
        rns = {r[ic].lower(): rn for rn, r in enumerate(input_sheet) if r[ic]}
        for iid in self.gen_pc_base_ids(sheet=input_sheet, nodes=nodes, ic=ic, pc=pc):
            node = nodes[iid]
            if justify_left and not reverse:
                row = deque()
                if detail_columns:
                    row = deque(input_sheet[rns[iid]][i] for i in detail_cols_idxs_names) + row
                row.appendleft(node.name)
            elif (justify_left and reverse) or (not justify_left and not reverse):
                row = [node.name]
                if detail_columns:
                    row.extend(input_sheet[rns[iid]][i] for i in detail_cols_idxs_names)
            elif not justify_left and reverse:
                row = []
                if detail_columns:
                    row.extend(input_sheet[rns[iid]][i] for i in detail_cols_idxs_names)
                row.append(node.name)
            if node.ps[pc]:
                iid = node.ps[pc]
                node = nodes[iid]
                lvl = 2
                while iid:
                    if justify_left and not reverse:
                        if detail_columns:
                            row.extendleft(input_sheet[rns[iid]][i] for i in reversed(detail_cols_idxs_names))
                        row.appendleft(node.name)
                    elif (justify_left and reverse) or (not justify_left and not reverse):
                        row.append(node.name)
                        if detail_columns:
                            row.extend(input_sheet[rns[iid]][i] for i in detail_cols_idxs_names)
                    elif not justify_left and reverse:
                        if detail_columns:
                            row.extend(input_sheet[rns[iid]][i] for i in detail_cols_idxs_names)
                        row.append(node.name)

                    if node.ps[pc]:
                        lvl += 1
                        if lvl > self.n_lvls:
                            self.n_lvls = lvl
                        iid = node.ps[pc]
                        node = nodes[iid]
                    else:
                        break

            if justify_left and not reverse:
                output_sheet.append(list(row))
            else:
                output_sheet.append(row)

        equalize_sublist_lens(output_sheet)

        if justify_left and not reverse:
            output_sheet = list(map(shift_elements_to_start, output_sheet))
            for i in range(self.n_lvls):
                output_headers.append(f"{pc_name}_{i}")
                if detail_columns:
                    output_headers.extend(f"{detail_name}_{i}" for detail_name in detail_cols_idxs_names.values())

        elif justify_left and reverse:
            for i in reversed(range(self.n_lvls)):
                output_headers.append(f"{pc_name}_{i}")
                if detail_columns:
                    output_headers.extend(f"{detail_name}_{i}" for detail_name in detail_cols_idxs_names.values())

        elif not justify_left and not reverse:
            output_sheet = [r[::-1] for r in output_sheet]
            for i in reversed(range(self.n_lvls)):
                output_headers.append(f"{pc_name}_{i}")
                if detail_columns:
                    output_headers.extend(f"{detail_name}_{i}" for detail_name in detail_cols_idxs_names.values())
            output_headers = output_headers[::-1]

        elif not justify_left and reverse:
            output_sheet = list(map(shift_elements_to_end, output_sheet))
            for i in range(self.n_lvls):
                output_headers.append(f"{pc_name}_{i}")
                if detail_columns:
                    output_headers.extend(f"{detail_name}_{i}" for detail_name in detail_cols_idxs_names.values())
            output_headers = output_headers[::-1]

        if empty_cells_to_none:
            for rn in range(len(output_sheet)):
                for cn in range(len(output_sheet[rn])):
                    if not output_sheet[rn][cn]:
                        output_sheet[rn][cn] = None

        if add_index:
            return [["Index"] + output_headers] + [[f"{rn}"] + r for rn, r in enumerate(output_sheet)]
        return [output_headers] + output_sheet

    def convert_flattened_to_normal(
        self,
        data: list[list[str]] | None = None,
        hier_cols: list[int] | None = None,
        rowlen: None | int = None,
        order: int = 1,  # 1. Top - Base, 2. Base - Top
        warnings: list[object] | None = None,
    ) -> tuple[list[list[str]], int, int, list[int]]:
        if data is None:
            data = []
        if hier_cols is None:
            hier_cols = []
        if warnings is None:
            warnings = []
        rowlen = max(map(len, data), default=0) if rowlen is None else rowlen
        added_ids, to_add, ids_parents_tally, rns = set(), {}, {}, {}
        detail_cols = sorted(set(range(rowlen)).difference(hier_cols))
        justify_left = (
            not hier_cols[0]
            or (hier_cols[0] and hier_cols[-1] < rowlen - 1)
            or (not hier_cols[0] and hier_cols[-1] == rowlen - 1)
        )
        hier_cols_detail_cols = {}
        if detail_cols and justify_left:
            hier_cols_detail_cols = {
                hier_col: (
                    list(range(hier_col + 1, hier_cols[i + 1]))
                    if i < len(hier_cols) - 1
                    else list(range(hier_col + 1, rowlen))
                )
                for i, hier_col in enumerate(hier_cols)
            }
        elif detail_cols and not justify_left:
            hier_cols_detail_cols = {
                hier_col: list(range(hier_cols[i - 1] + 1, hier_col)) if i else list(range(hier_col))
                for i, hier_col in enumerate(hier_cols)
            }
        if detail_cols:
            hier_col_with_most_detail_cols = max(
                hier_cols_detail_cols.items(), key=lambda kv: len(kv[1]), default=(hier_cols[0], [])
            )[0]
            num_detail_cols = sum(map(len, hier_cols_detail_cols.values()))
            detail_col_names = [
                data[0][detail_col] for detail_col in hier_cols_detail_cols[hier_col_with_most_detail_cols]
            ]
            num_detail_cols_to_be_added = len(detail_col_names)
            not_detail_cols_or_hier_cols = sorted(
                set(range(rowlen))
                - ({idx for detail_cols in hier_cols_detail_cols.values() for idx in detail_cols} | set(hier_cols))
            )
            ids_details_tally = {}

            if order == 2:
                for rn, r in enumerate(islice(data, 1, None), 1):
                    for idx in hier_cols:
                        if r[idx]:
                            ik = r[idx].lower()
                            if ik not in ids_details_tally:
                                ids_details_tally[ik] = {}
                            for det_col_enum, det_col in enumerate(hier_cols_detail_cols[idx]):
                                if det_col_enum not in ids_details_tally[ik]:
                                    ids_details_tally[ik][det_col_enum] = defaultdict(int)
                                ids_details_tally[ik][det_col_enum][r[det_col]] += 1
                            if ik not in rns:
                                rns[ik] = rn

                    for idx, (idcol, pcol) in enumerate(zip(hier_cols, islice(hier_cols, 1, None))):
                        ID = r[idcol]
                        ik = ID.lower()
                        par = r[pcol]
                        pk = par.lower()
                        if ik:
                            if not par:
                                try:
                                    par = next(r[i] for i in islice(hier_cols, idx + 1, None) if r[i])
                                    pk = par.lower()
                                    warnings.append(f" - Missing ID in hierarchy column {data[0][pcol]} row #{rn + 1}")
                                except Exception:
                                    pass
                            if ik not in ids_parents_tally:
                                ids_parents_tally[ik] = defaultdict(int)
                            ids_parents_tally[ik][pk] += 1
                            if ik not in added_ids:
                                added_ids.add(ik)
                                to_add[ik] = (ID, par)
                        if pcol == hier_cols[-1] and pk not in added_ids and par:
                            added_ids.add(pk)
                            to_add[pk] = (par, "")

            elif order == 1:
                for rn, r in enumerate(islice(data, 1, None), 1):
                    for idx in reversed(hier_cols):
                        if r[idx]:
                            ik = r[idx].lower()
                            if ik not in ids_details_tally:
                                ids_details_tally[ik] = {}
                            for det_col_enum, det_col in enumerate(hier_cols_detail_cols[idx]):
                                if det_col_enum not in ids_details_tally[ik]:
                                    ids_details_tally[ik][det_col_enum] = defaultdict(int)
                                ids_details_tally[ik][det_col_enum][r[det_col]] += 1
                            if ik not in rns:
                                rns[ik] = rn

                    for idx, (idcol, pcol) in enumerate(zip(reversed(hier_cols), islice(reversed(hier_cols), 1, None))):
                        ID = r[idcol]
                        ik = ID.lower()
                        par = r[pcol]
                        pk = par.lower()
                        if ik:
                            if not par:
                                try:
                                    par = next(r[i] for i in islice(reversed(hier_cols), idx + 1, None) if r[i])
                                    pk = par.lower()
                                    warnings.append(f" - Missing ID in hierarchy column {data[0][pcol]} row #{rn + 1}")
                                except Exception:
                                    pass
                            if ik not in ids_parents_tally:
                                ids_parents_tally[ik] = defaultdict(int)
                            ids_parents_tally[ik][pk] += 1
                            if ik not in added_ids:
                                added_ids.add(ik)
                                to_add[ik] = (ID, par)
                        if pcol == hier_cols[0] and par.lower() not in added_ids and par:
                            added_ids.add(par.lower())
                            to_add[pk] = (par, "")

            for ik, dct in ids_details_tally.items():
                for det_col_enum, detail_dct in dct.items():
                    if len(detail_dct) > 1:
                        tallies = "\n\t".join(f"{det}: {tally}" for det, tally in detail_dct.items())
                        warnings.append(
                            f" - {to_add[ik][0]} has multiple details in column '{detail_col_names[det_col_enum]}', "
                            f"using detail with highest tally '{max(detail_dct.items(), key=itemgetter(1))[0]}':\n\t"
                            f"{tallies}"
                        )
            for ik, dct in ids_parents_tally.items():
                if len(dct) > 1:
                    tallies = "\n\t".join(f"{to_add[pk][0]}: {num}" for pk, num in dct.items() if pk)
                    chosen_pk = max(dct.items(), key=itemgetter(1))[0]
                    chosen_par = to_add[chosen_pk][0] if chosen_pk else ""
                    to_add[ik] = (to_add[ik][0], chosen_par)
                    warnings.append(
                        f" - {to_add[ik][0]} has multiple parents, "
                        f"using parent with highest tally '{chosen_par}':\n\t"
                        f"{tallies}"
                    )
            output = []
            added_ids = set()
            other_cols_len = rowlen - num_detail_cols - len(hier_cols)
            output.append(["ID", "PARENT"] + detail_col_names + [data[0][i] for i in not_detail_cols_or_hier_cols])
            for ID, par in to_add.values():
                if (ik := ID.lower()) in ids_details_tally:
                    other_cols = (
                        [data[rns[ik]][i] for i in not_detail_cols_or_hier_cols]
                        if ik in rns
                        else list(repeat("", other_cols_len))
                    )
                    if ids_details_tally[ik]:
                        details = [
                            max(detail_dct.items(), key=itemgetter(1))[0]
                            for detail_dct in ids_details_tally[ik].values()
                        ]
                        if len(details) < num_detail_cols_to_be_added:
                            details += list(repeat("", num_detail_cols_to_be_added - len(details)))
                        output.append([ID, par] + details + other_cols)
                    else:
                        output.append([ID, par] + list(repeat("", num_detail_cols)) + other_cols)
                    added_ids.add(ik)
            for ID, par in to_add.values():
                if (ik := ID.lower()) not in added_ids:
                    other_cols = (
                        [data[rns[ik]][i] for i in not_detail_cols_or_hier_cols]
                        if ik in rns
                        else list(repeat("", other_cols_len))
                    )
                    output.append([ID, par] + list(repeat("", num_detail_cols)) + other_cols)
            return output, max(map(len, output), default=0), 0, [1]

        elif not detail_cols:
            if order == 2:
                for rn, r in enumerate(islice(data, 1, None), 1):
                    for idx in hier_cols:
                        if r[idx] and r[idx].lower() not in rns:
                            rns[r[idx].lower()] = rn
                    for idx, (idcol, pcol) in enumerate(zip(hier_cols, islice(hier_cols, 1, None))):
                        ID = r[idcol]
                        ik = ID.lower()
                        par = r[pcol]
                        pk = par.lower()
                        if ik:
                            if not par:
                                try:
                                    par = next(r[i] for i in islice(hier_cols, idx + 1, None) if r[i])
                                    pk = par.lower()
                                    warnings.append(f" - Missing ID in hierarchy column {data[0][pcol]} row #{rn + 1}")
                                except Exception:
                                    pass
                            if ik not in ids_parents_tally:
                                ids_parents_tally[ik] = defaultdict(int)
                            ids_parents_tally[ik][pk] += 1
                            if ik not in added_ids:
                                added_ids.add(ik)
                                to_add[ik] = (ID, par)
                        if pcol == hier_cols[-1] and par.lower() not in added_ids and par:
                            added_ids.add(par.lower())
                            to_add[pk] = (par, "")

            elif order == 1:
                for rn, r in enumerate(islice(data, 1, None), 1):
                    for idx in reversed(hier_cols):
                        if r[idx] and r[idx].lower() not in rns:
                            rns[r[idx].lower()] = rn
                    for idx, (idcol, pcol) in enumerate(zip(reversed(hier_cols), islice(reversed(hier_cols), 1, None))):
                        ID = r[idcol]
                        ik = ID.lower()
                        par = r[pcol]
                        pk = par.lower()
                        if ik:
                            if not par:
                                try:
                                    par = next(r[i] for i in islice(reversed(hier_cols), idx + 1, None) if r[i])
                                    pk = par.lower()
                                    warnings.append(f" - Missing ID in hierarchy column {data[0][pcol]} row #{rn + 1}")
                                except Exception:
                                    pass
                            if ik not in ids_parents_tally:
                                ids_parents_tally[ik] = defaultdict(int)
                            ids_parents_tally[ik][pk] += 1
                            if ik not in added_ids:
                                added_ids.add(ik)
                                to_add[ik] = (ID, par)
                        if pcol == hier_cols[0] and par.lower() not in added_ids and par:
                            added_ids.add(par.lower())
                            to_add[pk] = (par, "")

            for ik, dct in ids_parents_tally.items():
                if len(dct) > 1:
                    lp = "\n\t".join(f"{to_add[pk][0]}: {num}" for pk, num in dct.items() if pk)
                    chosen_pk = max(dct.items(), key=itemgetter(1))[0]
                    chosen_par = to_add[chosen_pk][0] if chosen_pk else ""
                    to_add[ik] = (to_add[ik][0], chosen_par)
                    warnings.append(
                        f" - {to_add[ik][0]} has multiple different parents, "
                        f"using parent with highest tally ({chosen_par}):\n\t"
                        f"{lp}"
                    )
            output = []
            added_ids = set()
            qindices = set(hier_cols)
            other_cols_len = rowlen - len(qindices)
            output.append(["ID", "PARENT"] + [e for i, e in enumerate(data[0]) if i not in qindices])
            for ID, par in to_add.values():
                if (ik := ID.lower()) in rns:
                    output.append([ID, par] + [e for i, e in enumerate(data[rns[ik]]) if i not in qindices])
                    added_ids.add(ik)
            for ID, par in to_add.values():
                if (ik := ID.lower()) not in added_ids:
                    output.append([ID, par] + list(repeat("", other_cols_len)))
            return output, max(map(len, output), default=0), 0, [1]

    def convert_indented_tree_detail_adjacent_to_normal(
        self,
        data: list[list[str]] | None = None,
    ) -> tuple[list[list[str]], int, int, list[int]]:
        output, parents = [], []
        for row in data:
            for level, value in enumerate(row):
                if value:
                    parents = parents[
                        : next(
                            (
                                len(parents) - i
                                for i, (parent_level, _) in enumerate(reversed(parents))
                                if parent_level < level
                            ),
                            0,
                        )
                    ]
                    parent = parents[-1][1] if parents else ""
                    with contextlib.suppress(Exception):
                        detail = next((e for e in row[level + 1 :] if e), "")
                    if detail:
                        output.append([value, parent, detail])
                    else:
                        output.append([value, parent])
                    parents.append((level, value))
                    break
        if not output:
            return [], 0, 0, [1]
        rowlen = equalize_sublist_lens(output)
        output = [["ID", "Parent"]] + output if rowlen <= 2 else [["ID", "Parent", "Detail"]] + output
        return output, rowlen, 0, [1]

    def convert_indented_tree_details_adjacent_to_normal(
        self,
        data: list[list[str]] | None = None,
    ) -> tuple[list[list[str]], int, int, list[int]]:
        output, parents = [], []
        for row in data:
            for level, value in enumerate(row):
                if value:
                    parents = parents[
                        : next(
                            (
                                len(parents) - i
                                for i, (parent_level, _) in enumerate(reversed(parents))
                                if parent_level < level
                            ),
                            0,
                        )
                    ]
                    parent = parents[-1][1] if parents else ""
                    details = row[level + 1 :]
                    try:
                        details = details[: len(details) - next(i for i, e in enumerate(reversed(details)) if e)]
                    except Exception:
                        details = []
                    if details:
                        output.append([value, parent] + details)
                    else:
                        output.append([value, parent])
                    parents.append((level, value))
                    break
        if not output:
            return [], 0, 0, [1]
        rowlen = equalize_sublist_lens(output)
        output = [["ID", "Parent"] + [f"Detail_{n}" for n in range(1, len(output[0]) - 1)]] + output
        return output, rowlen, 0, [1]


class SearchResult:
    __slots__ = ("hierarchy", "text", "iid", "column", "term", "type_", "exact")

    def __init__(
        self,
        hierarchy: int,
        text: str,
        iid: str,
        column: int,
        term: str,
        type_: int,  # 0 for id, 1 for detail, 2 for either
        exact: bool,
    ):
        self.hierarchy = hierarchy
        self.text = text
        self.iid = iid
        self.column = column
        self.term = term
        self.type_ = type_
        self.exact = exact


class Node:
    __slots__ = ("name", "cn", "ps")

    def __init__(
        self,
        name: str,
        hrs: list[int],
        cn: dict[int, list[str]] | None = None,
        ps: dict[int, None | str] | None = None,
    ) -> None:
        self.name: str = name
        self.cn = cn if cn else {v: [] for v in hrs}
        self.ps = ps if ps else dict.fromkeys(hrs)


class Header:
    __slots__ = (
        "name",
        "type_",
        "formatting",
        "validation",
    )

    def __init__(
        self,
        name: str,
        type_: Literal[
            "ID",
            "Parent",
            "Text",
            "Number",
            "Date",
        ] = "Text",
        formatting: None | list[object] = None,
        validation: None | list[str] = None,
    ):
        self.name = name
        # backwards compatibility
        if type_ == "Numerical Detail":
            type_ = "Number"
        elif type_.endswith(" Detail"):
            type_ = type_[:-7]

        self.type_ = type_
        if formatting is None:
            self.formatting = []
        else:
            self.formatting = [tuple(x) for x in formatting]
        if validation is None:
            self.validation = []
        else:
            self.validation = validation


# t = type, deleted (1) or changed (0)
class Del_stre:
    __slots__ = ("t", "rn", "row")

    def __init__(self, t, rn, r):
        self.t = t
        self.rn = rn
        self.row = r


def tk_trees_api(
    api_action: Literal["flatten", "unflatten-top-base", "unflatten-base-top"],
    input_filepath: str,
    output_filepath: str,
    all_parent_column_indexes: list[int],
    input_sheet: str | int = 0,
    output_sheet: str | None = None,
    csv_delimiter: str | Literal["tab"] = ",",
    justify_left: bool = True,
    reverse: bool = True,
    detail_columns: bool = True,
    add_index: bool = False,
    overwrite_file: bool = True,
    flatten_id_column: int = 0,
    flatten_parent_column: int = 1,
) -> None:
    try:
        dialect = csv.excel_tab if csv_delimiter == "tab" else csv.excel

        overwrite_file = "w" if overwrite_file else "x"

        sheet = []
        row_len = 0

        # ___________ LOAD FILE AND DATA ___________________

        if not input_filepath.lower().endswith((".xlsx", ".xls", ".xlsm", ".csv", ".tsv", ".json")):
            raise Exception("Input file must be .xlsx / .xls / .xlsm / .csv / .tsv")

        json_format = (1, "records")
        if input_filepath.lower().endswith((".csv", ".tsv")):
            with open(input_filepath, "r") as fh:
                temp_data = fh.read()
            sheet = csv_str_x_data(temp_data)

        elif input_filepath.lower().endswith((".xlsx", ".xls", ".xlsm")):
            with open(input_filepath, "rb") as fh:
                in_mem = io.BytesIO(fh.read())
            wb = load_workbook(in_mem, read_only=True, data_only=True)
            if isinstance(input_sheet, int):
                input_sheet = wb.sheetnames[input_sheet]
            ws = wb[input_sheet]
            ws.reset_dimensions()
            sheet = ws_x_data(ws)
            wb.close()

        elif input_filepath.lower().endswith(".json"):
            j = get_json_from_file(input_filepath)
            if not (json_format := get_json_format(j)):
                raise Exception("Invalid json file")
            sheet = json_to_sheet(
                j,
                format_=json_format[0],
                key=json_format[1],
                get_format=False,
            )

        row_len = max(map(len, sheet), default=0)

        if api_action == "flatten":
            headers_orig = sheet.pop(0)
            sheet, nodes, warnings, _ = TreeBuilder().build(
                input_sheet=sheet,
                output_sheet=[],
                row_len=row_len,
                ic=flatten_id_column,
                hiers=all_parent_column_indexes,
                nodes={},
                warnings=[],
                add_warnings=True,
                skip_1st=False,
                fix_associate=True,
            )
            data = TreeBuilder().build_flattened(
                input_sheet=sheet,
                output_sheet=[],
                nodes=nodes,
                headers=headers_orig,
                ic=flatten_id_column,
                pc=flatten_parent_column,
                hiers=all_parent_column_indexes,
                detail_columns=detail_columns,
                justify_left=justify_left,
                reverse=reverse,
                add_index=add_index,
            )

        elif api_action.startswith("unflatten"):
            order = 1 if api_action.endswith("base") else 2
            data = TreeBuilder().convert_flattened_to_normal(
                data=sheet,
                hier_cols=all_parent_column_indexes,
                rowlen=row_len,
                order=order,
            )[0]
        if output_filepath.endswith((".csv", ".tsv")):
            to_csv(
                filepath=output_filepath,
                overwrite=overwrite_file,
                dialect=dialect,
                data=data,
            )
        elif output_filepath.endswith(".xlsx"):
            if output_sheet is None:
                output_sheet = input_sheet if isinstance(input_sheet, str) else "Sheet1"
            to_xlsx(
                filepath=output_filepath,
                sheetname=output_sheet,
                data=data,
            )

        elif output_filepath.endswith(".json"):
            to_json(
                filepath=output_filepath,
                data=data,
                format_=json_format[0],
            )

    except Exception as error:
        try_write_error_log(f"{error}")
