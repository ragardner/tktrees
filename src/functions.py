# SPDX-License-Identifier: AGPL-3.0-only
# Copyright (c) 2025 R. A. Gardner

from __future__ import annotations

import csv
import io
import json
import lzma
import os
import re
import tkinter as tk
import zlib
from base64 import b32decode as b32d
from base64 import b32encode as b32e
from collections import defaultdict
from contextlib import suppress
from itertools import islice, repeat
from math import ceil
from typing import Literal

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from tksheet import (
    DotDict,
    get_csv_str_dialect,
)

from .constants import (
    config_name,
    current_dir,
    from_clipboard_delimiters,
    green_add_fill,
    isfloatre,
    isintlikere,
    isintre,
    isrealre,
    red_remove_fill,
    tv_lvls_colors,
    upone_dir,
)


def try_write_error_log(error: str) -> bool:
    try:
        with open(upone_dir + "TKTREES-ERROR.txt", "w") as fh:
            fh.write(f"{error}")
    except Exception:
        pass


def to_clipboard(widget: tk.Misc, s: str) -> None:
    widget.clipboard_clear()
    widget.clipboard_append(s)
    widget.update()


def to_csv(filepath: str, overwrite: Literal["w", "x"], dialect: csv.Dialect, data: list[list[str]]) -> None:
    with open(filepath, overwrite, newline="") as fh:
        writer = csv.writer(fh, dialect=dialect, lineterminator="\n")
        writer.writerows(data)


def to_xlsx(filepath: str, sheetname: str, data: list[list[str]]) -> None:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheetname)
    for row in data:
        ws.append(row)
    wb.active = wb[sheetname]
    wb.save(filepath)


def str_io_csv_writer(dialect: csv.Dialect):
    s = io.StringIO()
    return s, csv.writer(s, dialect=dialect, lineterminator="\n")


def str_x_bool(s: str) -> bool:
    if (s := f"{s}".lower()) not in ("false", "true"):
        raise Exception(f"Argument {s} must be either True or False")
    return s == "true"


def load_cfg():
    try:
        with open(current_dir + config_name, "r") as f:
            d = f.read()
        return json.loads(d)
    except Exception as error_msg:
        return f"{error_msg}"


def write_cfg(d: dict) -> bool:
    try:
        with open(current_dir + config_name, "w") as f:
            f.write(json.dumps(d, indent=4))
    except Exception:
        return False
    return True


def sort_key(s: str):
    return tuple(int(e) if e.isdigit() else e for e in re.split("([0-9]+)", s))


def case_insensitive_replace(find_, repl, text):
    return re.sub("(?i)" + re.escape(find_), lambda m: repl, text)


def xlsx_changelog_header(ws):
    ws.column_dimensions["A"].width = 37
    ws.column_dimensions["B"].width = 37
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["D"].width = 60
    ws.freeze_panes = "A2"
    old_val_cell = WriteOnlyCell(ws, value="Old Value")
    old_val_cell.fill = red_remove_fill
    new_val_cell = WriteOnlyCell(ws, value="New Value")
    new_val_cell.fill = green_add_fill
    return [
        WriteOnlyCell(ws, value="Date"),
        WriteOnlyCell(ws, value="Type"),
        WriteOnlyCell(ws, value="ID/Name/Number"),
        old_val_cell,
        new_val_cell,
    ]


def xl_column_string(n):
    s = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        s = chr(65 + remainder) + s
    return s


def nchars(s: str, n: int = 24) -> str:
    return s[: n - 2] + ".." if len(s) > n else s.ljust(n)


def frame_w_to_nchars(frame_w, fixed_font_w, ncols):
    # points = [(2100, 52), (1600, 37), (1100, 22), (600, 9)]
    # # Calculate the slope m
    # # m = (y2 - y1) / (x2 - x1), but we'll use all points for better accuracy
    # x_sum = sum(point[0] for point in points)
    # y_sum = sum(point[1] for point in points)
    # n = len(points)
    # x_mean = x_sum / n
    # y_mean = y_sum / n

    # numerator = sum((points[i][0] - x_mean) * (points[i][1] - y_mean) for i in range(n))
    # denominator = sum((points[i][0] - x_mean) ** 2 for i in range(n))
    # m = numerator / denominator if denominator != 0 else 0  # Slope

    # # Calculate the y-intercept b
    # b = y_mean - m * x_mean  # y = mx + b solved for b
    # nchars = round(m * frame_w + b)
    # print(frame_w, "=", nchars)
    max_chars = ceil(frame_w / fixed_font_w)
    col_chars = ceil(max_chars / ncols)

    if ncols == 4:
        col_chars -= 4
    elif ncols == 3:
        col_chars -= 3
    elif ncols == 2:
        col_chars -= 1

    if col_chars < 9:
        col_chars += 1

    return col_chars


def search_results_max_column_chars(results, limit):
    max_lengths = defaultdict(int)

    for result in results:
        for idx, string in enumerate(result.text):
            max_lengths[idx] = max(max_lengths[idx], len(string))
    return [e if e <= limit else limit for e in max_lengths.values()]
    # return [max_lengths[i] for i in range(max(max_lengths, default=-1) + 1)]


def process_search_results(
    objects: list[object],
    max_lengths: list[int],
    limit: int,
    separator: str = " | ",
) -> None:
    shortfall = sum(limit - val for val in max_lengths if val < limit)
    n_to_add_to = sum(1 for val in max_lengths if val == limit)
    if n_to_add_to:
        to_add = shortfall // n_to_add_to
    max_lengths = [e + to_add if e == limit else e for e in max_lengths]
    for obj in objects:
        obj.text = separator.join(nchars(s, max_length) for s, max_length in zip(obj.text, max_lengths))


def level_to_color(level):
    x = f"{float(level / len(tv_lvls_colors))}"
    if x.endswith("0"):
        return 0
    elif x.endswith("2"):
        return 1
    elif x.endswith("4"):
        return 2
    elif x.endswith("6"):
        return 3
    elif x.endswith("8"):
        return 4


def center(
    toplevel,
    desired_width=None,
    desired_height=None,
    x=None,
    y=None,
    get=False,
    scaling: float | None = None,
):
    if isinstance(scaling, (int, float)):
        if isinstance(desired_width, (int, float)):
            desired_width *= scaling
        if isinstance(desired_height, (int, float)):
            desired_height *= scaling
    if x is not None:
        x = int(x)
    if y is not None:
        y = int(y)
    toplevel.update_idletasks()
    w = toplevel.winfo_screenwidth()
    h = toplevel.winfo_screenheight()
    if desired_width is not None and desired_width >= w:
        desired_width = w - 20
    if desired_height is not None and desired_height >= h:
        desired_height = h - 20
    if desired_width and desired_height:
        size = (desired_width, desired_height)
    else:
        size = tuple(map(int, toplevel.geometry().split("+")[0].split("x")))
    if x is None or x > w or x < 0:
        x = w / 2 - size[0] / 2
    if y is None or y > h or y < 0:
        y = h / 2 - size[1] / 2
    if not get:
        toplevel.geometry("%dx%d+%d+%d" % (size + (x, y)))
    return x, y


def bytes_io_wb(filepath):
    with open(filepath, "rb") as fh:
        in_mem = io.BytesIO(fh.read())
    return in_mem


def csv_str_x_data(s: str, discard_empty_rows: bool = True, paste: bool = False) -> list[list[str]]:
    dialect = get_csv_str_dialect(s, delimiters=from_clipboard_delimiters)
    if discard_empty_rows:
        data = []
        for r in csv.reader(
            io.StringIO(s),
            dialect=dialect,
            skipinitialspace=True,
        ):
            try:
                data.append(r[: len(r) - next(i for i, c in enumerate(reversed(r)) if c)])
            except Exception:
                continue
        return data
    else:
        if not paste or dialect.delimiter in s or "\n" in s:
            return list(
                csv.reader(
                    io.StringIO(s),
                    dialect=dialect,
                    skipinitialspace=True,
                )
            )
        else:
            return [[s]]


def ws_x_data(ws) -> list[list[str]]:
    data = []
    for r in ws.iter_rows(values_only=True):
        try:
            data.append(
                [
                    "" if x is None else f"{x}"
                    for x in islice(r, 0, len(r) - next(i for i, e in enumerate(reversed(r)) if e is not None))
                ]
            )
        except Exception:
            continue
    return data


def ws_x_program_data_str(ws) -> str:
    return "".join("" if r[0] is None else f"{r[0]}" for r in islice(ws.iter_rows(values_only=True), 1, None))


def is_json_one(data):
    if not isinstance(data, dict):
        return False
    for hdr, lst in data.items():
        if not isinstance(hdr, str):
            return False
        if not isinstance(lst, list):
            return False
    return True


def is_json_two(data):
    if not isinstance(data, list):
        return False
    return all(isinstance(dct, dict) for dct in data)


def is_json_three(data):
    if not isinstance(data, list):
        return False
    return all(isinstance(lst, list) for lst in data)


def is_json_four(data):
    return isinstance(data, str)


def get_json_format(j):
    try:
        records = {k: i for i, k in enumerate(j)}
    except Exception:
        return None
    if "program_data" in records:
        return "program_data", "records"
    for key in ("records", "sheet", "data", "table"):
        if key in records:
            try:
                if is_json_one(j[key]):
                    return 1, key
                elif is_json_two(j[key]):
                    return 2, key
                elif is_json_three(j[key]):
                    return 3, key
                elif is_json_four(j[key]):
                    return 4, key
            except Exception:
                continue
    return None


def json_to_sheet(
    j,
    format_=1,
    key="records",
    get_format=True,
    return_rowlen=False,
):
    new_sheet = []
    if get_format:
        format_, key = get_json_format(j)
    if format_ == "program_data":
        try:
            d = b32_x_dict(j["program_data"])
            if return_rowlen:
                return [[h["name"] for h in d["headers"]]] + d["records"], len(d["headers"])
            else:
                return [[h["name"] for h in d["headers"]]] + d["records"]
        except Exception:
            return new_sheet, 0
    elif format_ == 1:
        new_sheet = [list(j[key])]
        keys = new_sheet[0]
        rowlen, numrows = len(keys), max(map(len, j[key].values()), default=0)
        for hdr in keys:
            if len(j[key][hdr]) < numrows:
                j[key][hdr].extend(list(repeat("", numrows - len(j[key][hdr]))))
        for i in range(numrows):
            row = []
            for hdr in keys:
                if isinstance(j[key][hdr][i], str):
                    row.append(j[key][hdr][i])
                else:
                    try:
                        row.append(f"{j[key][hdr][i]}")
                    except Exception:
                        row.append("")
            new_sheet.append(row)
    elif format_ == 2:
        headers = {}
        for dct in j[key]:
            for k in dct:
                if k not in headers and isinstance(k, str):
                    headers[k] = len(headers)
        if not headers and "headers" in j:
            headers = {k: i for i, k in enumerate(json_get_header_strings(j["headers"]))}
        if not headers and "columns" in j:
            headers = {k: i for i, k in enumerate(json_get_header_strings(j["columns"]))}
        rowlen = len(headers)
        if rowlen >= 2:
            new_sheet = [list(headers)]
            for dct in j[key]:
                row = []
                for v in dct.values():
                    if isinstance(v, str):
                        row.append(v)
                    else:
                        try:
                            row.append(f"{v}")
                        except Exception:
                            row.append("")
                new_sheet.append(row)
    elif format_ == 3:
        for r in j[key]:
            row = []
            try:
                for v in islice(
                    r,
                    0,
                    len(r) - next(i for i, c in enumerate(reversed(r)) if c != ""),
                ):
                    if isinstance(v, str):
                        row.append(v)
                    else:
                        try:
                            row.append(f"{v}")
                        except Exception:
                            row.append("")
                new_sheet.append(row)
            except Exception:
                pass
        rowlen = equalize_sublist_lens(new_sheet)
    elif format_ == 4:
        new_sheet = csv_str_x_data(j[key])
        rowlen = equalize_sublist_lens(new_sheet)
    if return_rowlen:
        return new_sheet, rowlen
    return new_sheet


def json_get_header_strings(obj):
    if isinstance(obj, list) and obj:
        if all(isinstance(e, dict) for e in obj):
            try:
                return [f"{h['name']}" for h in obj]
            except Exception:
                pass
        elif all(isinstance(e, str) for e in obj):
            return list(obj)


def full_sheet_to_dict(
    headers,
    data,
    include_headers=False,
    key="records",
    format_=1,
) -> dict:
    if format_ == 1:
        return {key: {hdr: [row[i] for row in data] for i, hdr in enumerate(headers)}}
    elif format_ == 2:
        if include_headers:
            return {
                key: [{hdr: row[i] for i, hdr in enumerate(headers)} for row in data],
                "headers": headers,
            }
        else:
            return {key: [{hdr: row[i] for i, hdr in enumerate(headers)} for row in data]}
    elif format_ == 3:
        return {key: [headers] + data}
    elif format_ == 4:
        s, writer = str_io_csv_writer(dialect=csv.excel_tab)
        writer.writerow(headers)
        writer.writerows(data)
        return {key: s.getvalue().rstrip()}


def to_json(
    filepath,
    data,
    format_,
):
    if data:
        headers = data.pop(0)
        d = full_sheet_to_dict(
            headers,
            data,
            format_=format_,
        )
    else:
        d = full_sheet_to_dict(
            [],
            data,
            format_=format_,
        )
    with open(filepath, "w") as fh:
        fh.write(json.dumps(d, indent=4))


def path_without_numbers(full_path):
    if full_path.lower().endswith((".csv", ".xls", ".tsv")):
        ext = full_path[-4:]
        path = full_path[:-4]
    elif full_path.lower().endswith((".xlsx", ".json", ".xlsm")):
        ext = full_path[-5:]
        path = full_path[:-5]
    last_index = 0
    for i, c in enumerate(reversed(path), 1):
        if c.isdigit():
            last_index = i
        else:
            break
    if not last_index:
        return full_path
    else:
        return path[:-last_index] + ext


def path_numbers(full_path):
    if full_path.lower().endswith((".csv", ".xls", ".tsv")):
        path = full_path[:-4]
    elif full_path.lower().endswith((".xlsx", ".json", ".xlsm")):
        path = full_path[:-5]
    numbers = []
    for c in reversed(path):
        if c.isdigit():
            numbers.append(c)
        else:
            break
    if numbers:
        return int("".join(numbers[::-1]))
    else:
        return 0


def increment_file_version(full_path):
    if full_path.lower().endswith((".csv", ".xls", ".tsv")):
        ext = full_path[-4:]
        path = full_path[:-4]
    elif full_path.lower().endswith((".xlsx", ".json", ".xlsm")):
        ext = full_path[-5:]
        path = full_path[:-5]
    numbers = []
    last_index = 0
    for i, c in enumerate(reversed(path), 1):
        if c.isdigit():
            numbers.append(c)
            last_index = i
        else:
            break
    if numbers:
        numbers = numbers[::-1]
        numbers[len(numbers) - 1] = f"{int(numbers[len(numbers) - 1]) + 1}"
        numbers = "".join(numbers)
        newfile = path[:-last_index] + numbers + ext
    else:
        newfile = path + "1" + ext
    return newfile


def convert_old_xl_to_xlsx(path_):
    if not path_.lower().endswith(".xlsx"):
        filename, file_extension = os.path.splitext(path_)
        return filename + ".xlsx"
    else:
        return path_


def try_remove(remove_from, remove):
    with suppress(Exception):
        remove_from.remove(remove)


def type_int(o):
    return isinstance(o, int) and not isinstance(o, bool)


def int_or_float(o):
    return (isinstance(o, int) and not isinstance(o, bool)) or isinstance(o, float)


def isreal(inp, str_only=False, num_only=False, allow_nan=False, allow_inf=False):
    if str_only and not isinstance(inp, str):
        return False
    if num_only and not int_or_float(inp):
        return False
    try:
        x = bool(isrealre.match(inp))
    except TypeError:
        return isinstance(inp, (int, float))
    else:
        return bool(
            x
            or allow_inf
            and inp.lower().strip().lstrip("-+") in ("inf", "infinity")
            or allow_nan
            and inp.lower().strip().lstrip("-+") == "nan"
        )


def isfloat(inp, str_only=False, num_only=False, allow_nan=False, allow_inf=False):
    if str_only and not isinstance(inp, str):
        return False
    if num_only and not isinstance(inp, float):
        return False
    try:
        x = bool(isfloatre.match(inp))
    except TypeError:
        return isinstance(inp, float)
    else:
        return bool(
            x
            or allow_inf
            and inp.lower().strip().lstrip("-+") in ("inf", "infinity")
            or allow_nan
            and inp.lower().strip().lstrip("-+") == "nan"
        )


def isint(inp, str_only=False, num_only=False):
    if str_only and not isinstance(inp, str):
        return False
    if num_only and not type_int(inp):
        return False
    try:
        return bool(isintre.match(inp))
    except TypeError:
        return False


def isintlike(inp, str_only=False, num_only=False):
    if str_only and not isinstance(inp, str):
        return False
    if num_only and not int_or_float(inp):
        return False
    try:
        if isintre.match(inp):
            return True
        elif isintlikere.match(inp):
            return float(inp).is_integer()
        else:
            return False
    except TypeError:
        if isinstance(inp, float):
            return inp.is_integer()
        return bool(type_int(inp))


def equalize_sublist_lens(seq: list[list[object]], len_: int | None = None) -> list[list[object]]:
    if len_ is None:
        len_ = max(map(len, seq), default=0)
    for sl in seq:
        if len(sl) < len_:
            sl.extend(repeat("", len_ - len(sl)))
    return len_


def shift_elements_to_start(seq):
    return (eles := seq[next(i for i, e in enumerate(seq) if e) :]) + list(repeat("", len(seq) - len(eles)))


def shift_elements_to_end(seq):
    return (
        list(repeat("", len(seq) - len(eles := seq[: len(seq) - next(i for i, e in enumerate(reversed(seq)) if e)])))
        + eles
    )


def filter_empty_rows(data: list[list[object]]) -> list[list[object]]:
    return [r for r in data if any(r)]


def dict_x_b32(d: dict):
    return b32e(zlib.compress(json.dumps(d).encode())).decode()


def b32_x_dict(s: str) -> dict:
    b = b32d(s.encode())
    if comp_method(b) == "zlib":
        return DotDict(json.loads(zlib.decompress(b).decode()))
    # backwards compatible, old versions used lzma
    else:
        return DotDict(json.loads(lzma.decompress(b).decode()))


def comp_method(b: bytes):
    if b[0] == 0x78:
        return "zlib"
    else:
        return "lzma"


def get_json_from_file(fp):
    with open(fp, "r") as fh:
        j = json.loads(fh.read())
    return j


def new_scrolls(scrolls: None | tuple[float, float, float, float] = None) -> DotDict:
    if scrolls is None:
        scrolls = (0.0, 0.0, 0.0, 0.0)
    return DotDict(
        treex=scrolls[0],
        treey=scrolls[1],
        sheetx=scrolls[2],
        sheety=scrolls[3],
    )


def new_saved_info(hierarchies: list[int]) -> dict:
    saved_info = DotDict()
    for h in hierarchies:
        saved_info[h] = new_info_storage()
    return saved_info


def new_info_storage(
    scrolls: None | tuple[float, float, float, float] = None,
    opens: None | set[int] = None,
    boxes: None | tuple = None,
    selected: None | tuple = None,
    twidths: None | dict[str, int] = None,
    theights: None | dict[str, int] = None,
) -> DotDict:
    return DotDict(
        scrolls=new_scrolls(scrolls=scrolls),
        opens={} if opens is None else opens,
        boxes=() if boxes is None else boxes,
        selected=() if selected is None else selected,
        twidths={} if twidths is None else twidths,
        theights={} if theights is None else theights,
    )


def create_cell_align_selector_menu(
    parent,
    command,
    menu_kwargs,
    icons,
):
    menu = tk.Menu(parent, tearoff=0, **menu_kwargs)
    menu.add_command(
        label="Left",
        command=lambda: command("w"),
        image=icons["w"],
        compound="left",
        **menu_kwargs,
    )
    menu.add_command(
        label="Center",
        command=lambda: command("center"),
        image=icons["c"],
        compound="left",
        **menu_kwargs,
    )
    menu.add_command(
        label="Right",
        command=lambda: command("e"),
        image=icons["e"],
        compound="left",
        **menu_kwargs,
    )
    menu.add_command(
        label="Default",
        command=lambda: command("global"),
        **menu_kwargs,
    )
    return menu
