# SPDX-License-Identifier: GPL-3.0-only
# Copyright © R. A. Gardner

import csv
import datetime
import json
import os
import re
import tkinter as tk
import webbrowser
from itertools import chain, islice, repeat
from operator import itemgetter
from tkinter import filedialog, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from tksheet import (
    Sheet,
    convert_align,
    move_elements_by_mapping,
)

from .classes import (
    TreeBuilder,
)
from .constants import (
    BF,
    EF,
    EFB,
    ERR_ASK_FNT,
    TF,
    app_title,
    blue_fill,
    green_fill,
    changelog_header,
    ctrl_button,
    lge_font,
    lge_font_size,
    menu_kwargs,
    rc_button,
    sheet_header_font,
    std_font_size,
    themes,
    top_left_icon,
    upone_dir,
    validation_allowed_date_chars,
    validation_allowed_num_chars,
)
from .functions import (
    b32_x_dict,
    bytes_io_wb,
    case_insensitive_replace,
    center,
    csv_str_x_data,
    equalize_sublist_lens,
    full_sheet_to_dict,
    get_json_format,
    get_json_from_file,
    json_to_sheet,
    str_io_csv_writer,
    to_clipboard,
    ws_x_data,
    ws_x_program_data_str,
    xlsx_changelog_header,
)
from .widgets import (
    Auto_Add_Condition_Date_Frame,
    Auto_Add_Condition_Num_Frame,
    Button,
    Date_Entry,
    Display_Text,
    Edit_Condition_Frame,
    Entry_With_Scrollbar,
    Error_Frame,
    Ez_Dropdown,
    Flattened_Column_Selector,
    FlattenedToggleAndOrder,
    Frame,
    Id_Parent_Column_Selector,
    Label,
    Normal_Entry,
    Numerical_Entry_With_Scrollbar,
    Readonly_Entry_With_Scrollbar,
    Scrollbar,
    Single_Column_Selector,
    Status_Bar,
    Working_Text,
    Wrapped_Text_With_Find_And_Yscroll,
    X_Checkbutton,
)


def new_toplevel_chores(toplevel, parent, title, grab=True, resizable=False):
    toplevel.update()
    if grab:
        toplevel.grab_set()
    toplevel.withdraw()
    toplevel.resizable(resizable, resizable)
    toplevel.tk.call("wm", "iconphoto", toplevel._w, tk.PhotoImage(format="gif", data=top_left_icon))
    toplevel.title(title)
    if grab:
        toplevel.wm_transient(parent)
    toplevel.focus_force()
    return parent


class Export_Flattened_Popup(tk.Toplevel):
    def __init__(self, C, width=1280, height=800, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(toplevel=self, parent=C, title=f"{app_title} - Flatten sheet", resizable=True)

        self.protocol("WM_DELETE_WINDOW", self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.wb_ = None
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.selector = Single_Column_Selector(self, theme=theme)
        self.selector.enable_me()

        self.selector.set_columns([self.C.headers[h].name for h in self.C.hiers])
        if self.C.pc == -1:
            self.selector.set_col(0)
        else:
            self.selector.set_col(self.C.hiers.index(self.C.pc))
        self.selector.grid(row=0, column=0, sticky="nwe", pady=(10, 20), padx=10)

        self.include_details_button = X_Checkbutton(
            self,
            text="Include detail columns  ",
            style="x_button.Std.TButton",
            compound="right",
            checked=self.C.xlsx_flattened_detail_columns,
        )
        self.include_details_button.grid(row=1, column=0, sticky="new", pady=(10, 5), padx=10)

        self.justify_left_button = X_Checkbutton(
            self,
            text="Justify left  ",
            style="x_button.Std.TButton",
            compound="right",
            checked=self.C.xlsx_flattened_justify,
        )
        self.justify_left_button.grid(row=3, column=0, sticky="new", pady=5, padx=10)

        self.order_button = X_Checkbutton(
            self,
            text="Reverse Order  ",
            style="x_button.Std.TButton",
            compound="right",
            checked=self.C.xlsx_flattened_reverse_order,
        )
        self.order_button.grid(row=4, column=0, sticky="new", pady=5, padx=10)

        self.add_index_button = X_Checkbutton(
            self,
            text="Add index column  ",
            style="x_button.Std.TButton",
            compound="right",
            checked=self.C.xlsx_flattened_add_index,
        )
        self.add_index_button.grid(row=5, column=0, sticky="new", pady=(10, 5), padx=10)

        self.build_button = Button(self, text="  Flatten sheet  ", style="EF.Std.TButton", command=self.build_flattened)
        self.build_button.grid(row=7, column=0, pady=10, padx=10, sticky="nsew")

        self.sheetdisplay = Sheet(
            self,
            theme=theme,
            header_font=sheet_header_font,
            outline_thickness=0,
        )
        self.sheetdisplay.enable_bindings("all", "ctrl_select")
        self.sheetdisplay.extra_bindings("begin_edit_cell", self.begin_edit)
        self.sheetdisplay.extra_bindings("end_edit_cell", self.end_edit)
        self.sheetdisplay.headers(newheaders=0)
        self.sheetdisplay.grid(row=0, column=1, rowspan=6, sticky="nswe")

        self.button_frame = Frame(self, theme=theme)
        self.button_frame.grid_rowconfigure(0, weight=1)
        self.button_frame.grid(row=7, column=1, sticky="e")
        self.save_button = Button(self.button_frame, text="Save as", style="EF.Std.TButton", command=self.save_as)
        self.save_button.grid(row=0, column=0, padx=10, pady=20, sticky="e")
        self.clipboard_json_button = Button(
            self.button_frame,
            text=" Clipboard as json ",
            style="EF.Std.TButton",
            command=self.clipboard_json,
        )
        self.clipboard_json_button.grid(row=0, column=1, padx=10, pady=20, sticky="e")
        self.clipboard_indent_button = Button(
            self.button_frame,
            text=" Clipboard (indent separated) ",
            style="EF.Std.TButton",
            command=self.clipboard_indent,
        )
        self.clipboard_indent_button.grid(row=0, column=2, padx=10, pady=20, sticky="e")
        self.clipboard_comma_button = Button(
            self.button_frame,
            text=" Clipboard (comma separated) ",
            style="EF.Std.TButton",
            command=self.clipboard_comma,
        )
        self.clipboard_comma_button.grid(row=0, column=3, padx=10, pady=20, sticky="e")
        self.done_button = Button(self.button_frame, text="Done", style="EF.Std.TButton", command=self.cancel)
        self.done_button.grid(row=0, column=4, padx=(10, 20), pady=20, sticky="e")
        self.status_bar = Status_Bar(
            self, text="Use the parent column selector to change hierarchy output", theme=theme
        )
        self.status_bar.grid(row=9, column=0, columnspan=2, sticky="nswe")

        self.bind("<Escape>", self.cancel)
        self.build_flattened()

        center(self, width, height)
        self.deiconify()
        self.grab_set()
        self.wait_window()

    def end_edit(self, event=None):
        self.bind("<Escape>", self.cancel)

    def begin_edit(self, event=None):
        self.unbind("<Escape>")
        return event.value

    def start_work(self, msg=""):
        self.status_bar.change_text(msg)
        self.disable_widgets()

    def stop_work(self, msg=""):
        self.status_bar.change_text(msg)
        self.enable_widgets()

    def enable_widgets(self):
        self.sheetdisplay.enable_bindings("all", "ctrl_select")
        self.sheetdisplay.extra_bindings("begin_edit_cell", self.begin_edit)
        self.sheetdisplay.extra_bindings("end_edit_cell", self.end_edit)
        self.sheetdisplay.basic_bindings(True)
        self.save_button.config(state="normal")
        self.clipboard_indent_button.config(state="normal")
        self.clipboard_json_button.config(state="normal")
        self.clipboard_comma_button.config(state="normal")
        self.build_button.config(state="normal")
        self.selector.enable_me()

    def disable_widgets(self):
        self.build_button.config(state="disabled")
        self.sheetdisplay.disable_bindings()
        self.sheetdisplay.extra_bindings("begin_edit_cell", None)
        self.sheetdisplay.extra_bindings("end_edit_cell", None)
        self.sheetdisplay.basic_bindings(False)
        self.save_button.config(state="disabled")
        self.clipboard_json_button.config(state="disabled")
        self.clipboard_indent_button.config(state="disabled")
        self.clipboard_comma_button.config(state="disabled")
        self.selector.disable_me()
        self.update()

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except Exception:
            pass
        try:
            self.wb_ = None
        except Exception:
            pass

    def USER_HAS_CLOSED_WINDOW(self, callback=None):
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except Exception:
            pass
        self.destroy()

    def clipboard_json(self):
        self.start_work("Copying to clipboard...")
        to_clipboard(
            self.C.C,
            json.dumps(
                full_sheet_to_dict(
                    self.sheetdisplay.get_sheet_data()[0],
                    self.sheetdisplay.get_sheet_data()[1:],
                    include_headers=True,
                    format_=self.C.which_json(),
                ),
                indent=4,
            ),
        )
        self.stop_work("Sheet successfully copied to clipboard as json!")

    def clipboard_indent(self):
        self.start_work("Copying to clipboard...")
        s, writer = str_io_csv_writer(dialect=csv.excel_tab)
        writer.writerows(self.sheetdisplay.get_sheet_data())
        to_clipboard(self.C.C, s.getvalue().rstrip())
        self.stop_work("Sheet successfully copied to clipboard (indent separated)!")

    def clipboard_comma(self):
        self.start_work("Copying to clipboard...")
        s, writer = str_io_csv_writer(dialect=csv.excel)
        writer.writerows(self.sheetdisplay.get_sheet_data())
        to_clipboard(self.C.C, s.getvalue().rstrip())
        self.stop_work("Sheet successfully copied to clipboard (comma separated)!")

    def build_flattened(self):
        self.start_work("Flattening sheet...")
        self.sheetdisplay.deselect("all")
        self.sheetdisplay.set_sheet_data(
            data=TreeBuilder().build_flattened(
                input_sheet=self.C.sheet.MT.data,
                output_sheet=[],
                nodes=self.C.nodes,
                headers=[f"{hdr.name}" for hdr in self.C.headers],
                ic=int(self.C.ic),
                pc=int(self.C.hiers[self.selector.get_col()]),
                hiers=list(self.C.hiers),
                detail_columns=self.include_details_button.get_checked(),
                justify_left=self.justify_left_button.get_checked(),
                reverse=self.order_button.get_checked(),
                add_index=self.add_index_button.get_checked(),
            ),
            verify=False,
        )
        self.stop_work("Sheet successfully flattened!")

    def save_as(self):
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(
            parent=self,
            title="Save as",
            filetypes=[("Excel file", ".xlsx"), ("JSON File", ".json"), ("CSV File", ".csv"), ("TSV File", ".tsv")],
            defaultextension=".xlsx",
            confirmoverwrite=True,
        )
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv", ".xlsx", ".json", ".tsv")):
            self.grab_set()
            self.stop_work("Can only save .json/.csv/.xlsx file types")
            return
        self.status_bar.change_text("Saving...")
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook(write_only=True)
                ws = self.wb_.create_sheet(title="Sheet1")
                ws.freeze_panes = "A2"
                for row in self.sheetdisplay.get_sheet_data():
                    ws.append(row)
                self.wb_.save(newfile)
                self.try_to_close_wb()
            elif newfile.lower().endswith(".json"):
                with open(newfile, "w", newline="") as fh:
                    fh.write(
                        json.dumps(
                            full_sheet_to_dict(
                                self.sheetdisplay.get_sheet_data()[0],
                                self.sheetdisplay.get_sheet_data()[1:],
                                include_headers=True,
                                format_=self.C.which_json(),
                            ),
                            indent=4,
                        )
                    )
            elif newfile.lower().endswith((".tsv", ".csv")):
                with open(newfile, "w", newline="", encoding="utf-8") as fh:
                    writer = csv.writer(
                        fh,
                        dialect=csv.excel_tab if newfile.lower().endswith(".tsv") else csv.excel,
                        lineterminator="\n",
                    )
                    writer.writerows(self.sheetdisplay.get_sheet_data())
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work(f"Error saving file: {error_msg}")
            return
        self.stop_work("Success! Flattened sheet saved")

    def cancel(self, event=None):
        self.USER_HAS_CLOSED_WINDOW()


class Post_Import_Changes_Popup(tk.Toplevel):
    def __init__(self, C, changes, successful, width=1000, height=800, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Imported Changes", resizable=True)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.sheetdisplay = Sheet(
            self,
            theme=theme,
            header_font=sheet_header_font,
            outline_thickness=0,
            auto_resize_row_index=True,
        )
        self.sheetdisplay.enable_bindings(
            "single",
            "copy",
            "drag_select",
            "column_width_resize",
            "double_click_column_resize",
            "row_height_resize",
            "double_click_row_resize",
            "row_width_resize",
            "row_select",
            "arrowkeys",
            "ctrl_select",
        )
        self.sheetdisplay.headers(newheaders=changelog_header)
        self.sheetdisplay.row_index(0)
        self.sheetdisplay.data_reference(newdataref=changes, reset_col_positions=True, reset_row_positions=True)
        self.sheetdisplay.hide_columns(0)
        for i, b in enumerate(successful):
            self.sheetdisplay.highlight(
                self.sheetdisplay.span(i, index=True),
                bg="#40bd59" if b else "#db7463",
                fg="black",
            )
        self.sheetdisplay.grid(row=0, column=0, sticky="nswe")
        self.status_bar = Status_Bar(
            self,
            text=f"Successful changes: {sum(1 for b in successful if b)}/{len(successful)}",
            theme=theme,
        )
        self.status_bar.grid(row=1, column=0, sticky="nswe")
        self.bind("<Escape>", self.cancel)
        center(self, width, height)
        self.deiconify()
        self.wait_window()

    def cancel(self, event=None):
        self.destroy()


class Changelog_Popup(tk.Toplevel):
    def __init__(self, C, width=999, height=800, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Changelog", resizable=True)
        self.USER_HAS_QUIT = False
        self.protocol("WM_DELETE_WINDOW", self.USER_HAS_CLOSED_WINDOW)

        self.find_results = []
        self.results_number = 0
        self.wb_ = None
        self.total_changes = f"Total changes: {len(self.C.changelog)} | "

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.find_frame = Frame(self, theme=theme)
        self.find_frame.grid(row=0, column=0, columnspan=2, sticky="nswe")
        self.search_button = Button(self.find_frame, text=" Find:", command=self.find)
        self.search_button.pack(side="left", fill="x")
        self.find_window = Normal_Entry(self.find_frame, font=BF, theme=theme)
        self.find_window.bind("<Return>", self.find)
        self.find_window.pack(side="left", fill="x", expand=True)
        self.find_reset_button = Button(self.find_frame, text="X", command=self.find_reset)
        self.find_reset_button.pack(side="left", fill="x")
        self.find_results_label = Label(self.find_frame, "0/0", BF, theme=theme)
        self.find_results_label.pack(side="left", fill="x")
        self.find_up_button = Button(self.find_frame, text="▲", command=self.find_up)
        self.find_up_button.pack(side="left", fill="x")
        self.find_down_button = Button(self.find_frame, text="▼", command=self.find_down)
        self.find_down_button.pack(side="left", fill="x")

        self.sheetdisplay = Sheet(
            self,
            theme=theme,
            headers=changelog_header,
            row_index=0,
            startup_select=(len(self.C.changelog) - 1, len(self.C.changelog), "rows"),
            data=self.C.changelog,
            row_index_align="w",
            header_font=sheet_header_font,
            outline_thickness=0,
            auto_resize_row_index=True,
            default_column_width=200,
        )
        self.sheetdisplay.column_width(0, width=1)
        self.enable_bindings()
        self.red_bg = "#db7463"
        self.green_bg = "#40bd59"
        self.red_fg = "black"
        self.green_fg = "black"
        self.sheetdisplay.highlight_columns(columns=3, bg=self.red_bg, fg=self.red_fg)
        self.sheetdisplay.highlight_columns(columns=4, bg=self.green_bg, fg=self.green_fg)
        self.sheetdisplay.grid(row=1, column=0, sticky="nswe")
        self.status_bar = Status_Bar(self, text=self.total_changes, theme=theme)
        self.status_bar.grid(row=2, column=0, sticky="nswe")

        self.buttonframe = Frame(self, theme=theme)
        self.buttonframe.grid(row=3, column=0, sticky="nswe")
        self.done_button = Button(self.buttonframe, text="Done", style="EF.Std.TButton", command=self.cancel)
        self.done_button.pack(side="right", fill="x", padx=20, pady=20)
        self.save_text_button = Button(
            self.buttonframe, text="Export all", style="EF.Std.TButton", command=self.save_as
        )
        self.save_text_button.pack(side="right", fill="x", padx=20, pady=20)

        self.export_selected_button = Button(
            self.buttonframe, text="Export selected as", style="EF.Std.TButton", command=self.save_selected_as
        )
        self.export_selected_button.pack(side="right", fill="x", padx=20, pady=20)

        self.prune_button = Button(
            self.buttonframe, text="Prune up to selected", style="EF.Std.TButton", command=self.prune
        )
        self.prune_button.pack(side="right", fill="x", padx=20, pady=20)

        self.bind("<Escape>", self.cancel)
        center(self, width, height)
        self.deiconify()
        self.wait_window()

    def prune(self, event=None):
        selectedrows = self.sheetdisplay.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        if not selectedrows:
            return
        num = len(selectedrows)
        self.start_work(f"Pruning {num} changes...")
        up_to = min(selectedrows)
        if self.C.changelog[up_to][1].endswith(("|", "| ")):
            for i, entry in enumerate(islice(self.C.changelog, up_to, None), up_to):
                if not entry[1].endswith(("|", "| ")):
                    up_to = i
                    break
        self.C.snapshot_prune_changelog(up_to)
        self.C.changelog[: up_to + 1] = []
        self.sheetdisplay.headers(newheaders=changelog_header)
        self.sheetdisplay.row_index(newindex=0)
        self.sheetdisplay.data_reference(
            newdataref=self.C.changelog, reset_col_positions=False, reset_row_positions=True, redraw=False
        )
        self.total_changes = f"Total changes: {len(self.C.changelog)} | "
        self.status_bar.config(text=self.total_changes)
        self.C.C.status_bar.change_text(self.C.get_tree_editor_status_bar_text())
        self.sheetdisplay.refresh()
        self.stop_work(f"Success! Pruned {up_to + 1} changes")

    def start_work(self, msg=""):
        self.status_bar.change_text(self.total_changes + msg)
        self.disable_widgets()

    def stop_work(self, msg=""):
        self.status_bar.change_text(self.total_changes + msg)
        self.enable_widgets()

    def enable_bindings(self):
        self.sheetdisplay.enable_bindings(
            "single",
            "copy",
            "right_click_popup_menu",
            "drag_select",
            "select_all",
            "row_width_resize",
            "column_width_resize",
            "double_click_column_resize",
            "row_height_resize",
            "double_click_row_resize",
            "row_select",
            "column_select",
            "arrowkeys",
        )

    def enable_widgets(self):
        self.enable_bindings()
        self.find_window.bind("<Return>", self.find)
        self.find_reset_button.config(state="normal")
        self.find_up_button.config(state="normal")
        self.find_down_button.config(state="normal")
        self.save_text_button.config(state="normal")

    def disable_widgets(self):
        self.sheetdisplay.disable_bindings()
        self.find_window.unbind("<Return>")
        self.find_reset_button.config(state="disabled")
        self.find_up_button.config(state="disabled")
        self.find_down_button.config(state="disabled")
        self.save_text_button.config(state="disabled")
        self.update()

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except Exception:
            pass
        try:
            self.wb_ = None
        except Exception:
            pass

    def USER_HAS_CLOSED_WINDOW(self, callback=None):
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except Exception:
            pass
        self.destroy()

    def save_as(self):
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(
            parent=self,
            title="Save as",
            filetypes=[("CSV File", ".csv"), ("TSV File", ".tsv"), ("Excel file", ".xlsx"), ("JSON File", ".json")],
            defaultextension=".csv",
            confirmoverwrite=True,
        )
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv", ".xlsx", ".json", ".tsv")):
            self.grab_set()
            self.stop_work("Can only save .csv/.xlsx/.json file types")
            return
        self.status_bar.change_text(f"{self.total_changes}Saving...")
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook(write_only=True)
                ws = self.wb_.create_sheet(title="Changelog")
                ws.append(xlsx_changelog_header(ws))
                for row in self.C.changelog:
                    ws.append((e if e else None for e in row))
                self.wb_.save(newfile)
                self.try_to_close_wb()
            elif newfile.lower().endswith((".csv", ".tsv")):
                with open(newfile, "w", newline="", encoding="utf-8") as fh:
                    writer = csv.writer(
                        fh,
                        dialect=csv.excel_tab if newfile.lower().endswith(".tsv") else csv.excel,
                        lineterminator="\n",
                    )
                    writer.writerow(changelog_header)
                    writer.writerows(self.C.changelog)
            elif newfile.lower().endswith(".json"):
                with open(newfile, "w", newline="") as fh:
                    fh.write(
                        json.dumps(
                            full_sheet_to_dict(
                                changelog_header,
                                self.C.changelog,
                                include_headers=True,
                                format_=self.C.which_json(),
                            ),
                            indent=4,
                        )
                    )
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work(f"Error saving file: {error_msg}")
            return
        self.stop_work("Success! Changelog saved")

    def save_selected_as(self):
        selectedrows = self.sheetdisplay.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        if not selectedrows:
            return
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(
            parent=self,
            title="Save as",
            filetypes=[("CSV File", ".csv"), ("TSV File", ".tsv"), ("Excel file", ".xlsx"), ("JSON File", ".json")],
            defaultextension=".csv",
            confirmoverwrite=True,
        )
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv", ".xlsx", ".json", ".tsv")):
            self.grab_set()
            self.stop_work("Can only save .csv/.xlsx/.json file types")
            return
        from_row = min(selectedrows)
        to_row = max(selectedrows) + 1
        self.status_bar.change_text(f"{self.total_changes}Saving...")
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook(write_only=True)
                ws = self.wb_.create_sheet(title="Changelog")
                ws.append(xlsx_changelog_header(ws))
                for row in islice(self.C.changelog, from_row, to_row):
                    ws.append((e if e else None for e in row))
                self.wb_.save(newfile)
                self.try_to_close_wb()
            elif newfile.lower().endswith((".csv", ".tsv")):
                with open(newfile, "w", newline="", encoding="utf-8") as fh:
                    writer = csv.writer(
                        fh,
                        dialect=csv.excel_tab if newfile.lower().endswith(".tsv") else csv.excel,
                        lineterminator="\n",
                    )
                    writer.writerow(changelog_header)
                    writer.writerows(islice(self.C.changelog, from_row, to_row))
            elif newfile.lower().endswith(".json"):
                with open(newfile, "w", newline="") as fh:
                    fh.write(
                        json.dumps(
                            full_sheet_to_dict(
                                changelog_header,
                                self.C.changelog[from_row:to_row],
                                include_headers=True,
                                format_=self.C.which_json(),
                            ),
                            indent=4,
                        )
                    )
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work(f"Error saving file: {error_msg}")
            return
        self.stop_work("Success! Changelog saved")

    def find(self, event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        x = self.word.lower()
        for rn, row in enumerate(self.C.changelog):
            for colno, cell in enumerate(row):
                if x in cell.lower():
                    if colno == 0:
                        self.find_results.append((rn, 6))
                        break
                    else:
                        self.find_results.append((rn, colno))
        if self.find_results:
            for rn, colno in islice(self.find_results, 1, len(self.find_results)):
                if colno == 6:
                    for i in range(1, 6):
                        self.sheetdisplay.highlight_cells(row=rn, column=i, bg="yellow", fg="black")
                else:
                    self.sheetdisplay.highlight_cells(row=rn, column=colno, bg="yellow", fg="black")
            if self.find_results[self.results_number][1] == 6:
                for i in range(1, 6):
                    self.sheetdisplay.highlight_cells(
                        row=self.find_results[self.results_number][0], column=i, bg="orange", fg="black"
                    )
            else:
                self.sheetdisplay.highlight_cells(
                    row=self.find_results[self.results_number][0],
                    column=self.find_results[self.results_number][1],
                    bg="orange",
                    fg="black",
                )
            self.find_results_label.config(text=f"1/{len(self.find_results)}")
            self.sheetdisplay.see(row=self.find_results[0][0], column=0, keep_xscroll=True)
        self.sheetdisplay.refresh()

    def find_up(self, event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        if self.find_results[self.results_number][1] == 6:
            for i in range(1, 6):
                self.sheetdisplay.highlight_cells(
                    row=self.find_results[self.results_number][0], column=i, bg="yellow", fg="black"
                )
        else:
            self.sheetdisplay.highlight_cells(
                row=self.find_results[self.results_number][0],
                column=self.find_results[self.results_number][1],
                bg="yellow",
                fg="black",
            )
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=f"{self.results_number + 1}/{len(self.find_results)}")
        if self.find_results[self.results_number][1] == 6:
            for i in range(1, 6):
                self.sheetdisplay.highlight_cells(
                    row=self.find_results[self.results_number][0], column=i, bg="orange", fg="black"
                )
        else:
            self.sheetdisplay.highlight_cells(
                row=self.find_results[self.results_number][0],
                column=self.find_results[self.results_number][1],
                bg="orange",
                fg="black",
            )
        self.sheetdisplay.see(row=self.find_results[self.results_number][0], column=0, keep_xscroll=True)
        self.sheetdisplay.refresh()

    def find_down(self, event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        if self.find_results[self.results_number][1] == 6:
            for i in range(1, 6):
                self.sheetdisplay.highlight_cells(
                    row=self.find_results[self.results_number][0], column=i, bg="yellow", fg="black"
                )
        else:
            self.sheetdisplay.highlight_cells(
                row=self.find_results[self.results_number][0],
                column=self.find_results[self.results_number][1],
                bg="yellow",
                fg="black",
            )
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=f"{self.results_number + 1}/{len(self.find_results)}")
        if self.find_results[self.results_number][1] == 6:
            for i in range(1, 6):
                self.sheetdisplay.highlight_cells(
                    row=self.find_results[self.results_number][0], column=i, bg="orange", fg="black"
                )
        else:
            self.sheetdisplay.highlight_cells(
                row=self.find_results[self.results_number][0],
                column=self.find_results[self.results_number][1],
                bg="orange",
                fg="black",
            )
        self.sheetdisplay.see(row=self.find_results[self.results_number][0], column=0, keep_xscroll=True)
        self.sheetdisplay.refresh()

    def find_reset(self, newfind=False):
        self.find_results = []
        self.results_number = 0
        self.sheetdisplay.dehighlight_cells(all_=True, redraw=False)
        if not newfind:
            self.find_window.delete(0, "end")
        self.find_results_label.config(text="0/0")
        self.sheetdisplay.highlight_columns(columns=3, bg=self.red_bg, fg=self.red_fg)
        self.sheetdisplay.highlight_columns(columns=4, bg=self.green_bg, fg=self.green_fg)
        self.sheetdisplay.refresh()

    def cancel(self, event=None):
        self.USER_HAS_CLOSED_WINDOW()


class Compare_Report_Popup(tk.Toplevel):
    def __init__(self, C, width=1000, height=800, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Comparison Report", resizable=True)
        self.USER_HAS_QUIT = False
        self.protocol("WM_DELETE_WINDOW", self.USER_HAS_CLOSED_WINDOW)

        self.find_results = []
        self.results_number = 0
        self.wb_ = None
        self.sheet1name = self.C.sheetname_1
        self.sheet2name = self.C.sheetname_2

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.find_frame = Frame(self, theme=theme)
        self.find_frame.grid(row=0, column=0, columnspan=2, sticky="nswe")
        self.search_button = Button(self.find_frame, text=" Find:", command=self.find)
        self.search_button.pack(side="left", fill="x")
        self.find_window = Normal_Entry(self.find_frame, font=BF, theme=theme)
        self.find_window.bind("<Return>", self.find)
        self.find_window.pack(side="left", fill="x", expand=True)
        self.find_reset_button = Button(self.find_frame, text="X", command=self.find_reset)
        self.find_reset_button.pack(side="left", fill="x")
        self.find_results_label = Label(self.find_frame, "0/0", BF, theme=theme)
        self.find_results_label.pack(side="left", fill="x")
        self.find_up_button = Button(self.find_frame, text="▲", command=self.find_up)
        self.find_up_button.pack(side="left", fill="x")
        self.find_down_button = Button(self.find_frame, text="▼", command=self.find_down)
        self.find_down_button.pack(side="left", fill="x")

        self.sheetdisplay1 = Sheet(
            self,
            theme=theme,
            header_font=sheet_header_font,
            default_row_index=None,
            outline_thickness=1,
            default_column_width=250,
            display_selected_fg_over_highlights=True,
            treeview=True,
        )
        self.sheetdisplay1.enable_bindings(
            "single",
            "copy",
            "rc_popup_menu",
            "select_all",
            "drag_select",
            "column_width_resize",
            "double_click_column_resize",
            "row_height_resize",
            "double_click_row_resize",
            "row_select",
            "column_select",
            "arrowkeys",
            "ctrl_select",
        )
        self.sheetdisplay1.insert(text="", values=[self.C.report_header])
        self.sheetdisplay1.column_width(0, "text")
        for header_row, rows in self.C.report.items():
            row = len(self.sheetdisplay1.data)
            self.sheetdisplay1.insert(iid=header_row, text="Result", values=[header_row])
            self.sheetdisplay1.highlight(
                self.sheetdisplay1.span((row, 0)),
                bg="#648748",
                fg="white",
            )

            row = len(self.sheetdisplay1.data)
            self.sheetdisplay1.bulk_insert(data=rows, parent=header_row)
            if len(rows[0]) > 1:
                self.sheetdisplay1.highlight(
                    row,
                    bg="#0078d7",
                    fg="white",
                )

        self.sheetdisplay1.grid(row=1, column=0, sticky="nswe")

        self.buttonframe = Frame(self, theme=theme)
        self.buttonframe.grid(row=3, column=0, sticky="nswe")
        self.cancel_button = Button(self.buttonframe, text="Done", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.pack(side="right", padx=(20, 100), pady=20)
        self.save_text_button = Button(
            self.buttonframe,
            text="Save Report",
            style="EF.Std.TButton",
            command=self.save_report,
        )
        self.save_text_button.pack(side="right", padx=(50, 30), pady=20)

        self.bind("<Escape>", self.cancel)
        center(self, width, height)
        self.deiconify()
        self.wait_window()

    def start_work(self, msg=""):
        if msg:
            self.C.C.status_bar.change_text(msg)
        self.disable_widgets()

    def stop_work(self, msg=""):
        if msg:
            self.C.C.status_bar.change_text(msg)
        self.enable_widgets()

    def enable_widgets(self):
        self.sheetdisplay1.enable_bindings(
            "single",
            "copy",
            "rc_popup_menu",
            "select_all",
            "column_width_resize",
            "double_click_column_resize",
            "row_height_resize",
            "double_click_row_resize",
            "row_select",
            "column_select",
            "arrowkeys",
            "ctrl_select",
        )
        self.find_window.bind("<Return>", self.find)
        self.find_reset_button.config(state="normal")
        self.find_up_button.config(state="normal")
        self.find_down_button.config(state="normal")
        self.save_text_button.config(state="normal")

    def disable_widgets(self):
        self.sheetdisplay1.disable_bindings()
        self.find_window.unbind("<Return>")
        self.find_reset_button.config(state="disabled")
        self.find_up_button.config(state="disabled")
        self.find_down_button.config(state="disabled")
        self.save_text_button.config(state="disabled")
        self.update()

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except Exception:
            pass
        try:
            self.wb_ = None
        except Exception:
            pass

    def USER_HAS_CLOSED_WINDOW(self, callback=None):
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except Exception:
            pass
        self.destroy()

    def save_report(self):
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(
            parent=self,
            title="Save as",
            filetypes=[("Excel file", ".xlsx")],
            defaultextension=".xlsx",
            confirmoverwrite=True,
        )
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith(".xlsx"):
            self.grab_set()
            self.stop_work("Can only save .xlsx file type")
            return
        white_font = Font(color="00FFFFFF")
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook(write_only=True)
                ws = self.wb_.create_sheet(title="Report")
                row_ctr = 3
                for header_row, rows in self.C.report.items():
                    ws.row_dimensions.group(
                        row_ctr,
                        row_ctr + len(rows) - 1,
                        outline_level=1,
                        hidden=True,
                    )
                    row_ctr += 1 + len(rows)
                ws.append([self.C.report_header])
                for header_row, rows in self.C.report.items():
                    cell = WriteOnlyCell(ws, value=header_row)
                    cell.fill = green_fill
                    cell.font = white_font
                    ws.append([cell])
                    if len(rows[0]) > 1:
                        row = []
                        for val in rows[0]:
                            cell = WriteOnlyCell(ws, value=val)
                            cell.fill = blue_fill
                            cell.font = white_font
                            row.append(cell)
                        ws.append(row)
                        for row in islice(rows, 1, None):
                            ws.append((e if e != "" else None for e in row))
                    else:
                        for row in rows:
                            ws.append((e if e != "" else None for e in row))
                self.wb_.save(newfile)
                self.try_to_close_wb()
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work(f"Error saving file: {error_msg}")
            return
        self.stop_work("Success! Report saved")

    def go_to_result(self, row: int) -> None:
        self.sheetdisplay1.scroll_to_item(self.sheetdisplay1.rowitem(row, data_index=True))

    def find(self, event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        x = self.word.lower()

        for rn, row in enumerate(self.sheetdisplay1.get_sheet_data()):
            for colno, cell in enumerate(row):
                if x in cell.lower():
                    self.find_results.append((rn, colno))
        if self.find_results:
            for rn, colno in islice(self.find_results, 1, len(self.find_results)):
                self.sheetdisplay1.highlight_cells(row=rn, column=colno, bg="yellow", fg="black")
            self.sheetdisplay1.highlight_cells(
                row=self.find_results[self.results_number][0],
                column=self.find_results[self.results_number][1],
                bg="orange",
                fg="black",
            )
            self.find_results_label.config(text=f"1/{len(self.find_results)}")
            self.go_to_result(self.find_results[0][0])
        self.sheetdisplay1.refresh()

    def find_up(self, event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.sheetdisplay1.highlight_cells(
            row=self.find_results[self.results_number][0],
            column=self.find_results[self.results_number][1],
            bg="yellow",
            fg="black",
        )
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=f"{self.results_number + 1}/{len(self.find_results)}")
        self.sheetdisplay1.highlight_cells(
            row=self.find_results[self.results_number][0],
            column=self.find_results[self.results_number][1],
            bg="orange",
            fg="black",
        )
        self.go_to_result(self.find_results[self.results_number][0])
        self.sheetdisplay1.refresh()

    def find_down(self, event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.sheetdisplay1.highlight_cells(
            row=self.find_results[self.results_number][0],
            column=self.find_results[self.results_number][1],
            bg="yellow",
            fg="black",
        )
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=f"{self.results_number + 1}/{len(self.find_results)}")
        self.sheetdisplay1.highlight_cells(
            row=self.find_results[self.results_number][0],
            column=self.find_results[self.results_number][1],
            bg="orange",
            fg="black",
        )
        self.go_to_result(self.find_results[self.results_number][0])
        self.sheetdisplay1.refresh()

    def find_reset(self, newfind=False):
        try:
            self.find_results = []
            self.results_number = 0
            self.sheetdisplay1.dehighlight_cells(all_=True, redraw=True)
            self.sheetdisplay2.dehighlight_cells(all_=True, redraw=True)
            if not newfind:
                self.find_window.delete(0, "end")
            self.find_results_label.config(text="0/0")
        except Exception:
            pass

    def cancel(self, event=None):
        self.USER_HAS_CLOSED_WINDOW()


class Find_And_Replace_Popup(tk.Toplevel):
    def __init__(self, C, theme="dark", within=False, pars=False):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.withdraw()
        self.resizable(False, False)
        self.tk.call("wm", "iconphoto", self._w, tk.PhotoImage(format="gif", data=top_left_icon))
        self.title(f"{app_title} - Find & replace")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grid_columnconfigure(0, weight=1)
        self.protocol("WM_DELETE_WINDOW", self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.theme = theme
        self.starting_up = True

        self.notebook = ttk.Notebook(self)
        self.notebook.grid(row=0, column=0, sticky="nswe")

        self.f2 = Frame(self, theme=theme)
        self.f2.grid_columnconfigure(0, weight=1)
        self.f2.grid_columnconfigure(1, weight=1)
        self.notebook.add(self.f2, text="Find & Replace")

        self.f3 = Frame(self, theme=theme)
        # self.f3.grid_columnconfigure(0,weight=1)
        self.f3.grid_columnconfigure(1, weight=1)
        self.f3.grid_rowconfigure(2, weight=1)
        self.notebook.add(self.f3, text="Replace using mapping")

        self.notebook.select(self.f2)
        self.notebook.enable_traversal()
        self.notebook.bind("<<NotebookTabChanged>>", self.notebook_tab_click)

        self.frframe = Frame(self.f2, theme=theme)
        # self.frframe.grid_columnconfigure(1, weight=1)
        self.frframe.grid(row=0, column=1, pady=(10, 0), sticky="nswe")

        self.find_label = Label(self.frframe, text="Find", font=EF, theme=theme, anchor="w")
        self.find_label.grid(row=0, column=0, sticky="nswe", pady=(0, 14), padx=(20, 10))
        self.find_display = Entry_With_Scrollbar(self.frframe, theme=theme)
        self.find_display.grid(row=0, column=1, sticky="nswe", pady=10, padx=(0, 20))

        self.rep_label = Label(self.frframe, text="Replace with", font=EF, theme=theme, anchor="w")
        self.rep_label.grid(row=1, column=0, sticky="nswe", pady=(0, 17), padx=(20, 10))
        self.rep_display = Entry_With_Scrollbar(self.frframe, theme=theme)
        self.rep_display.grid(row=1, column=1, sticky="nswe", pady=10, padx=(0, 20))

        self.ids_button = X_Checkbutton(
            self.frframe,
            text="  Find IDs & Parents",
            style="wx_button.Std.TButton",
            checked=pars,
            compound="left",
        )
        self.ids_button.grid(row=2, column=1, padx=(30, 20), pady=5, sticky="we")

        self.details_button = X_Checkbutton(
            self.frframe,
            text="  Find Details",
            style="wx_button.Std.TButton",
            checked=True,
            compound="left",
        )
        self.details_button.grid(row=3, column=1, padx=(30, 20), pady=5, sticky="we")

        self.where = X_Checkbutton(
            self.frframe,
            text="  Find in Selection",
            style="wx_button.Std.TButton",
            checked=bool(within),
            compound="left",
        )
        self.where.grid(row=4, column=1, padx=(30, 20), pady=5, sticky="we")

        self.match_button = X_Checkbutton(
            self.frframe, text="  Exact Match", style="wx_button.Std.TButton", compound="left"
        )
        self.match_button.grid(row=5, column=1, padx=(30, 20), pady=(5, 30), sticky="we")

        self.bf2 = Frame(self.frframe, theme=theme)
        self.bf2.grid_columnconfigure(0, weight=1, uniform="x")
        self.bf2.grid_columnconfigure(1, weight=1, uniform="x")
        self.bf2.grid_columnconfigure(2, weight=1, uniform="x")
        self.bf2.grid_columnconfigure(3, weight=1, uniform="x")
        self.bf2.grid(row=7, column=0, columnspan=2, sticky="nswe")

        self.find_button = Button(self.bf2, text="Find next", style="EF.Std.TButton", command=self.find_next)
        self.find_button.grid(row=0, column=0, sticky="nswe", padx=(20, 5), pady=(15, 10))

        self.replace_button = Button(self.bf2, text="Replace next", style="EF.Std.TButton", command=self.replace_next)
        self.replace_button.grid(row=0, column=1, sticky="nswe", padx=5, pady=(15, 10))

        self.replace_all_button = Button(self.bf2, text="Replace all", style="EF.Std.TButton", command=self.replace_all)
        self.replace_all_button.grid(row=0, column=2, sticky="nswe", padx=5, pady=(15, 10))

        self.done_button = Button(self.bf2, text="Done", style="EF.Std.TButton", command=self.cancel)
        self.done_button.grid(row=0, column=3, sticky="nswe", padx=(5, 20), pady=(15, 10))

        self.status_bar = Readonly_Entry_With_Scrollbar(self, theme=theme, use_status_fg=True)
        self.status_bar.change_text(text="Search in current hierarchy or sheet, case insenitive")
        self.status_bar.my_entry.config(relief="flat", font=("Calibri", std_font_size))
        self.status_bar.grid(row=1, column=0, columnspan=2, sticky="we")

        self.open_file_display = Readonly_Entry_With_Scrollbar(self.f3, font=EF, theme=theme)
        self.open_file_display.grid(row=0, column=0, padx=2, pady=2, sticky="nswe")
        self.open_file_button = Button(
            self.f3, text="⯇ Open file", style="wx_button.Std.TButton", command=self.open_file
        )
        self.open_file_button.grid(row=0, column=1, padx=10, pady=10, sticky="nswe")
        self.sheet_dropdown = Ez_Dropdown(self.f3, font=EF)
        self.sheet_dropdown.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
        self.sheet_dropdown.grid(row=1, column=0, padx=10, pady=10, sticky="nswe")
        self.select_sheet_button = Button(
            self.f3, text="⯇ Load sheet", style="wx_button.Std.TButton", state="disabled", command=self.select_sheet
        )
        self.select_sheet_button.grid(row=1, column=1, padx=10, pady=10, sticky="nswe")

        self.sheetdisplay = Sheet(
            self.f3,
            theme=theme,
            headers=["Exact Find (case in-sensitive)", "Replace With"],
            # font=self.C.tree.font(),
            header_font=sheet_header_font,  # self.C.tree.header_font()
            expand_sheet_if_paste_too_big=True,
            paste_insert_column_limit=2,
            outline_thickness=1,
            data=[["", ""] for r in range(20)],
        )
        self.sheetdisplay.set_all_column_widths()
        self.sheetdisplay.enable_bindings("all", "ctrl_select")
        self.sheetdisplay.disable_bindings("rc_delete_column", "rc_insert_column")
        self.sheetdisplay.grid(row=2, column=0, sticky="ns")

        self.options_frame = Frame(self.f3, theme=theme)
        self.options_frame.grid(row=2, column=1, sticky="nswe")
        self.options_frame.grid_columnconfigure(0, weight=1)

        self.clipboard_button = Button(
            self.options_frame,
            text=" Get data from clipboard ",
            style="wx_button.Std.TButton",
            state="normal",
            command=self.get_clipboard_data,
        )
        self.clipboard_button.grid(row=0, column=0, padx=10, pady=(10, 20), sticky="nswe")

        self.where2 = X_Checkbutton(
            self.options_frame,
            text="  Find in Selection",
            style="wx_button.Std.TButton",
            checked=bool(within),
            compound="left",
        )
        self.where2.grid(row=1, column=0, padx=15, pady=(5, 30), sticky="we")

        self.confirm_button = Button(
            self.options_frame,
            text="Replace",
            style="EF.Std.TButton",
            command=lambda: self.replace_all(event="mapping"),
        )
        self.confirm_button.grid(row=3, column=0, padx=50, pady=(40, 5), sticky="we")

        self.find_display.my_entry.bind("<Return>", self.find_next)
        self.rep_display.my_entry.bind("<Return>", self.find_next)
        self.bind(f"<{ctrl_button}-g>", self.find_next)
        self.bind(f"<{ctrl_button}-G>", self.find_next)
        self.bind("<Escape>", self.cancel)
        self.bind(f"<{ctrl_button}-z>", self.C.undo)
        self.bind(f"<{ctrl_button}-Z>", self.C.undo)
        self.result = False
        center(self, 480, self.window_height(), move_left=True)
        self.deiconify()
        self.find_display.place_cursor()
        self.starting_up = False

    def window_height(self) -> int:
        return self.notebook.winfo_height() + self.status_bar.winfo_height()

    def notebook_tab_click(self, event=None):
        if not self.starting_up:
            if self.notebook.index(self.notebook.select()) != 1:
                self.geometry(f"480x{self.window_height()}")
            else:
                self.geometry(f"720x{self.window_height()}")
                self.sheetdisplay.MT.focus_set()

    def enable_widgets(self):
        self.bind(f"<{ctrl_button}-g>", self.find_next)
        self.bind(f"<{ctrl_button}-G>", self.find_next)
        self.bind("<Escape>", self.cancel)
        self.bind(f"<{ctrl_button}-z>", self.C.undo)
        self.bind(f"<{ctrl_button}-Z>", self.C.undo)
        self.open_file_display.change_my_state("readonly")
        self.open_file_button.config(state="normal")
        self.sheet_dropdown.config(state="readonly")
        self.where2.config(state="normal")
        self.confirm_button.config(state="normal")
        self.sheetdisplay.enable_bindings("all", "ctrl_select")
        self.sheetdisplay.disable_bindings("rc_delete_column", "rc_insert_column")

    def disable_widgets(self):
        self.unbind(f"<{ctrl_button}-g>")
        self.unbind(f"<{ctrl_button}-G>")
        self.unbind("<Escape>")
        self.unbind(f"<{ctrl_button}-z>")
        self.unbind(f"<{ctrl_button}-Z>")
        self.open_file_display.change_my_state("disabled")
        self.open_file_button.config(state="disabled")
        self.sheet_dropdown.config(state="disabled")
        self.select_sheet_button.config(state="disabled")
        self.where2.config(state="disabled")
        self.confirm_button.config(state="disabled")
        self.sheetdisplay.disable_bindings()
        self.update()

    def get_clipboard_data(self, event=None):
        self.start_work("Loading...")
        self.reset()
        try:
            temp_data = self.C.clipboard_get()
        except Exception as error_msg:
            self.stop_work(f"Error: Error getting data from clipboard: {error_msg}")
            return
        try:
            self.status_bar.change_text("Loading...")
            if temp_data.startswith("{") and temp_data.endswith("}"):
                self.C.new_sheet = json_to_sheet(json.loads(temp_data))
            else:
                self.C.new_sheet = csv_str_x_data(temp_data)
        except Exception as error_msg:
            self.stop_work(f"Error: Error transforming clipboard data: {error_msg}")
            return
        if not self.C.new_sheet:
            self.stop_work("Error: No data found on clipboard")
            return
        self.sheetdisplay.deselect("all")
        self.C.new_sheet = [r + list(repeat("", 2 - len(r))) if len(r) < 2 else r[:2] for r in self.C.new_sheet]
        self.sheetdisplay.data_reference(
            newdataref=self.C.new_sheet,
            reset_col_positions=True,
            reset_row_positions=True,
            redraw=False,
        )
        self.sheetdisplay.refresh()
        self.file_opened = "n/a - Data obtained from clipboard"
        self.sheet_opened = "n/a"
        self.C.new_sheet = []
        self.stop_work("Data successfully loaded from clipboard")

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except Exception:
            pass
        try:
            self.wb_ = None
        except Exception:
            pass

    def USER_HAS_CLOSED_WINDOW(self, callback=None):
        self.C.new_sheet = []
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except Exception:
            pass
        self.destroy()

    def open_file(self):
        self.start_work("Loading...   ")
        self.reset()
        filepath = filedialog.askopenfilename(parent=self, title="Select file")
        if not filepath:
            self.stop_work("Select a file")
            return
        try:
            filepath = os.path.normpath(filepath)
        except Exception:
            self.stop_work("Error: filepath invalid")
            return
        if not filepath.lower().endswith((".json", ".xlsx", ".xls", ".xlsm", ".csv", ".tsv")):
            self.stop_work("Error: select json/excel/csv   ")
            return
        check = os.path.isfile(filepath)
        if not check:
            self.stop_work("Error: filepath invalid")
            return
        try:
            self.status_bar.change_text("Loading...")
            if filepath.lower().endswith((".csv", ".tsv")):
                with open(filepath, "r") as fh:
                    temp_data = fh.read()
                self.C.new_sheet = [
                    r + list(repeat("", 2 - len(r))) if len(r) < 2 else r[:2] for r in csv_str_x_data(temp_data)
                ]
            elif filepath.lower().endswith(".json"):
                j = get_json_from_file(filepath)
                json_format = get_json_format(j)
                if not json_format:
                    self.C.new_sheet = []
                    self.stop_work("Error opening file, could not find data of correct format")
                    return
                self.C.new_sheet = json_to_sheet(
                    j,
                    format_=json_format[0],
                    key=json_format[1],
                    get_format=False,
                    return_rowlen=False,
                )
                if not self.C.new_sheet:
                    self.stop_work("Error: File contained no data")
                    self.select_sheet_button.config(state="disabled")
                    return
                self.C.new_sheet = [r + list(repeat("", 2 - len(r))) if len(r) < 2 else r[:2] for r in self.C.new_sheet]
                self.stop_work("Ready to merge sheets")
            elif filepath.lower().endswith((".xlsx", ".xls", ".xlsm")):
                in_mem = bytes_io_wb(filepath)
                self.wb_ = load_workbook(in_mem, read_only=True, data_only=True)
                wbsheets = self.wb_.sheetnames
                if not wbsheets:
                    self.stop_work("Error: File/sheet contained no data")
                    return
                self.sheet_dropdown["values"] = wbsheets
                self.sheet_dropdown.set_my_value(wbsheets[0])
                self.stop_work("Select a sheet to open")
                self.select_sheet_button.config(state="normal")
        except Exception as error_msg:
            self.try_to_close_wb()
            self.C.new_sheet = []
            self.stop_work(f"Error: {error_msg}")
            return
        if not self.C.new_sheet and not filepath.lower().endswith((".xlsx", ".xls", ".xlsm")):
            self.C.new_sheet = []
            self.stop_work("Error: File/sheet contained no data")
            return
        self.sheetdisplay.data_reference(
            newdataref=self.C.new_sheet,
            reset_col_positions=True,
            reset_row_positions=True,
            redraw=True,
        )
        self.open_file_display.set_my_value(filepath)
        self.file_opened = os.path.basename(self.open_file_display.get_my_value())
        self.C.new_sheet = []
        self.stop_work(f"Data successfully loaded from {self.file_opened}")

    def select_sheet(self):
        self.start_work("Loading...")
        self.sheet_opened = self.sheet_dropdown.get_my_value()
        ws = self.wb_[self.sheet_opened]
        ws.reset_dimensions()
        self.C.new_sheet = ws_x_data(ws)
        self.try_to_close_wb()
        if not self.C.new_sheet:
            self.stop_work("Error: File/sheet contained no data")
            self.select_sheet_button.config(state="disabled")
            return
        self.C.new_sheet = [r + list(repeat("", 2 - len(r))) if len(r) < 2 else r[:2] for r in self.C.new_sheet]
        self.select_sheet_button.config(state="disabled")
        self.sheetdisplay.data_reference(
            newdataref=self.C.new_sheet,
            reset_col_positions=True,
            reset_row_positions=True,
            redraw=True,
        )
        self.C.new_sheet = []
        self.stop_work(f"Loaded sheet: {self.sheet_opened}")

    def start_work(self, msg=""):
        if msg is not None:
            self.status_bar.change_text(msg)
        self.C.stop_work(self.C.get_tree_editor_status_bar_text())
        self.disable_widgets()

    def stop_work(self, msg=""):
        if msg is not None:
            self.status_bar.change_text(msg)
        self.C.stop_work(self.C.get_tree_editor_status_bar_text())
        self.enable_widgets()

    def reset(self):
        self.try_to_close_wb()
        self.C.new_sheet = []
        self.open_file_display.set_my_value("")
        self.sheet_dropdown["values"] = []
        self.sheet_dropdown.set("")
        self.select_sheet_button.config(state="disabled")

    def sheet_see_and_set(self, row, column, just_see: bool = False):
        """
        Uses data indexes
        """
        self.C.sheet.see(
            row=row,
            column=column,
            keep_yscroll=False,
            keep_xscroll=False,
            bottom_right_corner=False,
            check_cell_visibility=True,
        )
        if not just_see:
            if self.where.get_checked():
                self.C.sheet.set_currently_selected(row, column)
            else:
                self.C.sheet.select_cell(row=row, column=column)
        return row, column

    def tree_see_and_set(self, row, column, just_see=False):
        """
        Uses data indexes
        """
        self.C.tree.scroll_to_item(self.C.tree.rowitem(row, data_index=True))
        self.C.tree.see(
            row=None,
            column=column,
            keep_yscroll=True,
            keep_xscroll=False,
            bottom_right_corner=False,
            check_cell_visibility=True,
        )
        if not just_see:
            if self.where.get_checked():
                self.C.tree.set_currently_selected(self.C.tree.disprn(row), column)
            else:
                self.C.tree.select_cell(row=self.C.tree.disprn(row), column=column)
        return row, column

    def gen_all_cells(self, start_row, start_col, widget):
        return chain(
            (
                (r, c)
                for r, c in islice(
                    ((r, c) for r in range(0, len(widget.data)) for c in range(self.C.row_len)),
                    (start_row * self.C.row_len) + start_col,
                    None,
                )
            ),
            (
                (r, c)
                for r, c in islice(
                    ((r, c) for r in range(0, len(widget.data)) for c in range(self.C.row_len)),
                    0,
                    (start_row * self.C.row_len) + start_col,
                )
            ),
        )

    def get_start_coords(
        self,
        widget,
        plus_one: bool = True,
        within: None | list = None,
    ) -> tuple[int, int, int]:
        selected = widget.selected
        if not selected:
            rst, cst = 0, 0
        else:
            rst, cst = selected.row, selected.column
            if plus_one and within:
                curridx = next(i for i, t in enumerate(within) if t[0] == rst and t[1] == cst) + 1
                if curridx == len(within):
                    curridx = 0
                return rst, cst, curridx
            elif plus_one:
                cst += 1
                if cst == self.C.row_len:
                    cst = 0
                    rst += 1
                if rst == len(widget.data):
                    rst = 0
        return rst, cst, 0

    def find_next(self, event=None, set_status_bar: bool = True):
        if not self.C.sheet.data:
            return
        if self.C.sheet_has_focus:
            return self.sheet_find_next(event, set_status_bar=set_status_bar)
        return self.tree_find_next(event, set_status_bar=set_status_bar)

    def sheet_find_next(self, event=None, set_status_bar: bool = True):
        ids = self.ids_button.get_checked()
        dets = self.details_button.get_checked()
        if not ids and not dets:
            self.status_bar.change_text("Select a search option, IDs and Parents and/or Details")
            return
        search = self.find_display.get_my_value().lower()
        match = self.match_button.get_checked()
        ind = set(self.C.hiers) | {self.C.ic}
        where = self.where.get_checked()
        data = self.C.sheet.MT.data

        found_coords = tuple()
        if where:
            sels = self.C.sheet.get_selected_cells(
                get_rows=True,
                get_columns=True,
                sort_by_row=True,
                sort_by_column=True,
            )
            rst, cst, curridx = self.get_start_coords(widget=self.C.sheet, plus_one=True, within=sels)
            for r, c in chain(islice(sels, curridx, None), islice(sels, 0, curridx)):
                cell = data[r][c]
                if (
                    ids and c in ind and ((match and cell.lower() == search) or (not match and search in cell.lower()))
                ) or (
                    dets
                    and c not in ind
                    and ((match and cell.lower() == search) or (not match and search in cell.lower()))
                ):
                    found_coords = (r, c)
                    break
        else:
            rst, cst, _ = self.get_start_coords(self.C.sheet, plus_one=True)
            for r, c in self.gen_all_cells(start_row=rst, start_col=cst, widget=self.C.sheet):
                cell = data[r][c]
                if (
                    ids and c in ind and ((match and cell.lower() == search) or (not match and search in cell.lower()))
                ) or (
                    dets
                    and c not in ind
                    and ((match and cell.lower() == search) or (not match and search in cell.lower()))
                ):
                    found_coords = (r, c)
                    break

        if found_coords:
            self.sheet_see_and_set(*found_coords)

        if set_status_bar:
            if found_coords:
                self.status_bar.change_text(
                    f"Found {self.find_display.get_my_value()} for {data[found_coords[0]][self.C.ic]} in {self.C.headers[found_coords[1]].name}"
                )
            else:
                self.status_bar.change_text(f"Could not find {self.find_display.get_my_value()}")
        return found_coords

    def tree_find_next(self, event=None, set_status_bar: bool = True):
        if not self.C.sheet.data:
            return
        ids = self.ids_button.get_checked()
        dets = self.details_button.get_checked()
        if not ids and not dets:
            self.status_bar.change_text("Select a search option, IDs and Parents and/or Details")
            return
        search = self.find_display.get_my_value().lower()
        match = self.match_button.get_checked()
        ind = set(self.C.hiers) | {self.C.ic}
        where = self.where.get_checked()
        data = self.C.tree.MT.data
        datarn = self.C.tree.data_r

        found_coords = tuple()
        if where:
            sels = self.C.tree.get_selected_cells(
                get_rows=True, get_columns=True, sort_by_row=True, sort_by_column=True
            )
            rst, cst, curridx = self.get_start_coords(self.C.tree, plus_one=True, within=sels)
            for r, c in chain(islice(sels, curridx, None), islice(sels, 0, curridx)):
                r = datarn(r)
                cell = data[r][c]
                if (
                    ids and c in ind and ((match and cell.lower() == search) or (not match and search in cell.lower()))
                ) or (
                    dets
                    and c not in ind
                    and ((match and cell.lower() == search) or (not match and search in cell.lower()))
                ):
                    found_coords = (r, c)
                    break
        else:
            rst, cst, _ = self.get_start_coords(self.C.tree, plus_one=True)
            rst = datarn(rst)
            for r, c in self.gen_all_cells(start_row=rst, start_col=cst, widget=self.C.tree):
                cell = data[r][c]
                if (
                    ids and c in ind and ((match and cell.lower() == search) or (not match and search in cell.lower()))
                ) or (
                    dets
                    and c not in ind
                    and ((match and cell.lower() == search) or (not match and search in cell.lower()))
                ):
                    found_coords = (r, c)
                    break

        if found_coords:
            self.tree_see_and_set(*found_coords)

        if set_status_bar:
            if found_coords:
                self.status_bar.change_text(
                    f"Found {self.find_display.get_my_value()} for {data[found_coords[0]][self.C.ic]} in {self.C.headers[found_coords[1]].name}"
                )
            else:
                self.status_bar.change_text(f"Could not find {self.find_display.get_my_value()}")
        return found_coords

    def replace_next(self, event=None):
        if not self.C.sheet.data:
            return
        if self.C.sheet_has_focus:
            return self.sheet_replace_next(event)
        return self.tree_replace_next(event)

    def sheet_replace_next(self, event=None):
        if not self.C.sheet.data:
            return
        ids = self.ids_button.get_checked()
        dets = self.details_button.get_checked()
        search = self.find_display.get_my_value().lower()
        newtext = self.rep_display.get_my_value()
        if search == newtext:
            self.status_bar.change_text("Error: Find value is the same as replace value")
            return

        self.start_work("Replacing...")

        allow_spaces = self.C.allow_spaces_ids_var
        if allow_spaces:
            id_par_newtext = newtext
        else:
            id_par_newtext = re.sub(r"[\n\t\s]*", "", newtext)
        match = self.match_button.get_checked()
        where = self.where.get_checked()
        ind = set(self.C.hiers) | {self.C.ic}
        data = self.C.sheet.MT.data

        rst, cst, _ = self.get_start_coords(self.C.sheet, plus_one=False)
        found_coords = tuple()

        cell = data[rst][cst]
        cell_k = cell.lower()
        if (
            ids
            and cst in ind
            and (
                (match and cell_k == search and cell_k != id_par_newtext)
                or (
                    not match
                    and search in cell_k
                    and cell_k != case_insensitive_replace(find_=search, repl=id_par_newtext, text=cell_k)
                )
            )
        ) or (
            dets
            and cst not in ind
            and (
                (match and cell_k == search and cell != newtext)
                or (
                    not match
                    and search in cell_k
                    and cell != case_insensitive_replace(find_=search, repl=newtext, text=cell)
                )
            )
        ):
            found_coords = (rst, cst)

        if not found_coords:
            if where:
                sels = self.C.sheet.get_selected_cells(
                    get_rows=True, get_columns=True, sort_by_row=True, sort_by_column=True
                )
                rst, cst, curridx = self.get_start_coords(self.C.sheet, plus_one=True, within=sels)
                iterable = chain(islice(sels, curridx, None), islice(sels, 0, curridx))
            else:
                rst, cst, _ = self.get_start_coords(self.C.sheet, plus_one=True)
                iterable = self.gen_all_cells(start_row=rst, start_col=cst, widget=self.C.sheet)

            for r, c in iterable:
                cell = data[r][c]
                cell_k = cell.lower()
                if (
                    ids
                    and c in ind
                    and (
                        (match and cell_k == search and cell_k != id_par_newtext)
                        or (
                            not match
                            and search in cell_k
                            and cell_k != case_insensitive_replace(find_=search, repl=id_par_newtext, text=cell_k)
                        )
                    )
                ) or (
                    dets
                    and c not in ind
                    and (
                        (match and cell_k == search and cell != newtext)
                        or (
                            not match
                            and search in cell_k
                            and cell != case_insensitive_replace(find_=search, repl=newtext, text=cell)
                        )
                    )
                ):
                    found_coords = (r, c)
                    break

        if not found_coords:
            self.stop_work(
                f"Could not find a cell containing {self.find_display.get_my_value()} to replace with {self.rep_display.get_my_value()}"
            )
            return

        self.sheet_see_and_set(*found_coords)

        r, c = found_coords
        if not match:
            newtext = case_insensitive_replace(search, newtext, self.C.sheet.MT.data[r][c])
        event = self.C.sheet.new_tksheet_event()
        event.value = newtext
        event.row = r
        event.column = c
        event.loc = (r, c)
        old_value = f"{self.C.sheet.MT.data[r][c]}"
        if isinstance(self.C.tree_sheet_edit_cell(event=event), str):
            self.stop_work(f"Replaced {old_value} with {newtext}")
        else:
            self.stop_work(f"Failed to replace {old_value} with {newtext}")

    def tree_replace_next(self, event=None):
        if not self.C.sheet.data:
            return
        ids = self.ids_button.get_checked()
        dets = self.details_button.get_checked()
        search = self.find_display.get_my_value().lower()
        newtext = self.rep_display.get_my_value()
        if search == newtext:
            self.status_bar.change_text("Error: Find value is the same as replace value")
            return

        self.start_work("Replacing...")

        allow_spaces = self.C.allow_spaces_ids_var
        if allow_spaces:
            id_par_newtext = newtext
        else:
            id_par_newtext = re.sub(r"[\n\t\s]*", "", newtext)
        match = self.match_button.get_checked()
        where = self.where.get_checked()
        ind = set(self.C.hiers) | {self.C.ic}
        data = self.C.tree.MT.data
        datarn = self.C.tree.data_r

        rst, cst, _ = self.get_start_coords(self.C.tree, plus_one=False)
        rst = datarn(rst)
        found_coords = tuple()

        cell = data[rst][cst]
        cell_k = cell.lower()
        if (
            ids
            and cst in ind
            and (
                (match and cell_k == search and cell_k != id_par_newtext)
                or (
                    not match
                    and search in cell_k
                    and cell_k != case_insensitive_replace(find_=search, repl=id_par_newtext, text=cell_k)
                )
            )
        ) or (
            dets
            and cst not in ind
            and (
                (match and cell_k == search and cell != newtext)
                or (
                    not match
                    and search in cell_k
                    and cell != case_insensitive_replace(find_=search, repl=newtext, text=cell)
                )
            )
        ):
            found_coords = (rst, cst)

        if not found_coords:
            if where:
                sels = self.C.tree.get_selected_cells(
                    get_rows=True, get_columns=True, sort_by_row=True, sort_by_column=True
                )
                rst, cst, curridx = self.get_start_coords(self.C.tree, plus_one=True, within=sels)
                iterable = chain(islice(sels, curridx, None), islice(sels, 0, curridx))
            else:
                rst, cst, _ = self.get_start_coords(self.C.tree, plus_one=True)
                rst = datarn(rst)
                iterable = self.gen_all_cells(start_row=rst, start_col=cst, widget=self.C.tree)

            for r, c in iterable:
                if where:
                    r = datarn(r)
                cell = data[r][c]
                cell_k = cell.lower()
                if (
                    ids
                    and c in ind
                    and (
                        (match and cell_k == search and cell_k != id_par_newtext)
                        or (
                            not match
                            and search in cell_k
                            and cell_k != case_insensitive_replace(find_=search, repl=id_par_newtext, text=cell_k)
                        )
                    )
                ) or (
                    dets
                    and c not in ind
                    and (
                        (match and cell_k == search and cell != newtext)
                        or (
                            not match
                            and search in cell_k
                            and cell != case_insensitive_replace(find_=search, repl=newtext, text=cell)
                        )
                    )
                ):
                    found_coords = (r, c)
                    break

        if not found_coords:
            self.stop_work(
                f"Could not find a cell containing {self.find_display.get_my_value()} to replace with {self.rep_display.get_my_value()}"
            )
            return

        self.tree_see_and_set(*found_coords)

        r, c = found_coords
        r = self.C.rns[data[r][self.C.ic].lower()]
        if not match:
            newtext = case_insensitive_replace(search, newtext, self.C.sheet.MT.data[r][c])
        event = self.C.sheet.new_tksheet_event()
        event.value = newtext
        event.row = r
        event.column = c
        event.loc = (r, c)
        old_value = f"{self.C.sheet.MT.data[r][c]}"
        if isinstance(self.C.tree_sheet_edit_cell(event=event), str):
            self.stop_work(f"Replaced {old_value} with {newtext}")
        else:
            self.stop_work(f"Failed to replace {old_value} with {newtext}")

    def get_cells(self, where=True, widget=None):
        if not widget:
            widget = self.C.sheet if self.C.sheet_has_focus else self.C.tree

        if widget == self.C.sheet:
            if where:
                return self.C.sheet.get_selected_cells(
                    get_rows=True,
                    get_columns=True,
                    sort_by_row=True,
                    sort_by_column=True,
                )
            else:
                return [(r, c) for r, row in enumerate(self.C.sheet.data) for c in range(len(row))]
        else:
            ic = self.C.ic
            tree_data = self.C.tree.MT.data
            rns = self.C.rns
            if where:
                return sorted(
                    (
                        (rns[tree_data[r][ic].lower()], c)
                        for r, c in self.C.tree.get_selected_cells(
                            get_rows=True,
                            get_columns=True,
                            sort_by_row=True,
                            sort_by_column=True,
                        )
                    ),
                    key=itemgetter(0),
                )
            else:
                return sorted(
                    (
                        (rns[tree_data[r][ic].lower()], c)
                        for r, row in enumerate(self.C.tree.data)
                        for c in range(len(row))
                    ),
                    key=itemgetter(0),
                )

    def replace_all(self, event=None):
        if not self.C.sheet.data:
            return

        successful = set()
        if event == "mapping":
            ids = True
            dets = True
            exact_match = True
            mapping = {r[0].lower(): r[1] for r in self.sheetdisplay.get_sheet_data() if r[0] and r[1]}
            newtext = ""

        else:
            ids = self.ids_button.get_checked()
            dets = self.details_button.get_checked()
            if not ids and not dets:
                self.status_bar.change_text("Select a search option, IDs and Parents and/or Details")
                return
            search = self.find_display.get_my_value().lower()
            newtext = self.rep_display.get_my_value()
            if search == newtext:
                self.status_bar.change_text("Error: Find value is the same as replace value")
                return
            exact_match = self.match_button.get_checked()
            valids = {c: self.C.detail_is_valid_for_col(c, newtext) for c in range(len(self.C.headers))}

        self.start_work("Replacing...")
        widget = self.C.sheet if self.C.sheet_has_focus else self.C.tree
        where = self.where.get_checked()
        ind = set(self.C.hiers) | {self.C.ic}
        qic = self.C.ic

        ids_changed = 0
        details_changed = 0
        if ids:
            self.C.snapshot_ctrl_x_v_del_key_id_par()
        else:
            self.C.snapshot_ctrl_x_v_del_key()

        refresh_rows = set()
        newtext2 = ""
        if event == "mapping":
            failed_name_changes = set()
        allow_spaces = self.C.allow_spaces_ids_var
        sheet_data = self.C.sheet.MT.data

        for r, c in self.get_cells(where=where, widget=widget):
            do_replace = False
            cell = sheet_data[r][c]
            cell_k = cell.lower()
            if c in ind and ids:
                if exact_match:
                    if event == "mapping":
                        if cell_k in mapping and cell_k != mapping[cell_k].lower():
                            if not allow_spaces:
                                if re.sub(r"[\n\t\s]*", "", mapping[cell_k]):
                                    newtext2 = re.sub(r"[\n\t\s]*", "", mapping[cell_k])
                                    do_replace = True
                            else:
                                newtext2 = mapping[cell_k]
                                do_replace = True

                    elif cell_k == search and cell_k != newtext:
                        newtext2 = newtext
                        do_replace = True

                    if do_replace:
                        old_id = f"{sheet_data[r][self.C.ic]}"
                        self.C.changelog_append_no_unsaved(
                            "Edit cell |",
                            f"ID: {old_id} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                            f"{cell}",
                            newtext2,
                        )
                        if c == qic and self.C.change_ID_name(f"{old_id}", newtext2, snapshot=False, errors=False):
                            ik = old_id.lower()
                            new_ik = newtext2.lower()
                            if ik in self.C.tagged_ids:
                                self.C.tagged_ids.discard(ik)
                                self.C.tagged_ids.add(new_ik)
                            successful.add(cell_k)
                            ids_changed += 1
                        elif c == qic:
                            failed_name_changes.add(cell_k)
                        else:
                            sheet_data[r][c] = newtext2
                            successful.add(cell_k)
                            ids_changed += 1
                        do_replace = False

                elif not exact_match and search in cell_k and cell != newtext:
                    if not allow_spaces:
                        newtext2 = case_insensitive_replace(search, re.sub(r"[\n\t\s]*", "", newtext), cell)
                    else:
                        newtext2 = case_insensitive_replace(search, newtext, cell)
                    old_id = f"{sheet_data[r][self.C.ic]}"
                    self.C.changelog_append_no_unsaved(
                        "Edit cell |",
                        f"ID: {old_id} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                        f"{cell}",
                        newtext2,
                    )
                    if c == qic and self.C.change_ID_name(f"{old_id}", newtext2, snapshot=False, errors=False):
                        ik = old_id.lower()
                        new_ik = newtext2.lower()
                        if ik in self.C.tagged_ids:
                            self.C.tagged_ids.discard(ik)
                            self.C.tagged_ids.add(new_ik)
                    else:
                        sheet_data[r][c] = newtext2
                    ids_changed += 1

            elif c not in ind and dets:
                if exact_match:
                    if event == "mapping":
                        if (
                            cell_k in mapping
                            and cell_k != mapping[cell_k].lower()
                            and self.C.detail_is_valid_for_col(c, mapping[cell_k])
                        ):
                            newtext2 = mapping[cell_k]
                            successful.add(cell_k)
                            do_replace = True

                    elif cell_k == search and cell_k != newtext2 and valids[c]:
                        newtext2 = newtext
                        do_replace = True

                    if do_replace:
                        refresh_rows.add(r)
                        if not ids:
                            self.C.vs[-1]["cells"][(r, c)] = f"{cell}"
                        self.C.changelog_append_no_unsaved(
                            "Edit cell |",
                            f"ID: {sheet_data[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                            f"{cell}",
                            newtext2,
                        )
                        if self.C.headers[c].type_ == "Date Detail":
                            sheet_data[r][c] = self.C.convert_date(newtext2, self.C.DATE_FORM)
                        else:
                            sheet_data[r][c] = newtext2
                        details_changed += 1

                elif (
                    not exact_match
                    and search in cell_k
                    and cell_k != newtext2
                    and self.C.detail_is_valid_for_col(c, case_insensitive_replace(search, newtext, cell))
                ):
                    refresh_rows.add(r)
                    if not ids:
                        self.C.vs[-1]["cells"][(r, c)] = f"{cell}"
                    newtext2 = case_insensitive_replace(search, newtext, cell)
                    self.C.changelog_append_no_unsaved(
                        "Edit cell |",
                        f"ID: {sheet_data[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                        f"{cell}",
                        newtext2,
                    )
                    if self.C.headers[c].type_ == "Date Detail":
                        sheet_data[r][c] = self.C.convert_date(newtext2, self.C.DATE_FORM)
                    else:
                        sheet_data[r][c] = newtext2
                    details_changed += 1

        total_changed = ids_changed + details_changed

        if not total_changed:
            self.C.vp -= 1
            self.C.set_undo_label()
            self.C.vs.pop()
            self.C.redraw_sheets()
            if event == "mapping":
                self.stop_work("There were no successful replacements")
                Error(self, "There were no successful replacements", theme=self.theme)
            else:
                self.stop_work(
                    f"Could not find a cell containing {self.find_display.get_my_value()} to replace with {self.rep_display.get_my_value()}"
                )
            return

        if ids_changed:
            self.C.disable_paste()
            self.C.rebuild_tree(deselect=False)

        elif details_changed:
            self.C.disable_paste()
            self.C.refresh_all_formatting(rows=refresh_rows)
            for rn in refresh_rows:
                self.C.refresh_tree_item(sheet_data[rn][self.C.ic])
            self.C.redraw_sheets()

        if total_changed > 1:
            self.C.changelog_append(f"Edit {total_changed} cells", "", "", "")
        else:
            self.C.changelog_singular("Edit cell")

        if event == "mapping" and (
            num_unsuccessful := sum(1 for k in mapping if k not in successful or k in failed_name_changes)
        ):
            self.stop_work(f"Sucessfully replaced {total_changed} cells, {num_unsuccessful} unsuccessful")
            if failed_name_changes:
                failed_data = (
                    [["Failed to replace the following IDs"]]
                    + [[cell_k, mapping[cell_k]] for cell_k in failed_name_changes]
                    + [["Full list of failed replacements:"]]
                )
            else:
                failed_data = [["Full list of failed replacements:"]]
            failed_data += [[cell_k, mapping[cell_k]] for cell_k in mapping if cell_k not in successful]
            Error_Sheet(
                self,
                failed_data,
                theme=self.theme,
                highlight_rows=(0, len(failed_name_changes) + 1) if failed_name_changes else (0,),
            )

        elif event == "mapping" and not num_unsuccessful:
            self.stop_work(f"Sucessfully replaced {total_changed} cells")

        elif event != "mapping":
            self.stop_work(
                f"Replaced {total_changed} cells containing {self.find_display.get_my_value()} with {self.rep_display.get_my_value()}"
            )

    def cancel(self, event=None):
        self.destroy()


class Edit_Conditional_Formatting_Popup(tk.Toplevel):
    def __init__(self, C, column, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.window_destroyed = False
        self.C = new_toplevel_chores(self, C, f"{app_title} - Edit conditional formatting")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.column = column
        self.displayed_colors_dct = {
            "Yellow": "yellow",
            "Red, normal": "firebrick1",
            "Brown": "#734021",
            "Orange": "orange",
            "Green, bright": "lawn green",
            "Green, dark": "forest green",
            "Red, bright": "red",
            "Turquoise": "turquoise",
            "Purple": "DarkOrchid1",
            "Pink": "orchid1",
            "Red, soft": "salmon1",
            "Blue, bright": "cyan",
            "Scale 1 (green)": "#509f56",
            "Scale 2": "#64a85b",
            "Scale 3": "#78b160",
            "Scale 4": "#8cba66",
            "Scale 5": "#a0c36c",
            "Scale 6": "#b4cc71",
            "Scale 7": "#c8d576",
            "Scale 8": "#dcde7c",
            "Scale 9": "#f0e782",
            "Scale 10 (yellow)": "#ffec87",
            "Scale 11": "#ffe182",
            "Scale 12": "#ffdc7d",
            "Scale 13": "#ffd77b",
            "Scale 14": "#ffc873",
            "Scale 15": "#ffb469",
            "Scale 16": "#fea05f",
            "Scale 17": "#fc8c55",
            "Scale 18": "#fb784b",
            "Scale 19": "#fa6441",
            "Scale 20 (red)": "#f85037",
        }
        self.scale_colors = (
            "#509f56",
            "#64a85b",
            "#78b160",
            "#8cba66",
            "#a0c36c",
            "#b4cc71",
            "#c8d576",
            "#dcde7c",
            "#f0e782",
            "#ffec87",
            "#ffe182",
            "#ffdc7d",
            "#ffd77b",
            "#ffc873",
            "#ffb469",
            "#fea05f",
            "#fc8c55",
            "#fb784b",
            "#fa6441",
            "#f85037",
        )
        self.internal_colors = {v: k for k, v in self.displayed_colors_dct.items()}
        ak = lambda key: [int(c) if c.isdigit() else c.lower() for c in re.split("([0-9]+)", key)]  # noqa: E731
        self.displayed_colors = sorted(self.displayed_colors_dct, key=ak)

        self.formatting_view = Sheet(
            self,
            theme=theme,
            align="center",
            header_align="center",
            row_index_align="center",
            header_font=sheet_header_font,
            auto_resize_row_index=True,
            auto_resize_columns=200,
            headers=[f"{self.C.headers[self.column].name} Conditions", "Color"],
        )
        self.formatting_view.basic_bindings(True)
        self.formatting_view.enable_bindings(
            "single",
            "row_drag_and_drop",
            "drag_select",
            "column_width_resize",
            "double_click_column_resize",
            "row_select",
            "arrowkeys",
        )
        self.formatting_view.extra_bindings([("row_index_drag_drop", self.formatting_view_drag)])
        self.formatting_view.set_column_widths(column_widths=[500, 100])
        self.formatting_view.grid(row=0, column=0, sticky="nswe")
        self.formatting_view.bind(rc_button, self.formatting_view_rc)
        self.formatting_view.bind("<Double-Button-1>", self.formatting_view_double_b1)
        self.formatting_view.bind("<Delete>", self.del_condition)

        self.formatting_view_rc_menu = tk.Menu(self.formatting_view, tearoff=0, **menu_kwargs)
        self.formatting_view_rc_menu.add_command(label="Add condition", command=self.add_condition, **menu_kwargs)
        self.formatting_view_rc_menu.add_command(label="Edit condition", command=self.edit_condition, **menu_kwargs)
        self.formatting_view_rc_menu.add_separator()
        self.formatting_view_rc_menu.add_command(label="Del condition", command=self.del_condition, **menu_kwargs)
        self.formatting_view_rc_menu.add_command(
            label="Del all & add num scale",
            command=lambda: self.add_auto_conditions("num"),
            state="normal" if self.C.headers[self.column].type_ == "Numerical Detail" else "disabled",
            **menu_kwargs,
        )
        self.formatting_view_rc_menu.add_command(
            label="Del all & add date scale",
            command=lambda: self.add_auto_conditions("date"),
            state="normal" if self.C.headers[self.column].type_ == "Date Detail" else "disabled",
            **menu_kwargs,
        )
        self.redo_formatting_view()
        self.bind("<Escape>", self.USER_HAS_CLOSED_WINDOW)
        center(self, 1150, 600)
        self.deiconify()
        self.wait_window()

    def formatting_view_double_b1(self, event):
        region = self.formatting_view.identify_region(event)
        if region == "table":
            column = self.formatting_view.identify_column(event, allow_end=False)
            condition = self.formatting_view.identify_row(event, allow_end=False)
            if column is not None and condition is not None:
                self.formatting_view.select_row(condition)
                self.cond_sel = int(condition)
                self.edit_condition()

    def formatting_view_rc(self, event):
        column = self.formatting_view.identify_column(event, allow_end=False)
        condition = self.formatting_view.identify_row(event, allow_end=False)
        self.formatting_view_rc_menu.entryconfig("Add condition", state="normal")
        if column is not None and condition is not None:
            self.formatting_view.select_row(condition)
            self.cond_sel = int(condition)
            self.formatting_view_rc_menu.entryconfig("Edit condition", state="normal")
            self.formatting_view_rc_menu.entryconfig("Del condition", state="normal")
            if len(self.C.headers[self.column].formatting) > 35:
                self.formatting_view_rc_menu.entryconfig("Add condition", state="disabled")
        else:
            self.formatting_view.deselect()
            self.cond_sel = len(self.C.headers[self.column].formatting)
            self.formatting_view_rc_menu.entryconfig("Edit condition", state="disabled")
            self.formatting_view_rc_menu.entryconfig("Del condition", state="disabled")
        self.formatting_view_rc_menu.tk_popup(event.x_root, event.y_root)

    def formatting_view_drag(self, event_data):
        self.C.headers[self.column].formatting = move_elements_by_mapping(
            self.C.headers[self.column].formatting,
            self.formatting_view.full_move_rows_idxs(event_data["moved"]["rows"]["data"]),
        )
        self.redo_formatting_view()

    def enable_formatting_view(self):
        self.formatting_view.bind(rc_button, self.formatting_view_rc)
        self.formatting_view.bind("<Double-Button-1>", self.formatting_view_double_b1)
        self.formatting_view.bind("<Delete>", self.del_condition)
        self.formatting_view.basic_bindings(True)
        self.formatting_view.enable_bindings(
            "single",
            "row_drag_and_drop",
            "column_width_resize",
            "double_click_column_resize",
            "row_select",
            "arrowkeys",
        )
        self.formatting_view.extra_bindings([("row_index_drag_drop", self.formatting_view_drag)])

    def disable_formatting_view(self):
        self.formatting_view.unbind(rc_button)
        self.formatting_view.unbind("<Double-Button-1>")
        self.formatting_view.unbind("<Delete>")
        self.formatting_view.basic_bindings(False)
        self.formatting_view.disable_bindings()
        self.formatting_view.extra_bindings()

    def edit_condition(self, event=None):
        self.disable_formatting_view()
        header = self.C.headers[self.column]
        if header.formatting:
            cond_tuple = header.formatting[self.cond_sel]
        else:
            cond_tuple = ("", self.displayed_colors[0])
        self.new_frame = Edit_Condition_Frame(
            self,
            condition=cond_tuple[0],
            colors=self.displayed_colors,
            color=self.internal_colors[cond_tuple[1]],
            coltype=header.type_,
            theme=self.C.C.theme,
        )
        self.new_frame.grid(row=1, column=0, sticky="nswe")
        self.bind("<Return>", self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if not self.new_frame.result:
            self.enable_formatting_view()
            return
        condition = self.C.check_condition_validity(self.column, self.new_frame.new_condition)
        if condition.startswith("Error:"):
            self.new_frame = Error_Frame(
                self,
                f" {condition}   See 'Help' under the 'File' menu for instructions on conditional formatting   ",
                theme=self.C.C.theme,
            )
            self.new_frame.grid(row=1, column=0, sticky="nswe")
            self.bind("<Return>", self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_formatting_view()
            return
        color = self.displayed_colors_dct[self.new_frame.color]
        self.C.headers[self.column].formatting[self.cond_sel] = (condition, color)
        self.redo_formatting_view()
        self.enable_formatting_view()
        self.formatting_view.select_row(f"{self.cond_sel}")

    def add_auto_conditions(self, num_or_date="num"):
        self.disable_formatting_view()
        header = self.C.headers[self.column]
        if num_or_date == "num":
            self.new_frame = Auto_Add_Condition_Num_Frame(
                self,
                self.column,
                self.C.sheet.MT.data,
                theme=self.C.C.theme,
            )
        else:
            self.new_frame = Auto_Add_Condition_Date_Frame(
                self,
                self.column,
                self.C.sheet.MT.data,
                self.C.DATE_FORM,
                theme=self.C.C.theme,
            )
        self.new_frame.grid(row=1, column=0, sticky="nswe")
        self.bind("<Return>", self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if not self.new_frame.result:
            self.enable_formatting_view()
            return
        if num_or_date == "num":
            ac = {"0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "-", "."}
            min_v = "".join(c for c in self.new_frame.min_val if c in ac)
            max_v = "".join(c for c in self.new_frame.max_val if c in ac)
            if not min_v and not max_v:
                self.enable_formatting_view()
                return
            try:
                if not min_v:
                    min_v = 0
                else:
                    min_v = float(min_v)
                if not max_v:
                    max_v = 0
                else:
                    max_v = float(max_v)
            except Exception:
                self.enable_formatting_view()
                return
            if min_v >= max_v:
                self.new_frame = Error_Frame(
                    self,
                    "Error: Minimum value greater than or equal to maximum value - see 'Help' under the 'File' menu for instructions on conditional formatting   ",
                    theme=self.C.C.theme,
                )
                self.new_frame.grid(row=1, column=0, sticky="nswe")
                self.bind("<Return>", self.new_frame.confirm)
                self.new_frame.wait_window()
                if self.window_destroyed:
                    return
                self.unbind("<Return>")
                self.enable_formatting_view()
                return
            self.C.headers[self.column].formatting = []
            step = (max_v - min_v) / 20
            if header.type_ == "Numerical Detail":
                if self.new_frame.order == "ASCENDING":
                    v = float(min_v)
                    for i in range(1, 21):
                        if not i % 20:
                            self.C.headers[self.column].formatting.append(
                                ("".join((">= ", str(v), " and <= ", str(v + step))), self.scale_colors[i - 1])
                            )
                        else:
                            self.C.headers[self.column].formatting.append(
                                ("".join((">= ", str(v), " and < ", str(v + step))), self.scale_colors[i - 1])
                            )
                            v += step
                elif self.new_frame.order == "DESCENDING":
                    v = float(max_v)
                    for i in range(1, 21):
                        if not i % 20:
                            self.C.headers[self.column].formatting.append(
                                ("".join(("<= ", str(v), " and >= ", str(v - step))), self.scale_colors[i - 1])
                            )
                        else:
                            self.C.headers[self.column].formatting.append(
                                ("".join(("<= ", str(v), " and > ", str(v - step))), self.scale_colors[i - 1])
                            )
                            v -= step
            elif header.type_ == "Date Detail":
                if self.new_frame.order == "ASCENDING":
                    v = min_v
                    for i in range(1, 21):
                        if not i % 20:
                            self.C.headers[self.column].formatting.append(
                                (
                                    "".join((">= ", str(round(v)), " and <= ", str(round(v + step)))),
                                    self.scale_colors[i - 1],
                                )
                            )
                        else:
                            self.C.headers[self.column].formatting.append(
                                (
                                    "".join((">= ", str(round(v)), " and < ", str(round(v + step)))),
                                    self.scale_colors[i - 1],
                                )
                            )
                            v += step
                elif self.new_frame.order == "DESCENDING":
                    v = max_v
                    for i in range(1, 21):
                        if not i % 20:
                            self.C.headers[self.column].formatting.append(
                                (
                                    "".join(("<= ", str(round(v)), " and >= ", str(round(v - step)))),
                                    self.scale_colors[i - 1],
                                )
                            )
                        else:
                            self.C.headers[self.column].formatting.append(
                                (
                                    "".join(("<= ", str(round(v)), " and > ", str(round(v - step)))),
                                    self.scale_colors[i - 1],
                                )
                            )
                            v -= step
        elif num_or_date == "date":
            ac = {"0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "/", "-"}
            min_v = "".join(c for c in self.new_frame.min_val if c in ac).replace("-", "/")
            max_v = "".join(c for c in self.new_frame.max_val if c in ac).replace("-", "/")
            if not min_v and not max_v:
                self.enable_formatting_view()
                return
            DATE_FORM = self.C.convert_hyphen_to_slash_date_form(self.C.DATE_FORM)
            try:
                min_v = datetime.datetime.strptime(min_v, DATE_FORM)
                max_v = datetime.datetime.strptime(max_v, DATE_FORM)
            except Exception:
                self.enable_formatting_view()
                return
            if min_v >= max_v:
                self.new_frame = Error_Frame(
                    self,
                    "Error: Minimum value greater than or equal to maximum value - see 'Help' under the 'File' menu for instructions on conditional formatting   ",
                    theme=self.C.C.theme,
                )
                self.new_frame.grid(row=1, column=0, sticky="nswe")
                self.bind("<Return>", self.new_frame.confirm)
                self.new_frame.wait_window()
                if self.window_destroyed:
                    return
                self.unbind("<Return>")
                self.enable_formatting_view()
                return
            self.C.headers[self.column].formatting = []
            step = ((max_v - min_v).days) / 20
            step = datetime.timedelta(days=step)
            if self.new_frame.order == "ASCENDING":
                v = min_v  # strptime
                for i in range(1, 21):
                    s1 = datetime.datetime.strftime(v, DATE_FORM)
                    s2 = datetime.datetime.strftime(v + step, DATE_FORM)
                    if not i % 20:
                        self.C.headers[self.column].formatting.append(
                            ("".join((">= ", s1, " and <= ", s2)), self.scale_colors[i - 1])
                        )
                    else:
                        self.C.headers[self.column].formatting.append(
                            ("".join((">= ", s1, " and < ", s2)), self.scale_colors[i - 1])
                        )
                        v = v + step
            elif self.new_frame.order == "DESCENDING":
                v = max_v  # strptime
                for i in range(1, 21):
                    s1 = datetime.datetime.strftime(v, DATE_FORM)
                    s2 = datetime.datetime.strftime(v - step, DATE_FORM)
                    if not i % 20:
                        self.C.headers[self.column].formatting.append(
                            ("".join(("<= ", s1, " and >= ", s2)), self.scale_colors[i - 1])
                        )
                    else:
                        self.C.headers[self.column].formatting.append(
                            ("".join(("<= ", s1, " and > ", s2)), self.scale_colors[i - 1])
                        )
                        v = v - step
        self.redo_formatting_view()
        self.enable_formatting_view()

    def add_condition(self, event=None):
        self.disable_formatting_view()
        header = self.C.headers[self.column]
        cond_tuple = ("", self.displayed_colors[0])
        self.new_frame = Edit_Condition_Frame(
            self,
            condition=cond_tuple[0],
            colors=self.displayed_colors,
            color=cond_tuple[1],
            coltype=header.type_,
            confirm_text="Add condition",
            theme=self.C.C.theme,
        )
        self.new_frame.grid(row=1, column=0, sticky="nswe")
        self.bind("<Return>", self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if not self.new_frame.result:
            self.enable_formatting_view()
            return
        condition = self.C.check_condition_validity(self.column, self.new_frame.new_condition)
        if condition.startswith("Error:"):
            self.new_frame = Error_Frame(
                self,
                f" {condition}   See 'Help' under the 'File' menu for instructions on conditional formatting   ",
                theme=self.C.C.theme,
            )
            self.new_frame.grid(row=1, column=0, sticky="nswe")
            self.bind("<Return>", self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_formatting_view()
            return
        color = self.displayed_colors_dct[self.new_frame.color]
        self.C.headers[self.column].formatting.insert(self.cond_sel, (condition, color))
        self.redo_formatting_view()
        self.enable_formatting_view()
        self.formatting_view.select_row(self.cond_sel)

    def del_condition(self, event=None):
        elements = self.formatting_view.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        if not elements:
            return
        self.C.headers[self.column].formatting[elements[0] : elements[-1] + 1] = []
        self.cond_sel = None
        self.redo_formatting_view()

    def USER_HAS_CLOSED_WINDOW(self, event=None):
        self.window_destroyed = True
        try:
            self.destroy()
        except Exception:
            pass

    def redo_formatting_view(self):
        self.formatting_view.deselect("all")
        self.formatting_view.dehighlight_cells(all_=True, redraw=False)
        self.formatting_view.dehighlight_cells(canvas="row_index", all_=True, redraw=False)
        self.formatting_view.set_sheet_data(
            [[cond, self.internal_colors[color]] for cond, color in self.C.headers[self.column].formatting]
        )
        for i, (cond, color) in enumerate(self.C.headers[self.column].formatting):
            self.formatting_view.highlight_cells(row=i, column=1, bg=color, fg="black")
        self.formatting_view.set_column_widths(column_widths=[650, 200])
        self.formatting_view.refresh()


class View_Id_Popup(tk.Toplevel):
    def __init__(self, C, ids_row, width=800, height=800, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - {C.sheet.MT.data[ids_row['rn']][C.ic]}", resizable=True)

        self.USER_HAS_QUIT = False
        self.protocol("WM_DELETE_WINDOW", self.USER_HAS_CLOSED_WINDOW)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.ids_rn = ids_row["rn"]
        self.changes_made = 0

        self.sheetdisplay = Sheet(
            self,
            theme=theme,
            header_font=sheet_header_font,
            auto_resize_row_index=True,
            row_index_align="w",
            outline_thickness=0,
        )
        self.sheetdisplay.headers(newheaders=["Column Value"])
        self.sheetdisplay.row_index(
            newindex=[
                f"{c}    {' ' * (len(str(len(self.C.headers))) - len(str(c)))}{hdr.name}"
                for c, hdr in enumerate(self.C.headers, 1)
            ]
        )
        self.redo_display()
        for c, hdr in enumerate(self.C.headers):
            if hdr.validation:
                self.sheetdisplay.dropdown(
                    (c, 0),
                    values=hdr.validation,
                    edit_data=False,
                    redraw=False,
                )
        self.sheetdisplay.set_width_of_index_to_text()
        self.sheetdisplay.set_xview(0.0)
        self.sheetdisplay.set_yview(0.0)
        self.sheetdisplay.grid(row=1, column=0, sticky="nswe")
        self.status_bar = Status_Bar(
            self,
            text=f"ID - {self.C.sheet.MT.data[self.ids_rn][self.C.ic]} concise view",
            theme=theme,
        )
        self.status_bar.grid(row=2, column=0, sticky="nswe")
        self.bind("<Escape>", self.cancel)
        center(self, width, height)
        self.deiconify()
        self.enable_bindings()
        self.wait_window()

    def redo_display(self, event=None, scroll: bool = False):
        self.sheetdisplay.data_reference(
            newdataref=[[v] for v in self.C.sheet.MT.data[self.ids_rn]],
            reset_col_positions=False,
            reset_row_positions=False,
            redraw=False,
        )
        self.sheetdisplay.dehighlight_cells(all_=True)
        for tup1, tup2 in self.C.sheet.get_highlighted_cells().items():
            if tup1[0] == self.ids_rn:
                self.sheetdisplay.highlight_cells(row=tup1[1], column=0, bg=tup2[0], fg=tup2[1])
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.sheetdisplay.refresh()
        self.sheetdisplay.recreate_all_selection_boxes()
        if scroll:
            self.C.tree.scroll_to_item(self.sheetdisplay.data[self.C.ic][0].lower())

    def cut(self, event=None):
        pass

    def copy(self, event=None):
        self.sheetdisplay.copy()

    def paste(self, event=None):
        pass

    def undo(self, event=None):
        if not self.changes_made:
            return
        self.C.undo()
        self.redo_display()
        self.changes_made -= 1

    def delete(self, event=None):
        pass

    def sheet_begin_edit_cell(self, event=None):
        self.unbind("<Escape>")
        return event.value

    def sheet_end_edit_cell(self, event=None):
        r = event.row
        newtext = event.value
        y1 = int(self.ids_rn)
        x1 = int(r)
        ID = self.C.sheet.MT.data[self.ids_rn][self.C.ic]
        ik = ID.lower()
        currentdetail = self.C.sheet.MT.data[self.ids_rn][r]
        if newtext == currentdetail or newtext is None:
            self.bind("<Escape>", self.cancel)
            return
        if self.C.headers[x1].type_ == "ID":
            id_ = ID
            ik = id_.lower()
            if self.C.tree.selection():
                tree_sel = self.C.tree.selection()
            else:
                tree_sel = False
            success = self.C.change_ID_name(id_, newtext)
            if not success:
                return
            self.C.changelog_append(
                "Rename ID",
                id_,
                id_,
                f"{newtext}",
            )
            new_ik = newtext.lower()
            if ik in self.C.tagged_ids:
                self.C.tagged_ids.discard(ik)
                self.C.tagged_ids.add(new_ik)
                self.C.reset_tagged_ids_dropdowns()
            self.C.disable_paste()
            self.C.rns = {r[self.C.ic].lower(): i for i, r in enumerate(self.C.sheet.data)}
            self.C.redo_tree_display()
            self.C.refresh_all_formatting(rows=self.C.refresh_rows)
            self.C.refresh_rows = []
            self.C.redraw_sheets()
            if tree_sel:
                try:
                    self.C.tree.scroll_to_item(tree_sel[0])
                    self.C.tree.selection_set(tree_sel)
                except Exception:
                    self.C.tree.scroll_to_item(newtext.lower())
                    self.C.tree.selection_set(newtext.lower())
            else:
                self.C.move_tree_pos()
            self._changes_made()
            return newtext
        successful = False
        if self.C.headers[x1].type_ == "Parent":
            self.C.snapshot_paste_id()
            oldparent = f"{self.C.sheet.MT.data[y1][x1]}"
            if self.C.cut_paste_edit_cell(self.C.sheet.MT.data[y1][self.C.ic], oldparent, x1, newtext):
                successful = True
            if not successful:
                self.C.vs.pop()
                self.C.vp -= 1
                self.C.set_undo_label()
            else:
                self.C.changelog_append(
                    "Cut and paste ID + children" if self.C.nodes[ik].cn[x1] else "Cut and paste ID",
                    self.C.sheet.MT.data[y1][self.C.ic],
                    f"Old parent: {oldparent if oldparent else 'n/a - Top ID'} old column #{x1 + 1} named: {self.C.headers[x1].name}",
                    f"New parent: {newtext if newtext else 'n/a - Top ID'} new column #{x1 + 1} named: {self.C.headers[x1].name}",
                )
                self.C.redo_tree_display()
                self.C.refresh_all_formatting(rows=[y1])
                self.C.redraw_sheets()
                try:
                    self.C.tree.scroll_to_item(self.sheet[y1][self.ic].lower())
                    self.C.tree.selection_set(self.sheet[y1][self.ic].lower())
                except Exception:
                    pass
                self.C.disable_paste()
                self._changes_made(scroll=True)
                return newtext
        if not successful and self.C.headers[x1].type_ not in ("Text Detail", "Numerical Detail", "Date Detail"):
            self.C.changelog_append(
                "Edit cell",
                f"ID: {ID} column #{x1 + 1} named: {self.C.headers[x1].name} with type: {self.C.headers[x1].type_}",
                f"{self.C.sheet.MT.data[y1][x1]}",
                f"{newtext}",
            )
            self.C.snapshot_ctrl_x_v_del_key_id_par()
            self.C.sheet.MT.data[y1][x1] = newtext
            self.C.rebuild_tree()
            self._changes_made(scroll=True)
            return newtext
        else:
            if not self.C.detail_is_valid_for_col(x1, newtext):
                self.bind("<Escape>", self.cancel)
                return
            if self.C.headers[x1].type_ == "Date Detail":
                newtext = self.C.convert_date(newtext, self.C.DATE_FORM)
            currentdetail = self.C.sheet.MT.data[y1][x1]
            self.C.changelog_append(
                "Edit cell",
                f"ID: {ID} column #{x1 + 1} named: {self.C.headers[x1].name} with type: {self.C.headers[x1].type_}",
                f"{self.C.sheet.MT.data[y1][x1]}",
                f"{newtext}",
            )
            self.C.snapshot_ctrl_x_v_del_key()
            self.C.vs[-1]["cells"][(y1, x1)] = f"{self.C.sheet.MT.data[y1][x1]}"
            self.C.sheet.MT.data[y1][x1] = f"{newtext}"
            self.C.refresh_all_formatting(rows=[y1])
            self.C.refresh_tree_item(ID)
            self.C.disable_paste()
            self.C.redraw_sheets()
            self._changes_made()
            return newtext

    def _changes_made(self, scroll: bool = False):
        self.redo_display(scroll=scroll)
        self.changes_made += 1
        self.C.C.status_bar.change_text(self.C.get_tree_editor_status_bar_text())
        self.bind("<Escape>", self.cancel)

    def enable_bindings(self, event=None):
        self.sheetdisplay.basic_bindings(True)
        self.sheetdisplay.enable_bindings(
            "single",
            "copy",
            "edit_cell",
            "drag_select",
            "column_width_resize",
            "double_click_column_resize",
            "row_height_resize",
            "double_click_row_resize",
            "column_select",
            "row_select",
            "arrowkeys",
            "ctrl_select",
        )
        self.sheetdisplay.extra_bindings(
            [
                ("begin_edit_cell", self.sheet_begin_edit_cell),
            ]
        )
        self.sheetdisplay.edit_validation(self.sheet_end_edit_cell)
        # self.sheetdisplay.bind(f"<{ctrl_button}-x>", self.cut)
        # self.sheetdisplay.bind(f"<{ctrl_button}-X>", self.cut)
        self.sheetdisplay.bind(f"<{ctrl_button}-c>", self.copy)
        self.sheetdisplay.bind(f"<{ctrl_button}-C>", self.copy)
        # self.sheetdisplay.bind(f"{ctrl_button}-v>", self.paste)
        # self.sheetdisplay.bind(f"{ctrl_button}-V>", self.paste)
        self.sheetdisplay.bind(f"<{ctrl_button}-z>", self.undo)
        self.sheetdisplay.bind(f"<{ctrl_button}-Z>", self.undo)
        # self.sheetdisplay.bind("<Delete>", self.delete)

    def disable_bindings(self, event=None):
        self.sheetdisplay.basic_bindings(False)
        self.sheetdisplay.disable_bindings()
        self.sheetdisplay.extra_bindings()
        # self.sheetdisplay.unbind(f"<{ctrl_button}-x>")
        # self.sheetdisplay.unbind(f"<{ctrl_button}-X>")
        self.sheetdisplay.unbind(f"<{ctrl_button}-c>")
        self.sheetdisplay.unbind(f"<{ctrl_button}-C>")
        # self.sheetdisplay.unbind(f"{ctrl_button}-v>")
        # self.sheetdisplay.unbind(f"{ctrl_button}-V>")
        self.sheetdisplay.unbind(f"<{ctrl_button}-z>")
        self.sheetdisplay.unbind(f"<{ctrl_button}-Z>")
        # self.sheetdisplay.unbind("<Delete>")

    def USER_HAS_CLOSED_WINDOW(self, callback=None):
        self.USER_HAS_QUIT = True
        self.destroy()

    def cancel(self, event=None):
        self.USER_HAS_CLOSED_WINDOW()


class Merge_Sheets_Popup(tk.Toplevel):
    def __init__(self, C, theme="dark", add_rows=False):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Merge sheets", resizable=True)
        self.protocol("WM_DELETE_WINDOW", self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.grid_columnconfigure(0, weight=1, uniform="x")
        self.grid_columnconfigure(1, weight=1, uniform="x")
        self.grid_rowconfigure(0, weight=1)

        self.l_frame = Frame(self, theme=theme)
        self.l_frame.grid(row=0, column=0, sticky="nswe")
        self.r_frame = Frame(self, theme=theme)
        self.r_frame.grid(row=0, column=1, sticky="nswe")
        self.l_frame.grid_rowconfigure(3, weight=1)
        self.r_frame.grid_rowconfigure(0, weight=1)
        self.l_frame.grid_columnconfigure(0, weight=1)
        self.r_frame.grid_columnconfigure(1, weight=1)

        self.open_file_display = Readonly_Entry_With_Scrollbar(self.l_frame, font=EF, theme=theme)
        self.open_file_display.grid(row=0, column=0, padx=2, pady=2, sticky="nswe")
        self.open_file_button = Button(self.l_frame, text="⯇ Open file", style="EF.Std.TButton", command=self.open_file)
        self.open_file_button.grid(row=0, column=1, padx=(2, 10), pady=2, sticky="nswe")
        self.sheet_dropdown = Ez_Dropdown(self.l_frame, font=EF)
        self.sheet_dropdown.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
        self.sheet_dropdown.grid(row=1, column=0, padx=2, pady=2, sticky="nswe")
        self.select_sheet_button = Button(
            self.l_frame,
            text="⯇ Load sheet",
            style="EF.Std.TButton",
            state="disabled",
            command=self.select_sheet,
        )
        self.select_sheet_button.grid(row=1, column=1, padx=(2, 10), pady=2, sticky="nswe")

        self.selector = Id_Parent_Column_Selector(self.l_frame, theme=theme)
        self.selector.grid(row=2, column=0, rowspan=2, sticky="nswe")

        self.clipboard_button = Button(
            self.l_frame,
            text=" Get data from clipboard ",
            style="EF.Std.TButton",
            state="normal",
            command=self.get_clipboard_data,
        )
        self.clipboard_button.grid(row=2, column=1, padx=(2, 10), pady=(2, 20), sticky="nswe")

        self.options_frame = Frame(self.l_frame, theme=theme)
        self.options_frame.grid(row=3, column=1, sticky="nswe")

        self.add_new_ids_button = X_Checkbutton(
            self.options_frame,
            text="Add new IDs",
            style="x_button.Std.TButton",
            compound="right",
            checked=True,
        )
        self.add_new_ids_button.grid(row=0, column=0, padx=10, pady=5, sticky="we")
        self.add_new_dcols_button = X_Checkbutton(
            self.options_frame,
            text="Add new detail columns",
            style="x_button.Std.TButton",
            compound="right",
            checked=True,
        )
        self.add_new_dcols_button.grid(row=1, column=0, padx=10, pady=5, sticky="we")
        self.add_new_pcols_button = X_Checkbutton(
            self.options_frame,
            text="Add new parent columns",
            style="x_button.Std.TButton",
            compound="right",
            checked=True,
        )
        self.add_new_pcols_button.grid(row=2, column=0, padx=10, pady=5, sticky="we")
        self.overwrite_details_button = X_Checkbutton(
            self.options_frame,
            text="Overwrite details",
            style="x_button.Std.TButton",
            compound="right",
            checked=True,
        )
        self.overwrite_details_button.grid(row=3, column=0, padx=10, pady=5, sticky="we")
        self.overwrite_parents_button = X_Checkbutton(
            self.options_frame,
            text="Overwrite parents",
            style="x_button.Std.TButton",
            compound="right",
            checked=True,
        )
        self.overwrite_parents_button.grid(row=4, column=0, padx=10, pady=5, sticky="we")

        self.button_frame = Frame(self.l_frame, theme=theme)
        self.button_frame.grid(row=4, column=0, columnspan=2, sticky="e")
        self.button_frame.grid_rowconfigure(0, weight=1)
        self.confirm_button = Button(
            self.button_frame,
            text=" Confirm merge ",
            style="EF.Std.TButton",
            command=self.confirm,
        )
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=20, pady=(20, 20))
        self.cancel_button = Button(self.button_frame, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=20, pady=(20, 20))
        self.status = Status_Bar(self.l_frame, text="Note: Opening files resets current merge sheet", theme=theme)
        self.status.grid(row=5, column=0, columnspan=2, sticky="ew")
        self.result = False
        self.add_new_ids = True
        self.add_new_dcols = False
        self.add_new_pcols = False
        self.overwrite_details = False
        self.overwrite_parents = False
        self.file_opened = ""
        self.sheet_opened = "n/a"
        self.row_len = 0
        self.ic = None
        self.pcols = []
        self.wb_ = None
        self.C.new_sheet = []
        self.rowsel = 0
        self.colsel = 0
        self.region = "header"

        self.showing_left = True
        self.toggle_left_button = Button(
            self.options_frame,
            text="Show Sheet",
            style="BF.Std.TButton",
            command=self.toggle_left_panel,
        )
        self.toggle_left_button.grid(row=5, column=0, padx=10, pady=10, sticky="e")

        self.toggle_left_button2 = Button(
            self.r_frame,
            text="Sheet",
            style="BF.Std.TButton",
            command=self.toggle_left_panel,
        )

        self.sheetdisplay = Sheet(
            self.r_frame,
            theme=theme,
            expand_sheet_if_paste_too_big=True,
            header_font=sheet_header_font,
            outline_thickness=0,
        )
        self.selector.link_sheet(self.sheetdisplay)
        self.sheetdisplay.enable_bindings("all", "ctrl_select")
        self.sheetdisplay.extra_bindings(
            [
                ("begin_edit_cell", self.begin_edit_cell),
                ("end_edit_cell", self.end_edit_cell),
            ]
        )
        self.sheetdisplay.bind("<<SheetModified>>", self.sheet_modified)
        self.sheetdisplay.headers(newheaders=0)
        self.setup_selectors()
        self.selector.detect_id_col()
        self.selector.detect_par_cols()
        self.sheetdisplay.grid(row=0, column=1, sticky="nswe")
        if add_rows:
            self.toggle_left_panel()
            self.toggle_left_button2.config(text="Options")
        self.bind("<Escape>", self.cancel)
        center(self, 1280, 620)
        self.deiconify()
        self.wait_window()

    def setup_selectors(self, event=None):
        self.sheetdisplay.deselect("all")
        self.C.new_sheet = [[h.name for h in self.C.headers]] + [
            list(repeat("", len(self.C.headers))) for r in range(2000)
        ]
        self.C.new_sheet = self.sheetdisplay.set_sheet_data(self.C.new_sheet, verify=False)
        self.selector.set_columns(self.C.new_sheet[0] if self.C.new_sheet else [])

    def toggle_left_panel(self, event=None):
        if self.showing_left:
            self.grid_columnconfigure(0, weight=0, uniform="y")
            self.l_frame.grid_forget()
            self.showing_left = False
            self.toggle_left_button2.grid(row=0, column=0, sticky="ns")
            self.toggle_left_button2.config(text="Options")
        else:
            self.grid_columnconfigure(0, weight=1, uniform="x")
            self.toggle_left_button2.grid_forget()
            self.l_frame.grid(row=0, column=0, sticky="nswe")
            self.showing_left = True
        self.update_idletasks()

    def begin_edit_cell(self, event):
        self.unbind("<Escape>")
        return event.value

    def end_edit_cell(self, event):
        self.bind("<Escape>", self.cancel)

    def sheet_modified(self, event):
        self.C.new_sheet = self.sheetdisplay.MT.data
        self.sheetdisplay.refresh()
        if "move" in event.eventname:
            self.selector.set_columns(self.C.new_sheet[0])
            self.selector.detect_id_col()
            self.selector.detect_par_cols()
        else:
            self.reset_selectors()
        self.sheetdisplay.focus_set()

    def reset_selectors(self, event=None):
        idcol = self.selector.get_id_col()
        parcols = self.selector.get_par_cols()
        self.selector.set_columns(self.C.new_sheet[0] if self.C.new_sheet else [])
        try:
            if idcol is not None and self.C.new_sheet:
                self.selector.set_id_col(idcol)
        except Exception:
            pass
        try:
            if parcols and self.C.new_sheet:
                self.selector.set_par_cols(parcols)
        except Exception:
            pass

    def get_clipboard_data(self, event=None):
        self.start_work("Loading...")
        self.reset()
        try:
            temp_data = self.C.clipboard_get()
        except Exception as error_msg:
            self.stop_work(f"Error: Error getting clipboard data: {error_msg}", sels=True)
            return
        try:
            if temp_data.startswith("{") and temp_data.endswith("}"):
                self.C.new_sheet = json_to_sheet(json.loads(temp_data))
            else:
                self.C.new_sheet = csv_str_x_data(temp_data)
        except Exception as error_msg:
            self.stop_work(f"Error: Error parsing clipboard data: {error_msg}", sels=True)
            return
        if not self.C.new_sheet:
            self.stop_work("Error: No data found on clipboard", sels=True)
            return
        equalize_sublist_lens(self.C.new_sheet)
        self.ic = None
        self.pcols = []
        self.load_display(self.C.new_sheet[0])
        self.stop_work("Select ID column and Parent columns")
        self.sheetdisplay.deselect("all")
        self.sheetdisplay.data_reference(
            newdataref=self.C.new_sheet,
            reset_col_positions=True,
            reset_row_positions=True,
            redraw=False,
        )
        self.sheetdisplay.refresh()
        self.file_opened = "n/a - Data obtained from clipboard"
        self.sheet_opened = "n/a"

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except Exception:
            pass
        try:
            self.wb_ = None
        except Exception:
            pass

    def USER_HAS_CLOSED_WINDOW(self, callback=None):
        self.C.new_sheet = []
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except Exception:
            pass
        self.destroy()

    def open_file(self):
        self.start_work("Loading...   ")
        self.reset()
        filepath = filedialog.askopenfilename(parent=self, title="Select file")
        if not filepath:
            self.stop_work("Select a file", sels=True)
            return
        try:
            filepath = os.path.normpath(filepath)
        except Exception:
            self.stop_work("Error: filepath invalid", sels=True)
            return
        if not filepath.lower().endswith((".json", ".xlsx", ".xls", ".xlsm", ".csv", ".tsv")):
            self.stop_work("Error: select json/excel/csv   ", sels=True)
            return
        check = os.path.isfile(filepath)
        if not check:
            self.stop_work("Error: filepath invalid", sels=True)
            return
        try:
            if filepath.lower().endswith((".csv", ".tsv")):
                with open(filepath, "r") as fh:
                    temp_data = fh.read()
                self.C.new_sheet = csv_str_x_data(temp_data)
                equalize_sublist_lens(self.C.new_sheet)
                self.load_display(self.C.new_sheet[0])
                self.stop_work("Ready to merge sheets")
            elif filepath.lower().endswith(".json"):
                j = get_json_from_file(filepath)
                json_format = get_json_format(j)
                if not json_format:
                    self.C.new_sheet = []
                    self.stop_work("Error opening file, could not find data of correct format", sels=True)
                    return
                self.C.new_sheet = json_to_sheet(
                    j,
                    format_=json_format[0],
                    key=json_format[1],
                    get_format=False,
                    return_rowlen=False,
                )
                if not self.C.new_sheet:
                    self.stop_work("Error: File contained no data", sels=True)
                    self.select_sheet_button.config(state="disabled")
                    return
                equalize_sublist_lens(self.C.new_sheet)
                self.load_display(self.C.new_sheet[0])
                self.stop_work("Ready to merge sheets")
            elif filepath.lower().endswith((".xlsx", ".xls", ".xlsm")):
                in_mem = bytes_io_wb(filepath)
                self.wb_ = load_workbook(in_mem, read_only=True, data_only=True)
                wbsheets = self.wb_.sheetnames
                if not wbsheets:
                    self.stop_work("Error: File/sheet contained no data", sels=True)
                    return
                sheetnames = set(self.wb_.sheetnames)
                if "program_data" in sheetnames:
                    ws = self.wb_["program_data"]
                    ws.reset_dimensions()
                    try:
                        d = b32_x_dict(ws_x_program_data_str(ws))
                        self.C.new_sheet = [[h["name"] for h in d["headers"]]] + d["records"]
                        self.wb_.close()
                        self.select_sheet_button.config(state="disabled")
                        self.load_display(
                            cols=self.C.new_sheet[0],
                            idcol=(
                                next(c for c, h in enumerate(d["headers"]) if h["type"] == "ID")
                                if d["headers"]
                                else None
                            ),
                            parcols=(
                                [c for c, h in enumerate(d["headers"]) if h["type"] == "Parent"]
                                if d["headers"]
                                else None
                            ),
                        )
                        self.stop_work("Ready to merge sheets")
                    except Exception:
                        self.C.new_sheet = []
                        self.wb_.close()
                        self.wb_ = load_workbook(in_mem, read_only=True, data_only=True)
                        self.stop_work("Error: Error opening program data")
                        self.sheet_dropdown["values"] = wbsheets
                        self.sheet_dropdown.set_my_value(wbsheets[0])
                        self.stop_work("Error: Error opening program data. Select a sheet to open")
                        self.select_sheet_button.config(state="normal")
                else:
                    self.sheet_dropdown["values"] = wbsheets
                    self.sheet_dropdown.set_my_value(wbsheets[0])
                    self.stop_work("Select a sheet to open")
                    self.select_sheet_button.config(state="normal")
        except Exception as error_msg:
            self.try_to_close_wb()
            self.C.new_sheet = []
            self.stop_work(f"Error: {error_msg}", sels=True)
            return
        if not self.C.new_sheet and not filepath.lower().endswith((".xlsx", ".xls", ".xlsm")):
            self.C.new_sheet = []
            self.stop_work("Error: File/sheet contained no data", sels=True)
            return
        self.open_file_display.set_my_value(filepath)
        self.file_opened = os.path.basename(self.open_file_display.get_my_value())

    def select_sheet(self):
        self.start_work("Loading...   ")
        self.sheet_opened = self.sheet_dropdown.get_my_value()
        ws = self.wb_[self.sheet_opened]
        ws.reset_dimensions()
        self.C.new_sheet = ws_x_data(ws)
        self.try_to_close_wb()
        self.stop_work("Ready to merge sheets")
        if not self.C.new_sheet:
            self.stop_work("Error: File/sheet contained no data", sels=True)
            self.select_sheet_button.config(state="disabled")
            return
        equalize_sublist_lens(self.C.new_sheet)
        self.select_sheet_button.config(state="disabled")
        self.load_display(self.C.new_sheet[0])

    def load_display(self, cols, idcol=None, parcols=None, set_sheet=True):
        if set_sheet:
            self.sheetdisplay.data_reference(
                newdataref=self.C.new_sheet,
                reset_col_positions=True,
                reset_row_positions=True,
                redraw=True,
            )
        self.selector.set_columns(cols)
        if idcol is not None and parcols is not None:
            self.selector.set_id_col(idcol)
            self.selector.set_par_cols(parcols)
        else:
            self.selector.detect_id_col()
            self.selector.detect_par_cols()

    def start_work(self, msg=""):
        self.status.change_text(msg)
        self.disable_widgets()

    def stop_work(self, msg="", sels=False):
        self.status.change_text(msg)
        if sels:
            self.setup_selectors()
        self.enable_widgets()

    def enable_widgets(self):
        self.open_file_display.change_my_state("readonly")
        self.open_file_button.config(state="normal")
        self.sheet_dropdown.config(state="readonly")
        self.selector.enable_me()
        self.add_new_ids_button.config(state="normal")
        self.add_new_dcols_button.config(state="normal")
        self.add_new_pcols_button.config(state="normal")
        self.overwrite_details_button.config(state="normal")
        self.overwrite_parents_button.config(state="normal")
        self.confirm_button.config(state="normal")
        self.sheetdisplay.enable_bindings("all", "ctrl_select")
        self.sheetdisplay.extra_bindings(
            [
                ("begin_edit_cell", self.begin_edit_cell),
                ("end_edit_cell", self.end_edit_cell),
            ]
        )

    def disable_widgets(self):
        self.open_file_display.change_my_state("disabled")
        self.open_file_button.config(state="disabled")
        self.sheet_dropdown.config(state="disabled")
        self.select_sheet_button.config(state="disabled")
        self.selector.disable_me()
        self.add_new_ids_button.config(state="disabled")
        self.add_new_dcols_button.config(state="disabled")
        self.add_new_pcols_button.config(state="disabled")
        self.overwrite_details_button.config(state="disabled")
        self.overwrite_parents_button.config(state="disabled")
        self.confirm_button.config(state="disabled")
        self.sheetdisplay.disable_bindings()
        self.sheetdisplay.extra_bindings("unbind_all")
        self.update()

    def reset(self):
        self.try_to_close_wb()
        self.row_len = 0
        self.ic = None
        self.pcols = []
        self.C.new_sheet = []
        self.open_file_display.set_my_value("")
        self.sheet_dropdown["values"] = []
        self.sheet_dropdown.set("")
        self.select_sheet_button.config(state="disabled")
        self.selector.clear_displays()

    def confirm(self, event=None):
        self.add_new_ids = self.add_new_ids_button.get_checked()
        self.add_new_dcols = self.add_new_dcols_button.get_checked()
        self.add_new_pcols = self.add_new_pcols_button.get_checked()
        self.overwrite_details = self.overwrite_details_button.get_checked()
        self.overwrite_parents = self.overwrite_parents_button.get_checked()
        self.ic = self.selector.get_id_col()
        self.pcols = self.selector.get_par_cols()
        if not self.C.new_sheet:
            self.status.change_text("Open a file to load data")
            return
        self.row_len = len(max(self.C.new_sheet, key=len))
        if all(
            x is False
            for x in (
                self.add_new_ids,
                self.add_new_dcols,
                self.add_new_pcols,
                self.overwrite_details,
                self.overwrite_parents,
            )
        ):
            self.status.change_text("Select at least one option")
            return
        if self.ic in set(self.pcols):
            self.status.change_text("ID column must be different to all parent columns")
            return
        if self.ic is None:
            self.status.change_text("Select an ID column")
            return
        self.result = True
        self.destroy()

    def cancel(self, event=None):
        self.USER_HAS_CLOSED_WINDOW()


class Get_Clipboard_Data_Popup(tk.Toplevel):
    def __init__(self, C, cols, row_len, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(
            self,
            C,
            f"{app_title} - Overwrite the current sheet using data from the clipboard",
            resizable=True,
        )
        # self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        # self.grid_rowconfigure(2,weight=1,uniform="y")
        self.selector = Id_Parent_Column_Selector(self)
        self.selector.grid(row=0, column=0, sticky="nsew")

        self.flattened_choices = FlattenedToggleAndOrder(self, command=self.flattened_mode_toggle, theme=theme)
        self.flattened_choices.change_theme(theme)
        self.flattened_choices.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
        self.flattened_selector = Flattened_Column_Selector(self)
        self.flattened_selector.set_columns(cols)
        self.selector.change_theme(theme)
        self.flattened_selector.change_theme(theme)
        self.selector.set_columns(cols)
        self.sheetdisplay = Sheet(
            self,
            theme=theme,
            header_font=sheet_header_font,
            expand_sheet_if_paste_too_big=True,
        )
        self.selector.link_sheet(self.sheetdisplay)
        self.flattened_selector.link_sheet(self.sheetdisplay, self.flattened_choices)
        self.sheetdisplay.enable_bindings("all", "ctrl_select")
        self.sheetdisplay.extra_bindings(
            [
                ("begin_edit_cell", self.begin_edit_cell),
                ("end_edit_cell", self.end_edit_cell),
            ]
        )
        self.sheetdisplay.bind("<<SheetModified>>", self.sheet_modified)
        self.sheetdisplay.headers(newheaders=0)
        self.sheetdisplay.data_reference(newdataref=self.C.new_sheet, redraw=True)
        self.sheetdisplay.grid(row=0, column=1, rowspan=4, sticky="nswe")

        self.selector.detect_id_col()
        self.selector.detect_par_cols()

        self.button_frame = Frame(self, theme=theme)
        self.button_frame.grid(row=2, column=0, sticky="e")
        self.button_frame.grid_rowconfigure(0, weight=1)

        self.confirm_button = Button(
            self.button_frame, text=" Overwrite existing sheet ", style="EF.Std.TButton", command=self.confirm
        )
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=10, pady=(20, 20))
        self.cancel_button = Button(self.button_frame, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=20, pady=(20, 20))
        self.status = Status_Bar(self, text="Select ID and Parent columns", theme=theme)
        self.status.grid(row=3, column=0, sticky="we")
        self.result = False
        self.ic = None
        self.pcols = []
        self.bind("<Escape>", self.cancel)
        center(self, 1280, 620)
        self.selector.grid_forget()
        self.flattened_selector.grid(row=0, column=0, pady=(0, 35), sticky="nsew")
        self.flattened_selector.grid_forget()
        self.selector.grid(row=0, column=0, sticky="nsew")
        self.deiconify()
        self.wait_window()

    def begin_edit_cell(self, event):
        self.unbind("<Escape>")
        return event.value

    def end_edit_cell(self, event):
        self.bind("<Escape>", self.cancel)

    def sheet_modified(self, event):
        self.sheetdisplay.MT.data = self.C.new_sheet
        self.sheetdisplay.refresh()
        if "move" in event.eventname:
            self.selector.set_columns(self.C.new_sheet[0])
            self.flattened_selector.set_columns(self.C.new_sheet[0])
            self.selector.detect_id_col()
            self.selector.detect_par_cols()
        else:
            self.reset_selectors()
        self.sheetdisplay.focus_set()

    def reset_selectors(self, event=None):
        idcol = self.selector.get_id_col()
        parcols = self.selector.get_par_cols()
        ancparcols = self.flattened_selector.get_par_cols()
        self.selector.set_columns(self.C.new_sheet[0] if self.C.new_sheet else [])
        self.flattened_selector.set_columns(self.C.new_sheet[0] if self.C.new_sheet else [])
        try:
            if idcol is not None and self.C.new_sheet:
                self.selector.set_id_col(idcol)
        except Exception:
            pass
        try:
            if parcols and self.C.new_sheet:
                self.selector.set_par_cols(parcols)
        except Exception:
            pass
        try:
            if ancparcols and self.C.new_sheet:
                self.flattened_selector.set_par_cols(ancparcols)
        except Exception:
            pass

    def flattened_mode_toggle(self):
        if self.flattened_choices.flattened:
            self.selector.grid_forget()
            self.flattened_selector.grid(row=0, column=0, pady=(0, 35), sticky="nsew")
        else:
            self.flattened_selector.grid_forget()
            self.selector.grid(row=0, column=0, sticky="nsew")

    def confirm(self, event=None):
        self.ic = self.selector.get_id_col()
        self.pcols = self.selector.get_par_cols()
        self.flattened_pcols = self.flattened_selector.get_par_cols()
        self.flattened = self.flattened_choices.flattened
        self.order = self.flattened_choices.order
        self.C.new_sheet = self.sheetdisplay.get_sheet_data()
        if self.flattened:
            if not self.flattened_pcols:
                self.status.change_text("Select hierarchy columns")
                return
        else:
            if self.ic in set(self.pcols):
                self.status.change_text("ID column must be different to all parent columns")
                return
            if self.ic is None:
                self.status.change_text("Select an ID column")
                return
            if not self.pcols:
                self.status.change_text("Select parent columns")
                return
        self.result = True
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Ask_Confirm_Quit(tk.Toplevel):
    def __init__(self, C, changes, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Quit")
        self.action_label = Label(
            self,
            text=("Save your changes before quitting?" if changes > 1 else "Save your last change before quitting?"),
            font=EFB,
            theme=theme,
        )
        self.action_label.grid(row=0, column=0, sticky="w", pady=20, padx=20)

        self.button_frame = Frame(self, theme=theme)
        self.button_frame.grid(row=1, column=0, columnspan=2, sticky="e", padx=(20, 20), pady=(20, 20))
        self.button_frame.grid_rowconfigure(0, weight=True)

        self.save_button = Button(
            self.button_frame,
            text="Save",
            style="EF.Std.TButton",
            command=self.save,
        )
        self.save_button.grid(row=0, column=0, sticky="e", padx=(0, 20))

        self.dont_save_button = Button(
            self.button_frame,
            text="Don't Save",
            style="EF.Std.TButton",
            command=self.dont_save,
        )
        self.dont_save_button.grid(row=0, column=1, sticky="e", padx=(20, 10))

        self.cancel_button = Button(
            self.button_frame,
            text="Cancel",
            style="EF.Std.TButton",
            command=self.cancel,
        )
        self.cancel_button.grid(row=0, column=2, sticky="e", padx=(20, 0))

        self.bind("<Return>", self.save)
        self.bind("<Escape>", self.cancel)
        self.option = "cancel"
        center(self, 415 if changes > 1 else 430, 150)
        self.deiconify()
        self.wait_window()

    def save(self, event=None):
        self.option = "save"
        self.destroy()

    def dont_save(self, event=None):
        self.option = "dont_save"
        self.destroy()

    def cancel(self, event=None):
        self.option = "cancel"
        self.destroy()


class Ask_Confirm(tk.Toplevel):
    def __init__(self, C, action, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Confirm Action")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.action_display = Display_Text(parent=self, text=action, theme=theme, bold=True)
        self.action_display.grid(row=0, column=0, sticky="nswe", pady=(20, 5), padx=30)
        self.action_display.config(height=75)
        self.button_frame = Frame(self, theme=theme)
        self.button_frame.grid(row=1, column=0, sticky="e", padx=20, pady=(10, 20))
        self.button_frame.grid_rowconfigure(0, weight=True)
        self.confirm_button = Button(self.button_frame, text="Confirm", style="EF.Std.TButton", command=self.confirm)
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=(10, 20))
        self.cancel_button = Button(self.button_frame, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=10)
        self.bind("<Return>", self.confirm)
        self.bind("<Escape>", self.cancel)
        self.boolean = False
        center(self, 530, 168)
        self.deiconify()
        self.action_display.place_cursor()
        self.wait_window()

    def confirm(self, event=None):
        self.boolean = True
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Save_New_Version_Presave_Popup(tk.Toplevel):
    def __init__(self, C, file_location, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Save new version")
        self.grid_columnconfigure(1, weight=1)
        self.file_loc_label = Label(self, text="Your new version\nwill be saved here:", font=EF, theme=theme)
        self.file_loc_label.grid(row=0, column=0, sticky="nswe", padx=(20, 10))
        self.file_loc_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.file_loc_display.set_my_value(file_location)
        self.file_loc_display.grid(row=0, column=1, sticky="nswe", pady=(20, 0), padx=(0, 20))
        self.confirm_button = Button(
            self,
            text="Confirm",
            style="EF.Std.TButton",
            command=self.confirm,
        )
        self.confirm_button.grid(row=1, column=1, sticky="e", padx=20, pady=20)
        self.choose_loc_button = Button(
            self,
            text="Choose another location instead",
            style="EF.Std.TButton",
            command=self.choose_loc,
        )
        self.choose_loc_button.grid(row=2, column=1, sticky="e", padx=20, pady=(5, 10))
        self.bind("<Return>", self.confirm)
        self.bind("<Escape>", self.cancel)
        self.result = False
        center(self, 550, 185)
        self.deiconify()
        self.file_loc_display.place_cursor()
        self.wait_window()

    def choose_loc(self, event=None):
        folder = os.path.normpath(filedialog.askdirectory(parent=self, title="Select a folder"))
        if folder == ".":
            return
        self.file_loc_display.set_my_value(folder)

    def confirm(self, event=None):
        self.result = os.path.normpath(self.file_loc_display.get_my_value())
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Save_New_Version_Postsave_Popup(tk.Toplevel):
    def __init__(self, C, file_location, filename, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", background=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Success! New version saved")
        self.grid_columnconfigure(1, weight=1)
        self.file_loc_label = Label(self, text="Your new version\nwas saved here:", font=EF, theme=theme)
        self.file_loc_label.grid(row=0, column=0, sticky="nswe", padx=(20, 10))
        self.file_loc_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.file_loc_display.set_my_value(file_location)
        self.file_loc_display.grid(row=0, column=1, sticky="nswe", pady=(20, 20), padx=(0, 20))
        self.file_name_label = Label(self, text="This is the\n filename:", font=EF, theme=theme)
        self.file_name_label.grid(row=1, column=0, sticky="nswe", padx=(20, 10))
        self.file_name_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.file_name_display.set_my_value(filename)
        self.file_name_display.grid(row=1, column=1, sticky="nswe", pady=(20, 20), padx=(0, 20))
        self.confirm_button = Button(self, text="Okay", style="EF.Std.TButton", command=self.cancel)
        self.confirm_button.grid(row=2, column=1, sticky="e", padx=20, pady=(0, 20))
        self.bind("<Return>", self.cancel)
        self.bind("<Escape>", self.cancel)
        self.result = False
        center(self, 550, 210)
        self.deiconify()
        self.file_name_display.place_cursor()
        self.wait_window()

    def cancel(self, event=None):
        self.destroy()


class Save_New_Version_Error_Popup(tk.Toplevel):
    def __init__(self, C, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", background=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Error locating folder")
        self.grid_columnconfigure(0, weight=1)
        self.info_label = Label(
            self,
            text="There was an error locating the saving folder. \n - Would you like to choose where to save a new version?",
            font=EF,
            theme=theme,
        )
        self.info_label.grid(row=0, column=0, sticky="nswe", padx=20, pady=20)
        self.confirm_button = Button(
            self, text="Choose where to save a new version", style="EF.Std.TButton", command=self.confirm
        )
        self.confirm_button.grid(row=1, column=0, columnspan=2, sticky="nswe", padx=20, pady=(0, 20))
        self.bind("<Return>", self.confirm)
        self.bind("<Escape>", self.cancel)
        self.result = False
        center(self, 550, 130)
        self.deiconify()
        self.wait_window()

    def confirm(self, event=None):
        self.result = True
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Sort_Sheet_Popup(tk.Toplevel):
    def __init__(self, C, headers, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", background=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Sort sheet")
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.sort_decision = {"type": None, "col": None, "order": None}

        self.order_label = Label(self, text="Order:", font=EF, theme=theme)
        self.order_label.grid(row=0, column=0, sticky="nswe", pady=(20, 5), padx=(50, 10))
        self.order_dropdown = Ez_Dropdown(self, EF)
        self.order_dropdown["values"] = ["ASCENDING", "DESCENDING"]
        self.order_dropdown.set_my_value("ASCENDING")
        self.order_dropdown.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 70))
        self.col_label = Label(self, text="Column:", font=EF, theme=theme)
        self.col_label.grid(row=1, column=0, sticky="nswe", pady=5, padx=(50, 10))
        self.col_dropdown = Ez_Dropdown(self, EF)
        self.col_dropdown["values"] = headers
        self.col_dropdown.set_my_value(headers[0])
        self.col_dropdown.grid(row=1, column=1, sticky="nswe", pady=5, padx=(0, 70))
        self.sort_by_col_button = Button(self, style="EF.Std.TButton", text="Sort by column", command=self.sort_by_col)
        self.sort_by_col_button.grid(row=2, column=1, sticky="nswe", pady=(15, 5), padx=70)
        self.divider = Frame(self)
        self.divider.config(bg=themes[theme].table_fg)
        self.divider.config(height=5)
        self.divider.grid(row=3, column=0, columnspan=2, padx=20, pady=(10, 15), sticky="ew")
        self.sort_by_tree_button = Button(
            self, style="EF.Std.TButton", text="Sort by tree walk", command=self.sort_by_tree
        )
        self.sort_by_tree_button.grid(row=4, column=1, sticky="nswe", pady=(20, 20), padx=70)
        self.bind("<Escape>", self.go_back)
        self.order_dropdown.bind("<<ComboboxSelected>>", lambda event: self.focus_set())
        self.col_dropdown.bind("<<ComboboxSelected>>", lambda event: self.focus_set())
        center(self, 550, 235)
        self.deiconify()
        self.wait_window()

    def sort_by_col(self, event=None):
        self.sort_decision = {
            "type": "by column",
            "col": self.col_dropdown.get_my_value(),
            "order": self.order_dropdown.get_my_value(),
        }
        self.destroy()

    def sort_by_tree(self, event=None):
        self.sort_decision["type"] = "by tree"
        self.destroy()

    def go_back(self, event=None):
        self.destroy()


class Edit_Detail_Date_Popup(tk.Toplevel):
    def __init__(self, C, ID, column, current_detail, DATE_FORM, validation_values=[], set_value=None, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Change date detail")
        self.grid_columnconfigure(1, weight=1)
        self.id_label = Label(self, text="ID:", font=EF, theme=theme)
        self.id_label.grid(row=0, column=0, sticky="nswe", padx=20)
        self.id_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.col_label = Label(self, text="Column:", font=EF, theme=theme)
        self.col_label.grid(row=2, column=0, sticky="nswe", padx=20)
        self.col_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2, column=1, sticky="nswe", pady=5, padx=(0, 20))

        self.bf = Frame(self, theme=theme)
        self.bf.grid(row=4, column=0, columnspan=2, sticky="e")

        if validation_values:
            self.validation_dropdown = Ez_Dropdown(self, font=EF)
            self.validation_dropdown["values"] = validation_values
            if set_value is not None:
                self.validation_dropdown.set_my_value(set_value)
            else:
                self.validation_dropdown.set_my_value(validation_values[0])
            self.validation_dropdown.grid(row=3, column=0, columnspan=2, sticky="nswe", padx=20, pady=10)
            self.validation_dropdown.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
            width_ = 600
            height_ = 225
            self.bind("<Return>", self.confirm_validation)
        else:
            self.entries_frame = Frame(self, theme=theme)
            self.entries_frame.grid_columnconfigure(3, weight=1)
            self.entries_frame.grid(row=3, column=0, columnspan=2, sticky="nswe", pady=10)
            if DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
                self.date_label = Label(self.entries_frame, text="Set date DD/MM/YYYY:", font=EF, theme=theme)
            elif DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
                self.date_label = Label(self.entries_frame, text="Set date YYYY/MM/DD:", font=EF, theme=theme)
            elif DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
                self.date_label = Label(self.entries_frame, text="Set date MM/DD/YYYY:", font=EF, theme=theme)
            self.date_label.grid(row=0, column=0, sticky="nswe", padx=(20, 10), pady=10)
            self.date_entry_widget = Date_Entry(self.entries_frame, DATE_FORM, theme=theme)
            self.date_entry_widget.grid(row=0, column=1, sticky="nswe", padx=(0, 30), pady=10)
            self.numerical_label = Label(self.entries_frame, text="OR set Number:", font=EF, theme=theme)
            self.numerical_label.grid(row=0, column=2, sticky="nswe", padx=(0, 10), pady=10)
            self.numerical_entry_widget = Numerical_Entry_With_Scrollbar(self.entries_frame, theme=theme)
            self.numerical_entry_widget.grid(row=0, column=3, sticky="nswe", padx=(0, 20), pady=15)
            if "/" in current_detail or "-" in current_detail:
                self.date_entry_widget.set_my_value(current_detail)
            else:
                self.numerical_entry_widget.set_my_value(current_detail)
            self.numerical_entry_widget.my_entry.bind("<Return>", self.confirm_normal)
            self.date_entry_widget.entry_1.bind("<Return>", self.confirm_normal)
            self.date_entry_widget.entry_2.bind("<Return>", self.confirm_normal)
            self.date_entry_widget.entry_3.bind("<Return>", self.confirm_normal)
            width_ = 850
            height_ = 280

        self.confirm_button = Button(
            self.bf,
            text="Save",
            style="EF.Std.TButton",
            command=self.confirm_validation if validation_values else self.confirm_normal,
        )
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=20, pady=(0, 20))
        self.cancel_button = Button(self.bf, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=20, pady=(0, 20))

        self.result = False
        center(self, width_, height_)
        self.deiconify()
        self.bind("<Escape>", self.cancel)
        if not validation_values:
            self.date_entry_widget.place_cursor()
        self.wait_window()

    def confirm_normal(self, event=None):
        self.result = True
        x1 = self.date_entry_widget.get_my_value()
        x2 = self.numerical_entry_widget.get_my_value()
        if not all(c in ("/", "-") for c in x1):
            self.saved_string = x1
        elif x2:
            self.saved_string = x2
        else:
            self.saved_string = ""
        self.destroy()

    def confirm_validation(self, event=None):
        self.result = True
        self.saved_string = self.validation_dropdown.get_my_value()
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Edit_Detail_Numerical_Popup(tk.Toplevel):
    def __init__(self, C, ID, column, current_detail, validation_values=[], set_value=None, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Change numerical detail")
        self.grid_columnconfigure(1, weight=1)
        self.id_label = Label(self, text="ID:", font=EF, theme=theme)
        self.id_label.grid(row=0, column=0, sticky="nswe", padx=20)
        self.id_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.col_label = Label(self, text="Column:", font=EF, theme=theme)
        self.col_label.grid(row=2, column=0, sticky="nswe", padx=20)
        self.col_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2, column=1, sticky="nswe", pady=5, padx=(0, 20))

        self.bf = Frame(self, theme=theme)
        self.bf.grid(row=4, column=0, columnspan=2, sticky="e")

        if validation_values:
            self.validation_dropdown = Ez_Dropdown(self, font=EF)
            self.validation_dropdown["values"] = validation_values
            if set_value is not None:
                self.validation_dropdown.set_my_value(set_value)
            else:
                self.validation_dropdown.set_my_value(validation_values[0])
            self.validation_dropdown.grid(row=3, column=0, columnspan=2, sticky="nswe", padx=20, pady=10)
            self.validation_dropdown.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
            width_ = 600
            height_ = 225
            self.bind("<Return>", self.confirm_validation)
        else:
            self.entry_widget = Numerical_Entry_With_Scrollbar(self, theme=theme)
            self.entry_widget.set_my_value(current_detail)
            self.entry_widget.grid(row=3, column=0, columnspan=2, sticky="nswe", padx=20, pady=10)
            self.entry_widget.my_entry.bind("<Return>", self.confirm_normal)
            width_ = 600
            height_ = 240

        self.confirm_button = Button(
            self.bf,
            text="Save",
            style="EF.Std.TButton",
            command=self.confirm_validation if validation_values else self.confirm_normal,
        )
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=20, pady=20)
        self.cancel_button = Button(self.bf, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=20, pady=20)

        self.result = False
        center(self, width_, height_)
        self.deiconify()
        self.bind("<Escape>", self.cancel)
        if not validation_values:
            self.entry_widget.place_cursor()
        self.wait_window()

    def confirm_normal(self, event=None):
        self.result = True
        self.saved_string = self.entry_widget.get_my_value()
        self.destroy()

    def confirm_validation(self, event=None):
        self.result = True
        self.saved_string = self.validation_dropdown.get_my_value()
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Edit_Detail_Text_Popup(tk.Toplevel):
    def __init__(self, C, ID, column, current_detail, validation_values=[], set_value=None, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1")
        self.C = new_toplevel_chores(self, C, f"{app_title} - Edit cell")
        self.grid_columnconfigure(1, weight=1)
        self.id_label = Label(self, text="ID:", font=EF, theme=theme)
        self.id_label.grid(row=0, column=0, sticky="nswe", padx=20)
        self.id_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.col_label = Label(self, text="Column:", font=EF, theme=theme)
        self.col_label.grid(row=2, column=0, sticky="nswe", padx=20)
        self.col_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2, column=1, sticky="nswe", pady=5, padx=(0, 20))

        self.bf = Frame(self, theme=theme)
        self.bf.grid(row=4, column=1, sticky="e")
        if validation_values:
            self.validation_dropdown = Ez_Dropdown(self, font=EF)
            self.validation_dropdown["values"] = validation_values
            if set_value is not None:
                self.validation_dropdown.set_my_value(set_value)
            else:
                self.validation_dropdown.set_my_value(validation_values[0])
            self.validation_dropdown.grid(row=3, column=0, columnspan=2, sticky="nswe", padx=20, pady=10)
            self.validation_dropdown.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
            width_ = 620
            height_ = 225
            self.confirm_button = Button(self.bf, text="Save", style="EF.Std.TButton", command=self.confirm_validation)
            self.bind("<Return>", self.confirm_validation)
        else:
            self.grid_rowconfigure(3, weight=1)
            self.text_widget = Wrapped_Text_With_Find_And_Yscroll(self, current_detail, "normal", 15, theme=theme)
            self.text_widget.grid(row=3, column=0, sticky="nswe", columnspan=2)
            width_ = 800
            height_ = 595
            self.confirm_button = Button(self.bf, text="Save", style="EF.Std.TButton", command=self.confirm_normal)
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=20, pady=20)
        self.cancel_button = Button(self.bf, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=20, pady=20)
        center(self, width_, height_)
        self.result = False
        self.deiconify()
        self.grab_set()
        self.bind("<Escape>", self.cancel)
        if not validation_values:
            self.text_widget.place_cursor()
        self.wait_window()

    def confirm_normal(self, event=None):
        self.result = True
        self.saved_string = self.text_widget.get_my_value().rstrip()
        self.destroy()

    def confirm_validation(self, event=None):
        self.result = True
        self.saved_string = self.validation_dropdown.get_my_value()
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class View_Column_Text_Popup(tk.Toplevel):
    def __init__(self, C, ID, column, text, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - View text")
        self.grid_columnconfigure(1, weight=1)
        self.id_label = Label(self, text="ID:", font=EF, theme=theme)
        self.id_label.grid(row=0, column=0, sticky="nswe", padx=20)
        self.id_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.col_label = Label(self, text="Column:", font=EF, theme=theme)
        self.col_label.grid(row=2, column=0, sticky="nswe", padx=20)
        self.col_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2, column=1, sticky="nswe", pady=5, padx=(0, 20))
        self.text_widget = Wrapped_Text_With_Find_And_Yscroll(self, text, "disabled", 15, theme=theme)
        self.text_widget.grid(row=3, column=0, sticky="nswe", columnspan=2)
        self.cancel_button = Button(self, text="Close", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=4, column=0, columnspan=2, sticky="nswe", padx=220, pady=(25, 20))
        self.bind("<Escape>", self.cancel)
        self.result = False
        center(self, 850, 545)
        self.deiconify()
        self.text_widget.place_cursor()
        self.wait_window()

    def cancel(self, event=None):
        self.destroy()


class Add_Top_Id_Popup(tk.Toplevel):
    def __init__(self, C, sheet_selection, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Add top ID")
        self.grid_columnconfigure(1, weight=1)
        self.id_name_label = Label(self, text="ID name:", font=EF, theme=theme)
        self.id_name_label.grid(row=0, column=0, sticky="nswe", padx=20)
        self.id_name_display = Entry_With_Scrollbar(self, theme=theme)
        self.id_name_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.id_tv_label = Label(self, text="ID Treeview Label:", font=EF, theme=theme)
        self.id_tv_display = Entry_With_Scrollbar(self, theme=theme)
        if self.C.tv_label_col != self.C.ic:
            self.id_tv_label.grid(row=1, column=0, sticky="nswe", padx=20)
            self.id_tv_display.grid(row=1, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.enter_sheet_sel_button = Button(
            self,
            text="Enter current sheet selection",
            style="EF.Std.TButton",
            command=self.enter_sheet_sel,
        )
        self.enter_sheet_sel_button.grid(row=2, column=1, sticky="e", padx=(0, 20), pady=5)
        if not sheet_selection:
            self.enter_sheet_sel_button.config(state="disabled")
        else:
            self.sheet_sel = sheet_selection

        self.bf = Frame(self, theme=theme)
        self.bf.grid(row=3, column=0, columnspan=2, sticky="e")

        self.confirm_button = Button(self.bf, text="Add", style="EF.Std.TButton", command=self.confirm)
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=(0, 20), pady=(30, 20))
        self.cancel_button = Button(self.bf, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=(0, 20), pady=(30, 20))

        self.bind("<Return>", self.confirm)
        self.bind("<Escape>", self.cancel)
        self.result = False
        center(self, 600, 186 if self.C.tv_label_col == self.C.ic else 250)
        self.deiconify()
        self.id_name_display.place_cursor()
        self.wait_window()

    def confirm(self, event=None):
        if self.C.allow_spaces_ids_var:
            self.result = self.id_name_display.get_my_value()
            if self.C.tv_label_col != self.C.ic:
                self.id_label = self.id_tv_display.get_my_value()
            else:
                self.id_label = self.result
        else:
            self.result = "".join(self.id_name_display.get_my_value().strip().split())
            if self.C.tv_label_col != self.C.ic:
                self.id_label = "".join(self.id_tv_display.get_my_value().strip().split())
            else:
                self.id_label = self.result
        self.destroy()

    def enter_sheet_sel(self, event=None):
        self.id_name_display.set_my_value(self.sheet_sel)
        if self.C.tv_label_col != self.C.ic:
            detail = self.C.sheet.MT.data[self.C.rns[self.sheet_sel.lower()]][self.C.tv_label_col]
            ni = detail.find("\n")
            if ni == -1:
                self.id_tv_display.set_my_value(detail)
            else:
                self.id_tv_display.set_my_value(detail[:ni])

    def cancel(self, event=None):
        self.destroy()


class Add_Child_Or_Sibling_Id_Popup(tk.Toplevel):
    def __init__(self, C, chld_or_sib, desired_parent, sheet_selection, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", background=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(
            self, C, f"{app_title} - Add sibling ID" if chld_or_sib == "sibling" else f"{app_title} - Add child ID"
        )
        self.grid_columnconfigure(1, weight=1)
        self.parent_label = Label(self, text="Parent:", font=EF, theme=theme)
        self.parent_label.grid(row=0, column=0, sticky="nswe", padx=20)
        self.parent_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        if desired_parent:
            self.parent_display.set_my_value(desired_parent)
        else:
            self.parent_display.set_my_value("** No parent - Top ID **")
        self.parent_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.id_name_label = Label(self, text="ID name:", font=EF, theme=theme)
        self.id_name_label.grid(row=1, column=0, sticky="nswe", padx=20)
        self.id_name_display = Entry_With_Scrollbar(self, theme=theme)
        self.id_name_display.grid(row=1, column=1, sticky="nswe", pady=(5, 10), padx=(0, 20))
        self.id_tv_label = Label(self, text="ID Treeview Label:", font=EF, theme=theme)
        self.id_tv_display = Entry_With_Scrollbar(self, theme=theme)
        if self.C.tv_label_col != self.C.ic:
            self.id_tv_label.grid(row=2, column=0, sticky="nswe", padx=20)
            self.id_tv_display.grid(row=2, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.enter_sheet_sel_button = Button(
            self,
            text="Enter current sheet selection",
            style="EF.Std.TButton",
            command=self.enter_sheet_sel,
        )
        self.enter_sheet_sel_button.grid(row=3, column=1, sticky="e", padx=(0, 20), pady=5)
        if not sheet_selection:
            self.enter_sheet_sel_button.config(state="disabled")
        else:
            self.sheet_sel = sheet_selection

        self.bf = Frame(self, theme=theme)
        self.bf.grid(row=4, column=0, columnspan=2, sticky="e")

        self.confirm_button = Button(self.bf, text="Add", style="EF.Std.TButton", command=self.confirm)
        self.confirm_button.grid(row=0, column=0, sticky="nswe", padx=(0, 20), pady=(30, 20))
        self.cancel_button = Button(self.bf, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="nswe", padx=(0, 20), pady=(30, 20))

        self.bind("<Return>", self.confirm)
        self.bind("<Escape>", self.cancel)
        self.result = False
        center(self, 600, 237 if self.C.tv_label_col == self.C.ic else 305)
        self.deiconify()
        self.id_name_display.place_cursor()
        self.wait_window()

    def confirm(self, event=None):
        if self.C.allow_spaces_ids_var:
            self.result = self.id_name_display.get_my_value()
            if self.C.tv_label_col != self.C.ic:
                self.id_label = self.id_tv_display.get_my_value()
            else:
                self.id_label = self.result
        else:
            self.result = "".join(self.id_name_display.get_my_value().strip().split())
            if self.C.tv_label_col != self.C.ic:
                self.id_label = "".join(self.id_tv_display.get_my_value().strip().split())
            else:
                self.id_label = self.result
        self.destroy()

    def enter_sheet_sel(self, event=None):
        self.id_name_display.set_my_value(self.sheet_sel)
        if self.C.tv_label_col != self.C.ic:
            detail = self.C.sheet.MT.data[self.C.rns[self.sheet_sel.lower()]][self.C.tv_label_col]
            ni = detail.find("\n")
            if ni == -1:
                self.id_tv_display.set_my_value(detail)
            else:
                self.id_tv_display.set_my_value(detail[:ni])

    def cancel(self, event=None):
        self.destroy()


class Edit_Validation_Popup(tk.Toplevel):
    def __init__(self, C, coltype, colname, validation, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Edit validation")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        if coltype == "Numerical Detail":
            self.allowed_chars = validation_allowed_num_chars
        elif coltype == "Date Detail":
            self.allowed_chars = validation_allowed_date_chars
        else:
            self.allowed_chars = set()  # all chars allowed
        self.validation_display = Sheet(
            self,
            name="validation",
            auto_resize_row_index=True,
            header=[f"Valid {colname} values:"],
            header_font=sheet_header_font,
            theme=theme,
        )
        if validation:
            self.validation_display.data = [[v] for v in validation]
        self.validation_display.insert_rows(100, create_selections=False)
        self.validation_display.edit_validation(self.edit_validation)
        self.validation_display.enable_bindings("all", "ctrl_select")
        self.validation_display.disable_bindings(
            "insert_columns",
            "delete_columns",
        )
        self.validation_display.set_all_cell_sizes_to_text()
        self.validation_display.grid(row=0, column=0, sticky="nswe", pady=(0, 20))
        self.button_frame = Frame(self, theme=theme)
        self.button_frame.grid(row=1, column=0, columnspan=2, sticky="e")
        self.button_frame.grid_rowconfigure(0, weight=1)
        self.confirm_button = Button(
            self.button_frame,
            text=" Save validation ",
            style="EF.Std.TButton",
            command=self.confirm,
        )
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=(20, 10), pady=(0, 20))
        self.cancel_button = Button(self.button_frame, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=(10, 20), pady=(0, 20))
        self.new_validation = ""
        center(self, 600, 500)
        self.deiconify()
        self.validation_display.select_cell(0, 0)
        self.validation_display.focus_set()
        self.wait_window()

    def edit_validation(self, event):
        if self.allowed_chars:
            return "".join(filter(self.allowed_chars.__contains__, event.value))
        return event.value

    def confirm(self, event=None):
        self.result = True
        self.new_validation = ",".join(filter(None, map(lambda row: row[0], self.validation_display.data)))
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Rename_Column_Popup(tk.Toplevel):
    def __init__(self, C, current_col_name, type_of_col, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"Rename {type_of_col} column")
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.col_label = Label(self, text="Current column\nname:", font=EF, theme=theme)
        self.col_label.grid(row=0, column=0, sticky="nswe", padx=20)
        self.col_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.col_display.set_my_value(current_col_name)
        self.col_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.new_name_label = Label(self, text="New column\nname:", font=EF, theme=theme)
        self.new_name_label.grid(row=1, column=0, sticky="nswe", padx=20)
        self.new_name_display = Entry_With_Scrollbar(self, theme=theme)
        self.new_name_display.grid(row=1, column=1, sticky="nswe", pady=5, padx=(0, 20))
        self.button_frame = Frame(self, theme=theme)
        self.button_frame.grid(row=2, column=0, columnspan=2, sticky="e")
        self.button_frame.grid_rowconfigure(0, weight=1)
        self.confirm_button = Button(self.button_frame, text="Confirm", style="EF.Std.TButton", command=self.confirm)
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=(20, 10), pady=(10, 20))
        self.cancel_button = Button(self.button_frame, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=(10, 20), pady=(10, 20))
        self.result = False
        self.bind("<Return>", self.confirm)
        self.bind("<Escape>", self.cancel)
        center(self, 600, 180)
        self.deiconify()
        self.new_name_display.place_cursor()
        self.wait_window()

    def confirm(self, event=None):
        if self.C.allow_spaces_columns_var:
            self.result = self.new_name_display.get_my_value()
        else:
            self.result = "".join(self.new_name_display.get_my_value().strip().split())
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Add_Hierarchy_Column_Popup(tk.Toplevel):
    def __init__(self, C, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Add Hierarchy Column")
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.hier_name_label = Label(self, text="New hierarchy\nname:", font=EF, theme=theme)
        self.hier_name_label.grid(row=0, column=0, sticky="nswe", padx=20)
        self.hier_name_display = Entry_With_Scrollbar(self, theme=theme)
        self.hier_name_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.button_frame = Frame(self, theme=theme)
        self.button_frame.grid(row=1, column=1, sticky="nse")
        self.button_frame.grid_rowconfigure(0, weight=1)
        self.confirm_button = Button(
            self.button_frame,
            text="Add Hierarchy Column",
            style="EF.Std.TButton",
            command=self.confirm,
        )
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=(20, 10), pady=(10, 20))
        self.cancel_button = Button(self.button_frame, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=(10, 20), pady=(10, 20))
        self.result = False
        self.bind("<Return>", self.confirm)
        self.bind("<Escape>", self.cancel)
        center(self, 600, 150)
        self.deiconify()
        self.hier_name_display.place_cursor()
        self.wait_window()

    def confirm(self, event=None):
        if self.C.allow_spaces_columns_var:
            self.hier_name_display.get_my_value()
        else:
            self.result = "".join(self.hier_name_display.get_my_value().strip().split())
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Add_Detail_Column_Popup(tk.Toplevel):
    def __init__(self, C, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Add Detail Column")
        self.grid_columnconfigure(2, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.type_display = Ez_Dropdown(self, EF)
        self.type_display["values"] = ("Text Detail", "Numerical Detail", "Date Detail")
        self.type_display.set_my_value("Text Detail")
        self.type_display.grid(row=0, column=0, sticky="nswe", padx=(20, 0), pady=(20, 5))
        self.type_display.bind("<<ComboboxSelected>>", lambda focus: self.detail_name_display.place_cursor())
        self.detail_name_label = Label(self, text="New detail\ncolumn name:", font=EF, theme=theme)
        self.detail_name_label.grid(row=0, column=1, sticky="nswe", padx=20, pady=(20, 5))
        self.detail_name_display = Entry_With_Scrollbar(self, theme=theme)
        self.detail_name_display.grid(row=0, column=2, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.button_frame = Frame(self, theme=theme)
        self.button_frame.grid(row=1, column=1, columnspan=2, sticky="nse")
        self.button_frame.grid_rowconfigure(0, weight=1)
        self.confirm_button = Button(
            self.button_frame,
            text="Add Detail Column",
            style="EF.Std.TButton",
            command=self.confirm,
        )
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=(20, 10), pady=(10, 20))
        self.cancel_button = Button(self.button_frame, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=(10, 20), pady=(10, 20))
        self.result = False
        self.type_ = "Text Detail"
        self.bind("<Return>", self.confirm)
        self.bind("<Escape>", self.cancel)
        center(self, 600, 155)
        self.deiconify()
        self.detail_name_display.place_cursor()
        self.wait_window()

    def confirm(self, event=None):
        if self.C.allow_spaces_columns_var:
            self.result = self.detail_name_display.get_my_value()
        else:
            self.result = "".join(self.detail_name_display.get_my_value().strip().split())
        self.type_ = self.type_display.get()
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Rename_Id_Popup(tk.Toplevel):
    def __init__(self, C, ID, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Rename ID")
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.id_label = Label(self, text="Current ID\nname:", font=EF, theme=theme)
        self.id_label.grid(row=0, column=0, sticky="nswe", padx=20)
        self.id_display = Readonly_Entry_With_Scrollbar(self, theme=theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.new_name_label = Label(self, text="New ID\nname:", font=EF, theme=theme)
        self.new_name_label.grid(row=1, column=0, sticky="nswe", padx=20)
        self.new_name_display = Entry_With_Scrollbar(self, theme=theme)
        self.new_name_display.grid(row=1, column=1, sticky="nswe", pady=5, padx=(0, 20))

        self.bf = Frame(self, theme=theme)
        self.bf.grid(row=2, column=0, columnspan=2, sticky="e")

        self.confirm_button = Button(self.bf, text="Rename", style="EF.Std.TButton", command=self.confirm)
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=20, pady=(20, 20))
        self.cancel_button = Button(self.bf, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=20, pady=(20, 20))

        self.bind("<Return>", self.confirm)
        self.bind("<Escape>", self.cancel)
        self.result = False
        center(self, 600, 185)
        self.deiconify()
        self.new_name_display.place_cursor()
        self.wait_window()

    def confirm(self, event=None):
        if self.C.allow_spaces_ids_var:
            self.result = self.new_name_display.get_my_value()
        else:
            self.result = "".join(self.new_name_display.get_my_value().strip().split())
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Enter_Sheet_Name_Popup(tk.Toplevel):
    def __init__(self, C, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Enter sheet name")
        self.grid_columnconfigure(1, weight=1)
        self.sheet_name_label = Label(self, text="New sheet\nname:", font=EF, theme=theme)
        self.sheet_name_label.grid(row=0, column=0, sticky="nswe", padx=20)
        self.sheet_entry = Entry_With_Scrollbar(self, theme=theme)
        self.sheet_entry.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))

        self.bf = Frame(self, theme=theme)
        self.bf.grid(row=1, column=0, columnspan=2, sticky="e")

        self.confirm_button = Button(self.bf, text="Confirm", style="EF.Std.TButton", command=self.confirm)
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=20, pady=(20, 20))
        self.cancel_button = Button(self.bf, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=20, pady=(20, 20))

        self.bind("<Return>", self.confirm)
        self.bind("<Escape>", self.cancel)
        self.result = False
        center(self, 600, 137)
        self.deiconify()
        self.sheet_entry.place_cursor()
        self.wait_window()

    def confirm(self, event=None):
        self.result = self.sheet_entry.get_my_value()
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Error(tk.Toplevel):
    def __init__(self, C, msg, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Error")
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.errorlabel = Label(self, text="Error\nmessage:", font=ERR_ASK_FNT, theme=theme)
        self.errorlabel.config(background="red", foreground="white")
        self.errorlabel.grid(row=0, column=0, sticky="nswe", pady=(20, 5), padx=20)
        self.error_display = Display_Text(parent=self, text=msg, theme=theme, bold=True)
        self.error_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.error_display.config(height=75)
        self.confirm_button = Button(self, text="Okay", style="EF.Std.TButton", command=self.cancel)
        self.confirm_button.grid(row=1, column=1, sticky="e", padx=20, pady=(10, 20))
        self.bind("<Return>", self.cancel)
        self.bind("<Escape>", self.cancel)
        center(self, 600, 180)
        self.deiconify()
        self.wait_window()

    def cancel(self, event=None):
        self.destroy()


class Error_Sheet(tk.Toplevel):
    def __init__(self, C, sheet_data, theme="dark", highlight_rows=None):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Error")
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.errorlabel = Label(self, text="Error\nmessage:", font=EF, theme=theme)
        self.errorlabel.config(background="red", foreground="white")
        self.errorlabel.grid(row=0, column=0, sticky="nswe", pady=(20, 5), padx=20)
        self.error_display = Sheet(
            self,
            theme=theme,
            headers=["To Find (items listed in lower case)", "Replace With"],
            header_font=sheet_header_font,
            data=sheet_data,
            outline_thickness=0,
        )
        self.error_display.grid(row=0, column=1, sticky="nswe", pady=(20, 5), padx=(0, 20))
        self.error_display.enable_bindings(
            (
                "single",
                "copy",
                "drag_select",
                "column_width_resize",
                "double_click_column_resize",
                "row_height_resize",
                "double_click_row_resize",
                "row_width_resize",
                "row_select",
                "arrowkeys",
            )
        )
        self.error_display.column_width(column=0, width="text", only_set_if_too_small=False, redraw=False)
        self.error_display.highlight_rows(rows=highlight_rows, bg="#fc8c55", fg="black")
        self.confirm_button = Button(self, text="Okay", style="EF.Std.TButton", command=self.cancel)
        self.confirm_button.grid(row=1, column=1, sticky="e", padx=20, pady=(10, 20))
        self.bind("<Return>", self.cancel)
        self.bind("<Escape>", self.cancel)
        center(self, 600, 420)
        self.deiconify()
        self.wait_window()

    def cancel(self, event=None):
        self.destroy()


class Treeview_Id_Finder(tk.Toplevel):
    def __init__(self, C, hiers, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - ID is in multiple hierarchies")
        self.GO = False
        self.selected = 0
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.dd_1 = Ez_Dropdown(self, EF)
        self.dd_1["values"] = hiers
        self.dd_1.set_my_value(hiers[0])
        self.dd_1.grid(row=0, column=0, columnspan=2, sticky="nswe", padx=20, pady=(20, 5))

        self.bf = Frame(self, theme=theme)
        self.bf.grid(row=1, column=1, sticky="e")

        self.confirm_button = Button(self.bf, text="Go to hierarchy", style="EF.Std.TButton", command=self.confirm)
        self.confirm_button.grid(row=0, column=0, sticky="e", padx=20, pady=(20, 20))
        self.cancel_button = Button(self.bf, text="Cancel", style="EF.Std.TButton", command=self.cancel)
        self.cancel_button.grid(row=0, column=1, sticky="e", padx=20, pady=(20, 20))

        self.bind("<Escape>", self.cancel)
        center(self, 400, 120)
        self.deiconify()
        self.wait_window()

    def confirm(self, event=None):
        self.selected = self.dd_1.displayed.get()
        self.GO = True
        self.destroy()

    def cancel(self, event=None):
        self.destroy()


class Text_Popup(tk.Toplevel):
    def __init__(
        self,
        C,
        text,
        width_=800,
        height_=650,
        theme="dark",
        use_entry_bg=False,
        wrap="none",
        show_finder=True,
        heading="",
    ):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title}")
        self.theme = theme
        self.word = ""
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        if show_finder:
            self.find_frame = Frame(self, theme=theme)
            self.find_frame.grid(row=0, column=0, columnspan=2, sticky="nswe")
            self.search_button = Button(self.find_frame, text=" Find:", command=self.find)
            self.search_button.pack(side="left", fill="x")
            self.find_window = Normal_Entry(self.find_frame, font=BF, theme=theme)
            self.find_window.bind("<Return>", self.find)
            self.find_window.pack(side="left", fill="x", expand=True)
            self.find_reset_button = Button(self.find_frame, text="X", command=self.find_reset)
            self.find_reset_button.pack(side="left", fill="x")
            self.find_results_label = Label(self.find_frame, "0/0", BF, theme=theme)
            self.find_results_label.pack(side="left", fill="x")
            self.find_up_button = Button(self.find_frame, text="▲", command=self.find_up)
            self.find_up_button.pack(side="left", fill="x")
            self.find_down_button = Button(self.find_frame, text="▼", command=self.find_down)
            self.find_down_button.pack(side="left", fill="x")
        if heading:
            self.heading_label = Label(self, text=f"{heading}", font=EF, theme=theme)
            self.heading_label.grid(row=1, column=0, columnspan=2, sticky="we")
        self.textbox = Working_Text(
            self,
            wrap=wrap,
            theme=theme,
            use_entry_bg=use_entry_bg,
            override_bg=None,
        )
        self.yscrollb = Scrollbar(self, self.textbox.yview, "vertical", self.textbox)
        self.xscrollb = Scrollbar(self, self.textbox.xview, "horizontal", self.textbox)
        self.textbox.delete(1.0, "end")
        self.textbox.insert(1.0, text)
        self.textbox.config(state="disabled")
        self.textbox.grid(row=2, column=0, padx=(5, 0), sticky="nswe")
        self.yscrollb.grid(row=2, column=1, sticky="nswe")
        if wrap == "none":
            self.xscrollb.grid(row=3, column=0, columnspan=2, sticky="nswe")
        self.buttonframe = Frame(self, theme=theme)
        self.buttonframe.grid(row=4, column=0, columnspan=2, sticky="nswe")
        self.copy_text_button = Button(
            self.buttonframe,
            text="Copy",
            style="EF.Std.TButton",
            command=lambda: self.copy_text(text),
        )
        self.copy_text_button.pack(side="right", fill="x", padx=20, pady=20)
        self.save_text_button = Button(
            self.buttonframe,
            text="Save as",
            style="EF.Std.TButton",
            command=lambda: self.save_text(text),
        )
        self.save_text_button.pack(side="right", fill="x", padx=20, pady=20)
        self.cancel_button = Button(
            self.buttonframe,
            text="Done",
            style="EF.Std.TButton",
            command=self.cancel,
        )
        self.cancel_button.pack(side="right", fill="x", padx=20, pady=20)
        self.bind("<Escape>", self.cancel)
        center(self, width_, height_)
        self.deiconify()
        self.wait_window()

    def copy_text(self, text):
        to_clipboard(self, text)

    def save_text(self, text):
        newfile = filedialog.asksaveasfilename(
            parent=self,
            title="Save as",
            filetypes=[("Text File", ".txt"), ("CSV File", ".csv")],
            defaultextension=".txt",
            confirmoverwrite=True,
        )
        if not newfile:
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv", ".txt")):
            Error(self, "Can only save .csv/.txt files", theme=self.theme)
            self.grab_set()
            return
        try:
            with open(newfile, "w") as fh:
                fh.writelines(text)
        except Exception:
            Error(self, "Error saving file", theme=self.theme)
            self.grab_set()
            return

    def find(self, event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        self.addon = f"+{len(self.word)}c"
        start = "1.0"
        while start:
            start = self.textbox.search(self.word, index=start, nocase=1, stopindex="end")
            if start:
                end = start + self.addon
                self.find_results.append(start)
                self.textbox.tag_add("i", start, end)
                start = end
        if self.find_results:
            self.textbox.tag_config("i", background="Yellow")
            self.find_results_label.config(text=f"1/{len(self.find_results)}")
            self.textbox.tag_add(
                "c",
                self.find_results[self.results_number],
                self.find_results[self.results_number] + self.addon,
            )
            self.textbox.tag_config("c", background="Orange")
            self.textbox.see(self.find_results[self.results_number])

    def find_up(self, event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove(
            "c",
            self.find_results[self.results_number],
            self.find_results[self.results_number] + self.addon,
        )
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=f"{self.results_number + 1}/{len(self.find_results)}")
        self.textbox.tag_add(
            "c",
            self.find_results[self.results_number],
            self.find_results[self.results_number] + self.addon,
        )
        self.textbox.tag_config("c", background="Orange")
        self.textbox.see(self.find_results[self.results_number])

    def find_down(self, event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove(
            "c",
            self.find_results[self.results_number],
            self.find_results[self.results_number] + self.addon,
        )
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=f"{self.results_number + 1}/{len(self.find_results)}")
        self.textbox.tag_add(
            "c",
            self.find_results[self.results_number],
            self.find_results[self.results_number] + self.addon,
        )
        self.textbox.tag_config("c", background="Orange")
        self.textbox.see(self.find_results[self.results_number])

    def find_reset(self, newfind=False):
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        if not newfind:
            self.find_window.delete(0, "end")
        for tag in self.textbox.tag_names():
            self.textbox.tag_delete(tag)
        self.find_results_label.config(text="0/0")

    def cancel(self, event=None):
        self.destroy()


class First_Start_Popup(tk.Toplevel):
    def __init__(
        self,
        C,
        text,
        width_=720,
        height_=570,
        theme="dark",
        use_entry_bg=False,
        wrap="word",
    ):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title}")
        self.theme = theme
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.textbox = Working_Text(
            self,
            wrap=wrap,
            font=("Calibri", lge_font_size, "normal"),
            theme=theme,
            use_entry_bg=use_entry_bg,
            override_bg=None,
            highlightthickness=0,
        )
        self.textbox.delete(1.0, "end")
        self.textbox.insert(1.0, text)
        self.textbox.config(state="disabled")
        self.textbox.grid(row=1, column=0, padx=20, sticky="nswe")
        self.buttonframe = Frame(self, theme=theme)
        self.buttonframe.grid(row=3, column=0, columnspan=2, sticky="nswe")
        self.cancel_button = Button(
            self.buttonframe,
            text="Okay",
            style="EF.Std.TButton",
            command=self.cancel,
        )
        self.cancel_button.pack(side="right", fill="x", padx=20, pady=20)
        self.bind("<Escape>", self.cancel)
        center(self, width_, height_)
        self.deiconify()
        self.wait_window()

    def cancel(self, event=None):
        self.destroy()


class License_Popup(tk.Toplevel):
    def __init__(self, C, text, show_buttons=True, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} license")
        self.has_agreed = False
        self.grid_columnconfigure(0, weight=1, uniform="y")
        self.grid_columnconfigure(1, weight=1, uniform="y")
        self.grid_rowconfigure(0, weight=1)
        self.textbox = Working_Text(
            self,
            wrap="word",
            theme=theme,
            use_entry_bg=False,
            override_bg=None,
        )
        self.textbox.config(height=18)
        self.yscrollb = Scrollbar(self, self.textbox.yview, "vertical", self.textbox)
        self.textbox.delete(1.0, "end")
        self.textbox.insert(1.0, text)
        self.textbox.config(state="disabled")
        self.textbox.grid(row=0, column=0, columnspan=2, padx=(20, 0), pady=20, sticky="nswe")
        self.yscrollb.grid(row=0, column=2, padx=(0, 20), pady=20, sticky="nswe")
        if show_buttons:
            self.bframe = Frame(self, theme=theme)
            self.bframe.grid(row=1, column=0, columnspan=3, padx=20, pady=(0, 20), sticky="e")
            self.agree_button = Button(
                self.bframe,
                text="Agree",
                underline=0,
                style="EF.Std.TButton",
                command=self.agree,
            )
            self.agree_button.grid(row=0, column=0, sticky="e", padx=20)
            self.disagree_button = Button(
                self.bframe,
                text="Disagree",
                underline=0,
                style="EF.Std.TButton",
                command=self.disagree,
            )
            self.disagree_button.grid(row=0, column=1, sticky="e", padx=10)
            self.bind("<A>", self.agree)
            self.bind("<a>", self.agree)
            self.bind("<D>", self.disagree)
            self.bind("<d>", self.disagree)
        self.bind("<Escape>", self.disagree)
        center(self, 700, 650)
        self.deiconify()
        self.wait_window()

    def agree(self, event=None):
        self.has_agreed = True
        self.destroy()

    def disagree(self, event=None):
        self.destroy()


class Settings_Popup(tk.Toplevel):
    def __init__(self, C, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} - Settings")
        self.grid_rowconfigure(0, weight=1)

        self.general = Frame(self, theme=theme)
        self.general.grid(row=0, column=1, pady=20, padx=20, sticky="nswe")

        self.general_header = Label(self.general, text="General Settings", font=TF, theme=theme, anchor="nw")
        self.general_header.pack(side="top", anchor="nw", fill="x", pady=(0, 20))

        self.auto_resize_indexes_button = X_Checkbutton(
            self.general,
            text="Auto-resize row indexes ",
            style="x_button.Std.TButton",
            command=self.toggle_auto_resize_indexes,
            checked=self.C.auto_resize_indexes,
            compound="right",
        )
        self.auto_resize_indexes_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.auto_select_id_button = X_Checkbutton(
            self.general,
            text="Auto-select sheet ID ",
            style="x_button.Std.TButton",
            command=self.toggle_auto_select,
            checked=self.C.mirror_var,
            compound="right",
        )
        self.auto_select_id_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.allow_spaces_ids_button = X_Checkbutton(
            self.general,
            text="Allow spaces in ID names ",
            style="x_button.Std.TButton",
            command=self.toggle_allow_spaces_ids,
            checked=self.C.allow_spaces_ids_var,
            compound="right",
        )
        self.allow_spaces_ids_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.allow_spaces_columns_button = X_Checkbutton(
            self.general,
            text="Allow spaces in column names ",
            style="x_button.Std.TButton",
            command=self.toggle_allow_spaces_columns,
            checked=self.C.allow_spaces_columns_var,
            compound="right",
        )
        self.allow_spaces_columns_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.auto_sort_tree_button = X_Checkbutton(
            self.general,
            text="Auto-sort tree IDs ",
            style="x_button.Std.TButton",
            command=self.toggle_auto_sort_tree_button,
            checked=self.C.auto_sort_nodes_bool,
            compound="right",
        )
        self.auto_sort_tree_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.show_tv_lvls_button = X_Checkbutton(
            self.general,
            text="Show ID level numbering ",
            style="x_button.Std.TButton",
            command=self.toggle_show_tv_lvls,
            checked=self.C.tv_lvls_bool,
            compound="right",
        )
        self.show_tv_lvls_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.layout_label = Label(self.general, text="Layout: ", font=EF, theme=theme, anchor="nw")
        self.layout_label.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.layout_dropdown = Ez_Dropdown(self.general, font=EF)
        self.layout_dropdown["values"] = [
            "Display Only Tree",
            "Display Only Sheet",
            "50/50 Tree/Sheet",
            "Adjustable Tree/Sheet",
        ]
        if self.C.full_left_bool.get():
            self.layout_dropdown.set_my_value("Display Only Tree")

        elif self.C.full_right_bool.get():
            self.layout_dropdown.set_my_value("Display Only Sheet")

        elif self.C._50_50_bool.get():
            self.layout_dropdown.set_my_value("50/50 Tree/Sheet")

        elif self.C.adjustable_bool.get():
            self.layout_dropdown.set_my_value("Adjustable Tree/Sheet")

        self.layout_dropdown.bind("<<ComboboxSelected>>", self.set_layout)
        self.layout_dropdown.pack(side="top", anchor="nw", fill="x", pady=10)

        self.indent_label = Label(self.general, text="Treeview Level Indent: ", font=EF, theme=theme, anchor="nw")
        self.indent_label.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.indent_dropdown = Ez_Dropdown(self.general, font=EF)
        self.indent_dropdown["values"] = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]
        self.indent_dropdown.set_my_value(self.C.tree.ops.treeview_indent)

        self.indent_dropdown.bind("<<ComboboxSelected>>", self.set_treeview_indent)
        self.indent_dropdown.pack(side="top", anchor="nw", fill="x", pady=10)

        self.xlsx = Frame(self, theme=theme)
        self.xlsx.grid(row=0, column=1, pady=20, padx=20, sticky="nswe")

        self.xlsx_header = Label(self.xlsx, text="xlsx Save Settings", font=TF, theme=theme, anchor="nw")
        self.xlsx_header.pack(side="top", anchor="nw", fill="x", pady=(0, 20))

        self.xlsx_app_data_button = X_Checkbutton(
            self.xlsx,
            text="Save with app data ",
            style="x_button.Std.TButton",
            command=self.toggle_xlsx_app_data,
            checked=self.C.save_xlsx_with_program_data,
            compound="right",
        )
        self.xlsx_app_data_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.xlsx_changelog_button = X_Checkbutton(
            self.xlsx,
            text="Save with changelog ",
            style="x_button.Std.TButton",
            command=self.toggle_xlsx_changelog,
            checked=self.C.save_xlsx_with_changelog,
            compound="right",
        )
        self.xlsx_changelog_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.xlsx_treeview_button = X_Checkbutton(
            self.xlsx,
            text="Save with tree ",
            style="x_button.Std.TButton",
            command=self.toggle_xlsx_treeview,
            checked=self.C.save_xlsx_with_treeview,
            compound="right",
        )
        self.xlsx_treeview_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.xlsx_flattened_button = X_Checkbutton(
            self.xlsx,
            text="Save with flattened sheet ",
            style="x_button.Std.TButton",
            command=self.toggle_xlsx_flattened,
            checked=self.C.save_xlsx_with_flattened,
            compound="right",
        )
        self.xlsx_flattened_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.flattened_divider = Frame(self.xlsx)
        self.flattened_divider.config(bg=themes[theme].table_fg, height=4)
        self.flattened_divider.pack(side="top", anchor="nw", fill="x", pady=(20, 0))

        self.flattened_header = Label(
            self.xlsx,
            text="xlsx Flatten Settings",
            font=TF,
            theme=theme,
            anchor="nw",
        )
        self.flattened_header.pack(side="top", anchor="nw", fill="x", pady=20)

        self.xlsx_flattened_details_button = X_Checkbutton(
            self.xlsx,
            text="Include detail columns ",
            style="x_button.Std.TButton",
            command=self.toggle_xlsx_flattened_details,
            checked=self.C.xlsx_flattened_detail_columns,
            compound="right",
        )
        self.xlsx_flattened_details_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.xlsx_flattened_justify_button = X_Checkbutton(
            self.xlsx,
            text="Justify cells left ",
            style="x_button.Std.TButton",
            command=self.toggle_xlsx_flattened_justify,
            checked=self.C.xlsx_flattened_justify,
            compound="right",
        )
        self.xlsx_flattened_justify_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.xlsx_flattened_reverse_button = X_Checkbutton(
            self.xlsx,
            text="Reverse order ",
            style="x_button.Std.TButton",
            command=self.toggle_xlsx_flattened_reverse,
            checked=self.C.xlsx_flattened_reverse_order,
            compound="right",
        )
        self.xlsx_flattened_reverse_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.xlsx_flattened_index_button = X_Checkbutton(
            self.xlsx,
            text="Add index column ",
            style="x_button.Std.TButton",
            command=self.toggle_xlsx_flattened_index,
            checked=self.C.xlsx_flattened_add_index,
            compound="right",
        )
        self.xlsx_flattened_index_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.json = Frame(self, theme=theme)
        self.json.grid(row=0, column=1, pady=20, padx=20, sticky="nswe")

        self.json_header = Label(self.json, text="json Save Settings", font=TF, theme=theme, anchor="nw")
        self.json_header.pack(side="top", anchor="nw", fill="x", pady=(0, 20))

        self.json_app_data_button = X_Checkbutton(
            self.json,
            text="Save with app data ",
            style="x_button.Std.TButton",
            command=self.toggle_json_app_data,
            checked=self.C.save_json_with_program_data,
            compound="right",
        )
        self.json_app_data_button.pack(side="top", anchor="nw", fill="x", pady=10)

        self.json_format_label = Label(self.json, text="json Format: ", font=EF, theme=theme, anchor="nw")
        self.json_format_label.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.json_format_dropdown = Ez_Dropdown(self.json, font=EF)
        self.json_format_dropdown["values"] = [
            """1) {"Header": [Column], ...}""",
            """2) [{"Header": value,..}, ...]""",
            """3) [["Header", "Header"], ["id1", "par1"]]""",
            """4) 'tab delimited csv'""",
        ]
        if self.C.json_format_one:
            self.json_format_dropdown.set_my_value("""1) {"Header": [Column], ...}""")

        elif self.C.json_format_two:
            self.json_format_dropdown.set_my_value("""2) [{"Header": value,..}, ...]""")

        elif self.C.json_format_three:
            self.json_format_dropdown.set_my_value("""3) [["Header", "Header"], ["id1", "par1"]]""")

        elif self.C.json_format_four:
            self.json_format_dropdown.set_my_value("""4) 'tab delimited csv'""")

        self.json_format_dropdown.bind("<<ComboboxSelected>>", self.set_json_format)
        self.json_format_dropdown.pack(side="top", anchor="nw", fill="x", pady=10)

        self.date_format = Frame(self, theme=theme)
        self.date_format.grid(row=0, column=1, pady=20, padx=20, sticky="nswe")

        self.date_format_header = Label(self.date_format, text="Date Format", font=TF, theme=theme, anchor="nw")
        self.date_format_header.pack(side="top", anchor="nw", fill="x", pady=(0, 20))

        self.date_format_label = Label(
            self.date_format,
            text="Date Format (per file): ",
            font=EF,
            theme=theme,
            anchor="nw",
        )
        self.date_format_label.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.date_format_dropdown = Ez_Dropdown(self.date_format, font=EF)
        self.date_format_dropdown["values"] = [
            "DD/MM/YYYY",
            "DD-MM-YYYY",
            "MM/DD/YYYY",
            "MM-DD-YYYY",
            "YYYY/MM/DD",
            "YYYY-MM-DD",
        ]
        if self.C.DATE_FORM == "%d/%m/%Y":
            self.date_format_dropdown.set_my_value("DD/MM/YYYY")

        elif self.C.DATE_FORM == "%d-%m-%Y":
            self.date_format_dropdown.set_my_value("DD-MM-YYYY")

        elif self.C.DATE_FORM == "%m/%d/%Y":
            self.date_format_dropdown.set_my_value("MM/DD/YYYY")

        elif self.C.DATE_FORM == "%m-%d-%Y":
            self.date_format_dropdown.set_my_value("MM-DD-YYYY")

        elif self.C.DATE_FORM == "%Y/%m/%d":
            self.date_format_dropdown.set_my_value("YYYY/MM/DD")

        elif self.C.DATE_FORM == "%Y-%m-%d":
            self.date_format_dropdown.set_my_value("YYYY-MM-DD")

        self.date_format_dropdown.bind("<<ComboboxSelected>>", self.set_date_format)
        self.date_format_dropdown.pack(side="top", anchor="nw", fill="x", pady=10)

        self.appearance = Frame(self, theme=theme)
        self.appearance.grid(row=0, column=1, pady=20, padx=20, sticky="nswe")

        self.appearance_header = Label(
            self.appearance,
            text="Appearance Settings",
            font=TF,
            theme=theme,
            anchor="nw",
        )
        self.appearance_header.pack(side="top", anchor="nw", fill="x", pady=(0, 20))

        self.theme_label = Label(self.appearance, text="Theme:", font=EF, theme=theme, anchor="nw")
        self.theme_label.pack(side="top", anchor="nw", fill="x")

        self.theme_dropdown = Ez_Dropdown(self.appearance, font=EF)
        self.theme_dropdown["values"] = [
            "Light Green",
            "Light Blue",
            "Dark",
            "Dark Blue",
            "Black",
        ]
        if self.C.light_green_theme_bool:
            self.theme_dropdown.set_my_value("Light Green")

        elif self.C.light_blue_theme_bool:
            self.theme_dropdown.set_my_value("Light Blue")

        elif self.C.dark_theme_bool:
            self.theme_dropdown.set_my_value("Dark")

        elif self.C.dark_blue_theme_bool:
            self.theme_dropdown.set_my_value("Dark Blue")

        elif self.C.black_theme_bool:
            self.theme_dropdown.set_my_value("Black")

        self.theme_dropdown.bind("<<ComboboxSelected>>", self.set_theme)
        self.theme_dropdown.pack(side="top", anchor="nw", fill="x", pady=10)

        self.alignments_label = Label(
            self.appearance,
            text="Cell alignments",
            font=lge_font,
            theme=theme,
            anchor="nw",
        )
        self.alignments_label.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.table_alignment_label = Label(
            self.appearance,
            text="Main table default cell alignment:",
            font=EF,
            theme=theme,
            anchor="nw",
        )
        self.table_alignment_label.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.table_alignment_dropdown = Ez_Dropdown(self.appearance, font=EF)
        self.table_alignment_dropdown.bind("<<ComboboxSelected>>", self.set_table_alignment)
        self.table_alignment_dropdown["values"] = ("Left", "Center", "Right")
        x = self.C.tree.table_align()
        if x == "w":
            self.table_alignment_dropdown.set_my_value("Left")
        elif x == "center":
            self.table_alignment_dropdown.set_my_value("Center")
        elif x == "e":
            self.table_alignment_dropdown.set_my_value("Right")
        self.table_alignment_dropdown.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.index_alignment_label = Label(
            self.appearance,
            "Row index align:",
            font=BF,
            anchor="nw",
            theme=theme,
        )
        self.index_alignment_label.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.index_alignment_dropdown = Ez_Dropdown(self.appearance, font=EF)
        self.index_alignment_dropdown.bind("<<ComboboxSelected>>", self.set_index_alignment)
        self.index_alignment_dropdown["values"] = ("Left", "Center", "Right")
        x = self.C.sheet.row_index_align()
        if x == "w":
            self.index_alignment_dropdown.set_my_value("Left")
        elif x == "center":
            self.index_alignment_dropdown.set_my_value("Center")
        elif x == "e":
            self.index_alignment_dropdown.set_my_value("Right")
        self.index_alignment_dropdown.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.header_alignment_label = Label(
            self.appearance,
            "Headers align:",
            font=BF,
            anchor="nw",
            theme=theme,
        )
        self.header_alignment_label.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.header_alignment_dropdown = Ez_Dropdown(self.appearance, font=EF)
        self.header_alignment_dropdown.bind("<<ComboboxSelected>>", self.set_header_alignment)
        self.header_alignment_dropdown["values"] = ("Left", "Center", "Right")
        x = self.C.tree.header_align()
        if x == "w":
            self.header_alignment_dropdown.set_my_value("Left")
        elif x == "center":
            self.header_alignment_dropdown.set_my_value("Center")
        elif x == "e":
            self.header_alignment_dropdown.set_my_value("Right")
        self.header_alignment_dropdown.pack(side="top", anchor="nw", fill="x", pady=(10, 0))

        self.page_chooser_frame = Frame(self, theme=theme)
        self.page_chooser_frame.config(bg=themes[theme].top_left_fg)
        self.page_chooser_frame.grid(row=0, column=0, sticky="nswe")

        self.general_button = Button(
            self.page_chooser_frame,
            text="General",
            style="EF.Std.TButton",
            command=self.goto_general,
        )
        self.general_button.pack(side="top", padx=20, pady=(20, 0), fill="x")

        self.xlsx_button = Button(
            self.page_chooser_frame,
            text="xlsx Save Options",
            style="EF.Std.TButton",
            command=self.goto_xlsx,
        )
        self.xlsx_button.pack(side="top", padx=20, pady=(20, 0), fill="x")

        self.json_button = Button(
            self.page_chooser_frame,
            text="json Save Options",
            style="EF.Std.TButton",
            command=self.goto_json,
        )
        self.json_button.pack(side="top", padx=20, pady=(20, 0), fill="x")

        self.date_format_button = Button(
            self.page_chooser_frame,
            text="Date Format",
            style="EF.Std.TButton",
            command=self.goto_date_format,
        )
        self.date_format_button.pack(side="top", padx=20, pady=(20, 0), fill="x")

        self.appearance_button = Button(
            self.page_chooser_frame,
            text="Appearance",
            style="EF.Std.TButton",
            command=self.goto_appearance,
        )
        self.appearance_button.pack(side="top", padx=20, pady=(20, 0), fill="x")

        self.general.tkraise()
        self.bind("<Escape>", self.cancel)
        center(self, 540, 650)
        self.deiconify()
        self.wait_window()

    def toggle_auto_resize_indexes(self):
        self.C.toggle_auto_resize_index(self.auto_resize_indexes_button.get_checked())

    def toggle_auto_select(self):
        self.C.toggle_mirror(self.auto_select_id_button.get_checked())

    def toggle_allow_spaces_ids(self):
        self.C.allow_spaces_ids_var = self.allow_spaces_ids_button.get_checked()

    def toggle_allow_spaces_columns(self):
        self.C.allow_spaces_columns_var = self.allow_spaces_columns_button.get_checked()

    def toggle_auto_sort_tree_button(self):
        self.C.toggle_sort_all_nodes(self.auto_sort_tree_button.get_checked())

    def toggle_show_tv_lvls(self):
        self.C.tv_lvls_bool = self.show_tv_lvls_button.get_checked()
        self.C.redo_tree_display()

    def set_layout(self, event=None):
        layout = self.layout_dropdown.get_my_value()
        if "Only Tree" in layout:
            self.C.set_display_option("left")
        elif "Only Sheet" in layout:
            self.C.set_display_option("right")
        elif "50" in layout:
            self.C.set_display_option("50/50")
        elif "Adjustable" in layout:
            self.C.set_display_option("adjustable")

    def set_treeview_indent(self, event=None):
        indent = self.indent_dropdown.get_my_value()
        self.C.tree.set_options(treeview_indent=indent)

    def toggle_xlsx_app_data(self):
        self.C.save_xlsx_with_program_data = self.xlsx_app_data_button.get_checked()
        self.C.C.save_cfg()

    def toggle_xlsx_changelog(self):
        self.C.save_xlsx_with_changelog = self.xlsx_changelog_button.get_checked()
        self.C.C.save_cfg()

    def toggle_xlsx_treeview(self):
        self.C.save_xlsx_with_treeview = self.xlsx_treeview_button.get_checked()
        self.C.C.save_cfg()

    def toggle_xlsx_flattened(self):
        self.C.save_xlsx_with_flattened = self.xlsx_flattened_button.get_checked()
        self.C.C.save_cfg()

    def toggle_xlsx_flattened_details(self):
        self.C.xlsx_flattened_detail_columns = self.xlsx_flattened_details_button.get_checked()
        self.C.C.save_cfg()

    def toggle_xlsx_flattened_justify(self):
        self.C.xlsx_flattened_justify = self.xlsx_flattened_justify_button.get_checked()
        self.C.C.save_cfg()

    def toggle_xlsx_flattened_reverse(self):
        self.C.xlsx_flattened_reverse_order = self.xlsx_flattened_reverse_button.get_checked()
        self.C.C.save_cfg()

    def toggle_xlsx_flattened_index(self):
        self.C.xlsx_flattened_add_index = self.xlsx_flattened_index_button.get_checked()
        self.C.C.save_cfg()

    def toggle_json_app_data(self):
        self.C.save_json_with_program_data = self.json_app_data_button.get_checked()
        self.C.C.save_cfg()

    def set_json_format(self, event=None):
        fmt = self.json_format_dropdown.get_my_value()
        if fmt.startswith("1"):
            self.C.change_json_format_one()

        elif fmt.startswith("2"):
            self.C.change_json_format_two()

        elif fmt.startswith("3"):
            self.C.change_json_format_three()

        elif fmt.startswith("4"):
            self.C.change_json_format_four()

        self.C.C.save_cfg()

    def set_theme(self, event=None):
        theme = self.theme_dropdown.get_my_value().lower().replace(" ", "_")
        self.page_chooser_frame.config(bg=themes[theme].top_left_fg)
        self.config(bg=themes[theme].top_left_bg)
        self.general.config(bg=themes[theme].top_left_bg)
        self.xlsx.config(bg=themes[theme].top_left_bg)
        self.json.config(bg=themes[theme].top_left_bg)
        self.appearance.config(bg=themes[theme].top_left_bg)
        self.general_header.change_theme(theme)
        self.xlsx_header.change_theme(theme)
        self.json_header.change_theme(theme)
        self.appearance_header.change_theme(theme)
        self.json_format_label.change_theme(theme)
        self.theme_label.change_theme(theme)
        self.flattened_header.change_theme(theme)
        self.flattened_divider.config(bg=themes[theme].table_fg)
        self.alignments_label.change_theme(theme)
        self.table_alignment_label.change_theme(theme)
        self.index_alignment_label.change_theme(theme)
        self.header_alignment_label.change_theme(theme)
        self.date_format.config(bg=themes[theme].top_left_bg)
        self.date_format_header.change_theme(theme)
        self.date_format_label.change_theme(theme)
        self.layout_label.change_theme(theme)
        self.indent_label.change_theme(theme)
        self.C.change_theme(theme)

    def set_table_alignment(self, event=None):
        align = convert_align(self.table_alignment_dropdown.get_my_value())
        self.C.tree.table_align(align, redraw=False)
        self.C.sheet.table_align(align, redraw=False)
        self.C.redraw_sheets()

    def set_index_alignment(self, event=None):
        align = convert_align(self.index_alignment_dropdown.get_my_value())
        # self.C.tree.index_align(align, redraw=False)
        self.C.sheet.index_align(align, redraw=False)
        self.C.redraw_sheets()

    def set_header_alignment(self, event=None):
        align = convert_align(self.header_alignment_dropdown.get_my_value())
        self.C.tree.header_align(align, redraw=False)
        self.C.sheet.header_align(align, redraw=False)
        self.C.redraw_sheets()

    def set_date_format(self, event=None):
        fmt = self.date_format_dropdown.get_my_value()
        if fmt == "DD/MM/YYYY":
            self.C.change_date_format("%d/%m/%Y")
        elif fmt == "DD-MM-YYYY":
            self.C.change_date_format("%d-%m-%Y")
        elif fmt == "MM/DD/YYYY":
            self.C.change_date_format("%m/%d/%Y")
        elif fmt == "MM-DD-YYYY":
            self.C.change_date_format("%m-%d-%Y")
        elif fmt == "YYYY/MM/DD":
            self.C.change_date_format("%Y/%m/%d")
        elif fmt == "YYYY-MM-DD":
            self.C.change_date_format("%Y-%m-%d")

    def goto_general(self):
        self.general.tkraise()

    def goto_xlsx(self):
        self.xlsx.tkraise()

    def goto_json(self):
        self.json.tkraise()

    def goto_date_format(self):
        self.date_format.tkraise()

    def goto_appearance(self):
        self.appearance.tkraise()

    def cancel(self, event=None):
        self.destroy()


class Help_Popup(tk.Toplevel):
    def __init__(self, C, text, theme="dark"):
        tk.Toplevel.__init__(self, C, width="1", height="1", bg=themes[theme].top_left_bg)
        self.C = new_toplevel_chores(self, C, f"{app_title} help", grab=False)
        self.word = ""
        self.findpos_start = "1.0"
        self.findpos_end = "1.0"
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.find_frame = Frame(self, theme=theme)
        self.find_frame.grid(row=0, column=0, columnspan=3, sticky="nswe")
        self.search_button = Button(self.find_frame, text=" Find:", command=self.find)
        self.search_button.pack(side="left", fill="x")
        self.find_window = Normal_Entry(self.find_frame, font=BF, theme=theme)
        self.find_window.bind("<Return>", self.find)
        self.find_window.pack(side="left", fill="x", expand=True)
        self.find_reset_button = Button(self.find_frame, text="X", command=self.find_reset)
        self.find_reset_button.pack(side="left", fill="x")
        self.find_results_label = Label(self.find_frame, "0/0", BF, theme=theme)
        self.find_results_label.pack(side="left", fill="x")
        self.find_up_button = Button(self.find_frame, text="▲", command=self.find_up)
        self.find_up_button.pack(side="left", fill="x")
        self.find_down_button = Button(self.find_frame, text="▼", command=self.find_down)
        self.find_down_button.pack(side="left", fill="x")
        self.buttonframe = Frame(self, theme=theme)
        self.buttonframe.grid(row=1, column=0, rowspan=2, padx=10, pady=10, sticky="nswe")

        self.open_in_browser = Button(
            self.buttonframe,
            text="Open in Browser",
            style="EF.Std.TButton",
            command=self.open_docs_in_browser,
        )
        self.open_in_browser.pack(side="top", pady=2, fill="x")

        self.basics = Button(
            self.buttonframe,
            text="Basics",
            style="EF.Std.TButton",
            command=lambda: self.scrollto("programbasics"),
        )
        self.basics.pack(side="top", pady=2, fill="x")

        self.tutorials = Button(
            self.buttonframe,
            text="Tutorials",
            style="EF.Std.TButton",
            command=lambda: self.scrollto("tutorials"),
        )
        self.tutorials.pack(side="top", pady=2, fill="x")

        self.menubar = Button(
            self.buttonframe,
            text="Tree Menu",
            style="EF.Std.TButton",
            command=lambda: self.scrollto("treeviewmenu"),
        )
        self.menubar.pack(side="top", pady=2, fill="x")

        self.columns = Button(
            self.buttonframe,
            text="Columns",
            style="EF.Std.TButton",
            command=lambda: self.scrollto("columns"),
        )
        self.columns.pack(side="top", pady=2, fill="x")

        self.buttons = Button(
            self.buttonframe,
            text="Tree Buttons",
            style="EF.Std.TButton",
            command=lambda: self.scrollto("treeviewbuttons"),
        )
        self.buttons.pack(side="top", pady=2, fill="x")

        self.functions = Button(
            self.buttonframe,
            text="Tree Functions",
            style="EF.Std.TButton",
            command=lambda: self.scrollto("treeviewfunctions"),
        )
        self.functions.pack(side="top", pady=2, fill="x")

        self.treecomparehelp = Button(
            self.buttonframe,
            text="Tree Compare",
            style="EF.Std.TButton",
            command=lambda: self.scrollto("treecompare"),
        )
        self.treecomparehelp.pack(side="top", pady=2, fill="x")

        self.xlsx_files = Button(
            self.buttonframe,
            text="XLSX Files",
            style="EF.Std.TButton",
            command=lambda: self.scrollto("xlsxfiles"),
        )
        self.xlsx_files.pack(side="top", pady=2, fill="x")

        self.program_data = Button(
            self.buttonframe,
            text="JSON FILES",
            style="EF.Std.TButton",
            command=lambda: self.scrollto("jsonfiles"),
        )
        self.program_data.pack(side="top", pady=2, fill="x")

        self.api = Button(
            self.buttonframe,
            text="API",
            style="EF.Std.TButton",
            command=lambda: self.scrollto("api"),
        )
        self.api.pack(side="top", pady=2, fill="x")

        self.textbox = Working_Text(
            self,
            font=("Mono", std_font_size, "normal"),
            wrap="word",
            theme=theme,
            use_entry_bg=False,
            override_bg=None,
        )
        self.yscrollb = Scrollbar(self, self.textbox.yview, "vertical", self.textbox)
        self.textbox.delete(1.0, "end")
        self.textbox.insert(1.0, text)
        self.textbox.config(state="disabled")
        self.textbox.grid(row=1, column=1, sticky="nswe")
        self.yscrollb.grid(row=1, column=2, sticky="nswe")
        self.textbox.focus_set()
        self.bind("<Escape>", self.cancel)
        center(self, 975, 650)
        self.deiconify()
        self.wait_window()

    def open_docs_in_browser(self):
        try:
            url = f"file://{upone_dir + 'DOCUMENTATION.html'}"
            webbrowser.open(url)
        except Exception:
            pass

    def scrollto(self, option):
        if option == "programbasics":
            self.textbox.see(self.textbox.search("## PROGRAM BASICS", "1.0").split(".")[0] + ".0")
        elif option == "tutorials":
            self.textbox.see(self.textbox.search("## HELPFUL TIPS AND TUTORIALS", "1.0").split(".")[0] + ".0")
        elif option == "columns":
            self.textbox.see(self.textbox.search("## MANAGING COLUMNS", "1.0").split(".")[0] + ".0")
        elif option == "treeviewmenu":
            self.textbox.see(self.textbox.search("## MENU BAR", "1.0").split(".")[0] + ".0")
        elif option == "treeviewbuttons":
            self.textbox.see(self.textbox.search("## TREE BUTTONS", "1.0").split(".")[0] + ".0")
        elif option == "treeviewfunctions":
            self.textbox.see(self.textbox.search("## TREE FUNCTIONS", "1.0").split(".")[0] + ".0")
        elif option == "treecompare":
            self.textbox.see(self.textbox.search("## TREE COMPARE", "1.0").split(".")[0] + ".0")
        elif option == "xlsxfiles":
            self.textbox.see(self.textbox.search("## XLSX FILES", "1.0").split(".")[0] + ".0")
        elif option == "jsonfiles":
            self.textbox.see(self.textbox.search("## JSON FILES", "1.0").split(".")[0] + ".0")
        elif option == "api":
            self.textbox.see(self.textbox.search("## USING THE API", "1.0").split(".")[0] + ".0")

    def find(self, event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        self.addon = f"+{len(self.word)}c"
        start = "1.0"
        while start:
            start = self.textbox.search(self.word, index=start, nocase=1, stopindex="end")
            if start:
                end = start + self.addon
                self.find_results.append(start)
                self.textbox.tag_add("i", start, end)
                start = end
        if self.find_results:
            self.textbox.tag_config("i", background="Yellow")
            self.find_results_label.config(text=f"1/{len(self.find_results)}")
            self.textbox.tag_add(
                "c",
                self.find_results[self.results_number],
                self.find_results[self.results_number] + self.addon,
            )
            self.textbox.tag_config("c", background="Orange")
            self.textbox.see(self.find_results[self.results_number])

    def find_up(self, event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove(
            "c",
            self.find_results[self.results_number],
            self.find_results[self.results_number] + self.addon,
        )
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=f"{self.results_number + 1}/{len(self.find_results)}")
        self.textbox.tag_add(
            "c",
            self.find_results[self.results_number],
            self.find_results[self.results_number] + self.addon,
        )
        self.textbox.tag_config("c", background="Orange")
        self.textbox.see(self.find_results[self.results_number])

    def find_down(self, event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove(
            "c",
            self.find_results[self.results_number],
            self.find_results[self.results_number] + self.addon,
        )
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=f"{self.results_number + 1}/{len(self.find_results)}")
        self.textbox.tag_add(
            "c",
            self.find_results[self.results_number],
            self.find_results[self.results_number] + self.addon,
        )
        self.textbox.tag_config("c", background="Orange")
        self.textbox.see(self.find_results[self.results_number])

    def find_reset(self, newfind=False):
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        if not newfind:
            self.find_window.delete(0, "end")
        for tag in self.textbox.tag_names():
            self.textbox.tag_delete(tag)
        self.find_results_label.config(text="0/0")

    def cancel(self, event=None):
        self.destroy()
