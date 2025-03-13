# SPDX-License-Identifier: AGPL-3.0-only
# Copyright © R. A. Gardner

from __future__ import annotations

import csv
import datetime
import json
import os
import pickle
import re
import tkinter as tk
import zlib
from bisect import bisect_left
from collections import defaultdict, deque
from collections.abc import Generator, Iterator, Sequence
from contextlib import suppress
from itertools import cycle, filterfalse, islice, repeat
from locale import getdefaultlocale
from math import floor
from operator import attrgetter, itemgetter
from tkinter import filedialog, font, ttk
from typing import Literal

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from tksheet import (
    DotDict,
    Highlight,
    Sheet,
    is_contiguous,
    move_elements_by_mapping,
    push_n,
)
from tksheet import (
    num2alpha as _n2a,
)

from .classes import (
    Del_stre,
    Header,
    Node,
    SearchResult,
    TreeBuilder,
)
from .constants import (
    BF,
    EF,
    align_c_icon,
    align_e_icon,
    align_w_icon,
    changelog_header,
    ctrl_button,
    date_formats_usable,
    detail_column_types,
    menu_kwargs,
    rc_button,
    rc_motion,
    rc_press,
    rc_release,
    remove_nrt,
    sheet_bindings,
    sheet_header_font,
    software_version_number,
    themes,
    tree_bindings,
    tv_lvls_colors,
    validation_allowed_date_chars,
    validation_allowed_num_chars,
    warnings_header,
)
from .functions import (
    bytes_io_wb,
    convert_old_xl_to_xlsx,
    create_cell_align_selector_menu,
    csv_str_x_data,
    dict_x_b32,
    equalize_sublist_lens,
    frame_w_to_nchars,
    full_sheet_to_dict,
    get_json_format,
    get_json_from_file,
    increment_file_version,
    isfloat,
    isint,
    isintlike,
    isreal,
    json_to_sheet,
    level_to_color,
    new_info_storage,
    new_saved_info,
    path_numbers,
    path_without_numbers,
    process_search_results,
    search_results_max_column_chars,
    sort_key,
    str_io_csv_writer,
    to_clipboard,
    try_remove,
    ws_x_data,
    xlsx_changelog_header,
)
from .toplevels import (
    Add_Child_Or_Sibling_Id_Popup,
    Add_Detail_Column_Popup,
    Add_Hierarchy_Column_Popup,
    Add_Top_Id_Popup,
    Ask_Confirm,
    Changelog_Popup,
    Edit_Conditional_Formatting_Popup,
    Edit_Detail_Date_Popup,
    Edit_Detail_Number_Popup,
    Edit_Detail_Text_Popup,
    Edit_Validation_Popup,
    Enter_Sheet_Name_Popup,
    Error,
    Export_Flattened_Popup,
    Get_Clipboard_Data_Popup,
    Merge_Sheets_Popup,
    Post_Import_Changes_Popup,
    Rename_Column_Popup,
    Rename_Id_Popup,
    Replace_Popup,
    Save_New_Version_Error_Popup,
    Save_New_Version_Postsave_Popup,
    Save_New_Version_Presave_Popup,
    Settings_Popup,
    Sort_Sheet_Popup,
    Text_Popup,
    Treeview_Id_Finder,
    View_Id_Popup,
)
from .widgets import (
    Button,
    Ez_Dropdown,
    Frame,
    Normal_Entry,
)

# OVERRIDE LOCALE DETECTION FOR DATE FORMAT
override_locale = None

# DEFAULT SETTING FOR SAVING WITH PROGRAM DATA
save_xlsx_and_json_with_program_data = True
# "normal" to allowed user choice or "disabled" to disallow user choice
user_option_save_with_program_data = "normal"


class Tree_Editor(tk.Frame):
    def __init__(self, parent, C):
        tk.Frame.__init__(self, parent)
        self.C = C
        # try:
        #     self.monitor_scale = self.C.call("tk", "scaling")
        # except Exception:
        #     self.monitor_scale = 1
        self.undo_in_progress = False
        self.l_frame_proportion = float(0.50)
        self.last_width = 0
        self.last_height = 0
        self.currently_adjusting_divider = False
        self.tree_has_focus = True
        self.sheet_has_focus = False
        self.sheet_changes = 0
        self.nodes = {}
        self.topnodes_order = {}
        self.levels = defaultdict(list)
        self.row_len = 0
        self.headers = []
        self.changelog = []
        self.treecolsel = 0
        self.ic = 0
        self.tv_label_col = 0
        self.pc = 0
        self.hiers = []
        self.warnings = []
        self.reset_tree_drag_vars()
        self.row_cut_updated = False
        self.mirror_sels_disabler = False
        self.tagged_ids = set()
        self.date_split_regex = "|".join(map(re.escape, ("/", "-")))
        self.find_popup = None
        self.fixed_font_w = font.nametofont("TkFixedFont").measure("0")

        self.auto_resize_indexes = True
        self.mirror_var = False
        self.allow_spaces_ids_var = False
        self.allow_spaces_columns_var = False
        self.save_xlsx_with_program_data = bool(save_xlsx_and_json_with_program_data)
        self.save_json_with_program_data = bool(save_xlsx_and_json_with_program_data)
        self.save_xlsx_with_changelog = False
        self.save_xlsx_with_treeview = False
        self.save_xlsx_with_flattened = False
        self.xlsx_flattened_detail_columns = True
        self.xlsx_flattened_justify = True
        self.xlsx_flattened_reverse_order = False
        self.xlsx_flattened_add_index = False
        self.json_format = 1
        self.black_theme_bool = self.C.theme == "black"
        self.dark_blue_theme_bool = self.C.theme == "dark_blue"
        self.dark_theme_bool = self.C.theme == "dark"
        self.light_green_theme_bool = self.C.theme == "light_green"
        self.light_blue_theme_bool = self.C.theme == "light_blue"
        self.auto_sort_nodes_bool = True
        self.tv_lvls_bool = False

        if override_locale is not None:
            self.user_locale = f"{override_locale}"
        else:
            self.user_locale = f"{getdefaultlocale()[0]}".lower()

        if self.user_locale == "en_us":
            self.DATE_FORM = "%m-%d-%Y"
        elif self.user_locale == "en_ca" or "zh" in self.user_locale:
            self.DATE_FORM = "%Y-%m-%d"
        else:
            self.DATE_FORM = "%d-%m-%Y"

        self.warnings_filepath = ""
        self.warnings_sheet = ""

        # cell alignment menu images
        self.align_icons = {
            "w": tk.PhotoImage(format="png", data=align_w_icon),
            "c": tk.PhotoImage(format="png", data=align_c_icon),
            "e": tk.PhotoImage(format="png", data=align_e_icon),
        }

        self.C.file.entryconfig("Save", command=self.save_)
        self.C.file.entryconfig(
            "Save as",
            accelerator="Ctrl+Shift+S",
            command=self.save_as,
        )
        self.C.file.entryconfig("Save new version", command=self.save_new_vrsn)
        self.C.file.entryconfig("Settings", command=self.settings)

        self.edit_menu = tk.Menu(self.C.menubar, tearoff=0, **menu_kwargs)
        self.C.menubar.add_cascade(
            label="Edit",
            menu=self.edit_menu,
            state="disabled",
            **menu_kwargs,
        )
        self.edit_menu.add_command(
            label="Undo  0/30",
            accelerator="Ctrl+Z",
            state="disabled",
            command=self.undo,
            **menu_kwargs,
        )
        self.edit_menu.add_separator()
        self.copy_clipboard_menu = tk.Menu(self.edit_menu, tearoff=0, **menu_kwargs)
        self.copy_clipboard_menu.add_command(
            label="Copy sheet to clipboard (indent separated)",
            command=self.clipboard_sheet_indent,
            **menu_kwargs,
        )
        self.copy_clipboard_menu.add_command(
            label="Copy sheet to clipboard (comma separated)",
            command=self.clipboard_sheet,
            **menu_kwargs,
        )
        self.copy_clipboard_menu.add_command(
            label="Copy sheet to clipboard as json",
            command=self.clipboard_sheet_json,
            **menu_kwargs,
        )
        self.edit_menu.add_command(
            label="Sort sheet",
            command=self.sort_sheet_choice,
            **menu_kwargs,
        )
        self.edit_menu.add_cascade(
            label="Copy to clipboard",
            menu=self.copy_clipboard_menu,
            state="normal",
            **menu_kwargs,
        )
        self.edit_menu.add_command(
            label="Tag/Untag IDs",
            command=self.tag_ids,
            accelerator="Ctrl+T",
            **menu_kwargs,
        )
        self.edit_menu.add_separator()
        self.edit_menu.add_command(
            label="Replace using mapping",
            command=self.replace_using_mapping,
            **menu_kwargs,
        )
        self.edit_menu.add_separator()
        self.edit_menu.add_command(
            label="Clear copied/cut",
            command=self.clear_copied_details,
            **menu_kwargs,
        )
        self.edit_menu.add_command(
            label="Clear panel selections",
            command=self.remove_selections,
            **menu_kwargs,
        )
        self.edit_menu.add_command(
            label="Clear all tagged IDs",
            command=self.clear_tagged_ids,
            **menu_kwargs,
        )

        # view menu
        self.view_menu = tk.Menu(self.C.menubar, tearoff=0, **menu_kwargs)
        self.C.menubar.add_cascade(label="View", menu=self.view_menu, state="disabled", **menu_kwargs)
        self.view_menu.add_command(
            label="View changelog",
            accelerator="Ctrl+L",
            command=self.show_changelog,
            **menu_kwargs,
        )
        self.view_menu.add_command(
            label="View build warnings",
            command=lambda: self.show_warnings(show_regardless=True),
            **menu_kwargs,
        )
        # self.view_menu.add_command(label="View organizational chart", command=self.show_org_chart,**menu_kwargs)
        self.view_menu.add_separator()
        self.view_menu.add_command(
            label="Treeview IDs information",
            command=self.show_ids_full_info_tree,
            **menu_kwargs,
        )
        self.view_menu.add_command(
            label="Sheet IDs information",
            command=self.show_ids_full_info_sheet,
            **menu_kwargs,
        )
        self.view_menu.add_separator()
        self.view_menu.add_command(
            label="Expand ID",
            accelerator="Ctrl+E",
            command=self.expand_id,
            **menu_kwargs,
        )
        self.view_menu.add_command(
            label="Collapse ID",
            accelerator="Ctrl+R",
            command=self.collapse_id,
            **menu_kwargs,
        )
        self.view_menu.add_separator()
        self.view_menu.add_command(
            label="Zoom in",
            accelerator="Ctrl++",
            command=self.zoom_in,
            **menu_kwargs,
        )
        self.view_menu.add_command(
            label="Zoom out",
            accelerator="Ctrl+-",
            command=self.zoom_out,
            **menu_kwargs,
        )
        self.view_menu.add_separator()
        self.adjustable_bool = tk.BooleanVar()
        self.adjustable_bool.set(False)
        self._50_50_bool = tk.BooleanVar()
        self._50_50_bool.set(False)
        self.full_left_bool = tk.BooleanVar()
        self.full_left_bool.set(True)
        self.full_right_bool = tk.BooleanVar()
        self.full_right_bool.set(False)
        self.display_menu = tk.Menu(self.view_menu, tearoff=0, **menu_kwargs)
        self.display_menu.add_checkbutton(
            label="Display Only Tree",
            variable=self.full_left_bool,
            command=self.option_full_left,
            **menu_kwargs,
        )
        self.display_menu.add_checkbutton(
            label="Display Only Sheet",
            variable=self.full_right_bool,
            command=self.option_full_right,
            **menu_kwargs,
        )
        self.display_menu.add_checkbutton(
            label="50/50 Tree/Sheet",
            variable=self._50_50_bool,
            command=self.option_50_50,
            **menu_kwargs,
        )
        self.display_menu.add_checkbutton(
            label="Adjustable Display",
            variable=self.adjustable_bool,
            command=self.option_adjustable,
            **menu_kwargs,
        )
        self.view_menu.add_cascade(
            label="Layout",
            menu=self.display_menu,
            state="normal",
            **menu_kwargs,
        )
        self.view_menu.add_command(
            label="Set all column widths",
            command=self.set_all_col_widths,
            **menu_kwargs,
        )

        # import menu
        self.import_menu = tk.Menu(self.C.menubar, tearoff=0, **menu_kwargs)
        self.C.menubar.add_cascade(
            label="Import",
            menu=self.import_menu,
            state="disabled",
            **menu_kwargs,
        )
        self.import_menu.add_command(
            label="Import changes",
            command=self.import_changes,
            **menu_kwargs,
        )
        self.import_menu.add_command(
            label="Get sheet from clipboard and overwrite",
            command=self.get_clipboard_data,
            **menu_kwargs,
        )
        self.import_menu.add_command(
            label="Merge sheets / Add rows",
            command=self.merge_sheets,
            **menu_kwargs,
        )

        # export menu
        self.export_menu = tk.Menu(self.C.menubar, tearoff=0, **menu_kwargs)
        self.C.menubar.add_cascade(
            label="Export",
            menu=self.export_menu,
            state="disabled",
            **menu_kwargs,
        )
        self.export_menu.add_command(
            label="Export specific changes",
            accelerator="Ctrl+L",
            command=lambda: self.show_changelog("specific"),
            **menu_kwargs,
        )
        self.export_menu.add_command(
            label="Export file session changes",
            command=lambda: self.show_changelog("sheet"),
            **menu_kwargs,
        )
        self.export_menu.add_command(
            label="Export all changes",
            command=lambda: self.show_changelog("all"),
            **menu_kwargs,
        )
        self.export_menu.add_command(
            label="Export flattened sheet",
            command=self.export_flattened,
            **menu_kwargs,
        )

        # help menu
        self.help_menu = tk.Menu(self.C.menubar, tearoff=0, **menu_kwargs)
        self.C.menubar.add_cascade(label="Help", menu=self.help_menu, state="normal", **menu_kwargs)
        self.help_menu.add_command(label="View Help", command=self.C.help_func, **menu_kwargs)
        self.help_menu.add_command(label="View License", command=self.C.license_func, **menu_kwargs)
        self.help_menu.add_command(label="About", command=self.C.about_func, **menu_kwargs)

        # MAIN CANVAS
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.main_canvas = tk.Canvas(self, highlightthickness=0)
        self.main_canvas.grid(row=0, column=0, sticky="nswe")

        # ======================= LEFT FRAME ======================================================
        self.l_frame = tk.Frame(
            self.main_canvas,
            highlightbackground="white",
            highlightcolor="white",
            highlightthickness=2,
        )

        # frames for left frame
        self.btns_tree = Frame(self.l_frame)
        self.btns_tree.pack(side="top", fill="x")
        self.btns_tree.grid_rowconfigure(0, weight=1)
        self.btns_tree.grid_rowconfigure(1, weight=1)
        self.btns_tree.grid_columnconfigure(4, weight=1)
        # self.btns_tree.grid_columnconfigure(4, weight=1)

        self.treeframe = Frame(self.l_frame)
        self.treeframe.pack(side="top", fill="both", expand=True)
        self.treeframe.grid_rowconfigure(0, weight=1)
        self.treeframe.grid_columnconfigure(0, weight=1)

        self.tree = Sheet(
            self.treeframe,
            name="tree",
            header_font=sheet_header_font,
            theme=self.C.theme,
            default_column_width=200,
            treeview=True,
            row_drag_and_drop_perform=False,
            alternate_color="#f6f9fb",
            allow_cell_overflow=True,
            max_undos=0,
        )
        self.tree.grid(row=0, column=0, sticky="nswe")

        # status bar tree
        self.sts_tree = Frame(self.l_frame)
        self.sts_tree.pack(side="top", fill="x")
        self.sts_tree.grid_rowconfigure(0, weight=1)

        # buttons for bottom left frame
        # switch hierarchy dropdown
        self.switch_label = Button(
            self.btns_tree,
            text="Hierarchy:",
            command=self.next_hier,
        )
        self.switch_label.grid(row=0, column=0, sticky="nswe")
        self.switch_displayed = tk.StringVar(self.btns_tree)
        self.switch_displayed.set("")
        self.switch_hier_dropdown = ttk.Combobox(
            self.btns_tree,
            textvariable=self.switch_displayed,
            state="readonly",
            font=BF,
        )
        self.switch_hier_dropdown.grid(row=0, column=1, sticky="nswe")
        self.switch_hier_dropdown.bind("<<ComboboxSelected>>", self.switch_hier)

        # tag ID tree
        self.tree_tag_id_button = Button(self.btns_tree, text="Tag ID:", underline=0, command=self.tag_ids)
        self.tree_tag_id_button.grid(row=1, column=0, ipady=1, sticky="nswe")
        self.tree_tagged_ids_dropdown = Ez_Dropdown(self.btns_tree, font=BF)
        self.tree_tagged_ids_dropdown.grid(row=1, column=1, sticky="nswe")
        self.tree_tagged_ids_dropdown.bind("<<ComboboxSelected>>", self.tree_go_to_tagged_id)

        # buttons for top left frame
        # tree search function
        self.search_displayed = tk.StringVar(self.btns_tree)
        self.search_displayed.set("")
        self.search_button = Button(self.btns_tree, text=" Find:", command=self.search_choice)
        self.search_button.grid(row=0, column=2, sticky="nswe")
        self.search_choice_displayed = tk.StringVar(self.btns_tree)
        self.search_choice_displayed.set("Non-exact")
        self.search_choice_dropdown = ttk.Combobox(
            self.btns_tree,
            textvariable=self.search_choice_displayed,
            state="readonly",
            font=BF,
        )
        self.search_choice_dropdown.config(width=13)
        self.search_choice_dropdown["values"] = [
            "Non-exact",
            "ID non-exact",
            "ID exact",
            "Detail non-exact",
            "Detail exact",
        ]
        self.search_choice_dropdown.grid(row=0, column=3, sticky="nswe")
        self.search_entry = Normal_Entry(self.btns_tree, font=BF, theme="light_blue")
        self.search_entry.grid(row=0, column=4, sticky="nswe")
        self.search_entry.bind("<Return>", self.search_choice)
        self.search_dropdown = ttk.Combobox(
            self.btns_tree,
            textvariable=self.search_displayed,
            state="readonly",
            font=EF,
        )
        self.search_dropdown["values"] = []
        self.search_dropdown.bind("<<ComboboxSelected>>", self.show_search_result)
        self.search_choice_dropdown.bind("<<ComboboxSelected>>", lambda focus: self.search_entry.focus_set())
        self.search_dropdown.grid(row=1, column=2, columnspan=3, sticky="nswe")

        # ======================= RIGHT FRAME ======================================================
        self.r_frame = tk.Frame(
            self.main_canvas,
            highlightbackground="white",
            highlightcolor="white",
            highlightthickness=2,
        )

        # frames for right frame
        self.btns_sheet = Frame(self.r_frame)
        self.btns_sheet.pack(side="top", fill="x")
        self.btns_sheet.grid_rowconfigure(0, weight=1)
        self.btns_sheet.grid_rowconfigure(1, weight=1)
        self.btns_sheet.grid_columnconfigure(3, weight=1)

        self.sheetframe = Frame(self.r_frame)
        self.sheetframe.pack(side="top", fill="both", expand=True)

        self.sheet = Sheet(
            self.sheetframe,
            name="sheet",
            theme=self.C.theme,
            row_index_align="w",
            auto_resize_row_index=True,
            header_font=sheet_header_font,
            allow_cell_overflow=True,
            max_undos=0,
        )
        self.sheet.pack(side="right", fill="both", expand=True)

        # buttons for top right frame
        # tag ID
        self.sheet_tag_id_button = Button(self.btns_sheet, text="↓Tag ID", underline=0, command=self.tag_ids)
        self.sheet_tag_id_button.grid(row=0, column=0, ipady=1, sticky="nswe")
        self.sheet_tagged_ids_dropdown = Ez_Dropdown(self.btns_sheet, font=BF)
        self.sheet_tagged_ids_dropdown.grid(row=1, column=0, sticky="nswe")
        self.sheet_tagged_ids_dropdown.bind("<<ComboboxSelected>>", self.sheet_go_to_tagged_id)

        # sheet search function
        self.sheet_search_displayed = tk.StringVar(self.btns_sheet)
        self.sheet_search_displayed.set("")
        self.sheet_search_button = Button(self.btns_sheet, text="Find:", command=self.sheet_search_choice)
        self.sheet_search_button.grid(row=0, column=1, sticky="nswe")
        self.sheet_search_choice_displayed = tk.StringVar(self.btns_sheet)
        self.sheet_search_choice_displayed.set("Non-exact")
        self.sheet_search_choice_dropdown = ttk.Combobox(
            self.btns_sheet,
            textvariable=self.sheet_search_choice_displayed,
            state="readonly",
            font=BF,
        )
        self.sheet_search_choice_dropdown.config(width=13)
        self.sheet_search_choice_dropdown["values"] = [
            "Non-exact",
            "ID non-exact",
            "ID exact",
            "Detail non-exact",
            "Detail exact",
        ]
        self.sheet_search_choice_dropdown.grid(row=0, column=2, sticky="nswe")
        self.sheet_search_entry = Normal_Entry(self.btns_sheet, font=BF, theme="light_blue")
        self.sheet_search_entry.grid(row=0, column=3, sticky="nswe")
        self.sheet_search_entry.bind("<Return>", self.sheet_search_choice)
        self.sheet_search_dropdown = ttk.Combobox(
            self.btns_sheet,
            textvariable=self.sheet_search_displayed,
            state="readonly",
            font=EF,
        )
        self.sheet_search_dropdown["values"] = []
        self.sheet_search_dropdown.bind("<<ComboboxSelected>>", self.sheet_show_search_result)
        self.sheet_search_choice_dropdown.bind(
            "<<ComboboxSelected>>", lambda focus: self.sheet_search_entry.focus_set()
        )
        self.sheet_search_dropdown.grid(row=1, column=1, columnspan=3, sticky="nswe")

        # RIGHT CLICK MENUS

        # SINGLE CELL MENU - SHEET AND TREE
        self.tree_sheet_rc_menu_single_cell = tk.Menu(self.sheet, tearoff=0, **menu_kwargs)
        self.tree_sheet_rc_menu_single_cell.add_command(
            label="Detail",
            state="disabled",
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_cell.add_command(
            label="Edit",
            command=self.tree_sheet_edit_detail,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_cell.add_command(
            label="Cut",
            accelerator="Ctrl+X",
            command=self.cut_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_cell.add_command(
            label="Copy",
            accelerator="Ctrl+C",
            command=self.copy_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_cell.add_command(
            label="Paste",
            accelerator="Ctrl+V",
            command=self.paste_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_cell.add_command(
            label="Clear contents",
            accelerator="Del",
            command=self.del_key,
            **menu_kwargs,
        )

        # MULTI CELL MENU - SHEET AND TREE
        self.tree_sheet_rc_menu_multi_cell = tk.Menu(
            self.sheet,
            tearoff=0,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_multi_cell.add_command(
            label="Cut",
            accelerator="Ctrl+X",
            command=self.cut_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_multi_cell.add_command(
            label="Copy",
            accelerator="Ctrl+C",
            command=self.copy_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_multi_cell.add_command(
            label="Paste",
            accelerator="Ctrl+V",
            command=self.paste_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_multi_cell.add_command(
            label="Clear contents",
            accelerator="Del",
            command=self.del_key,
            **menu_kwargs,
        )

        # SINGLE COLUMN MENU - SHEET AND TREE
        self.tree_sheet_rc_menu_single_col = tk.Menu(
            self.sheet,
            tearoff=0,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col_align = create_cell_align_selector_menu(
            parent=self.tree_sheet_rc_menu_single_col,
            command=self.tree_sheet_align,
            menu_kwargs=menu_kwargs,
            icons=self.align_icons,
        )
        self.tree_sheet_rc_menu_single_col.add_cascade(
            label="Alignment",
            menu=self.tree_sheet_rc_menu_single_col_align,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Cut",
            accelerator="Ctrl+X",
            command=self.cut_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Copy",
            accelerator="Ctrl+C",
            command=self.copy_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Paste",
            accelerator="Ctrl+V",
            command=self.paste_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Clear contents",
            accelerator="Del",
            command=self.del_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Delete column",
            command=self.del_cols_rc,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_separator()
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Add detail",
            command=self.rc_add_col,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Add hierarchy",
            command=self.rc_add_hier_col,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Rename column",
            command=self.rc_rename_col,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_separator()
        self.tree_sheet_rc_menu_single_col_type = tk.Menu(
            self.tree_sheet_rc_menu_single_col,
            tearoff=0,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_cascade(
            label="Type",
            menu=self.tree_sheet_rc_menu_single_col_type,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col_type.add_command(
            label="Text",
            command=self.rc_change_coltype_text,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col_type.add_command(
            label="Number",
            command=self.rc_change_coltype_number,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col_type.add_command(
            label="Date",
            command=self.rc_change_coltype_date,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Validation",
            command=self.rc_edit_validation,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Conditional Formatting",
            command=self.rc_edit_formatting,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_separator()
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Sort sheet A → Z",
            command=self.sort_sheet_rc_asc,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Sort sheet Z → A",
            command=self.sort_sheet_rc_desc,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Sort sheet tree walk",
            command=self.sort_sheet_walk,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_single_col.add_separator()
        self.tree_sheet_rc_menu_single_col.add_command(
            label="Set as treeview label",
            command=self.sheet_rc_tv_label,
            **menu_kwargs,
        )

        # MULTI COLUMN MENU - SHEET AND TREE
        self.tree_sheet_rc_menu_multi_col = tk.Menu(
            self.sheet,
            tearoff=0,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_multi_col_align = create_cell_align_selector_menu(
            parent=self.tree_sheet_rc_menu_multi_col,
            command=self.tree_sheet_align,
            menu_kwargs=menu_kwargs,
            icons=self.align_icons,
        )
        self.tree_sheet_rc_menu_multi_col.add_cascade(
            label="Alignment",
            menu=self.tree_sheet_rc_menu_multi_col_align,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_multi_col.add_command(
            label="Cut",
            accelerator="Ctrl+X",
            command=self.cut_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_multi_col.add_command(
            label="Copy",
            accelerator="Ctrl+C",
            command=self.copy_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_multi_col.add_command(
            label="Paste",
            accelerator="Ctrl+V",
            command=self.paste_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_multi_col.add_command(
            label="Clear contents",
            accelerator="Del",
            command=self.del_key,
            **menu_kwargs,
        )
        self.tree_sheet_rc_menu_multi_col.add_command(
            label="Delete columns",
            command=self.del_cols_rc,
            **menu_kwargs,
        )

        # MULTI ROW MENU - TREE
        self.tree_rc_menu_multi_row = tk.Menu(self.treeframe, tearoff=0, **menu_kwargs)
        self.tree_rc_menu_multi_row.add_command(
            label="Cut",
            accelerator="Ctrl+X",
            command=self.cut_ids,
            **menu_kwargs,
        )
        self.tree_rc_menu_multi_row.add_command(
            label="Copy",
            accelerator="Ctrl+C",
            command=self.copy_key,
            **menu_kwargs,
        )
        self.tree_rc_menu_multi_row.add_command(
            label="Paste",
            accelerator="Ctrl+V",
            command=self.paste_key,
            **menu_kwargs,
        )
        self.tree_rc_menu_multi_row.add_command(
            label="Clipboard",
            command=self.copy_ID_row,
            **menu_kwargs,
        )
        self.tree_rc_menu_multi_row.add_command(
            label="Clipboard selected + children",
            command=self.copy_ID_children_rows,
            **menu_kwargs,
        )
        self.tree_rc_menu_multi_row.add_command(
            label="Paste details",
            state="disabled",
            command=self.paste_details,
            **menu_kwargs,
        )
        self.tree_rc_menu_multi_row.add_separator()
        self.tree_rc_menu_multi_row.add_command(
            label="Tag/Untag IDs",
            accelerator="Ctrl+T",
            command=self.tag_ids,
            **menu_kwargs,
        )
        self.tree_rc_menu_multi_row.add_separator()
        self.tree_rc_menu_multi_row.add_command(
            label="Clear IDs details",
            command=self.del_all_details,
            **menu_kwargs,
        )
        self.tree_rc_menu_multi_row.add_command(
            label="Delete IDs",
            accelerator="Del",
            command=self.del_key,
            **menu_kwargs,
        )
        self.tree_rc_menu_multi_row.add_command(
            label="Delete IDs all hierarchies",
            command=self.del_id_all,
            **menu_kwargs,
        )

        # SINGLE ROW MENU - TREE
        self.tree_rc_menu_single_row = tk.Menu(self.treeframe, tearoff=0, **menu_kwargs)
        self.tree_rc_menu_single_row_add = tk.Menu(self.tree_rc_menu_single_row, tearoff=0, **menu_kwargs)
        self.tree_rc_menu_single_row.add_cascade(label="Add", menu=self.tree_rc_menu_single_row_add, **menu_kwargs)
        self.tree_rc_menu_single_row_add.add_command(label="Add child", command=self.add_child_node, **menu_kwargs)
        self.tree_rc_menu_single_row_add.add_command(label="Add sibling", command=self.add_sibling_node, **menu_kwargs)
        self.tree_rc_menu_single_row_add.add_command(label="Add top ID", command=self.add_top_node, **menu_kwargs)

        # cut
        self.tree_rc_menu_single_row_cut = tk.Menu(self.tree_rc_menu_single_row, tearoff=0, **menu_kwargs)
        self.tree_rc_menu_single_row_cut.add_command(
            label="Detach ID",
            accelerator="Ctrl+X",
            command=self.cut_ids,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row_cut.add_command(
            label="Detach children",
            command=self.cut_children,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row.add_cascade(
            label="Cut",
            menu=self.tree_rc_menu_single_row_cut,
            **menu_kwargs,
        )

        # copy
        self.tree_rc_menu_single_row_copy = tk.Menu(self.tree_rc_menu_single_row, tearoff=0, **menu_kwargs)
        self.tree_rc_menu_single_row_copy.add_command(
            label="Copy ID",
            accelerator="Ctrl+C",
            command=self.copy_key,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row_copy.add_command(label="Clipboard row", command=self.copy_ID_row, **menu_kwargs)
        self.tree_rc_menu_single_row_copy.add_command(
            label="Clipboard ID + descendants",
            command=self.copy_ID_children_rows,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row_copy.add_command(label="Copy details", command=self.copy_details, **menu_kwargs)
        self.tree_rc_menu_single_row.add_cascade(label="Copy", menu=self.tree_rc_menu_single_row_copy, **menu_kwargs)

        # paste options
        self.tree_rc_menu_single_row_paste = tk.Menu(self.tree_rc_menu_single_row, tearoff=0, **menu_kwargs)
        self.tree_rc_menu_single_row.add_cascade(
            label="Paste",
            state="normal",
            menu=self.tree_rc_menu_single_row_paste,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row_paste.add_command(label="Paste IDs as sibling", **menu_kwargs)
        self.tree_rc_menu_single_row_paste.add_command(label="Paste IDs as child", **menu_kwargs)
        self.tree_rc_menu_single_row_paste.add_separator()
        self.tree_rc_menu_single_row_paste.add_command(label="Paste IDs and children as sibling", **menu_kwargs)
        self.tree_rc_menu_single_row_paste.add_command(label="Paste IDs and children as child", **menu_kwargs)
        self.tree_rc_menu_single_row_paste.add_separator()
        self.tree_rc_menu_single_row_paste.add_command(label="Attach children", **menu_kwargs)
        self.tree_rc_menu_single_row_paste.add_command(
            label="Paste details",
            state="disabled",
            command=self.paste_details,
            **menu_kwargs,
        )

        # delete options
        self.tree_rc_menu_single_row_del = tk.Menu(self.tree_rc_menu_single_row, tearoff=0, **menu_kwargs)
        self.tree_rc_menu_single_row.add_cascade(label="Delete", menu=self.tree_rc_menu_single_row_del, **menu_kwargs)
        self.tree_rc_menu_single_row_del.add_command(
            label="Clear IDs details",
            command=self.del_all_details,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row_del.add_separator()
        self.tree_rc_menu_single_row_del.add_command(
            label="Delete ID",
            accelerator="Del",
            command=self.del_key,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row_del.add_command(
            label="Delete ID, orphan children",
            command=self.del_id_orphan,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row_del.add_separator()
        self.tree_rc_menu_single_row_del.add_command(
            label="Delete ID all hierarchies",
            command=self.del_id_all,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row_del.add_command(
            label="Delete ID all hierarchies, orphan children",
            command=self.del_id_all_orphan,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row_del.add_separator()
        self.tree_rc_menu_single_row_del.add_command(
            label="Delete ID + children",
            command=self.del_id_children,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row_del.add_command(
            label="Delete ID + children, all hierarchies",
            command=self.del_id_children_all,
            **menu_kwargs,
        )

        self.tree_rc_menu_single_row.add_separator()
        self.tree_rc_menu_single_row.add_command(
            label="ID concise view",
            command=self.show_ids_details_tree,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row.add_command(
            label="Tag/Untag ID",
            accelerator="Ctrl+T",
            command=self.tag_ids,
            **menu_kwargs,
        )
        self.tree_rc_menu_single_row.add_command(label="Rename ID", command=self.rename_node, **menu_kwargs)

        # EMPTY MENU - TREE
        self.tree_rc_menu_empty = tk.Menu(self.treeframe, tearoff=0, **menu_kwargs)
        self.tree_rc_menu_empty.add_command(label="Paste IDs", state="disabled", **menu_kwargs)
        self.tree_rc_menu_empty.add_command(label="Paste IDs and children", state="disabled", **menu_kwargs)
        self.tree_rc_menu_empty.add_command(label="Attach children", state="disabled", **menu_kwargs)
        self.tree_rc_menu_empty.add_separator()
        self.tree_rc_menu_empty.add_command(label="Add top ID", command=self.add_top_node, **menu_kwargs)
        self.tree_rc_menu_empty.add_command(label="Add rows", command=self.add_rows_rc, **menu_kwargs)
        self.tree_rc_menu_empty.add_command(label="Add detail", command=self.rc_add_col, **menu_kwargs)
        self.tree_rc_menu_empty.add_command(label="Add hierarchy", command=self.rc_add_hier_col, **menu_kwargs)

        # SINGLE ROW MENU - SHEET
        self.sheet_rc_menu_single_row = tk.Menu(self.sheet, tearoff=0, **menu_kwargs)
        self.sheet_rc_menu_single_row.add_command(
            label="Tag/Untag ID",
            accelerator="Ctrl+T",
            command=self.tag_ids,
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_command(
            label="Go to ID in Treeview",
            command=self.select_id_in_treeview_from_sheet,
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_command(
            label="ID concise view",
            command=self.show_ids_details_sheet,
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_separator()
        self.sheet_rc_menu_single_row.add_command(
            label="Clipboard",
            accelerator="Ctrl+C",
            command=self.copy_key,
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_command(
            label="Paste",
            accelerator="Ctrl+V",
            command=self.paste_key,
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_separator()
        self.sheet_rc_menu_single_row.add_command(
            label="Copy details",
            command=self.sheet_copy_details,
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_command(
            label="Paste details",
            command=self.sheet_paste_details,
            state="disabled",
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_command(
            label="Clear IDs details",
            command=self.sheet_del_all_details,
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_separator()
        self.sheet_rc_menu_single_row.add_command(
            label="Add top ID",
            command=self.sheet_add_top_node,
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_command(
            label="Insert rows",
            command=lambda: self.add_rows_rc(True),
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_command(
            label="Rename ID",
            command=self.sheet_rename_node,
            **menu_kwargs,
        )
        self.sheet_rc_menu_single_row.add_separator()
        self.sheet_rc_menu_single_row.add_command(
            label="Del IDs, all hierarchies",
            command=self.del_key,
            **menu_kwargs,
        )

        # MULTI ROW MENU - SHEET
        self.sheet_rc_menu_multi_row = tk.Menu(
            self.sheet,
            tearoff=0,
            **menu_kwargs,
        )
        self.sheet_rc_menu_multi_row.add_command(
            label="Tag/Untag IDs",
            accelerator="Ctrl+T",
            command=self.tag_ids,
            **menu_kwargs,
        )
        self.sheet_rc_menu_multi_row.add_separator()
        self.sheet_rc_menu_multi_row.add_command(
            label="Clear all details",
            command=self.sheet_del_all_details,
            **menu_kwargs,
        )
        self.sheet_rc_menu_multi_row.add_command(
            label="Paste details",
            command=self.sheet_paste_details,
            state="disabled",
            **menu_kwargs,
        )
        self.sheet_rc_menu_multi_row.add_separator()
        self.sheet_rc_menu_multi_row.add_command(
            label="Clipboard",
            accelerator="Ctrl+C",
            command=self.copy_key,
            **menu_kwargs,
        )
        self.sheet_rc_menu_multi_row.add_command(
            label="Paste",
            accelerator="Ctrl+V",
            command=self.paste_key,
            **menu_kwargs,
        )
        self.sheet_rc_menu_multi_row.add_separator()
        self.sheet_rc_menu_multi_row.add_command(
            label="Del IDs, all hierarchies",
            accelerator="Del",
            command=self.del_key,
            **menu_kwargs,
        )

        # EMPTY MENU - SHEET
        self.sheet_rc_menu_empty = tk.Menu(
            self.sheet,
            tearoff=0,
            **menu_kwargs,
        )
        self.sheet_rc_menu_empty.add_command(
            label="Add top ID",
            command=self.sheet_add_top_node,
            **menu_kwargs,
        )
        self.sheet_rc_menu_empty.add_command(
            label="Add rows",
            command=self.add_rows_rc,
            **menu_kwargs,
        )
        self.sheet_rc_menu_empty.add_command(
            label="Add detail",
            command=self.rc_add_col,
            **menu_kwargs,
        )
        self.sheet_rc_menu_empty.add_command(
            label="Add hierarchy",
            command=self.rc_add_hier_col,
            **menu_kwargs,
        )

        self.l_frame_id = self.main_canvas.create_window(
            (0, 0),
            window=self.l_frame,
            anchor="nw",
            state="normal",
        )
        self.r_frame_id = self.main_canvas.create_window(
            (0, 0),
            window=self.r_frame,
            anchor="nw",
            state="normal",
        )
        self.main_canvas.create_rectangle(0, 1, 0, 1, fill="gray60", outline="", tag="div")

    def populate(self, program_data=None):
        if program_data:
            self.sheet.MT.data = program_data.records
            self.ic = int(program_data.ic)
            self.pc = int(program_data.pc)
            self.hiers = [int(h) for h in program_data.hiers]
            self.headers = [
                Header(
                    h["name"],
                    h["type"],
                    [tuple(x) for x in h["formatting"]],
                    h["validation"],
                )
                for h in program_data.headers
            ]
            self.row_len = len(self.headers)
            self.changelog = program_data.changelog
            if self.changelog and len(self.changelog[0]) > 5:
                self.changelog = []
            self.sheet.align(program_data.sheet_table_align, redraw=False)
            self.sheet.row_index_align(program_data.sheet_index_align, redraw=False)
            self.sheet.header_align(program_data.sheet_header_align, redraw=False)
            self.tree.align(program_data.sheet_table_align, redraw=False)
            self.tree.header_align(program_data.sheet_header_align, redraw=False)
            self.nodes = self.nodes_json_x_dict(program_data.nodes, hiers=self.hiers)
            self.topnodes_order = {int(h): v for h, v in program_data.topnodes_order.items()}
            self.auto_sort_nodes_bool = bool(program_data.auto_sort_nodes_bool)
            self.tv_label_col = int(program_data.tv_label_col)
            self.sheet.set_row_heights(
                row_heights=map(self.sheet.valid_row_height, map(int, program_data.row_heights)),
            )
            self.sheet.set_column_widths(
                column_widths=map(int, program_data.column_widths),
            )
            self.saved_info = DotDict({int(k): v for k, v in program_data.saved_info.items()})
            for dct in self.saved_info.values():
                dct["theights"] = {k: self.tree.valid_row_height(int(v)) for k, v in dct["theights"].items()}
                dct["twidths"] = {k: int(v) for k, v in dct["twidths"].items()}
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            for c, align in program_data.sheet_column_alignments.items():
                self.sheet.align_columns(int(c), align=align, redraw=False)
                self.tree.align_columns(int(c), align=align, redraw=False)
            self.allow_spaces_ids_var = bool(program_data.allow_spaces_ids)
            self.allow_spaces_columns_var = bool(program_data.allow_spaces_columns)
            self.set_headers()
            self.tag_ids(selection=set(program_data.tagged_ids), toggle=False, do_tree=False)
            if "date_format" in program_data:
                self.DATE_FORM = program_data.date_format
        else:
            self.set_headers()
            self.tagged_ids = set()
            self.pc = int(self.hiers[0])
            self.tv_label_col = int(self.ic)
            self.saved_info = new_saved_info(self.hiers)
            self.topnodes_order = {}
            if not self.C.created_new:
                self.fix_associate_sort()
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            self.remake_topnodes_order()
            self.tree.set_column_widths()
            self.sheet.set_row_heights().set_column_widths()
        self.saved_sheet_row_heights = {}
        self.reset_tree_search_dropdown()
        self.reset_sheet_search_dropdown()
        self.selected_ID = ""
        self.selected_PAR = ""
        self.new_sheet = []
        self.reset_tree_drag_vars()
        self.search_results = []
        self.sheet_search_results = []
        self.sort_later_dct = None
        self.reset_tagged_ids_dropdowns()
        self.C.file.entryconfig("Compare sheets", command=self.compare_from_within_treeframe)
        self.C.file.entryconfig("New", command=self.create_new_from_within_treeframe)
        self.C.file.entryconfig("Open", command=self.open_from_within_treeframe)
        self.refresh_hier_dropdown(self.hiers.index(self.pc))
        self.edit_menu.entryconfig(0, label="Undo  0/30", state="disabled")
        self.tree.unbind("<z>")
        self.tree.unbind("<Z>")
        self.sheet.unbind("<z>")
        self.sheet.unbind("<Z>")
        self.copied_details = {"copied": [], "id": ""}
        self.copied_detail = {"copied": "", "id": ""}
        self.vs = deque(maxlen=30)
        self.vp = 0
        self.cut = []
        self.copied = []
        self.cut_children_dct = {}
        self.sheet.row_index(self.ic)
        self.tree.set_options(show_horizontal_grid=not self.tree.ops.alternate_color)
        self.refresh_formatting()
        self.redo_tree_display()
        self.disable_paste()
        self.refresh_dropdowns()
        if program_data:
            self.move_sheet_pos()
            self.move_tree_pos()
        else:
            self.sheet.set_xview(0.0).set_yview(0.0)
            self.tree.set_xview(0.0).set_yview(0.0)
        self.C.show_frame("tree_edit", start=False, msg=self.get_tree_editor_status_bar_text())
        self.C.file.entryconfig("Settings", state="normal")
        self.WINDOW_DIMENSIONS_CHANGED()
        self.focus_tree()

    def settings(self):
        Settings_Popup(self, theme=self.C.theme)

    def change_theme(self, theme="light_green", write=True):
        self.C.theme = theme
        self.dark_blue_theme_bool = self.C.theme == "dark_blue"
        self.black_theme_bool = self.C.theme == "black"
        self.dark_theme_bool = self.C.theme == "dark"
        self.light_green_theme_bool = self.C.theme == "light_green"
        self.light_blue_theme_bool = self.C.theme == "light_blue"
        self.config(bg=themes[theme].top_left_bg)
        self.C.config(bg=themes[theme].top_left_bg)
        self.main_canvas.config(bg=themes[theme].top_left_bg)
        self.l_frame.config(bg=themes[theme].top_left_bg)
        self.treeframe.config(bg=themes[theme].top_left_bg)
        self.sts_tree.config(bg=themes[theme].top_left_bg)
        self.r_frame.config(bg=themes[theme].top_left_bg)
        self.sheetframe.config(bg=themes[theme].top_left_bg)
        # if USER_OS == "darwin":
        # button_kwargs = {
        #     "background": themes["light_green"].table_bg,
        #     "darkcolor": themes["light_green"].table_bg,
        #     "bordercolor": themes["light_green"].table_grid_fg,
        #     "lightcolor": themes["light_green"].table_bg,
        #     "highlightcolor": themes["light_green"].table_bg,
        #     "foreground": themes["light_green"].table_fg,
        #     "borderwidth": 1 if theme.startswith("light") else 0,
        # }
        # self.C.style.configure("Std.TButton", **button_kwargs)
        # self.C.style.configure("EF.Std.TButton", **button_kwargs)
        # self.C.style.configure("TF.Std.TButton", **button_kwargs)
        # self.C.style.configure("STSF.Std.TButton", **button_kwargs)
        # self.C.style.configure("EFB.Std.TButton", **button_kwargs)
        # self.C.style.configure("ERR_ASK_FNT.Std.TButton", **button_kwargs)
        # self.C.style.configure("x_button.Std.TButton", **button_kwargs)
        # for s in (
        #     "Std.TButton",
        #     "EF.Std.TButton",
        #     "TF.Std.TButton",
        #     "STSF.Std.TButton",
        #     "EFB.Std.TButton",
        #     "ERR_ASK_FNT.Std.TButton",
        #     "x_button.Std.TButton",
        # ):
        #     self.C.style.map(
        #         s,
        #         foreground=[
        #             ("!active", themes["light_green"].table_fg),
        #             ("pressed", themes["light_green"].table_fg),
        #             ("active", themes["light_green"].table_fg),
        #         ],
        #         background=[
        #             ("!active", themes["light_green"].table_bg),
        #             ("pressed", themes["light_green"].table_grid_fg),
        #             ("active", "#91c9f7"),
        #         ],
        #     )

        self.search_entry.config(
            background=themes["light_green"].table_bg,
            foreground=themes["light_green"].table_fg,
            disabledbackground=themes["light_green"].table_bg,
            disabledforeground=themes["light_green"].table_fg,
            insertbackground=themes["light_green"].table_fg,
            readonlybackground=themes["light_green"].table_bg,
        )
        self.sheet_search_entry.config(
            background=themes["light_green"].table_bg,
            foreground=themes["light_green"].table_fg,
            disabledbackground=themes["light_green"].table_bg,
            disabledforeground=themes["light_green"].table_fg,
            insertbackground=themes["light_green"].table_fg,
            readonlybackground=themes["light_green"].table_bg,
        )
        self.C.selection_info.change_theme(theme)

        self.btns_tree.config(bg=themes[theme].top_left_bg)
        self.btns_sheet.config(bg=themes[theme].top_left_bg)

        self.C.status_bar.config(bg=themes[theme].top_left_bg, fg=themes[theme].table_selected_box_cells_fg)
        self.C.status_frame.config(bg=themes[theme].top_left_bg)

        self.C.frames["column_selection"].sheet_selector.config(bg=themes[theme].top_left_bg)
        self.C.frames["column_selection"].sheet_selector.sheets_label.config(
            bg=themes[theme].top_left_bg,
            fg=themes[theme].table_fg,
        )
        self.C.frames["column_selection"].config(bg=themes[theme].top_left_bg)
        self.C.frames["column_selection"].flattened_choices.change_theme(theme)
        self.C.frames["column_selection"].selector.change_theme(theme)
        self.C.frames["column_selection"].flattened_selector.change_theme(theme)
        self.C.frames["tree_compare"].sheet_filename1.change_theme(theme)
        self.C.frames["tree_compare"].sheet_filename2.change_theme(theme)
        self.C.frames["tree_compare"].l_frame.config(
            highlightbackground=themes[theme].table_fg, background=themes[theme].top_left_bg
        )
        self.C.frames["tree_compare"].l_frame_btns.config(background=themes[theme].top_left_bg)
        self.C.frames["tree_compare"].r_frame.config(
            highlightbackground=themes[theme].table_fg, background=themes[theme].top_left_bg
        )
        self.C.frames["tree_compare"].r_frame_btns.config(background=themes[theme].top_left_bg)
        self.C.frames["tree_compare"].selector_1.change_theme(theme)
        self.C.frames["tree_compare"].selector_2.change_theme(theme)
        self.C.frames["tree_compare"].sheetdisplay1.change_theme(theme)
        self.C.frames["tree_compare"].sheetdisplay2.change_theme(theme)
        self.C.frames["tree_compare"].file_label1.change_theme(theme)
        self.C.frames["tree_compare"].file_label2.change_theme(theme)
        self.sheet.change_theme(theme)
        self.tree.change_theme(theme)
        self.C.frames["column_selection"].sheetdisplay.change_theme(theme)
        if write:
            self.C.save_cfg()
        self.focus_tree()

    def destroy_find_popup(self, event=None):
        with suppress(Exception):
            self.find_popup.destroy()
        self.find_popup = None

    def reset_tree(self, extra=True):
        self.destroy_find_popup()
        if extra:
            self.C.file.entryconfig("Compare sheets", command=self.C.compare_at_start)
            self.C.file.entryconfig("Open", command=self.C.open_file_at_start)
            self.C.file.entryconfig("New", command=self.C.create_new_at_start)
            self.C.menubar_state("disabled")
            self.bind_or_unbind_save("disabled")
        self.C.unsaved_changes = False
        self.sheet_changes = 0
        self.tv_label_col = 0
        self.selected_ID = ""
        self.selected_PAR = ""
        self.disable_paste()
        self.changelog = []
        self.search_results = []
        self.sheet_search_results = []
        self.tree.reset()
        self.sheet.data_reference(newdataref=[], redraw=True)
        self.sheet.deselect("all", redraw=False)
        self.sheet.reset_all_options()
        self.headers = []
        self.set_headers()
        self.auto_sort_nodes_bool = True
        self.topnodes_order = {}
        self.nodes = {}
        self.rns = {}
        self.sheet.MT.data = []
        self.new_sheet = []
        self.vs = deque(maxlen=30)
        self.vp = 0
        self.row_len = 0
        self.headers = []
        self.ic = 0
        self.pc = 0
        self.sheet.row_index(newindex=self.ic)
        self.hiers = []
        self.warnings = []
        self.tagged_ids = set()
        self.C.created_new = False
        self.C.change_app_title(title=None)

    def bind_or_unbind_save(self, save_menu_state: Literal["normal", "save as", "disabled"] | None = None):
        if isinstance(save_menu_state, str):
            self.C.save_menu_state = save_menu_state
        self.C.unbind_class("all", f"<{ctrl_button}-s>")
        self.C.unbind_class("all", f"<{ctrl_button}-S>")
        self.C.unbind_class("all", f"<{ctrl_button}-Shift-S>")
        self.C.unbind_class("all", f"<{ctrl_button}-Shift-s>")
        self.C.file.entryconfig("Save", state="disabled")
        self.C.file.entryconfig("Save as", state="disabled")
        self.C.file.entryconfig("Save new version", state="disabled")
        self.C.file.entryconfig("Settings", state="disabled")
        if self.C.save_menu_state == "normal":
            self.C.file.entryconfig("Save", state="normal")
            self.C.file.entryconfig("Save as", state="normal")
            self.C.file.entryconfig("Save new version", state="normal")
            self.C.bind_class("all", f"<{ctrl_button}-s>", self.save_)
            self.C.bind_class("all", f"<{ctrl_button}-S>", self.save_)
            self.C.bind_class("all", f"<{ctrl_button}-Shift-s>", self.save_as)
            self.C.bind_class("all", f"<{ctrl_button}-Shift-S>", self.save_as)
            self.C.file.entryconfig("Settings", state="normal")
        elif self.C.save_menu_state == "save as":
            self.C.file.entryconfig("Save as", state="normal")
            self.C.bind_class("all", f"<{ctrl_button}-s>", self.save_as)
            self.C.bind_class("all", f"<{ctrl_button}-S>", self.save_as)
            self.C.bind_class("all", f"<{ctrl_button}-Shift-s>", self.save_as)
            self.C.bind_class("all", f"<{ctrl_button}-Shift-S>", self.save_as)
            self.C.file.entryconfig("Settings", state="normal")

    def enable_widgets(self, widgets=True, menubar=True):
        self.C.menubar_state("normal")
        for widget in (self.tree, self.sheet):
            widget.bind(f"<{ctrl_button}-e>", self.expand_id)
            widget.bind(f"<{ctrl_button}-E>", self.expand_id)
            widget.bind(f"<{ctrl_button}-r>", self.collapse_id)
            widget.bind(f"<{ctrl_button}-R>", self.collapse_id)
            widget.bind(f"<{ctrl_button}-z>", self.undo)
            widget.bind(f"<{ctrl_button}-Z>", self.undo)
            widget.bind(f"<{ctrl_button}-l>", self.show_changelog)
            widget.bind(f"<{ctrl_button}-L>", self.show_changelog)
            widget.bind(f"<{ctrl_button}-v>", self.paste_key)
            widget.bind(f"<{ctrl_button}-V>", self.paste_key)
            widget.bind(f"<{ctrl_button}-x>", self.cut_key)
            widget.bind(f"<{ctrl_button}-X>", self.cut_key)
            widget.bind(f"<{ctrl_button}-c>", self.copy_key)
            widget.bind(f"<{ctrl_button}-C>", self.copy_key)
            widget.bind(f"<{ctrl_button}-t>", self.tag_ids)
            widget.bind(f"<{ctrl_button}-T>", self.tag_ids)
            widget.bind("<Delete>", self.del_key)
            widget.bind("<Double-Button-1>", self.tree_sheet_double_left)
            widget.extra_bindings(
                [
                    ("begin_column_header_drag_drop", self.snapshot_begin_drag_cols),
                    ("column_header_drag_drop", self.snapshot_drag_cols),
                ]
            )
        self.sheet.extra_bindings(
            [
                ("begin_row_index_drag_drop", self.snapshot_begin_drag_rows),
                ("row_index_drag_drop", self.snapshot_drag_rows),
            ]
        )
        self.tree.extra_bindings(
            [
                ("begin_row_index_drag_drop", self.begin_tree_drag_drop_ids),
                ("row_index_drag_drop", self.tree_drag_drop_ids),
            ]
        )
        self.sheet.enable_bindings(sheet_bindings).basic_bindings(True)
        self.tree.enable_bindings(tree_bindings).basic_bindings(True)
        self.sheet.bind(rc_release, self.sheet_rc_release)
        self.sheet.bind("<<SheetSelect>>", self.sheet_select_event)
        self.tree.bind("<<SheetSelect>>", self.tree_select_event)
        self.tree.bind(rc_press, self.tree_rc_press)
        # self.tree.bind(rc_motion, self.tree_rc_motion)
        # self.tree.bind(rc_release, self.tree_rc_release)
        self.tree.bind("<FocusIn>", self.tree_focus_enter).bind("<FocusOut>", self.tree_focus_leave)
        self.sheet.bind("<FocusIn>", self.sheet_focus_enter).bind("<FocusOut>", self.sheet_focus_leave)
        self.sheet.bulk_table_edit_validation(self.tree_sheet_edit_table)
        self.tree.bulk_table_edit_validation(self.tree_sheet_edit_table)
        self.sheet_tag_id_button.config(state="normal")
        self.sheet_tagged_ids_dropdown.config(state="readonly")
        self.sheet_tagged_ids_dropdown.bind("<<ComboboxSelected>>", self.sheet_go_to_tagged_id)
        self.tree_tag_id_button.config(state="normal")
        self.tree_tagged_ids_dropdown.config(state="readonly")
        self.tree_tagged_ids_dropdown.bind("<<ComboboxSelected>>", self.tree_go_to_tagged_id)
        self.switch_label.config(state="normal")
        self.search_button.config(state="normal")
        self.search_choice_dropdown.config(state="readonly")
        self.search_choice_dropdown.bind("<<ComboboxSelected>>", lambda focus: self.search_entry.focus_set())
        self.search_entry.config(state="normal")
        self.search_entry.bind("<Return>", self.search_choice)
        self.search_dropdown.config(state="readonly")
        self.search_dropdown.bind("<<ComboboxSelected>>", self.show_search_result)
        self.switch_hier_dropdown.config(state="readonly")
        self.switch_hier_dropdown.bind("<<ComboboxSelected>>", self.switch_hier)
        self.sheet_search_button.config(state="normal")
        self.sheet_search_choice_dropdown.config(state="readonly")
        self.sheet_search_entry.enable_me()
        self.sheet_search_entry.bind("<Return>", self.sheet_search_choice)
        self.sheet_search_dropdown.config(state="readonly")
        self.sheet_search_dropdown.bind("<<ComboboxSelected>>", self.sheet_show_search_result)
        self.sheet_search_choice_dropdown.bind(
            "<<ComboboxSelected>>",
            lambda focus: self.sheet_search_entry.focus_set(),
        )
        self.bind_or_unbind_save("save as" if self.C.created_new else "normal")

    def disable_widgets(self):
        self.C.menubar_state("disabled")
        self.bind_or_unbind_save("disabled")
        for x in (self.tree, self.sheet):
            x.unbind(f"<{ctrl_button}-e>")
            x.unbind(f"<{ctrl_button}-E>")
            x.unbind(f"<{ctrl_button}-r>")
            x.unbind(f"<{ctrl_button}-R>")
            x.unbind(f"<{ctrl_button}-z>")
            x.unbind(f"<{ctrl_button}-Z>")
            x.unbind(f"<{ctrl_button}-l>")
            x.unbind(f"<{ctrl_button}-L>")
            x.unbind(f"<{ctrl_button}-t>")
            x.unbind(f"<{ctrl_button}-T>")
            x.unbind(f"<{ctrl_button}-c>")
            x.unbind(f"<{ctrl_button}-C>")
            x.unbind(f"<{ctrl_button}-v>")
            x.unbind(f"<{ctrl_button}-V>")
            x.unbind(f"<{ctrl_button}-x>")
            x.unbind(f"<{ctrl_button}-X>")
            x.unbind("<Delete>")
            # x.disable_bindings().basic_bindings(False)
            x.unbind("<Double-Button-1>")
            x.unbind("<FocusIn>")
            x.unbind("<FocusOut>")
        self.C.unbind(f"<{ctrl_button}-s>")
        self.C.unbind(f"<{ctrl_button}-S>")
        self.sheet.unbind(rc_button)
        self.sheet.extra_bindings(
            [
                ("row_index_drag_drop", None),
                ("all_select_events", None),
                ("column_header_drag_drop", None),
            ]
        )
        self.tree.unbind(rc_press)
        self.tree.unbind(rc_motion)
        self.tree.unbind(rc_release)
        self.sheet_tag_id_button.config(state="disabled")
        self.sheet_tagged_ids_dropdown.config(state="disabled")
        self.sheet_tagged_ids_dropdown.unbind("<<ComboboxSelected>>")
        self.tree_tag_id_button.config(state="disabled")
        self.tree_tagged_ids_dropdown.config(state="disabled")
        self.tree_tagged_ids_dropdown.unbind("<<ComboboxSelected>>")
        self.switch_label.config(state="disabled")
        self.search_button.config(state="disabled")
        self.search_choice_dropdown.config(state="disabled")
        self.search_choice_dropdown.unbind("<<ComboboxSelected>>")
        self.search_entry.config(state="disabled")
        self.search_entry.unbind("<Return>")
        self.search_dropdown.config(state="disabled")
        self.search_dropdown.unbind("<<ComboboxSelected>>")
        self.switch_hier_dropdown.config(state="disabled")
        self.switch_hier_dropdown.bind("<<ComboboxSelected>>")
        self.sheet_search_button.config(state="disabled")
        self.sheet_search_choice_dropdown.config(state="disabled")
        self.sheet_search_entry.disable_me()
        self.sheet_search_entry.unbind("<Return>")
        self.sheet_search_dropdown.config(state="disabled")
        self.sheet_search_dropdown.unbind("<<ComboboxSelected>>")
        self.sheet_search_choice_dropdown.unbind("<<ComboboxSelected>>")

    def toggle_sort_all_nodes(self, enabled, snapshot=True):
        if enabled:
            if snapshot:
                self.snapshot_auto_sort_nodes()
            self.auto_sort_nodes_bool = enabled
            self.sort_all_children()
            self.redo_tree_display()
        else:
            self.auto_sort_nodes_bool = enabled
            self.remake_topnodes_order()

    def sort_all_children(self):
        for n in self.nodes.values():
            for h, cn in n.cn.items():
                if cn:
                    n.cn[h] = self.sort_node_cn(cn, h)

    def copy_ID_row(self, event=None):
        selections = self.tree.selection(cells=True)
        if not selections:
            return
        s, writer = str_io_csv_writer(dialect=csv.excel_tab)
        writer.writerow(h.name for h in self.headers)
        writer.writerows(self.sheet.data[self.rns[iid.lower()]] for iid in selections)
        to_clipboard(self.C, s.getvalue().rstrip())

    def copy_ID_children_rows(self, event=None):
        iids = set(self.tree.selection(cells=True))
        if not iids:
            return
        h = int(self.pc)
        tc = set()
        for iid in iids:
            if self.nodes[iid.lower()].ps[h]:
                if all(pk not in iids and pk not in tc for pk in self.check_ps(self.nodes[iid.lower()].ps[h], h)):
                    tc.add(iid)
            elif iid not in tc:
                tc.add(iid)
        s, writer = str_io_csv_writer(dialect=csv.excel_tab)
        writer.writerow(h.name for h in self.headers)
        for iid in sorted(tc, key=lambda x: self.rns[x]):
            iid_lower = iid.lower()
            stack = [iid_lower]
            while stack:
                ik = stack.pop()
                writer.writerow(self.sheet.data[self.rns[ik]])
                children = self.nodes[ik].cn[self.pc]
                if children:
                    stack.extend(reversed(children))
        to_clipboard(self.C, s.getvalue().rstrip())

    def clipboard_sheet(self, event=None):
        s, writer = str_io_csv_writer(dialect=csv.excel)
        writer.writerow(h.name for h in self.headers)
        writer.writerows(self.sheet.data)
        to_clipboard(self.C, s.getvalue().rstrip())

    def clipboard_sheet_indent(self, event=None):
        s, writer = str_io_csv_writer(dialect=csv.excel_tab)
        writer.writerow(h.name for h in self.headers)
        writer.writerows(self.sheet.data)
        to_clipboard(self.C, s.getvalue().rstrip())

    def clipboard_sheet_json(self, event=None):
        to_clipboard(
            self.C,
            json.dumps(
                full_sheet_to_dict(
                    [h.name for h in self.headers],
                    self.sheet.MT.data,
                    include_headers=True,
                    format_=self.json_format,
                )
            ),
        )

    def changelog_singular(self, text):
        self.changelog[-1] = self.changelog[-1][:1] + (text,) + self.changelog[-1][2:]
        self.increment_unsaved()

    def changelog_append(self, change, id_, old, new):
        self.changelog.append(
            (
                self.get_datetime_changelog(increment_unsaved=True),
                change,
                id_,
                old,
                new,
            )
        )

    def changelog_append_no_unsaved(self, change, id_, old, new):
        self.changelog.append(
            (
                self.get_datetime_changelog(increment_unsaved=False),
                change,
                id_,
                old,
                new,
            )
        )

    def edit_cell_rebuild(self, r, c, value) -> object:
        self.snapshot_ctrl_x_v_del_key_id_par()
        self.edit_cell_single(r, c, value)
        self.rebuild_tree()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        return value

    def edit_cell_single(self, r: int, c: int, value: object) -> None:
        if self.headers[c].type_ == "Number":
            value = self.convert_num(value)
        elif self.headers[c].type_ == "Date":
            value = self.convert_date(value, self.DATE_FORM)
        self.changelog_append(
            "Edit cell",
            f"ID: {self.sheet.MT.data[r][self.ic]} column #{c + 1} named: {self.headers[c].name} with type: {self.headers[c].type_}",
            f"{self.sheet.MT.data[r][c]}",
            value,
        )
        self.sheet.MT.data[r][c] = value
        return value

    def edit_cell_multiple(self, r: int, c: int, value: object) -> None:
        if self.headers[c].type_ == "Number":
            value = self.convert_num(value)
        elif self.headers[c].type_ == "Date":
            value = self.convert_date(value, self.DATE_FORM)
        self.changelog_append_no_unsaved(
            "Edit cell |",
            f"ID: {self.sheet.MT.data[r][self.ic]} column #{c + 1} named: {self.headers[c].name} with type: {self.headers[c].type_}",
            f"{self.sheet.MT.data[r][c]}",
            value,
        )
        self.sheet.MT.data[r][c] = value
        return value

    def tree_sheet_edit_table(self, event=None):
        if not event:
            return None
        if len(event.data) == 1:
            y1, x1 = next(iter(event.data))
            newtext = event.data[(y1, x1)]
            if event.sheetname == "tree":
                y1 = self.rns[self.tree.rowitem(y1, data_index=True)]

            if self.headers[x1].type_ in ("ID", "Parent") and not self.allow_spaces_ids_var:
                newtext = re.sub(r"[\n\t\s]*", "", newtext)

            if newtext == self.sheet.data[y1][x1]:
                event.data = {}
                return None

            ID = self.sheet.data[y1][self.ic]
            ik = ID.lower()

            if self.headers[x1].type_ == "ID":
                id_ = ID
                ik = id_.lower()
                tree_sel = self.tree.selection()
                if not self.change_ID_name(id_, newtext, errors=False):
                    self.edit_cell_rebuild(y1, x1, newtext)
                    event.data = {}
                    return None

                self.changelog_append(
                    "Rename ID",
                    id_,
                    id_,
                    f"{newtext}",
                )
                new_ik = newtext.lower()
                if ik in self.tagged_ids:
                    self.tagged_ids.discard(ik)
                    self.tagged_ids.add(new_ik)
                    self.reset_tagged_ids_dropdowns()
                self.disable_paste()
                self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                self.refresh_formatting(rows=self.refresh_rows)
                self.redo_tree_display()
                self.refresh_rows = set()
                if tree_sel:
                    if self.tree.exists(tree_sel[0]):
                        self.tree.scroll_to_item(tree_sel[0])
                        self.tree.selection_set(tree_sel[0])
                    else:
                        self.tree.scroll_to_item(newtext.lower())
                        self.tree.selection_set(newtext.lower())
                else:
                    self.move_tree_pos()
                self.sheet.set_cell_size_to_text(y1, x1, only_set_if_too_small=True)
                self.tree_set_cell_size_to_text(y1, x1)
                self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

            elif self.headers[x1].type_ == "Parent":
                self.snapshot_paste_id()
                oldparent = f"{self.sheet.data[y1][x1]}"
                tree_sel = self.tree.selection()
                if self.cut_paste_edit_cell(ID, oldparent, x1, newtext):
                    self.changelog_append(
                        "Cut and paste ID + children" if self.nodes[ik].cn[x1] else "Cut and paste ID",
                        ID,
                        f"Old parent: {oldparent if oldparent else 'n/a - Top ID'} old column #{x1 + 1} named: {self.headers[x1].name}",
                        f"New parent: {newtext if newtext else 'n/a - Top ID'} new column #{x1 + 1} named: {self.headers[x1].name}",
                    )
                    self.refresh_formatting(rows=y1, columns=x1)
                    self.redo_tree_display()
                    self.sheet.set_cell_size_to_text(y1, x1, only_set_if_too_small=True)
                    self.tree_set_cell_size_to_text(y1, x1)
                    if tree_sel:
                        self.tree.scroll_to_item(tree_sel[0])
                        self.tree.selection_set(tree_sel)
                    self.disable_paste()
                    self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
                else:
                    self.vs.pop()
                    self.vp -= 1
                    self.set_undo_label()
                    self.edit_cell_rebuild(y1, x1, newtext)
            else:
                if self.detail_is_valid_for_col(x1, newtext):
                    self.snapshot_ctrl_x_v_del_key()
                    self.vs[-1]["cells"][(y1, x1)] = f"{self.sheet.MT.data[y1][x1]}"
                    newtext = self.edit_cell_single(y1, x1, newtext)
                    self.refresh_formatting(rows=y1, columns=x1)
                    self.refresh_tree_item(ID)
                    self.sheet.set_cell_size_to_text(y1, x1, only_set_if_too_small=True)
                    self.tree_set_cell_size_to_text(y1, x1)
                    self.disable_paste()
                    self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
                else:
                    Error(
                        self,
                        f"Entered text invalid for column type - {self.why_isnt_detail_valid(x1, newtext)}   ",
                        theme=self.C.theme,
                    )
            if event.sheetname == "tree" and event.loc:
                self.tree.next_cell(*event.loc, event.key)
            elif event.sheetname == "sheet" and event.loc:
                self.sheet.next_cell(*event.loc, event.key)

        else:
            self.start_work("Editing table...")
            idcols = set(self.hiers) | {self.ic}
            need_rebuild = any(k[1] in idcols for k in event["data"])
            refresh_rows = set()
            refresh_cols = set()
            edit_ctr = 0
            tree = event.sheetname == "tree"
            if need_rebuild:
                self.snapshot_ctrl_x_v_del_key_id_par()
            else:
                self.snapshot_ctrl_x_v_del_key()
            for (r, c), value in event["data"].items():
                if tree:
                    r = self.rns[self.tree.rowitem(row=r, data_index=True)]
                if (need_rebuild and c in idcols) or (
                    self.detail_is_valid_for_col(c, value) and self.sheet.MT.data[r][c] != value
                ):
                    if not need_rebuild:
                        self.vs[-1]["cells"][(r, c)] = f"{self.sheet.MT.data[r][c]}"
                    self.edit_cell_multiple(r, c, value)
                    refresh_rows.add(r)
                    refresh_cols.add(c)
                    edit_ctr += 1
            self.disable_paste()
            if edit_ctr:
                if need_rebuild:
                    self.rebuild_tree()
                else:
                    self.refresh_formatting(rows=refresh_rows, columns=refresh_cols)
                    for rn in refresh_rows:
                        self.refresh_tree_item(self.sheet.MT.data[rn][self.ic])
                if edit_ctr > 1:
                    self.changelog_append(
                        f"Edit {edit_ctr} cells",
                        "",
                        "",
                        "",
                    )
                else:
                    self.changelog_singular("Edit cell")
                self.redraw_sheets()
                self.stop_work(self.get_tree_editor_status_bar_text())
            else:
                self.vp -= 1
                self.set_undo_label()
                self.vs.pop()
                self.redraw_sheets()
                self.stop_work(self.get_tree_editor_status_bar_text())
        event.data = {}

    def tree_set_cell_size_to_text(self, sheet_r, sheet_c):
        if self.tree.exists(self.sheet.data[sheet_r][self.ic].lower()) and self.tree.item_displayed(
            self.sheet.data[sheet_r][self.ic].lower()
        ):
            self.tree.set_cell_size_to_text(
                self.tree.displayed_rows.index(self.tree.itemrow(self.sheet.data[sheet_r][self.ic].lower())),
                sheet_c,
                only_set_if_too_small=True,
            )

    def replace_using_mapping(self, event=None) -> None:
        Replace_Popup(self, theme=self.C.theme)

    def copy_key(self, event: object = None) -> None:
        if self.tree.has_focus():
            iids = tuple(
                self.tree.rowitem(row)
                for box in self.tree.boxes
                for row in range(box.coords.from_r, box.coords.upto_r)
                if box[1] == "rows"
            )
            if iids:
                self.copy_ID(iids=iids)
            else:
                self.tree.copy()
        elif self.sheet.has_focus():
            self.sheet.copy()

    def del_key(self, event: object = None) -> None:
        if self.tree.has_focus():
            if iids := tuple(
                self.tree.rowitem(row)
                for box in self.tree.boxes
                for row in range(box.coords.from_r, box.coords.upto_r)
                if box[1] == "rows"
            ):
                self.del_id(iids=iids)
            elif self.tree.boxes:
                self.tree.delete()
        elif self.sheet.has_focus():
            if rows := self.sheet.get_selected_rows():
                self.del_id_all(iids=[self.sheet.MT.data[r][self.ic].lower() for r in rows])
            elif self.sheet.boxes:
                self.sheet.delete()

    def rebuild_tree(self, deselect=True, redraw=False):
        if deselect:
            self.sheet.deselect("all", redraw=False)
        self.nodes = {}
        self.clear_copied_details()
        self.auto_sort_nodes_bool = True
        self.save_info_get_saved_info()
        self.sheet.MT.data, self.nodes = TreeBuilder().build(
            input_sheet=self.sheet.MT.data,
            output_sheet=self.new_sheet,
            row_len=self.row_len,
            ic=self.ic,
            hiers=self.hiers,
            nodes=self.nodes,
            add_warnings=False,
            strip=not self.allow_spaces_ids_var,
        )
        self.new_sheet = []
        self.fix_associate_sort_edit_cells()
        self.rns = {}
        rhs = []
        default_row_height = self.sheet.MT.get_default_row_height()
        for i, r in enumerate(self.sheet.data):
            ik = r[self.ic].lower()
            self.rns[ik] = i
            if ik in self.saved_sheet_row_heights:
                rhs.append(self.saved_sheet_row_heights[ik])
            else:
                rhs.append(default_row_height)
        self.sheet.set_row_heights(rhs)
        self.reset_tagged_ids_dropdowns()
        self.rehighlight_tagged_ids()
        self.refresh_rows = set()
        self.refresh_formatting()
        if redraw:
            self.redraw_sheets()
        self.redo_tree_display()
        self.sheet.recreate_all_selection_boxes()
        return

    def cut_key(self, event: object = None) -> None:
        if self.tree.has_focus():
            if iids := tuple(
                self.tree.rowitem(row)
                for box in self.tree.boxes
                for row in range(box.coords.from_r, box.coords.upto_r)
                if box[1] == "rows"
            ):
                self.cut_ids(iids=iids)
            elif self.tree.ctrl_boxes:
                self.tree.cut()
        elif self.sheet.has_focus() and self.sheet.ctrl_boxes:
            self.sheet.cut()

    def paste_key(self, event: object = None) -> None:
        if self.tree.has_focus():
            if event:
                if self.tree.selected:
                    rows = sorted(self.tree.get_selected_rows())
                    if rows and len(rows) == 1 and self.winfo_containing(event.x_root, event.y_root) is not None:
                        self.tree_rc_menu_single_row_paste.tk_popup(self.C.winfo_pointerx(), self.C.winfo_pointery())
                    else:
                        self.tree.paste()
                elif self.winfo_containing(event.x_root, event.y_root) is not None:
                    self.tree_rc_menu_empty.tk_popup(self.C.winfo_pointerx(), self.C.winfo_pointery())
            elif self.tree.selected:
                self.tree.paste()
        elif self.sheet.has_focus() and self.sheet.selected:
            self.sheet.paste()

    def select_id_in_treeview_from_sheet(self, event=None):
        ik = self.sheet.MT.data[self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)[0]][
            self.ic
        ].lower()
        self.go_to_treeview_id_finder(ik)

    def start_work(self, msg="", outside_treeframe=False):
        self.C.working = True
        self.C.save_menu_state = "disabled"
        if not outside_treeframe:
            self.disable_widgets()
        self.C.status_bar.change_text(msg)

    def stop_work(self, msg="", outside_treeframe=False):
        self.C.working = False
        self.C.save_menu_state = "normal"
        if self.C.USER_HAS_QUIT:
            self.C.USER_HAS_CLOSED_WINDOW()  # user still wants to quit
            if self.C.USER_HAS_QUIT:
                return
        if not outside_treeframe:
            self.after_idle(self.enable_widgets)
        self.C.status_bar.change_text(msg)

    def hide_frames(self, l_frame=False, r_frame=False, set_dimensions=True):
        if l_frame:
            self.main_canvas.itemconfig(
                self.l_frame_id,
                state="hidden",
            )
        if r_frame:
            self.main_canvas.itemconfig(
                self.r_frame_id,
                state="hidden",
            )
        if set_dimensions:
            self.WINDOW_DIMENSIONS_CHANGED()

    def unhide_frames(self, l_frame=False, r_frame=False, set_dimensions=True):
        if l_frame:
            self.main_canvas.itemconfig(
                self.l_frame_id,
                state="normal",
            )
        if r_frame:
            self.main_canvas.itemconfig(
                self.r_frame_id,
                state="normal",
            )
        if set_dimensions:
            self.WINDOW_DIMENSIONS_CHANGED()

    def ask_continue_unsaved(self):
        if self.C.unsaved_changes:
            confirm = Ask_Confirm(
                self,
                "You have unsaved changes, continue anyway?",
                theme=self.C.theme,
            )
            return confirm.boolean
        else:
            return True

    def compare_from_within_treeframe(self):
        if not self.ask_continue_unsaved():
            return
        self.reset_tree()
        self.bind_or_unbind_save("disabled")
        self.C.frames["tree_compare"].populate()

    def open_from_within_treeframe(self, event=None):
        if not self.ask_continue_unsaved():
            return
        fp = filedialog.askopenfilename(parent=self.C, title="Select file")
        if not fp:
            return
        try:
            fp = os.path.normpath(fp)
        except Exception:
            Error(self, "Filepath invalid   ", theme=self.C.theme)
            return
        if not fp.lower().endswith((".json", ".xlsx", ".xls", ".xlsm", ".csv", ".tsv")):
            Error(self, "Please select json/excel/csv   ", theme=self.C.theme)
            return
        self.disable_widgets()
        if os.path.isfile(fp):
            self.C.open_dict["filepath"] = fp
            self.reset_tree()
            self.C.load_from_file()
        else:
            Error(self, "Filepath invalid   ", theme=self.C.theme)
            self.enable_widgets()

    def create_new_from_within_treeframe(self, event=None):
        if not self.ask_continue_unsaved():
            return
        self.reset_tree(False)
        self.headers = [
            Header("ID", "ID"),
            Header("DETAIL_1"),
            Header("PARENT_1", "Parent"),
        ]
        self.ic = 0
        self.tv_label_col = 0
        self.pc = 2
        self.hiers = [2]
        self.row_len = 3
        self.C.created_new = True
        self.C.open_dict["filepath"] = "New sheet"
        self.C.change_app_title(title="New sheet")
        self.C.open_dict["sheet"] = "Sheet1"
        self.warnings_filepath = "n/a - CREATED NEW"
        self.warnings_sheet = "n/a"
        self.populate()

    def enter_divider(self, event):
        if not self.currently_adjusting_divider:
            self.main_canvas.config(cursor="sb_h_double_arrow")

    def leave_divider(self, event):
        if not self.currently_adjusting_divider:
            self.main_canvas.config(cursor="")

    def divider_b1_press(self, event):
        self.currently_adjusting_divider = True

    def divider_b1_motion(self, event):
        if self.currently_adjusting_divider:
            self.l_frame_proportion = float(round(event.x / self.winfo_width(), 2))
            if self.l_frame_proportion < 0.01:
                self.l_frame_proportion = 0.01
            elif self.l_frame_proportion > 0.99:
                self.l_frame_proportion = 0.99
            self.WINDOW_DIMENSIONS_CHANGED(place_left_panel=False)

    def divider_b1_release(self, event):
        self.currently_adjusting_divider = False

    def unhide_adjustable_divider(self):
        self.main_canvas.itemconfig("div", state="normal")
        self.main_canvas.tag_bind("div", "<Enter>", self.enter_divider)
        self.main_canvas.tag_bind("div", "<Leave>", self.leave_divider)
        self.main_canvas.tag_bind("div", "<ButtonPress-1>", self.divider_b1_press)
        self.main_canvas.tag_bind("div", "<B1-Motion>", self.divider_b1_motion)
        self.main_canvas.tag_bind("div", "<ButtonRelease-1>", self.divider_b1_release)

    def hide_adjustable_divider(self):
        self.main_canvas.itemconfig("div", state="hidden")
        self.main_canvas.tag_unbind("div", "<Enter>")
        self.main_canvas.tag_unbind("div", "<Leave>")
        self.main_canvas.tag_unbind("div", "<ButtonPress-1>")
        self.main_canvas.tag_unbind("div", "<B1-Motion>")
        self.main_canvas.tag_unbind("div", "<ButtonRelease-1>")

    def get_display_option(self):
        if self.full_left_bool.get():
            return "left"
        if self.adjustable_bool.get():
            return "adjustable"
        if self.full_right_bool.get():
            return "right"
        if self._50_50_bool.get():
            return "50/50"

    def set_display_option(self, option: Literal["left", "adjustable", "right", "50/50"]) -> None:
        if option == "left":
            self.option_full_left(event="config")
        elif option == "right":
            self.option_full_right(event="config")
        elif option == "adjustable":
            self.option_adjustable(event="config")
        elif option == "50/50":
            self.option_50_50(event="config")
        self.WINDOW_DIMENSIONS_CHANGED()

    def option_adjustable(self, event=None):
        if event is None:
            if not (
                self.full_left_bool.get()
                or self.full_right_bool.get()
                or self._50_50_bool.get()
                or self.adjustable_bool.get()
            ):
                self.adjustable_bool.set(True)
                return
            self.unhide_adjustable_divider()
            if self.full_left_bool.get():
                self.unhide_frames(r_frame=True)
                self.full_left_bool.set(False)
                self.focus_tree()
            elif self.full_right_bool.get():
                self.unhide_frames(l_frame=True)
                self.full_right_bool.set(False)
                self.focus_sheet()
            elif self._50_50_bool.get():
                self._50_50_bool.set(False)
            self.WINDOW_DIMENSIONS_CHANGED()
        elif event == "config":
            self.full_left_bool.set(False)
            self.full_right_bool.set(False)
            self._50_50_bool.set(False)
            self.adjustable_bool.set(True)
            self.unhide_adjustable_divider()
            self.unhide_frames(l_frame=True, r_frame=True)
        else:
            if self.adjustable_bool.get():
                self.adjustable_bool.set(False)
            else:
                self.adjustable_bool.set(True)
            self.option_adjustable()

    def option_50_50(self, event=None):
        if event is None:
            if not (
                self.full_left_bool.get()
                or self.full_right_bool.get()
                or self._50_50_bool.get()
                or self.adjustable_bool.get()
            ):
                self._50_50_bool.set(True)
                return
            if self.full_left_bool.get():
                self.unhide_frames(r_frame=True)
                self.full_left_bool.set(False)
                self.focus_tree()
            elif self.full_right_bool.get():
                self.unhide_frames(l_frame=True)
                self.full_right_bool.set(False)
                self.focus_sheet()
            elif self.adjustable_bool.get():
                self.adjustable_bool.set(False)
            self.hide_adjustable_divider()
            self.WINDOW_DIMENSIONS_CHANGED()
        elif event == "config":
            self.full_left_bool.set(False)
            self.full_right_bool.set(False)
            self._50_50_bool.set(True)
            self.adjustable_bool.set(False)
            self.hide_adjustable_divider()
            self.unhide_frames(l_frame=True, r_frame=True)
        else:
            if self._50_50_bool.get():
                self._50_50_bool.set(False)
            else:
                self._50_50_bool.set(True)
            self.option_50_50()

    def option_full_left(self, event=None):
        if event is None:
            if not (
                self.full_left_bool.get()
                or self.full_right_bool.get()
                or self._50_50_bool.get()
                or self.adjustable_bool.get()
            ):
                self.full_left_bool.set(True)
                return
            if self._50_50_bool.get():
                self._50_50_bool.set(False)
            elif self.full_right_bool.get():
                self.unhide_frames(l_frame=True, set_dimensions=False)
                self.full_right_bool.set(False)
            elif self.adjustable_bool.get():
                self.adjustable_bool.set(False)
            self.hide_adjustable_divider()
            self.hide_frames(r_frame=True)
            self.focus_tree()
        elif event == "config":
            self.full_left_bool.set(True)
            self.full_right_bool.set(False)
            self._50_50_bool.set(False)
            self.adjustable_bool.set(False)
            self.hide_adjustable_divider()
            self.hide_frames(r_frame=True, set_dimensions=False)
            self.unhide_frames(l_frame=True)
            self.focus_tree()
        else:
            if self.full_left_bool.get():
                self.full_left_bool.set(False)
            else:
                self.full_left_bool.set(True)
            self.option_full_left()

    def option_full_right(self, event=None):
        if event is None:
            if not (
                self.full_left_bool.get()
                or self.full_right_bool.get()
                or self._50_50_bool.get()
                or self.adjustable_bool.get()
            ):
                self.full_right_bool.set(True)
                return
            if self._50_50_bool.get():
                self._50_50_bool.set(False)
            elif self.full_left_bool.get():
                self.unhide_frames(r_frame=True, set_dimensions=False)
                self.full_left_bool.set(False)
            elif self.adjustable_bool.get():
                self.adjustable_bool.set(False)
            self.hide_adjustable_divider()
            self.hide_frames(l_frame=True)
            self.focus_sheet()
        elif event == "config":
            self.full_left_bool.set(False)
            self.full_right_bool.set(True)
            self._50_50_bool.set(False)
            self.adjustable_bool.set(False)
            self.hide_adjustable_divider()
            self.hide_frames(l_frame=True, set_dimensions=False)
            self.unhide_frames(r_frame=True)
            self.focus_sheet()
        else:
            if self.full_right_bool.get():
                self.full_right_bool.set(False)
            else:
                self.full_right_bool.set(True)
            self.option_full_right()

    def WINDOW_DIMENSIONS_CHANGED(self, event=None, place_left_panel=True):
        if event is not None:
            if event.height == self.last_height and event.width == self.last_width:
                return
            self.last_width = event.width
            self.last_height = event.height
        if self.C.current_frame == "tree_edit":
            if self.adjustable_bool.get():
                width = self.winfo_width()
                height = self.winfo_height()
                if self.l_frame_proportion == 0.01:
                    l_frame_width = 1
                    l_frame_x = -1
                    r_frame_width = int(width) - 5
                    r_frame_x = 5
                elif self.l_frame_proportion == 0.99:
                    l_frame_width = int(width) - 5
                    l_frame_x = 0
                    r_frame_width = 1
                    r_frame_x = width + 1
                else:
                    l_frame_width = int(width * self.l_frame_proportion)
                    r_frame_width = int(width - l_frame_width) - 5
                    l_frame_x = 0
                    r_frame_x = l_frame_width + 5
                self.main_canvas.itemconfig(self.l_frame_id, width=l_frame_width, height=height)
                self.main_canvas.itemconfig(self.r_frame_id, width=r_frame_width, height=height)
                self.btns_tree.update_idletasks()
                self.btns_sheet.update_idletasks()
                self.main_canvas.update_idletasks()
                self.main_canvas.coords("div", l_frame_x + l_frame_width, 0, l_frame_width + 5, height)
                self.main_canvas.coords(self.l_frame_id, l_frame_x, 0)
                self.main_canvas.coords(self.r_frame_id, r_frame_x, 0)
                self.btns_tree.update_idletasks()
                self.btns_sheet.update_idletasks()
                self.main_canvas.update_idletasks()
            elif self._50_50_bool.get():
                width = floor(self.winfo_width() / 2) - 1
                height = self.winfo_height()
                self.main_canvas.coords(self.l_frame_id, 0, 0)
                self.main_canvas.coords(self.r_frame_id, width + 1, 0)
                self.main_canvas.itemconfig(self.l_frame_id, width=width, height=height)
                self.main_canvas.itemconfig(self.r_frame_id, width=self.winfo_width() - width - 1, height=height)
            elif self.full_left_bool.get():
                self.main_canvas.coords(self.l_frame_id, 0, 0)
                self.main_canvas.itemconfig(
                    self.l_frame_id,
                    width=self.winfo_width(),
                    height=self.winfo_height(),
                )
            elif self.full_right_bool.get():
                self.main_canvas.coords(self.r_frame_id, 0, 0)
                self.main_canvas.itemconfig(
                    self.r_frame_id,
                    width=self.winfo_width(),
                    height=self.winfo_height(),
                )

    def fix_headers(self, headers, row_len, warnings=True):
        if len(headers) < row_len:
            headers += list(repeat("", row_len - len(headers)))
        tally_of_headers = defaultdict(lambda: -1)
        allow_whitespace = self.allow_spaces_columns_var
        for coln in range(len(headers)):
            cell = headers[coln]
            if not cell:
                cell = f"_MISSING_{coln + 1}"
                if warnings:
                    self.warnings.append(f" - Missing header in column #{coln + 1}")
            if not allow_whitespace:
                if warnings:
                    if " " in cell:
                        self.warnings.append(f" - Spaces in header column #{coln + 1}")
                    if "\n" in cell:
                        self.warnings.append(f" - Newlines in header column #{coln + 1}")
                    if "\r" in cell:
                        self.warnings.append(f" - Carriage returns in header column #{coln + 1}")
                    if "\t" in cell:
                        self.warnings.append(f" - Tabs in header column #{coln + 1}")
                cell = "".join(cell.strip().split())
            hk = cell.lower()
            tally_of_headers[hk] += 1
            if tally_of_headers[hk] > 0:
                if warnings:
                    self.warnings.append(f" - Duplicate header in column #{coln + 1}")
                orig = cell
                x = 1
                while hk in tally_of_headers:
                    cell = f"{orig}_DUPLICATED_{x}"
                    hk = cell.lower()
                    x += 1
                tally_of_headers[hk] += 1
            headers[coln] = cell
        return headers

    def remove_selections(self, event=None):
        self.sheet.deselect()
        self.tree.deselect()

    def tree_select_event(self, event):
        selected = event.selected
        if selected and self.tree.data:
            iid = self.tree.rowitem(selected.row)
            self.selected_ID = self.nodes[iid].name
            pariid = self.tree.parent(iid)
            if pariid == "":
                self.selected_PAR = ""
            else:
                self.selected_PAR = self.nodes[pariid].name
            if self.mirror_var and not self.mirror_sels_disabler:
                self.go_to_row()
            self.mirror_sels_disabler = False
        else:
            self.selected_ID = ""
            self.selected_PAR = ""
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        self.C.selection_info.set_my_value(self.get_tree_selection_info())

    def sheet_select_event(self, event=None):
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        self.C.selection_info.set_my_value(self.get_sheet_selection_info())

    def get_tree_selection_info(self):
        count = 0
        _sum = 0.0
        quick_data = self.tree.MT.data
        quick_displayed_rows = self.tree.MT.displayed_rows
        for r, c in self.tree.gen_selected_cells(get_rows=True, get_columns=True):
            count += 1
            try:
                _sum += float(quick_data[quick_displayed_rows[r]][c])
            except Exception:
                continue
        s = f"Count {count} Sum {int(_sum) if _sum.is_integer() else _sum}"
        self.C.selection_info.config(width=len(s))
        return s

    def get_sheet_selection_info(self):
        count = 0
        _sum = 0.0
        quick_data = self.sheet.MT.data
        for r, c in self.sheet.gen_selected_cells(get_rows=True, get_columns=True):
            count += 1
            try:
                _sum += float(quick_data[r][c])
            except Exception:
                continue
        s = f"Count {count} Sum {int(_sum) if _sum.is_integer() else _sum}"
        self.C.selection_info.config(width=len(s))
        return s

    def get_tree_editor_status_bar_text(self):
        if self.tree.selected:
            sels = self.tree.selection()
            box = next(reversed(self.tree.boxes))
            if box.type_ == "rows":
                tree_addition = f"|   Tree {len(sels)} IDs selected   "
            elif box.type_ == "columns":
                tree_addition = f"|   Tree Columns: {_n2a(box.coords.from_c)}:{_n2a(box.coords.upto_c - 1)}   "
            else:
                if box.coords.upto_r - box.coords.from_r == 1 and box.coords.upto_c - box.coords.from_c == 1:
                    tree_addition = f"|   Tree Cells: {_n2a(box.coords.from_c)}{box.coords.from_r + 2}   "
                else:
                    tree_addition = f"|   Tree Cells: {_n2a(box.coords.from_c)}{box.coords.from_r + 2}:{_n2a(box.coords.upto_c - 1)}{box.coords.upto_r + 1}   "
        else:
            tree_addition = ""
        if self.sheet.selected:
            box = next(reversed(self.sheet.boxes))
            if box.type_ == "rows":
                sheet_addition = f"|   Sheet Rows: {box.coords.from_r + 2}:{box.coords.upto_r + 1}   "
            elif box.type_ == "columns":
                sheet_addition = f"|   Sheet Columns: {_n2a(box.coords.from_c)}:{_n2a(box.coords.upto_c - 1)}   "
            else:
                if box.coords.upto_r - box.coords.from_r == 1 and box.coords.upto_c - box.coords.from_c == 1:
                    sheet_addition = f"|   Sheet Cells: {_n2a(box.coords.from_c)}{box.coords.from_r + 2}   "
                else:
                    sheet_addition = f"|   Sheet Cells: {_n2a(box.coords.from_c)}{box.coords.from_r + 2}:{_n2a(box.coords.upto_c - 1)}{box.coords.upto_r + 1}   "
        else:
            sheet_addition = ""
        if self.copied:
            cc_add = (
                f"|   Copied {len(self.copied)} IDs   "
                if len(self.copied) > 3
                else f"|   Copied: {', '.join(self.nodes[dct['id']].name for dct in self.copied)}   "
            )
        elif self.cut:
            cc_add = (
                f"|   Cut {len(self.cut)} IDs   "
                if len(self.cut) > 3
                else f"|   Cut: {', '.join(self.nodes[dct['id']].name for dct in self.cut)}   "
            )
        else:
            cc_add = ""
        if self.changelog:
            end = f"|   Last Edit: {self.changelog[-1][1]}   {cc_add}"
        else:
            end = f"|   No Changes Made   {cc_add}"
        return f"{len(self.sheet.MT.data)} IDs   {tree_addition}{sheet_addition}{end}"

    def tree_rc_press(self, event):
        self.focus_tree()
        row = self.tree.identify_row(event, allow_end=False)
        col = self.tree.identify_column(event, allow_end=False)
        region = self.tree.identify_region(event)
        if region == "header" and isinstance(col, int):
            if not self.tree.column_selected(col):
                self.tree.select_column(col)
        elif region == "table" and isinstance(row, int) and isinstance(col, int):
            if not self.tree.cell_selected(
                row,
                col,
                rows=True,
                columns=True,
            ):
                self.tree.select_cell(row, col)
        elif region == "index" and isinstance(row, int):
            rows = self.tree.get_selected_rows()
            if row not in rows:
                self.tree.select_row(row)
                rows = {row}
            if len(rows) == 1:
                iid = self.tree.rowitem(rows.pop())
                self.drag_pc = int(self.pc)
                self.drag_iid = iid
                self.drag_pariid = self.tree.parent(iid)
                self.last_rced = iid
                self.drag_start_index = self.tree.index(iid)
        else:
            self.tree.deselect()
        self.tree_rc_release(event)

    def tree_rc_motion(self, event):
        pass
        # if self.auto_sort_nodes_bool or self.drag_iid is None:
        #     return
        # iid = self.tree.rowitem(self.tree.identify_row(event, allow_end=False))
        # if not iid or iid == self.last_rced:
        #     return
        # selections = self.tree.selection()
        # if not selections or selections[0] != self.drag_iid:
        #     self.tree_drop_iid()
        #     return
        # if not selections or selections[0] != self.drag_iid:
        #     self.tree_drop_iid()
        #     return
        # if self.pc != self.drag_pc or len(selections) > 1:
        #     self.reset_tree_drag_vars()
        #     return
        # if iid:
        #     pariid = self.tree.parent(iid)
        #     if pariid != self.drag_pariid:
        #         return
        #     # try:
        #     move_to_index = self.tree.index(iid)
        #     parik = self.drag_pariid.lower()
        #     if parik:
        #         pop_index = self.nodes[parik].cn[self.pc].index(self.drag_iid)
        #         self.nodes[parik].cn[self.pc].insert(
        #             move_to_index,
        #             self.nodes[parik].cn[self.pc].pop(pop_index),
        #         )
        #     else:
        #         pop_index = self.topnodes_order[self.pc].index(self.drag_iid)
        #         self.topnodes_order[self.pc].insert(
        #             move_to_index,
        #             self.topnodes_order[self.pc].pop(pop_index),
        #         )
        #     self.tree.move(self.drag_iid, self.drag_pariid, move_to_index)
        #     self.tree.selection_set(self.drag_iid)
        #     # except Exception:
        #     #     self.tree_drop_iid()

    def tree_sheet_rc_menu_option_enabler_disabler(self, col: int):
        if col == self.ic or col in self.hiers:
            self.tree_sheet_rc_menu_single_col.entryconfig("Type", state="disabled")
            self.tree_sheet_rc_menu_single_col.entryconfig("Validation", state="disabled")
        else:
            self.tree_sheet_rc_menu_single_col.entryconfig("Type", state="normal")
            self.tree_sheet_rc_menu_single_col.entryconfig("Validation", state="normal")

    def tree_rc_release(self, event):
        if self.drag_iid is not None:
            self.drag_end_index = self.tree.index(self.drag_iid)
        self.drag_iid = None  # del if enabling rc move
        if self.auto_sort_nodes_bool or self.drag_iid is None or self.drag_end_index == self.drag_start_index:
            row = self.tree.identify_row(event, allow_end=False)
            col = self.tree.identify_column(event, allow_end=False)
            self.tree_sheet_rc_menu_option_enabler_disabler(col)
            region = self.tree.identify_region(event)
            if region == "header":
                if isinstance(col, int):
                    self.treecolsel = col
                    if len(self.tree.get_selected_columns()) > 1:
                        self.tree_sheet_rc_menu_multi_col.tk_popup(event.x_root, event.y_root)
                    else:
                        self.tree_sheet_rc_menu_single_col.tk_popup(event.x_root, event.y_root)
                else:
                    self.tree_rc_menu_empty.tk_popup(event.x_root, event.y_root)
            elif region == "index":
                if self.auto_sort_nodes_bool:
                    with suppress(Exception):
                        self.tree_rc_menu_single_row.delete("Sort children")
                    with suppress(Exception):
                        self.tree_rc_menu_multi_row.delete("Sort children")
                else:
                    if self.tree_rc_menu_single_row.entrycget("end", "label") != "Sort children":
                        self.tree_rc_menu_single_row.add_command(
                            label="Sort children",
                            command=self.tree_sort_children,
                            **menu_kwargs,
                        )
                    if self.tree_rc_menu_multi_row.entrycget("end", "label") != "Sort children":
                        self.tree_rc_menu_multi_row.add_command(
                            label="Sort children",
                            command=self.tree_sort_children,
                            **menu_kwargs,
                        )
                self.treecolsel = self.ic
                if isinstance(row, int):
                    if len(self.tree.get_selected_rows()) > 1:
                        self.tree_rc_menu_multi_row.tk_popup(event.x_root, event.y_root)
                    else:
                        self.tree_rc_menu_single_row.tk_popup(event.x_root, event.y_root)
                else:
                    self.tree_rc_menu_empty.tk_popup(event.x_root, event.y_root)
            elif region == "table":
                if isinstance(row, int) and isinstance(col, int):
                    self.treecolsel = col
                    if len(self.tree.get_selected_cells()) > 1:
                        self.tree_sheet_rc_menu_multi_cell.tk_popup(event.x_root, event.y_root)
                    else:
                        self.tree_sheet_rc_menu_single_cell.entryconfig(
                            0,
                            label=self.headers[self.treecolsel].name[:25],
                        )
                        self.tree_sheet_rc_menu_single_cell.tk_popup(event.x_root, event.y_root)
                else:
                    self.tree_rc_menu_empty.tk_popup(event.x_root, event.y_root)
        self.reset_tree_drag_vars()

    def tree_sort_children(self, event=None):
        for iid in self.tree.selection():
            self.nodes[iid].cn[self.pc] = self.sort_node_cn(self.nodes[iid].cn[self.pc], self.pc)
        self.save_info_get_saved_info()
        self.redo_tree_display()

    def reset_tree_drag_vars(self):
        self.drag_pc = None
        self.drag_iid = None
        self.drag_pariid = None
        self.drag_start_index = None
        self.drag_end_index = None
        self.last_rced = None

    def tree_drop_iid(self):
        self.reset_tree_drag_vars()
        self.redo_tree_display()

    def sheet_rc_release(self, event):
        self.focus_sheet()
        row = self.sheet.identify_row(event, allow_end=False)
        col = self.sheet.identify_column(event, allow_end=False)
        self.tree_sheet_rc_menu_option_enabler_disabler(col)
        region = self.sheet.identify_region(event)
        if region == "header":
            if isinstance(col, int):
                self.treecolsel = col
                if len(self.sheet.get_selected_columns()) > 1:
                    self.tree_sheet_rc_menu_multi_col.tk_popup(event.x_root, event.y_root)
                else:
                    self.tree_sheet_rc_menu_single_col.tk_popup(event.x_root, event.y_root)
            else:
                self.sheet_rc_menu_empty.tk_popup(event.x_root, event.y_root)
        elif region == "index":
            self.treecolsel = self.ic
            if isinstance(row, int):
                if len(self.sheet.get_selected_rows()) > 1:
                    self.sheet_rc_menu_multi_row.tk_popup(event.x_root, event.y_root)
                else:
                    self.sheet_rc_menu_single_row.tk_popup(event.x_root, event.y_root)
            else:
                self.sheet_rc_menu_empty.tk_popup(event.x_root, event.y_root)
        elif region == "table":
            if isinstance(row, int) and isinstance(col, int):
                self.treecolsel = col
                if len(self.sheet.get_selected_cells()) > 1:
                    self.tree_sheet_rc_menu_multi_cell.tk_popup(event.x_root, event.y_root)
                else:
                    self.tree_sheet_rc_menu_single_cell.entryconfig(
                        0,
                        label=self.headers[self.treecolsel].name[:25],
                    )
                    self.tree_sheet_rc_menu_single_cell.tk_popup(event.x_root, event.y_root)
            else:
                self.sheet_rc_menu_empty.tk_popup(event.x_root, event.y_root)

    def tree_sheet_double_left(self, event):
        if (
            self.tree.event_widget_is_sheet(event)
            and (column := self.tree.identify_column(event, allow_end=False)) is not None
        ) or (
            self.sheet.event_widget_is_sheet(event)
            and (column := self.sheet.identify_column(event, allow_end=False)) is not None
        ):
            self.treecolsel = column

    def switch_hier(self, event=None, hier: int | None = None):
        if isinstance(hier, int):
            self.switch_displayed.set(self.headers[hier].name)
            index = self.hiers.index(hier)
        else:
            index = self.switch_hier_dropdown.current()
            if self.hiers[index] == self.pc:
                self.focus_tree()
                return
        self.save_info_get_saved_info()
        self.pc = int(self.hiers[index])
        self.tree.close_dropdown()
        self.redo_tree_display()
        self.move_tree_pos()
        self.mirror_sels_disabler = True
        self.refresh_tree_dropdowns()
        self.focus_tree()

    def next_hier(self, event=None):
        if self.pc == self.hiers[-1]:
            self.switch_hier(hier=self.hiers[0])
        else:
            self.switch_hier(hier=self.hiers[self.switch_hier_dropdown.current() + 1])

    def check_cn(self, iid: str, h: int) -> Generator[str]:
        stack = [iid]
        while stack:
            current = stack.pop()
            yield current
            stack.extend(reversed(self.nodes[current].cn[h]))

    def check_ps(self, iid: str, h: int) -> Generator[str]:
        current = iid
        while True:
            yield current
            if not self.nodes[current].ps[h]:
                break
            current = self.nodes[current].ps[h]

    def add(self, ID, parent, insert_row=None, snapshot=True, errors=True):
        ik = ID.lower()
        pk = parent.lower()
        if ik in self.nodes and self.nodes[ik].ps[self.pc] is not None:
            if errors:
                Error(self, "ID already in hierarchy   ", theme=self.C.theme)
            return False
        if snapshot:
            self.snapshot_add_id()
        if ik not in self.nodes:
            self.nodes[ik] = Node(ID, self.hiers)
            newrow = list(repeat("", self.row_len))
            newrow[self.ic] = ID
            newrow[self.pc] = parent
            if insert_row is None:
                self.sheet.insert_row(newrow)
                rn = len(self.sheet.MT.data) - 1
                self.rns[ik] = rn
            else:
                self.sheet.insert_row(newrow, insert_row)
                rn = int(insert_row)
            if snapshot:
                self.vs[-1]["row"]["added_or_changed"] = "added"
                self.vs[-1]["row"]["rn"] = rn
        else:
            rn = self.rns[ik]
            if snapshot:
                self.vs[-1]["row"]["added_or_changed"] = "changed"
                self.vs[-1]["row"]["rn"] = rn
                self.vs[-1]["row"]["stored"] = self.sheet.MT.data[rn].copy()
            self.sheet.MT.data[rn][self.pc] = parent
        if parent == "":
            self.nodes[ik].ps[self.pc] = ""
        else:
            self.nodes[ik].ps[self.pc] = pk
            self.nodes[pk].cn[self.pc].append(ik)
            if self.auto_sort_nodes_bool:
                self.nodes[pk].cn[self.pc] = self.sort_node_cn(self.nodes[pk].cn[self.pc], self.pc)
                if self.nodes[pk].ps[self.pc]:
                    parent_parent_node = self.nodes[self.nodes[pk].ps[self.pc]]
                    parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)
        if not self.auto_sort_nodes_bool and parent == "":
            self.topnodes_order[self.pc].append(ik)
        if insert_row is not None and snapshot:
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        if snapshot:
            self.refresh_formatting(rows=len(self.sheet.data) - 1 if insert_row is None else insert_row)
        return True

    def change_ID_name(self, ID, new_name, snapshot=True, errors=True):
        self.refresh_rows = set()
        ik = ID.lower()
        if ik not in self.nodes:
            if errors:
                Error(self, "ID doesn't exist   ", theme=self.C.theme)
            return False
        nnk = new_name.lower()
        if nnk in self.nodes and ik != nnk:
            if errors:
                Error(self, "New name already exists   ", theme=self.C.theme)
            return False
        if not nnk:
            if errors:
                Error(self, "New name cannot be empty   ", theme=self.C.theme)
            return False
        if snapshot:
            self.snapshot_rename_id()
            qvsrwsapp = self.vs[-1]["rows"].append
        ik_rn = self.rns[ik]
        self.sheet.MT.data[ik_rn][self.ic] = new_name
        for h, cn in self.nodes[ik].cn.items():
            for ciid in cn:
                chld_rn = self.rns[ciid]
                self.refresh_rows.add(chld_rn)
                if snapshot:
                    qvsrwsapp(zlib.compress(pickle.dumps((chld_rn, h, self.sheet.MT.data[chld_rn][h]))))
                self.sheet.MT.data[chld_rn][h] = f"{new_name}"
                self.nodes[ciid].ps[h] = nnk
        for h, p in self.nodes[ik].ps.items():
            if p:
                self.nodes[p].cn[h][self.nodes[p].cn[h].index(ik)] = nnk
        if snapshot:
            self.vs[-1]["ikrow"] = (ik_rn, ik, self.nodes[ik].name, new_name)
        self.nodes[ik].name = new_name
        self.nodes[nnk] = self.nodes.pop(ik)
        self.rns[nnk] = self.rns.pop(ik)
        if self.auto_sort_nodes_bool:
            for h, p in self.nodes[nnk].ps.items():
                if p:
                    parent_node = self.nodes[self.nodes[nnk].ps[h]]
                    parent_node.cn[h] = self.sort_node_cn(parent_node.cn[h], h)

        else:
            for h in self.hiers:
                if self.nodes[nnk].ps[h] == "":
                    try:
                        self.topnodes_order[h][self.topnodes_order[h].index(ik)] = nnk
                    except Exception:
                        continue
        if ik in self.saved_info[self.pc].opens:
            self.saved_info[self.pc].opens[nnk] = self.saved_info[self.pc].opens.pop(ik)
        return True

    def cut_paste(
        self,
        ID,
        oldparent,
        hier,
        newparent,
        snapshot=True,
        errors=True,
        sort_later=False,
    ):
        self.refresh_rows = set()
        if self.sort_later_dct is None:
            self.sort_later_dct = {
                "filled": False,
                "old_parents_of_parents": set(),
                "old_hier": None,
                "new_parent": (),
                "new_parent_of_parent": (),
            }
        ik = ID.lower()
        pk = oldparent.lower()
        npk = newparent.lower()
        parent_of_ik: str = self.nodes[ik].ps[hier]
        if ik == npk:
            if errors:
                Error(self, "New parent is ID   ", theme=self.C.theme)
            return False
        if hier != self.pc and self.nodes[ik].ps[self.pc] is not None:
            if errors:
                Error(self, f"ID: {ID} already in hierarchy   ", theme=self.C.theme)
            return False
        if npk == "":
            if self.nodes[ik].ps[self.pc] == "":
                if errors:
                    Error(self, f"ID: {ID} already has this parent   ", theme=self.C.theme)
                return False
        else:
            if self.nodes[ik].ps[self.pc] and npk == self.nodes[ik].ps[self.pc]:
                if errors:
                    Error(self, f"ID: {ID} already has this parent   ", theme=self.C.theme)
                return False
        auto_sort_quick = self.auto_sort_nodes_bool
        for ciid in self.nodes[ik].cn[hier]:
            child = self.nodes[ciid]
            child.ps[hier] = parent_of_ik
            crow = self.rns[ciid]
            if snapshot:
                self.vs[-1]["rows"].append(
                    zlib.compress(
                        pickle.dumps(
                            (
                                crow,
                                hier,
                                self.sheet.MT.data[crow][hier],
                                self.pc,
                                self.sheet.MT.data[crow][self.pc],
                            )
                        )
                    )
                )
                self.refresh_rows.add(int(crow))
            self.sheet.MT.data[crow][hier] = self.nodes[parent_of_ik].name if parent_of_ik else ""
            if parent_of_ik:
                self.nodes[parent_of_ik].cn[hier].append(ciid)
            elif not parent_of_ik and not auto_sort_quick:
                self.topnodes_order[hier].append(ciid)
        self.nodes[ik].cn[hier] = []
        self.nodes[ik].ps[hier] = None
        if pk != "":
            self.nodes[pk].cn[hier].remove(ik)
        if npk == "":
            self.nodes[ik].ps[self.pc] = ""
        else:
            self.nodes[ik].ps[self.pc] = npk
            self.nodes[npk].cn[self.pc].append(ik)
            if auto_sort_quick:
                if sort_later and not self.sort_later_dct["filled"]:
                    self.sort_later_dct["new_parent"] = (npk, self.pc)
                    if self.nodes[npk].ps[self.pc]:
                        self.sort_later_dct["new_parent_of_parent"] = (
                            self.nodes[npk].ps[self.pc],
                            self.pc,
                        )
                elif not sort_later:
                    self.nodes[npk].cn[self.pc] = self.sort_node_cn(self.nodes[npk].cn[self.pc], self.pc)
                    if self.nodes[npk].ps[self.pc]:
                        parent_parent_node = self.nodes[self.nodes[npk].ps[self.pc]]
                        parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)
        if not auto_sort_quick:
            if pk == "":
                try_remove(self.topnodes_order[hier], ik)
            if npk == "":
                self.topnodes_order[self.pc].append(ik)
        idrow = self.rns[ik]
        if snapshot:
            self.vs[-1]["rows"].append(
                zlib.compress(
                    pickle.dumps(
                        (
                            idrow,
                            hier,
                            self.sheet.MT.data[idrow][hier],
                            self.pc,
                            self.sheet.MT.data[idrow][self.pc],
                        )
                    )
                )
            )
            self.refresh_rows.add(int(idrow))
        self.sheet.MT.data[idrow][hier] = ""
        self.sheet.MT.data[idrow][self.pc] = newparent
        if auto_sort_quick and parent_of_ik and self.nodes[parent_of_ik].ps[hier]:
            parent_parent_iid = self.nodes[parent_of_ik].ps[hier]
            if sort_later:
                self.sort_later_dct["old_parents_of_parents"].add(parent_parent_iid)
                if self.sort_later_dct["old_hier"] is None:
                    self.sort_later_dct["old_hier"] = hier
            elif not sort_later:
                parent_parent_node = self.nodes[parent_parent_iid]
                parent_parent_node.cn[hier] = self.sort_node_cn(parent_parent_node.cn[hier], hier)
        self.sort_later_dct["filled"] = True
        return True

    def cut_paste_all(
        self,
        ID,
        oldparent,
        hier,
        newparent,
        snapshot=True,
        errors=True,
        sort_later=False,
    ):
        self.refresh_rows = set()
        if self.sort_later_dct is None:
            self.sort_later_dct = {
                "filled": False,
                "old_parents_of_parents": set(),
                "old_hier": None,
                "new_parent": (),
                "new_parent_of_parent": (),
            }
        ik = ID.lower()
        pk = oldparent.lower()
        npk = newparent.lower()

        # Validation checks
        if hier != self.pc:
            if self.nodes[ik].ps[self.pc] is not None:
                if errors:
                    Error(self, f"ID: {ID} already in hierarchy   ", theme=self.C.theme)
                return False
            for ck in self.check_cn(ik, hier):
                if self.nodes[ck].ps[self.pc] is not None:
                    if errors:
                        Error(
                            self,
                            f"ID: {self.nodes[ck].name} is already in hierarchy   ",
                            theme=self.C.theme,
                        )
                    return False
        else:
            if any(npk == ck for ck in self.check_cn(ik, hier)):
                if errors:
                    Error(self, f"Cannot add ID: {ID} to same line   ", theme=self.C.theme)
                return False
        if npk == "":
            if self.nodes[ik].ps[self.pc] == "":
                if errors:
                    Error(self, f"ID: {ID} already has this parent   ", theme=self.C.theme)
                return False
        else:
            if self.nodes[ik].ps[self.pc] and npk == self.nodes[ik].ps[self.pc]:
                if errors:
                    Error(self, f"ID: {ID} already has this parent   ", theme=self.C.theme)
                return False

        # Update node relationships
        self.nodes[ik].ps[hier] = None
        if pk != "":
            self.nodes[pk].cn[hier].remove(ik)
        if npk == "":
            self.nodes[ik].ps[self.pc] = ""
        else:
            self.nodes[ik].ps[self.pc] = npk
            self.nodes[npk].cn[self.pc].append(ik)

        # Handle top-level node ordering
        if not self.auto_sort_nodes_bool:
            if pk == "":
                try_remove(self.topnodes_order[hier], ik)
            if npk == "":
                self.topnodes_order[self.pc].append(ik)

        # Update sheet data and snapshot for the root node
        idrow = self.rns[ik]
        if snapshot:
            self.vs[-1]["rows"].append(
                zlib.compress(
                    pickle.dumps(
                        (
                            idrow,
                            hier,
                            self.sheet.MT.data[idrow][hier],
                            self.pc,
                            self.sheet.MT.data[idrow][self.pc],
                        )
                    )
                )
            )
            self.refresh_rows.add(int(idrow))
        self.sheet.MT.data[idrow][hier] = ""
        self.sheet.MT.data[idrow][self.pc] = newparent

        # Process the subtree iteratively if hierarchies differ
        if hier != self.pc:
            self.nodes[ik].cn[self.pc] = list(self.nodes[ik].cn[hier])
            self.nodes[ik].cn[hier] = []
            # Initialize stack with the root node of the subtree
            stack = [self.nodes[ik]]
            while stack:
                node = stack.pop()
                for ciid in node.cn[self.pc]:
                    child = self.nodes[ciid]
                    # Update child’s hierarchy data
                    child.ps[self.pc] = child.ps[hier]
                    child.ps[hier] = None
                    child.cn[self.pc] = list(child.cn[hier])
                    child.cn[hier] = []
                    # Update sheet data and snapshot for the child
                    crow = self.rns[ciid]
                    if snapshot:
                        self.vs[-1]["rows"].append(
                            zlib.compress(
                                pickle.dumps(
                                    (
                                        crow,
                                        hier,
                                        self.sheet.MT.data[crow][hier],
                                        self.pc,
                                        self.sheet.MT.data[crow][self.pc],
                                    )
                                )
                            )
                        )
                        self.refresh_rows.add(int(crow))
                    self.sheet.MT.data[crow][self.pc] = f"{self.sheet.MT.data[crow][hier]}"
                    self.sheet.MT.data[crow][hier] = ""
                    # Add child to stack for further processing
                    stack.append(child)

        # Handle sorting
        if self.auto_sort_nodes_bool:
            if sort_later:
                if npk and not self.sort_later_dct["filled"]:
                    self.sort_later_dct["new_parent"] = (npk, self.pc)
                    if self.nodes[npk].ps[self.pc]:
                        self.sort_later_dct["new_parent_of_parent"] = (
                            self.nodes[npk].ps[self.pc],
                            self.pc,
                        )
                if pk and self.nodes[pk].ps[hier]:
                    self.sort_later_dct["old_parents_of_parents"].add(self.nodes[pk].ps[hier])
                    if self.sort_later_dct["old_hier"] is None:
                        self.sort_later_dct["old_hier"] = hier
            elif not sort_later:
                if npk:
                    self.nodes[npk].cn[self.pc] = self.sort_node_cn(self.nodes[npk].cn[self.pc], self.pc)
                    if self.nodes[npk].ps[self.pc]:
                        parent_parent_node = self.nodes[self.nodes[npk].ps[self.pc]]
                        parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)
                if pk and self.nodes[pk].ps[hier]:
                    parent_parent_node = self.nodes[self.nodes[pk].ps[hier]]
                    parent_parent_node.cn[hier] = self.sort_node_cn(parent_parent_node.cn[hier], hier)

        self.sort_later_dct["filled"] = True
        return True

    def copy_paste(self, ID, hier, newparent, snapshot=True, errors=True, sort_later=False):
        self.refresh_rows = set()
        if self.sort_later_dct is None:
            self.sort_later_dct = {
                "filled": False,
                "old_parents_of_parents": set(),
                "old_hier": None,
                "new_parent": (),
                "new_parent_of_parent": (),
            }
        ik = ID.lower()
        npk = newparent.lower()
        if hier == self.pc or self.nodes[ik].ps[self.pc] is not None:
            if errors:
                Error(self, f"ID {ID} already in hierarchy   ", theme=self.C.theme)
            return False
        if npk == "":
            self.nodes[ik].ps[self.pc] = ""
        else:
            self.nodes[ik].ps[self.pc] = npk
            self.nodes[npk].cn[self.pc].append(ik)
            if self.auto_sort_nodes_bool:
                if sort_later and not self.sort_later_dct["filled"]:
                    self.sort_later_dct["new_parent"] = (npk, self.pc)
                    if self.nodes[npk].ps[self.pc]:
                        self.sort_later_dct["new_parent_of_parent"] = (
                            self.nodes[npk].ps[self.pc],
                            self.pc,
                        )
                elif not sort_later:
                    self.nodes[npk].cn[self.pc] = self.sort_node_cn(self.nodes[npk].cn[self.pc], self.pc)
                    if self.nodes[npk].ps[self.pc]:
                        parent_parent_node = self.nodes[self.nodes[npk].ps[self.pc]]
                        parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)
        if not self.auto_sort_nodes_bool and npk == "":
            self.topnodes_order[self.pc].append(ik)
        rn = self.rns[ik]
        if snapshot:
            self.vs[-1]["rows"].append(
                zlib.compress(
                    pickle.dumps(
                        (
                            rn,
                            hier,
                            self.sheet.MT.data[rn][hier],
                            self.pc,
                            self.sheet.MT.data[rn][self.pc],
                        )
                    )
                )
            )
            self.refresh_rows.add(rn)
        self.sheet.MT.data[rn][self.pc] = newparent
        self.sort_later_dct["filled"] = True
        return True

    def copy_paste_all(self, ID, hier, newparent, snapshot=True, errors=True, sort_later=False):
        self.refresh_rows = set()
        if self.sort_later_dct is None:
            self.sort_later_dct = {
                "filled": False,
                "old_parents_of_parents": set(),
                "old_hier": None,
                "new_parent": (),
                "new_parent_of_parent": (),
            }
        ik = ID.lower()
        npk = newparent.lower()

        # Validation checks
        if hier == self.pc or self.nodes[ik].ps[self.pc] is not None:
            if errors:
                Error(self, f"ID {ID} already in hierarchy   ", theme=self.C.theme)
            return False
        for ck in self.check_cn(ik, hier):
            if self.nodes[ck].ps[self.pc] is not None:
                if errors:
                    Error(
                        self,
                        f"ID: {self.nodes[ck].name} is already in hierarchy   ",
                        theme=self.C.theme,
                    )
                return False

        # Update root node relationships
        if npk == "":
            self.nodes[ik].ps[self.pc] = ""
        else:
            self.nodes[ik].ps[self.pc] = npk
            self.nodes[npk].cn[self.pc].append(ik)
        if not self.auto_sort_nodes_bool and npk == "":
            self.topnodes_order[self.pc].append(ik)

        # Update sheet data and snapshot for the root node
        rn = self.rns[ik]
        if snapshot:
            self.vs[-1]["rows"].append(
                zlib.compress(
                    pickle.dumps(
                        (
                            rn,
                            hier,
                            self.sheet.MT.data[rn][hier],
                            self.pc,
                            self.sheet.MT.data[rn][self.pc],
                        )
                    )
                )
            )
            self.refresh_rows.add(rn)
        self.sheet.MT.data[rn][self.pc] = newparent

        # Process the subtree iteratively
        self.nodes[ik].cn[self.pc] = list(self.nodes[ik].cn[hier])
        # Initialize stack with the root node's children
        stack = list(self.nodes[ik].cn[hier])
        while stack:
            ciid = stack.pop()
            child = self.nodes[ciid]
            crow = self.rns[ciid]
            # Update child data and snapshot
            if snapshot:
                self.vs[-1]["rows"].append(
                    zlib.compress(
                        pickle.dumps(
                            (
                                crow,
                                hier,
                                self.sheet.MT.data[crow][hier],
                                self.pc,
                                self.sheet.MT.data[crow][self.pc],
                            )
                        )
                    )
                )
                self.refresh_rows.add(int(crow))
            child.ps[self.pc] = child.ps[hier]
            child.cn[self.pc] = list(child.cn[hier])
            self.sheet.MT.data[crow][self.pc] = f"{self.sheet.MT.data[crow][hier]}"
            # Add the child's children to the stack
            stack.extend(child.cn[hier])

        # Handle sorting
        if npk and self.auto_sort_nodes_bool:
            if sort_later and not self.sort_later_dct["filled"]:
                self.sort_later_dct["new_parent"] = (npk, self.pc)
                if self.nodes[npk].ps[self.pc]:
                    self.sort_later_dct["new_parent_of_parent"] = (
                        self.nodes[npk].ps[self.pc],
                        self.pc,
                    )
            elif not sort_later:
                self.nodes[npk].cn[self.pc] = self.sort_node_cn(self.nodes[npk].cn[self.pc], self.pc)
                if self.nodes[npk].ps[self.pc]:
                    parent_parent_node = self.nodes[self.nodes[npk].ps[self.pc]]
                    parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)

        self.sort_later_dct["filled"] = True
        return True

    def cut_paste_children(self, oldparent, newparent, hier, snapshot=True, errors=True):
        self.refresh_rows = set()
        pk = oldparent.lower()
        npk = newparent.lower()
        if not len(self.nodes[pk].cn[hier]):
            if errors:
                Error(
                    self,
                    f"{self.nodes[pk].name} has no children   ",
                    theme=self.C.theme,
                )
            return
        already_in = set()
        if hier != self.pc:
            for ciid in self.nodes[pk].cn[hier]:
                for diid in self.check_cn(ciid, hier):
                    if self.nodes[diid].ps[self.pc] is not None:
                        already_in.add(ciid)
                        break
            if len(already_in) == len(self.nodes[pk].cn[hier]):
                if errors:
                    Error(
                        self,
                        f"Unable to move children, key IDs are already in {self.headers[self.pc].name}   ",
                        theme=self.C.theme,
                    )
                return
        else:
            if any(npk == ck for ck in self.check_cn(pk, hier)):
                if errors:
                    Error(self, "Cannot add ID to same line   ", theme=self.C.theme)
                return False
            if pk == npk:
                if errors:
                    Error(self, "Children already have this parent   ", theme=self.C.theme)
                return False
        if already_in and errors and snapshot:
            confirm = Ask_Confirm(
                self,
                f"Move {oldparent}'s children\nCannot move the following IDs to {self.headers[self.pc].name}, skip?\n{', '.join(already_in)}",
                theme=self.C.theme,
            )
            if not confirm.boolean:
                return False

        for ciid in tuple(self.nodes[pk].cn[hier]):
            if ciid not in already_in:
                # Move the direct child
                if not self.auto_sort_nodes_bool and npk == "":
                    self.topnodes_order[self.pc].append(ciid)
                crow = self.rns[ciid]
                if snapshot:
                    self.vs[-1]["rows"].append(
                        zlib.compress(
                            pickle.dumps(
                                (
                                    crow,
                                    hier,
                                    self.sheet.MT.data[crow][hier],
                                    self.pc,
                                    self.sheet.MT.data[crow][self.pc],
                                )
                            )
                        )
                    )
                    self.refresh_rows.add(int(crow))
                self.sheet.MT.data[crow][hier] = ""
                self.nodes[ciid].ps[hier] = None
                if npk:
                    self.sheet.MT.data[crow][self.pc] = self.nodes[npk].name
                    self.nodes[ciid].ps[self.pc] = npk
                    self.nodes[npk].cn[self.pc].append(ciid)
                else:
                    self.sheet.MT.data[crow][self.pc] = ""
                    self.nodes[ciid].ps[self.pc] = ""
                self.nodes[pk].cn[hier].remove(ciid)

                # Process the child's subtree iteratively if hier != self.pc
                if hier != self.pc:
                    stack = [ciid]
                    while stack:
                        current_iid = stack.pop()
                        # Get children not already in self.pc
                        children = [
                            child_iid for child_iid in self.nodes[current_iid].cn[hier] if child_iid not in already_in
                        ]
                        self.nodes[current_iid].cn[self.pc] = children
                        children_to_remove = set(children)
                        self.nodes[current_iid].cn[hier] = [
                            child for child in self.nodes[current_iid].cn[hier] if child not in children_to_remove
                        ]
                        # Push children in reverse to maintain DFS order (first child processed first)
                        for child_iid in reversed(children):
                            child = self.nodes[child_iid]
                            crow = self.rns[child_iid]
                            if snapshot:
                                self.vs[-1]["rows"].append(
                                    zlib.compress(
                                        pickle.dumps(
                                            (
                                                crow,
                                                hier,
                                                self.sheet.MT.data[crow][hier],
                                                self.pc,
                                                self.sheet.MT.data[crow][self.pc],
                                            )
                                        )
                                    )
                                )
                                self.refresh_rows.add(int(crow))
                            self.sheet.MT.data[crow][self.pc] = f"{self.sheet.MT.data[crow][hier]}"
                            self.sheet.MT.data[crow][hier] = ""
                            child.ps[self.pc] = current_iid  # Parent in self.pc is the current node
                            child.ps[hier] = None
                            stack.append(child_iid)

            if self.auto_sort_nodes_bool:
                if self.nodes[pk].ps[hier]:
                    parent_parent_node = self.nodes[self.nodes[pk].ps[hier]]
                    parent_parent_node.cn[hier] = self.sort_node_cn(parent_parent_node.cn[hier], hier)
                if npk:
                    if self.nodes[npk].ps[self.pc]:
                        parent_parent_node = self.nodes[self.nodes[npk].ps[self.pc]]
                        parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)
                    self.nodes[npk].cn[self.pc] = self.sort_node_cn(self.nodes[npk].cn[self.pc], self.pc)
        return True

    def cut_paste_edit_cell(self, ID, oldparent, hier, newparent, snapshot=True):
        ik = ID.lower()
        pk = oldparent.lower()
        npk = newparent.lower()
        if ik == npk:
            return False
        if npk != "":
            if npk not in self.nodes:
                return False
            if self.nodes[npk].ps[hier] is None:
                return False
            if self.nodes[ik].ps[hier] and npk == self.nodes[ik].ps[hier]:
                return False
        else:
            if self.nodes[ik].ps[hier] == "":
                return False
        if any(npk == ck for ck in self.check_cn(ik, hier)):
            return False
        if oldparent == "" and self.nodes[ik].ps[hier] is None and newparent:
            for ck in self.check_cn(ik, hier):
                if self.nodes[ck].ps[hier] is not None:
                    return False
        self.nodes[ik].ps[hier] = None
        if pk != "":
            self.nodes[pk].cn[hier].remove(ik)
        if npk == "":
            self.nodes[ik].ps[hier] = ""
        else:
            self.nodes[ik].ps[hier] = npk
            self.nodes[npk].cn[hier].append(ik)
            if self.auto_sort_nodes_bool:
                self.nodes[npk].cn[hier] = self.sort_node_cn(self.nodes[npk].cn[hier], hier)
                if self.nodes[npk].ps[hier]:
                    parent_parent_node = self.nodes[self.nodes[npk].ps[hier]]
                    parent_parent_node.cn[hier] = self.sort_node_cn(parent_parent_node.cn[hier], hier)
        if not self.auto_sort_nodes_bool:
            if pk == "":
                try_remove(self.topnodes_order[hier], ik)
            if npk == "":
                self.topnodes_order[hier].append(ik)
        idrow = self.rns[ik]
        if snapshot:
            self.vs[-1]["rows"].append(
                zlib.compress(
                    pickle.dumps(
                        (
                            idrow,
                            hier,
                            self.sheet.MT.data[idrow][hier],
                            hier,
                            self.sheet.MT.data[idrow][hier],
                        )
                    )
                )
            )
        self.sheet.MT.data[idrow][hier] = newparent
        return True

    def _del_id_core(self, name: str, to_del: list[str] | None = None, snapshot: bool = True) -> list[str]:
        if to_del is None:
            to_del = []
        iid = name.lower()
        pk = self.get_ids_parent(iid)
        if pk:
            self.nodes[pk].cn[self.pc].remove(iid)
        if not self.auto_sort_nodes_bool:
            if pk == "":
                self.topnodes_order[self.pc].remove(iid)
                for ciid in self.nodes[iid].cn[self.pc]:
                    self.topnodes_order[self.pc].append(ciid)
            else:
                for ciid in self.nodes[iid].cn[self.pc]:
                    self.nodes[pk].cn[self.pc].append(ciid)
        else:
            if pk:
                for ciid in self.nodes[iid].cn[self.pc]:
                    self.nodes[pk].cn[self.pc].append(ciid)
                self.nodes[pk].cn[self.pc] = self.sort_node_cn(self.nodes[pk].cn[self.pc], self.pc)
        if pk:
            for ciid in self.nodes[iid].cn[self.pc]:
                rn = self.rns[ciid]
                if snapshot:
                    self.vs[-1]["rows"].append(
                        Del_stre(
                            0,
                            rn,
                            zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h] for h in self.hiers])),
                        )
                    )
                self.nodes[ciid].ps[self.pc] = pk
                self.sheet.MT.data[rn][self.pc] = self.nodes[pk].name
                self.refresh_rows.add(ciid)
        elif pk == "":
            for ciid in self.nodes[iid].cn[self.pc]:
                rn = self.rns[ciid]
                if snapshot:
                    self.vs[-1]["rows"].append(
                        Del_stre(
                            0,
                            rn,
                            zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h] for h in self.hiers])),
                        )
                    )
                self.nodes[ciid].ps[self.pc] = ""
                self.sheet.MT.data[rn][self.pc] = ""
                self.refresh_rows.add(ciid)
        rn = self.rns[iid]
        if sum(1 for v in self.nodes[iid].ps.values() if v is not None) < 2:
            if snapshot:
                self.vs[-1]["rows"].append(Del_stre(1, rn, self.sheet.MT.data[rn]))
            del self.nodes[iid]
            self.untag_id(iid)
            to_del.append(iid)
        else:
            if snapshot:
                self.vs[-1]["rows"].append(
                    Del_stre(
                        0,
                        rn,
                        zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h] for h in self.hiers])),
                    )
                )
            self.nodes[iid].cn[self.pc] = []
            self.nodes[iid].ps[self.pc] = None
            self.sheet.MT.data[rn][self.pc] = ""
            self.refresh_rows.add(iid)
        if self.auto_sort_nodes_bool and pk and self.nodes[pk].ps[self.pc]:
            parent_parent_node = self.nodes[self.nodes[pk].ps[self.pc]]
            parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)
        return to_del

    def _del_id_all_core(self, name: str, to_del: list[str] | None = None, snapshot: bool = True) -> list[str]:
        if to_del is None:
            to_del = []
        iid = name.lower()
        piid = self.get_ids_parent(iid)
        self.untag_id(iid)
        if not self.auto_sort_nodes_bool:
            for h, pk in self.nodes[iid].ps.items():
                if pk == "":
                    self.topnodes_order[h].remove(iid)
                    for ciid in self.nodes[iid].cn[h]:
                        child = self.nodes[ciid]
                        self.topnodes_order[h].append(ciid)
                        child.ps[h] = ""
                        rn = self.rns[ciid]
                        if snapshot:
                            self.vs[-1]["rows"].append(
                                Del_stre(
                                    0,
                                    rn,
                                    zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h_] for h_ in self.hiers])),
                                )
                            )
                            self.refresh_rows.add(ciid)
                        self.sheet.MT.data[rn][h] = ""
                elif pk:
                    for ciid in self.nodes[iid].cn[h]:
                        self.nodes[pk].cn[h].append(ciid)
                        child = self.nodes[ciid]
                        child.ps[h] = pk
                        rn = self.rns[ciid]
                        if snapshot:
                            self.vs[-1]["rows"].append(
                                Del_stre(
                                    0,
                                    rn,
                                    zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h_] for h_ in self.hiers])),
                                )
                            )
                            self.refresh_rows.add(ciid)
                        self.sheet.MT.data[rn][h] = self.nodes[pk].name
                    self.nodes[pk].cn[h].remove(iid)
        else:
            for h, pk in self.nodes[iid].ps.items():
                if pk:
                    for ciid in self.nodes[iid].cn[h]:
                        self.nodes[pk].cn[h].append(ciid)
                        child = self.nodes[ciid]
                        child.ps[h] = pk
                        rn = self.rns[ciid]
                        if snapshot:
                            self.vs[-1]["rows"].append(
                                Del_stre(
                                    0,
                                    rn,
                                    zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h_] for h_ in self.hiers])),
                                )
                            )
                            self.refresh_rows.add(ciid)
                        self.sheet.MT.data[rn][h] = self.nodes[pk].name
                    self.nodes[pk].cn[h].remove(iid)
                    self.nodes[pk].cn[h] = self.sort_node_cn(self.nodes[pk].cn[h], h)
                elif pk == "":
                    for ciid in self.nodes[iid].cn[h]:
                        child = self.nodes[ciid]
                        child.ps[h] = ""
                        rn = self.rns[ciid]
                        if snapshot:
                            self.vs[-1]["rows"].append(
                                Del_stre(
                                    0,
                                    rn,
                                    zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h_] for h_ in self.hiers])),
                                )
                            )
                            self.refresh_rows.add(ciid)
                        self.sheet.MT.data[rn][h] = ""
        rn = self.rns[iid]
        if snapshot:
            self.vs[-1]["rows"].append(Del_stre(1, rn, self.sheet.MT.data[rn]))
        del self.nodes[iid]
        to_del.append(iid)
        if self.auto_sort_nodes_bool and piid and self.nodes[piid].ps[self.pc]:
            parent_parent_node = self.nodes[self.nodes[piid].ps[self.pc]]
            parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)
        return to_del

    def _del_id_orphan_core(self, name: str, parent: str, snapshot: bool = True) -> None:
        ik = name.lower()
        pk = parent.lower()
        self.refresh_rows = set()
        if pk:
            self.nodes[pk].cn[self.pc].remove(ik)
        if not self.auto_sort_nodes_bool:
            if pk == "":
                self.topnodes_order[self.pc].remove(ik)
            for ciid in self.nodes[ik].cn[self.pc]:
                self.topnodes_order[self.pc].append(ciid)
        for ciid in self.nodes[ik].cn[self.pc]:
            rn = self.rns[ciid]
            child = self.nodes[ciid]
            if snapshot:
                self.vs[-1]["rows"].append(
                    Del_stre(
                        0,
                        rn,
                        zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h] for h in self.hiers])),
                    )
                )
                self.refresh_rows.add(ciid)
            child.ps[self.pc] = ""
            self.sheet.MT.data[rn][self.pc] = ""
        rn = self.rns[ik]
        if sum(1 for v in self.nodes[ik].ps.values() if v is not None) < 2:
            if snapshot:
                self.vs[-1]["rows"].append(Del_stre(1, rn, self.sheet.MT.data[rn]))
            del self.nodes[ik]
            self.sheet.delete_row(rn, redraw=False)
            self.untag_id(ik)
        else:
            if snapshot:
                self.vs[-1]["rows"].append(
                    Del_stre(
                        0,
                        rn,
                        zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h] for h in self.hiers])),
                    )
                )
                self.refresh_rows.add(ik)
            self.nodes[ik].cn[self.pc] = []
            self.nodes[ik].ps[self.pc] = None
            self.sheet.MT.data[rn][self.pc] = ""
        if self.auto_sort_nodes_bool and pk and self.nodes[pk].ps[self.pc]:
            parent_parent_node = self.nodes[self.nodes[pk].ps[self.pc]]
            parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)

    def _del_id_all_orphan_core(self, name: str, snapshot: bool = True) -> None:
        ik = name.lower()
        pk = self.nodes[ik].ps[self.pc]
        self.refresh_rows = set()
        self.untag_id(ik)
        if not self.auto_sort_nodes_bool:
            for h, p in self.nodes[ik].ps.items():
                if p == "":
                    self.topnodes_order[h].remove(ik)
                for ciid in self.nodes[ik].cn[h]:
                    self.topnodes_order[h].append(ciid)
        for h, p in self.nodes[ik].ps.items():
            if p:
                self.nodes[p].cn[h].remove(ik)
        for h, cn in self.nodes[ik].cn.items():
            for ciid in cn:
                child = self.nodes[ciid]
                child.ps[h] = ""
                rn = self.rns[ciid]
                if snapshot:
                    self.vs[-1]["rows"].append(
                        Del_stre(
                            0,
                            rn,
                            zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h] for h in self.hiers])),
                        )
                    )
                    self.refresh_rows.add(ciid)
                self.sheet.MT.data[rn][h] = ""
        rn = self.rns[ik]
        if snapshot:
            self.vs[-1]["rows"].append(Del_stre(1, rn, self.sheet.MT.data[rn]))
        del self.nodes[ik]
        self.sheet.delete_row(rn, redraw=False)
        if self.auto_sort_nodes_bool and pk and self.nodes[pk].ps[self.pc]:
            parent_parent_node = self.nodes[self.nodes[pk].ps[self.pc]]
            parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)

    def get_lvls(self, iid: str, lvl=1):
        # Initialize stack with the initial node at lvl - 1
        stack = [(iid, lvl - 1)]

        while stack:
            # Pop the current node and its level
            current_iid, current_lvl = stack.pop()

            # Get the children of the current node
            children = self.nodes[current_iid].cn[self.pc]

            # The children's level is the next level
            next_lvl = current_lvl + 1

            # Ensure the level exists in self.levels
            if next_lvl not in self.levels:
                self.levels[next_lvl] = []

            # Process each child
            for child in children:
                self.levels[next_lvl].append(child)
                stack.append((child, next_lvl))

    def _del_id_children_core(self, name: str, parent: str, snapshot: bool = True) -> None:
        if snapshot:
            qvsapp = self.vs[-1]["rows"].append
        ik = name.lower()
        to_del = []
        self.refresh_rows = set()
        self.levels = defaultdict(list)
        self.get_lvls(ik)
        for lvl in sorted(((k, v) for k, v in self.levels.items()), key=itemgetter(0), reverse=True):
            for ik_ in lvl[1]:
                rn = self.rns[ik_]
                if sum(1 for v in self.nodes[ik_].ps.values() if v is not None) < 2:
                    if snapshot:
                        qvsapp(Del_stre(1, rn, self.sheet.MT.data[rn]))
                    del self.nodes[ik_]
                    to_del.append(rn)
                    self.untag_id(ik)
                else:
                    if snapshot:
                        qvsapp(
                            Del_stre(
                                0,
                                rn,
                                zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h] for h in self.hiers])),
                            )
                        )
                        self.refresh_rows.add(ik_)
                    self.nodes[ik_].cn[self.pc] = []
                    self.nodes[ik_].ps[self.pc] = None
                    self.sheet.MT.data[rn][self.pc] = ""
        pk = parent.lower()
        rn = self.rns[ik]
        if pk:
            self.nodes[pk].cn[self.pc].remove(ik)
        if sum(1 for v in self.nodes[ik].ps.values() if v is not None) < 2:
            if snapshot:
                qvsapp(Del_stre(1, rn, self.sheet.MT.data[rn]))
            del self.nodes[ik]
            to_del.append(rn)
            self.untag_id(ik)
        else:
            if snapshot:
                qvsapp(
                    Del_stre(
                        0,
                        rn,
                        zlib.compress(pickle.dumps([self.sheet.MT.data[rn][h] for h in self.hiers])),
                    )
                )
                self.refresh_rows.add(ik)
            self.nodes[ik].cn[self.pc] = []
            self.nodes[ik].ps[self.pc] = None
            self.sheet.MT.data[rn][self.pc] = ""
        if to_del:
            self.sheet.del_rows(to_del)
        self.levels = defaultdict(list)
        if self.auto_sort_nodes_bool:
            if pk and self.nodes[pk].ps[self.pc]:
                parent_parent_node = self.nodes[self.nodes[pk].ps[self.pc]]
                parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)
        elif not self.auto_sort_nodes_bool and pk == "":
            try_remove(self.topnodes_order[self.pc], ik)

    def _del_id_children_all_core(self, name: str, parent: str, snapshot: bool = True) -> None:
        if snapshot:
            qvsapp = self.vs[-1]["rows"].append
        ik = name.lower()
        to_del = []
        self.levels = defaultdict(list)
        self.get_lvls(ik)
        for lvl in sorted(((k, v) for k, v in self.levels.items()), key=itemgetter(0), reverse=True):
            for ik_ in lvl[1]:
                rn = self.rns[ik_]
                if snapshot:
                    qvsapp(Del_stre(1, rn, self.sheet.MT.data[rn]))
                to_del.append(rn)
                self.untag_id(ik)
                del self.nodes[ik_]
        pk = parent.lower()
        rn = self.rns[ik]
        if snapshot:
            qvsapp(Del_stre(1, rn, self.sheet.MT.data[rn]))
        to_del.append(rn)
        self.untag_id(ik)
        if not self.auto_sort_nodes_bool:
            for h, p in self.nodes[ik].ps.items():
                if p == "":
                    self.topnodes_order[h].remove(ik)
                elif p:
                    self.nodes[p].cn[h].remove(ik)
        else:
            for h, p in self.nodes[ik].ps.items():
                if p:
                    self.nodes[p].cn[h].remove(ik)
                    self.nodes[p].cn[h] = self.sort_node_cn(self.nodes[p].cn[h], h)
        del self.nodes[ik]
        self.sheet.del_rows(to_del, redraw=False)
        self.levels = defaultdict(list)
        with suppress(Exception):
            if self.auto_sort_nodes_bool:
                if pk and self.nodes[pk].ps[self.pc]:
                    parent_parent_node = self.nodes[self.nodes[pk].ps[self.pc]]
                    parent_parent_node.cn[self.pc] = self.sort_node_cn(parent_parent_node.cn[self.pc], self.pc)
            elif pk == "":
                self.topnodes_order[self.pc].remove(ik)

    def details(self, ik):
        allrows = []
        spaces = " " * 10
        string = f"\n ID:   {self.nodes[ik].name}"
        allrows.append(string)
        allrows.append("\n\n Parents across all hierarchies:")
        for h, p in self.nodes[ik].ps.items():
            allrows.append(f"    Column #{h + 1} {self.headers[h].name}: ")
            if p == "":
                allrows.append(spaces + "Appears as top ID")
            elif p is not None:
                allrows.append(spaces + self.nodes[p].name)
        allrows.append("\n\n Children across all hierarchies:")
        for h in self.nodes[ik].cn:
            allrows.append(f"    Column #{h + 1} {self.headers[h].name}: ")
            for ciid in self.nodes[ik].cn[h]:
                allrows.append(spaces + self.nodes[ciid].name)
        if len(self.hiers) + 1 == self.row_len:
            allrows.append(spaces + "\n\n No detail columns in sheet")
        else:
            idcol_hiers = set(self.hiers) | {self.ic}
            allrows.append("\n\n Details:")
            for index, cell in enumerate(self.sheet.MT.data[self.rns[ik]]):
                if index not in idcol_hiers:
                    allrows.append(f"    Column #{index + 1} {self.headers[index].name}:")
                    allrows.append(spaces + cell)
        return "\n".join(allrows)

    def fix_associate_sort(self, startup=True):
        first_hier = self.hiers[0]
        quick_hiers = self.hiers[1:]
        lh = len(self.hiers)
        if startup and self.auto_sort_nodes_bool:
            for node in self.nodes.values():
                if all(p is None for p in node.ps.values()):
                    node.ps = {h: "" if node.cn[h] else None for h in self.hiers}
                    newrow = list(repeat("", self.row_len))
                    newrow[self.ic] = node.name
                    self.sheet.MT.data.append(newrow)
                    self.warnings.append(f" - ID ({node.name}) missing from ID column, new row added")
                tlly = 0
                for k, v in node.cn.items():
                    if v:
                        node.cn[k] = self.sort_node_cn(v, k)
                    elif not node.ps[k]:
                        node.ps[k] = None
                        tlly += 1
                if tlly == lh:
                    node.ps[first_hier] = ""
                    for h in quick_hiers:
                        node.ps[h] = None

        elif not startup and self.auto_sort_nodes_bool:
            to_insert = []
            for node in self.nodes.values():
                if all(p is None for p in node.ps.values()):
                    node.ps = {h: "" if node.cn[h] else None for h in self.hiers}
                    newrow = list(repeat("", self.row_len))
                    newrow[self.ic] = node.name
                    to_insert.append(newrow)
                tlly = 0
                for k, v in node.cn.items():
                    if v:
                        node.cn[k] = self.sort_node_cn(v, k)
                    elif not node.ps[k]:
                        node.ps[k] = None
                        tlly += 1
                if tlly == lh:
                    node.ps[first_hier] = ""
                    for h in quick_hiers:
                        node.ps[h] = None
            if to_insert:
                self.sheet.insert_rows(rows=to_insert)

        elif not startup and not self.auto_sort_nodes_bool:
            st_check_topnodes_order = {k: set(v) for k, v in self.topnodes_order.items()}
            to_insert = []
            for iid, node in self.nodes.items():
                if all(p is None for p in node.ps.values()):
                    node.ps = {h: "" if node.cn[h] else None for h in self.hiers}
                    newrow = list(repeat("", self.row_len))
                    newrow[self.ic] = node.name
                    to_insert.append(newrow)
                tlly = 0
                for k, v in node.cn.items():
                    if not v and not node.ps[k]:
                        node.ps[k] = None
                        tlly += 1
                if tlly == lh:
                    if all(iid not in h for h in st_check_topnodes_order.values()):
                        node.ps[first_hier] = ""
                        for h in quick_hiers:
                            node.ps[h] = None
                        self.topnodes_order[first_hier].append(iid)
                    else:
                        for h, v in st_check_topnodes_order.items():
                            if iid in v:
                                node.ps[h] = ""
            if to_insert:
                self.sheet.insert_rows(rows=to_insert)

    def fix_associate_sort_edit_cells(self):
        first_hier = self.hiers[0]
        quick_hiers = self.hiers[1:]
        lh = len(self.hiers)
        to_insert = []
        if self.auto_sort_nodes_bool:
            for n in self.nodes.values():
                if all(p is None for p in n.ps.values()):
                    n.ps = {h: "" if n.cn[h] else None for h in self.hiers}
                    newrow = list(repeat("", self.row_len))
                    newrow[self.ic] = n.name
                    to_insert.append(newrow)
                tlly = 0
                for k, v in n.cn.items():
                    if v:
                        n.cn[k] = self.sort_node_cn(v, k)
                    elif not n.ps[k]:
                        n.ps[k] = None
                        tlly += 1
                if tlly == lh:
                    n.ps[first_hier] = ""
                    for h in quick_hiers:
                        n.ps[h] = None
        else:
            for n in self.nodes.values():
                if all(p is None for p in n.ps.values()):
                    n.ps = {h: "" if n.cn[h] else None for h in self.hiers}
                    newrow = list(repeat("", self.row_len))
                    newrow[self.ic] = n.name
                    to_insert.append(newrow)
                tlly = 0
                for k, v in n.cn.items():
                    if not v and not n.ps[k]:
                        n.ps[k] = None
                        tlly += 1
                if tlly == lh:
                    n.ps[first_hier] = ""
                    for h in quick_hiers:
                        n.ps[h] = None
        if to_insert:
            self.sheet.insert_rows(rows=to_insert)
        return "break"

    def associate(self):
        first_hier = self.hiers[0]
        quick_hiers = self.hiers[1:]
        lh = len(self.hiers)
        for node in self.nodes.values():
            tlly = 0
            for k, v in node.cn.items():
                if not (v or node.ps[k]):
                    node.ps[k] = None
                    tlly += 1
            if tlly == lh:
                node.ps[first_hier] = ""
                for h in quick_hiers:
                    node.ps[h] = None
        if not self.auto_sort_nodes_bool:
            current_nodes = {n: None for n in self.topnodes_order[self.hiers[0]]}
            wc = []
            woc = []
            for iid, node in self.nodes.items():
                if iid not in current_nodes and node.ps[self.hiers[0]] == "":
                    if node.cn[self.hiers[0]]:
                        wc.append(iid)
                    else:
                        woc.append(iid)
            self.topnodes_order[self.hiers[0]] = (
                list(current_nodes) + sorted(wc, key=sort_key) + sorted(woc, key=sort_key)
            )

    def sort_node_cn(self, cn: list[str], h: int):
        wc = []
        woc = []
        for ciid in cn:
            if self.nodes[ciid].cn[h]:
                wc.append(ciid)
            else:
                woc.append(ciid)
        return sorted(wc, key=sort_key) + sorted(woc, key=sort_key)

    def top_iids(self):
        pc = self.pc
        if self.auto_sort_nodes_bool:
            wc = []
            woc = []
            for iid, node in self.nodes.items():
                if node.ps[pc] == "":
                    if node.cn[pc]:
                        wc.append(iid)
                    else:
                        woc.append(iid)
            yield from sorted(wc, key=sort_key)
            yield from sorted(woc, key=sort_key)
        else:
            yield from self.topnodes_order[pc]

    def pc_iids(self) -> Generator[str]:
        top_iter = iter(self.top_iids())
        stack = []
        while True:
            if not stack:
                try:
                    top_iid = next(top_iter)
                    yield top_iid
                    stack.extend(reversed(self.nodes[top_iid].cn[self.pc]))
                except StopIteration:  # If no more top nodes, exit
                    break
            else:
                iid = stack.pop()
                yield iid
                stack.extend(reversed(self.nodes[iid].cn[self.pc]))

    def remake_topnodes_order(self):
        self.topnodes_order = {}
        for h in self.hiers:
            wc = []
            woc = []
            for iid, node in self.nodes.items():
                if node.ps[h] == "":
                    if node.cn[h]:
                        wc.append(iid)
                    else:
                        woc.append(iid)
            self.topnodes_order[h] = sorted(wc, key=sort_key) + sorted(woc, key=sort_key)

    def gen_sheet_w_headers(self):
        yield (h.name for h in self.headers)
        yield from ((e if e else None for e in r) for r in self.sheet.MT.data)

    def check_validation_validity(self, col: int, validation: list[str]) -> str | list[str]:
        if not validation:
            return validation
        if self.headers[col].type_ == "Number":
            for e in validation:
                if e != "" and not isreal(e):
                    return f"Error: Only numbers are allowed in Number columns. Error caused by: {e}"
                for c in e:
                    if c not in validation_allowed_num_chars:
                        return f"Error: Invalid character in validation for Number column. Error caused by: {c}"
        elif self.headers[col].type_ == "Date":
            for e in validation:
                for c in e:
                    if c not in validation_allowed_date_chars:
                        return f"Error: Invalid character in validation for Date columns. Error caused by: {c}"
            for i in range(len(validation)):
                e = validation[i]
                if not isint(e):
                    x = self.detect_date_form(e)
                    if x and len(x) == 1 and x[0] != self.DATE_FORM:
                        e = datetime.datetime.strftime(datetime.datetime.strptime(e, x[0]), self.DATE_FORM)
                    elif not x:
                        return f"Error: Only dates are allowed in Date columns. Error caused by: {e}"
                validation[i] = e
        elif self.headers[col].type_ == "Text":
            pass
        else:
            return "Error: Only Detail columns can have validation"
        return validation if "" in validation else [""] + validation

    def apply_validation_to_col(self, col):
        validset = set(self.headers[col].validation)
        for rn in range(len(self.sheet.MT.data)):
            if not self.is_in_validation(validset, self.sheet.MT.data[rn][col]):
                self.sheet.MT.data[rn][col] = ""

    def check_condition_validity(self, col, condition, input_headers=None):
        if not condition:
            return ""
        if input_headers is None:
            input_headers = []
        heads = input_headers if input_headers else self.headers
        if heads[col].type_ in ("Number", "Date"):
            all_allowed_chars = {
                "a",
                "n",
                "d",
                "o",
                "r",
                "A",
                "N",
                "D",
                "O",
                "R",
                "0",
                "1",
                "2",
                "3",
                "4",
                "5",
                "6",
                "7",
                "8",
                "9",
                "!",
                ">",
                "<",
                "=",
                " ",
                "c",
                "C",
                "/",
                "-",
                ".",
            }
            condition = "".join(c.lower() for c in condition.replace("  ", " ") if c in all_allowed_chars)
            if not condition:
                return "Error:"
            if not condition.startswith(" "):
                condition = " " + condition
            if len(condition) < 3:
                return "Error: Condition too short"

            nums = {"0", "1", "2", "3", "4", "5", "6", "7", "8", "9"}
            syms = {"!", ">", "<", "=", "-", "."}

            # detect single equals character
            for last_char, char, next_char in zip(
                islice(condition, 0, len(condition)),
                islice(condition, 1, len(condition)),
                islice(condition, 2, len(condition)),
            ):
                if char == "=" and last_char not in syms and next_char != "=":
                    return "Error: Incorrect use of ="
                if char == "=" and last_char in syms and next_char != " ":
                    return "Error: Missing a space after ="

            d = defaultdict(int)
            for char in condition:
                d[char] += 1
            if d["a"] > 1:
                return "Error: Too many a characters"
            if d["n"] > 1:
                return "Error: Too many n characters"
            if d["o"] > 1:
                return "Error: Too many o characters"
            if d["r"] > 1:
                return "Error: Too many r characters"
            if d["c"] > 2:
                return "Error: Too many c characters"
            if d["d"] > 2:
                return "Error: Too many d characters"
            if d["."] > 2:
                return "Error: Too many . characters"
            if d["-"] > 2:
                return "Error: Too many - characters"
            if d["!"] > 2:
                return "Error: Too many ! characters"
            if d[">"] > 2:
                return "Error: Too many > characters"
            if d["<"] > 2:
                return "Error: Too many < characters"
            if d["/"] > 4:
                return "Error: Too many / characters"
            if d["="] > 4:
                return "Error: Too many = characters"

            # number larger or less than 10 trillion
            for n in re.findall("([0-9]+)", condition):
                x = float(n)
                if x > 10000000000000 or x < -10000000000000:
                    return "Error: Condition contains number larger or less than 10 trillion"

            # more than one . character in a number
            for n in re.findall("([0-9.]+)", condition):
                if n.count(".") > 1:
                    return "Error: A number contained more than one . character"

        if heads[col].type_ == "Number":
            for last_char, char in zip(
                islice(condition, 0, len(condition)),
                islice(condition, 1, len(condition)),
            ):
                # a
                if last_char == "a" and char != "n":
                    return "Error: and spelt incorrectly"
                if char == "a" and last_char != " ":
                    return "Error: and must follow a space"

                # n
                if last_char == "n" and char != "d":
                    return "Error: and spelt incorrectly"
                if char == "n" and last_char != "a":
                    return "Error: and spelt incorrectly"

                # d
                if last_char == "d" and char != " ":
                    return "Error: A space must follow and"
                if char == "d" and last_char != "n":
                    return "Error: and spelt incorrectly"

                # o
                if last_char == "o" and char != "r":
                    return "Error: or spelt incorrectly"
                if char == "o" and last_char != " ":
                    return "Error: or must follow a space"

                # r
                if last_char == "r" and char != " ":
                    return "Error: A space must follow or"
                if char == "r" and last_char != "o":
                    return "Error: or spelt incorrectly"

                # num
                if last_char in nums and char != " " and char not in nums and char != ".":
                    return "Error: A space or another number must follow a number"
                if char in nums and last_char not in nums and last_char not in ("c", ".", "-", " "):
                    return "Error: A number can only follow another number, a c character or a space"

                # !
                if last_char == "!" and char != "=":
                    return "Error: = must follow !"
                if char == "!" and last_char != " ":
                    return "Error: ! must follow a space"

                # >
                if last_char == ">" and char != " " and char != "=":
                    return "Error: A space or = must follow >"
                if char == ">" and last_char != " ":
                    return "Error: > must follow a space"

                # <
                if last_char == "<" and char != " " and char != "=":
                    return "Error: A space or = must follow <"
                if char == "<" and last_char != " ":
                    return "Error: < must follow a space"

                # =
                if last_char == "=" and char != "=" and char != " ":
                    return "Error: = or a space must follow ="
                if char == "=" and last_char not in ("!", "<", ">", "=", " "):
                    return "Error: = can only follow ! < > =  Equal to: ==  Not Equal to: !=  Less than or equal to: <=  Greater than or equal to: >="

                # /
                if last_char == "/" or char == "/":
                    return "Error: / not allowed in Number conditions"

                # .
                if last_char == "." and char not in nums:
                    return "Error: A number must follow a . character"
                if char == "." and last_char not in nums:
                    return "Error: A number must be before a . character"

                # -
                if last_char == "-" and char not in nums:
                    return "Error: A number must follow a - character"
                if char == "-" and last_char != " ":
                    return "Error: A space must be before - characters"

                # c
                if char == "c":
                    return "Error: c character not allowed in Number conditions"

            if condition[-1] not in nums:
                return "Error: Condition can only end in a number"

        elif heads[col].type_ == "Date":
            for last_char, char in zip(
                islice(condition, 0, len(condition)),
                islice(condition, 1, len(condition)),
            ):
                # a
                if last_char == "a" and char != "n":
                    return "Error: and spelt incorrectly"
                if char == "a" and last_char != " ":
                    return "Error: and must follow a space"

                # n
                if last_char == "n" and char != "d":
                    return "Error: and spelt incorrectly"
                if char == "n" and last_char != "a":
                    return "Error: and spelt incorrectly"

                # d
                if last_char == "d" and char != " ":
                    return "Error: A space must follow and"
                # different to Number check, added c
                if char == "d" and last_char != "n" and last_char != "c":
                    return "Error: and or current date (cd) spelt incorrectly"

                # o
                if last_char == "o" and char != "r":
                    return "Error: or spelt incorrectly"
                if char == "o" and last_char != " ":
                    return "Error: or must follow a space"

                # r
                if last_char == "r" and char != " ":
                    return "Error: A space must follow or"
                if char == "r" and last_char != "o":
                    return "Error: or spelt incorrectly"

                # num
                if last_char in nums and char != "/" and char != " " and char not in nums:
                    return "Error: A space or / must follow a number"
                if char in nums and last_char not in nums and last_char not in (" ", "/", "c"):
                    return "Error: A number can only follow another number, a / or a space"

                # !
                if last_char == "!" and char != "=":
                    return "Error: = must follow !"
                if char == "!" and last_char != " ":
                    return "Error: ! must follow a space"

                # >
                if last_char == ">" and char != " " and char != "=":
                    return "Error: A space or = must follow >"
                if char == ">" and last_char != " ":
                    return "Error: > must follow a space"

                # <
                if last_char == "<" and char != " " and char != "=":
                    return "Error: A space or = must follow <"
                if char == "<" and last_char != " ":
                    return "Error: < must follow a space"

                # =
                if last_char == "=" and char != "=" and char != " ":
                    return "Error: = or a space must follow ="
                if char == "=" and last_char not in ("!", "<", ">", "=", " "):
                    return "Error: = can only follow ! < > =  Equal to: ==  Not Equal to: !=  Less than or equal to: <=  Greater than or equal to: >="

                # /
                if last_char == "/" and char not in nums:
                    return "Error: A number must follow a / character"
                if char == "/" and last_char not in nums:
                    return "Error: A number must be before a / character"

                # .
                if char == ".":
                    return "Error: . characters not allowed in Date conditions"

                # -
                if char == "-":
                    return "Error: - characters not allowed in Date conditions"

                # c
                if last_char == "c" and char != "d":
                    return "Error: Current date (cd) spelt incorrectly"
                if char == "c" and last_char != " ":
                    return "Error: A space must be before current date (cd)"

            if condition[-1] == "d" and condition[-2] != "c":
                return "Error: Current date (cd) spelt incorrectly"

            if condition[-1] not in nums and condition[-1] != "d":
                return "Error: Condition can only end in a number or current date (cd)"
        # elif heads[col].type_ in ("ID", "Parent", "Text"):
        # pass
        return condition

    def format_str_number(self, s):
        if not s:
            return s
        elif isint(s):
            return int(s)
        elif isintlike(s):
            return int(float(s))
        elif isfloat(s):
            return float(s)
        return s

    def format_str_date(self, s):
        if not s:
            return s
        elif isint(s):
            return datetime.timedelta(days=int(s))
        else:
            try:
                return datetime.datetime.strptime(s, self.DATE_FORM)
            except Exception:
                pass
        return s

    def format_date_str(self, d):
        if isinstance(d, datetime.timedelta):
            return f"{d.days}"
        elif isinstance(d, datetime.datetime):
            return d.strftime(self.DATE_FORM)
        return f"{d}"

    def refresh_formatting(
        self,
        rows: int | Iterator | None = None,
        columns: int | Iterator | None = None,
        dehighlight: bool = False,
    ):
        if dehighlight:
            self.sheet.dehighlight_cells(all_=True, redraw=False)

        if rows is None:
            rows = range(len(self.sheet.MT.data))
        elif isinstance(rows, int):
            rows = (rows,)

        if columns is None:
            columns = tuple(range(len(self.headers)))
        elif isinstance(columns, int):
            columns = (columns,)

        if not rows:
            return

        # used within eval if a date column condition contains "cd"
        try:
            cd = datetime.datetime.strptime(
                datetime.datetime.today().strftime(self.DATE_FORM),
                self.DATE_FORM,
            )  # noqa: F841
        except Exception:
            cd = datetime.timedelta(days=0)  # noqa: F841

        all_conditions = {}
        number_cols = set()
        date_cols = set()
        for col, hdr in enumerate(self.headers):
            if hdr.type_ == "Number":
                number_cols.add(col)
            elif hdr.type_ == "Date":
                date_cols.add(col)

        for col in columns:
            if not self.headers[col].formatting:
                continue
            modified_conditions = []
            all_conditions[col] = {}

            if self.headers[col].type_ in ("ID", "Parent", "Text"):
                all_conditions[col] = self.headers[col].formatting

            elif self.headers[col].type_ == "Number":
                for condition in self.headers[col].formatting:
                    cond, color = condition
                    if cond:
                        cond = "cell " + cond.replace("and", "and cell").replace("or", "or cell")
                        modified_conditions.append((cond, color))
                    else:
                        modified_conditions.append(("not cell", color))
                all_conditions[col] = modified_conditions

            elif self.headers[col].type_ == "Date":
                for condition in self.headers[col].formatting:
                    cond, color = condition
                    if cond:
                        cond = cond.replace("and", "and cell").replace("or", "or cell")
                        cond = "cell " + "".join(
                            [
                                (
                                    f"datetime.timedelta(days=int({e}))"
                                    if isreal(e)
                                    else (
                                        f"datetime.datetime.strptime('{e}',self.convert_hyphen_to_slash_date_form(self.DATE_FORM))"
                                        if "/" in e
                                        else e
                                    )
                                )
                                for e in re.split("([0-9/]+)", cond)
                            ]
                        )
                        modified_conditions.append((cond, color))
                    else:
                        modified_conditions.append(("not cell", color))
                all_conditions[col] = modified_conditions

        quick_data = self.sheet.MT.data
        for col in filter(all_conditions.__contains__, columns):
            for rn in rows:
                self.sheet.dehighlight_cells(row=rn, column=col, redraw=False)
                cell = quick_data[rn][col]

                # convert cell to number/date
                if cell:
                    if col in number_cols:
                        cell = self.format_str_number(cell)
                    elif col in date_cols:
                        cell = self.format_str_date(cell)

                # apply highlights
                if col in number_cols or col in date_cols:
                    for cond, color in all_conditions[col]:
                        try:
                            if eval(cond):
                                self.sheet.highlight_cells(
                                    row=rn,
                                    column=col,
                                    bg=color,
                                    fg="black",
                                )
                                break
                        except Exception:
                            continue

                else:
                    for cond, color in all_conditions[col]:
                        if cell.lower() == cond.lower():
                            self.sheet.highlight_cells(row=rn, column=col, bg=color, fg="black")
                            break

                # convert cell back to string
                if cell != "":
                    if col in number_cols:
                        quick_data[rn][col] = f"{cell}"
                    elif col in date_cols:
                        quick_data[rn][col] = self.format_date_str(cell)

        self.refresh_rows = set()

    def rc_edit_validation(self, event=None):
        if (col := self.rc_selected_col()) is None:
            return
        popup = Edit_Validation_Popup(
            self,
            self.headers[col].type_,
            self.headers[col].name,
            self.headers[col].validation,
            self.C.theme,
        )
        if popup.new_validation is None:
            return
        if popup.new_validation:
            validation = self.check_validation_validity(col, popup.new_validation)
            if isinstance(validation, str):
                Error(
                    self,
                    f" {validation}     see 'Help' under the 'File' menu for instructions on validation   ",
                    theme=self.C.C.theme,
                )
                return
        else:
            validation = []
        if validation == self.headers[col].validation:
            return
        self.snapshot_edit_validation(col, validation)
        self.headers[col].validation = validation
        if validation:
            self.apply_validation_to_col(col)
        self.refresh_dropdowns()
        self.refresh_formatting(columns=col)
        self.redo_tree_display()
        self.redraw_sheets()

    def rc_edit_formatting(self, event=None):
        if (col := self.rc_selected_col(allow_hiers=True)) is None:
            return
        self.save_info_get_saved_info()
        Edit_Conditional_Formatting_Popup(self, column=col, theme=self.C.theme)
        self.headers[col].formatting = [tup for tup in self.headers[col].formatting if tup[1]]
        self.refresh_formatting(columns=col)
        self.redo_tree_display()
        self.redraw_sheets()

    def rc_selected_col(self, allow_hiers=False):
        widget = self.sheet if self.sheet.has_focus() else self.tree
        col = widget.get_selected_columns()
        if len(col) != 1:
            return
        col = widget.selected.column
        if not allow_hiers and (col == self.ic or col in self.hiers):
            return
        return col

    def rc_change_coltype_text(self, event=None):
        if (col := self.rc_selected_col()) is None:
            return
        self.snapshot_col_type_text(col)
        self.change_coltype_text(col)
        self.set_headers()

    def change_coltype_text(self, col):
        self.headers[col].type_ = "Text"
        self.headers[col].formatting = []

    def rc_change_coltype_number(self, event=None):
        if (col := self.rc_selected_col()) is None:
            return
        self.snapshot_col_type_num_date(col, "Number")
        self.headers[col].type_ = "Number"
        self.change_coltype_number(col)
        if isinstance(self.check_validation_validity(col, self.headers[col].validation), str):
            self.headers[col].validation = []
            self.refresh_dropdowns()
        self.headers[col].formatting = [
            tup
            for tup in self.headers[col].formatting
            if not self.check_condition_validity(col, tup[0]).startswith("Error:")
        ]
        self.set_headers()
        self.refresh_formatting(columns=col)
        self.redo_tree_display()
        self.redraw_sheets()

    def change_coltype_number(self, col, warnings=False):
        if warnings:
            for rn in range(len(self.sheet.MT.data)):
                cell = self.sheet.MT.data[rn][col]
                if cell and not isreal(cell):
                    self.warnings.append(
                        f" - Deleted cell row #{rn} column #{col} because {cell} was not valid for Number column"
                    )
                    self.sheet.MT.data[rn][col] = ""
                    self.refresh_tree_item(self.sheet.data[rn][self.ic])
        else:
            for rn in range(len(self.sheet.MT.data)):
                cell = self.sheet.MT.data[rn][col]
                if cell and not isreal(cell):
                    self.sheet.MT.data[rn][col] = ""
                    self.refresh_tree_item(self.sheet.data[rn][self.ic])

    def convert_num(self, num):
        if isint(num):
            return f"{int(num)}"
        return f"{float(num)}"

    def detect_date_form(self, date):
        forms = []
        for form in date_formats_usable:
            try:
                datetime.datetime.strptime(date, form).date()
                forms.append(form)
            except Exception:
                continue
        return forms

    def convert_date(self, date, new_form):
        if isint(date):
            return f"{int(date)}"
        for form in date_formats_usable:
            try:
                return datetime.datetime.strftime(datetime.datetime.strptime(date, form), new_form)
            except Exception:
                continue
        return date

    def convert_hyphen_to_slash_date_form(self, form):
        if form.startswith("%d"):
            return "%d/%m/%Y"
        elif form.startswith("%m"):
            return "%m/%d/%Y"
        return "%Y/%m/%d"

    def is_in_validation(self, validation, text):
        return text in validation

    def detail_is_valid_for_col(self, col, detail):
        t = self.headers[col].type_
        if self.headers[col].validation and not self.is_in_validation(self.headers[col].validation, detail):
            return False
        if t == "Text":
            return True
        elif t == "Number":
            if detail == "":
                return True
            return isreal(detail)
        elif t == "Date":
            if isint(detail):
                return True
            return bool(self.detect_date_form(detail))
        return False

    def why_isnt_detail_valid(self, col, detail):
        t = self.headers[col].type_
        if self.headers[col].validation and not self.is_in_validation(self.headers[col].validation, detail):
            return "Entered detail is not in column validation"
        if t == "Number":
            return "Entered detail is not a valid number"
        elif t == "Date":
            return "Entered detail is not a valid date or integer"

    def rc_change_coltype_date(self, event=None):
        if (col := self.rc_selected_col()) is None:
            return
        self.snapshot_col_type_num_date(col, "Date")
        self.headers[col].type_ = "Date"
        self.change_coltype_date(col, detect_date_form=True)
        if isinstance(self.check_validation_validity(col, self.headers[col].validation), str):
            self.headers[col].validation = []
            self.refresh_dropdowns()
        self.headers[col].formatting = [
            tup
            for tup in self.headers[col].formatting
            if not self.check_condition_validity(col, tup[0]).startswith("Error:")
        ]
        self.set_headers()
        self.refresh_formatting(columns=col)
        self.redo_tree_display()
        self.redraw_sheets()

    def change_coltype_date(self, col, detect_date_form=False, warnings=False):
        if detect_date_form:
            sheet_date_form = {
                form for row in self.sheet.MT.data if len(row[col]) == 10 for form in self.detect_date_form(row[col])
            }
            if len(sheet_date_form) == 1:
                sheet_date_form = tuple(sheet_date_form)[0]
                quick_data = self.sheet.MT.data
                for rn in range(len(quick_data)):
                    if quick_data[rn][col] and not isint(quick_data[rn][col]):
                        try:
                            quick_data[rn][col][rn][col] = datetime.datetime.strftime(
                                datetime.datetime.strptime(quick_data[rn][col], sheet_date_form),
                                self.DATE_FORM,
                            )
                        except Exception:
                            if warnings:
                                self.warnings.append(
                                    f" - Deleted cell row #{rn} column #{col} because {quick_data[rn][col]} was not valid for Date columns"
                                )
                            quick_data[rn][col][rn][col] = ""
                        self.refresh_tree_item(self.sheet.data[rn][self.ic])
            else:
                self.change_coltype_date_validate(col)
        else:
            self.change_coltype_date_validate(col)

    def change_coltype_date_validate(self, col):
        quick_data = self.sheet.MT.data
        for rn in range(len(quick_data)):
            if (
                quick_data[rn][col]
                and not isint(quick_data[rn][col])
                and not self.detect_date_form(quick_data[rn][col])
            ):
                quick_data[rn][col] = ""
                self.refresh_tree_item(quick_data[rn][self.ic])

    def increment_unsaved(self):
        self.C.unsaved_changes = True
        self.C.change_app_title(star="add")

    def get_datetime_changelog(self, increment_unsaved=True):
        if increment_unsaved:
            self.increment_unsaved()
        self.sheet_changes += 1
        return f"{datetime.datetime.today().strftime('%Y/%m/%d')}"

    def rc_rename_col(self, event=None):
        if (col := self.rc_selected_col(allow_hiers=True)) is None:
            return
        if col in self.hiers:
            popup = Rename_Column_Popup(self, self.headers[col].name, "hierarchy", theme=self.C.theme)
        elif col == self.ic:
            popup = Rename_Column_Popup(self, self.headers[col].name, "ID", theme=self.C.theme)
        else:
            popup = Rename_Column_Popup(self, self.headers[col].name, "detail", theme=self.C.theme)
        if not popup.result:
            return
        new_name = popup.result
        new_name_k = new_name.lower()
        if any(new_name_k == h.name.lower() for h in self.headers):
            Error(self, f"Name: {new_name} already exists", theme=self.C.theme)
            return
        self.rename_col(col, new_name)
        self.disable_paste()
        self.refresh_hier_dropdown(self.hiers.index(self.pc))
        self.set_headers()

    def rename_col(self, col, name, snapshot=True):
        if snapshot:
            self.snapshot_rename_col()
            self.changelog_append(
                "Column rename",
                f"Column #{col + 1} with type: {self.headers[col].type_}",
                f"{self.headers[col].name}",
                f"{name}",
            )
        self.headers[col].name = name
        if snapshot:
            self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def add_hier_col(self, col, name, snapshot=True):
        if snapshot:
            self.snapshot_add_col(col)
        self.ic = push_n(self.ic, [col])
        self.pc = push_n(self.pc, [col])
        self.tv_label_col = push_n(self.tv_label_col, [col])
        self.row_len += 1
        self.adjust_hiers_add_cols(cols=[col])
        self.hiers = sorted([col] + self.hiers)
        self.headers.insert(col, Header(name, "Parent"))
        self.tree.insert_columns(idx=col, add_row_heights=False)
        self.sheet.insert_columns(idx=col, add_row_heights=False)
        for node in self.nodes.values():
            node.ps[col] = None
            node.cn[col] = []
        self.saved_info[col] = new_info_storage()
        if not self.auto_sort_nodes_bool:
            self.topnodes_order[col] = []
        if snapshot:
            self.changelog_append(
                "Add new hierarchy column",
                f"Column #{col + 1} with name: {name}",
                "",
                "",
            )
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def rc_add_hier_col(self, event=None):
        if (col := self.rc_selected_col(allow_hiers=True)) is None:
            col = len(self.headers)
        popup = Add_Hierarchy_Column_Popup(self, theme=self.C.theme)
        if not popup.result:
            return
        name = popup.result
        namekey = name.lower()
        if any(namekey == h.name.lower() for h in self.headers):
            Error(self, f"Column {name} already exists.", theme=self.C.theme)
            return
        self.add_hier_col(col, name)
        self.disable_paste()
        self.refresh_hier_dropdown(self.hiers.index(self.pc))
        self.set_headers()

    def rc_add_col(self, event=None):
        if (col := self.rc_selected_col(allow_hiers=True)) is None:
            col = len(self.headers)
        popup = Add_Detail_Column_Popup(self, theme=self.C.theme)
        if not popup.result:
            return
        name = popup.result
        namekey = name.lower()
        type_ = popup.type_
        if any(namekey == h.name.lower() for h in self.headers):
            Error(self, f"Column {name} already exists.", theme=self.C.theme)
            return
        self.add_col(col, name, type_)
        self.disable_paste()
        self.set_headers()

    def add_col(self, col, name, type_, snapshot=True):
        if snapshot:
            self.snapshot_add_col(col)
        self.ic = push_n(self.ic, [col])
        self.pc = push_n(self.pc, [col])
        self.tv_label_col = push_n(self.tv_label_col, [col])
        self.row_len += 1
        self.headers.insert(col, Header(name, type_))
        self.tree.insert_columns(idx=col, add_row_heights=False)
        self.sheet.insert_columns(idx=col, add_row_heights=False)
        self.adjust_hiers_add_cols(cols=[col])
        if snapshot:
            self.changelog_append(
                "Add new detail column",
                f"Column #{col} with name: {name} and type: {type_}",
                "",
                "",
            )
        if snapshot:
            self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def del_cols(self, cols, snapshot=True):
        if snapshot:
            self.snapshot_del_cols()
            cols_dict = self.vs[-1]["cols"]
            for datacn in reversed(cols):
                for rn in range(len(self.sheet.MT.data)):
                    if datacn not in cols_dict:
                        cols_dict[datacn] = {}
                    try:
                        cols_dict[datacn][rn] = self.sheet.MT.data[rn][datacn]
                    except Exception:
                        continue
        self.sheet.del_columns(cols)
        self.tree.del_columns(cols)
        self.ic = self.ic if not (num := bisect_left(cols, self.ic)) else self.ic - num
        self.pc = self.pc if not (num := bisect_left(cols, self.pc)) else self.pc - num
        if self.tv_label_col == self.ic or self.tv_label_col in cols:
            self.tv_label_col = self.ic
        else:
            self.tv_label_col = (
                self.tv_label_col if not (num := bisect_left(cols, self.tv_label_col)) else self.tv_label_col - num
            )
        if snapshot:
            colnames = ", ".join(self.headers[col].name for col in cols)
            self.changelog_append(
                "Delete columns",
                f"Columns: {colnames}",
                "",
                "",
            )
        cols_set = set(cols)
        self.headers = [hdr for i, hdr in enumerate(self.headers) if i not in cols_set]
        self.hiers_orig = self.hiers.copy()
        self.hiers = list(filterfalse(cols_set.__contains__, self.hiers))
        if hiers_to_del := list(filter(cols_set.__contains__, reversed(self.hiers_orig))):
            for col in hiers_to_del:
                for node in self.nodes.values():
                    del node.ps[col]
                    del node.cn[col]
                del self.saved_info[col]
                if not self.auto_sort_nodes_bool:
                    del self.topnodes_order[col]
            self.associate()
        self.row_len -= len(cols)
        self.adjust_hiers_del_cols(cols)
        if snapshot:
            self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def del_cols_rc(self, event=None):
        if self.tree.has_focus():
            cols = self.tree.get_selected_columns()
        elif self.sheet.has_focus():
            cols = self.sheet.get_selected_columns()
        focused = self.tree if self.tree.has_focus() else self.sheet
        if self.ic in cols or self.pc in cols:
            Error(
                self,
                "Cannot delete selected columns, they contain either the ID column or the current hierarchy   ",
                theme=self.C.theme,
            )
            return
        cols = sorted(cols)
        self.save_info_get_saved_info()
        self.del_cols(cols)
        self.clear_copied_details()
        self.refresh_hier_dropdown(self.hiers.index(self.pc))
        self.sheet.row_index(newindex=self.ic)
        self.refresh_dropdowns()
        self.redo_tree_display()
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        if focused == self.tree:
            self.focus_tree()
        else:
            self.focus_sheet()

    def cut_cols(self, cols):
        self.cut_columns = cols

    def adjust_hiers_del_cols(self, cols):
        auto_sort_nodes_bool = self.auto_sort_nodes_bool
        self.hiers = [k if not (num := bisect_left(cols, k)) else k - num for k in self.hiers]
        for node in self.nodes.values():
            node.ps = {k if not (num := bisect_left(cols, k)) else k - num: v for k, v in node.ps.items()}
            node.cn = {k if not (num := bisect_left(cols, k)) else k - num: v for k, v in node.cn.items()}
        self.saved_info = {k if not (num := bisect_left(cols, k)) else k - num: v for k, v in self.saved_info.items()}
        if not auto_sort_nodes_bool:
            self.topnodes_order = {
                k if not (num := bisect_left(cols, k)) else k - num: v for k, v in self.topnodes_order.items()
            }

    def adjust_hiers_add_cols(self, cols):
        auto_sort_nodes_bool = self.auto_sort_nodes_bool
        self.hiers = [push_n(k, cols) for k in self.hiers]
        for node in self.nodes.values():
            node.ps = {push_n(k, cols): v for k, v in node.ps.items()}
            node.cn = {push_n(k, cols): v for k, v in node.cn.items()}
        self.saved_info = {push_n(k, cols): v for k, v in self.saved_info.items()}
        if not auto_sort_nodes_bool:
            self.topnodes_order = {push_n(k, cols): v for k, v in self.topnodes_order.items()}

    def refresh_hier_dropdown(self, idx):
        switch_values = [f"{self.headers[h].name}" for h in self.hiers]
        self.switch_hier_dropdown["values"] = switch_values
        self.switch_displayed.set(switch_values[idx])

    def refresh_dropdowns(self):
        self.tree.del_dropdown("A:")
        self.sheet.del_dropdown("A:")
        for c, hdr in enumerate(self.headers):
            if hdr.validation:
                self.sheet.dropdown(
                    _n2a(c),
                    values=hdr.validation,
                    edit_data=False,
                )
                self.tree.dropdown(
                    _n2a(c),
                    values=hdr.validation,
                    edit_data=False,
                )
        self.redraw_sheets()

    def refresh_tree_dropdowns(self):
        self.tree.del_dropdown("A:")
        for c, hdr in enumerate(self.headers):
            if hdr.validation:
                self.tree.dropdown(
                    _n2a(c),
                    values=hdr.validation,
                    edit_data=False,
                )
        self.redraw_sheets()

    def prev_change(self) -> int:
        return len(self.changelog) - next(
            i
            for i, row in enumerate(islice(reversed(self.changelog), 1, None), start=1)
            if not row[1].startswith(
                (
                    "Merge | ",
                    "Imported change |",
                    "Edit cell |",
                    "Delete ID from all hierarchies |",
                    "Delete ID |",
                    "Cut and paste ID + children |",
                    "Copy and paste ID |",
                    "Copy and paste ID + children |",
                    "Cut and paste ID |",
                )
            )
        )

    def undo_complete(self):
        self.undo_in_progress = False

    def undo(self, event=None):
        if self.undo_in_progress or not self.vs:
            return
        self.start_work("Undoing last action...")
        self.C.unsaved_changes = True
        self.C.change_app_title(star="add")
        self.undo_in_progress = True
        self.vp -= 1
        new_vs = self.vs.pop()
        self.ic = new_vs["required_data"]["ic"]
        self.pc = new_vs["required_data"]["pc"]
        self.hiers = new_vs["required_data"]["hiers"]
        self.nodes = pickle.loads(new_vs["required_data"]["nodes"])
        self.tv_label_col = new_vs["required_data"]["tv_label_col"]
        self.row_len = new_vs["required_data"]["row_len"]
        self.mirror_var = new_vs["required_data"]["mirror_bool"]
        self.auto_sort_nodes_bool = new_vs["required_data"]["auto_sort_nodes_bool"]
        self.topnodes_order = new_vs["required_data"]["topnodes_order"]
        self.saved_info = pickle.loads(new_vs["required_data"]["saved_info"])
        self.tagged_ids = new_vs["required_data"]["tagged_ids"]
        self.sheet.align_columns(
            columns=new_vs["required_data"]["sheet_column_alignments"],
            redraw=False,
        )
        self.tree.align_columns(
            columns=new_vs["required_data"]["sheet_column_alignments"],
            redraw=False,
        )
        self.reset_tagged_ids_dropdowns()
        self.clear_copied_details()
        self.headers = new_vs["required_data"]["headers"]

        if new_vs["type"] in (
            "full sheet",
            "ctrl x, v, del key",
            "ctrl x, v, del key id par",
            "paste id",
            "delete ids",
        ):
            try:
                self.changelog = self.changelog[: self.prev_change()]
            except Exception:
                self.changelog = []
        else:
            del self.changelog[-1]
        if new_vs["type"] == "add id":
            rn = new_vs["row"]["rn"]
            if new_vs["row"]["added_or_changed"] == "changed":
                self.sheet.MT.data[rn] = new_vs["row"]["stored"]
                self.refresh_formatting(rows=rn)
            elif new_vs["row"]["added_or_changed"] == "added":
                del self.sheet.MT.data[rn]
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}

        elif new_vs["type"] == "rename id":
            rows = {new_vs["ikrow"][0]}
            cols = {self.ic}
            for tup in new_vs["rows"]:
                rn, h, v = pickle.loads(zlib.decompress(tup))
                rows.add(rn)
                cols.add(h)
                self.sheet.MT.data[rn][h] = v
            self.sheet.MT.data[new_vs["ikrow"][0]][self.ic] = new_vs["ikrow"][2]
            self.refresh_formatting(rows=rows, columns=cols)
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}

        elif new_vs["type"] == "paste id":
            rows = set()
            cols = set()
            for tup in new_vs["rows"]:
                rn, fromcol, frompar, tocol, topar = pickle.loads(zlib.decompress(tup))
                self.sheet.MT.data[rn][fromcol] = frompar
                self.sheet.MT.data[rn][tocol] = topar
                rows.add(rn)
                cols.add(fromcol)
                cols.add(tocol)
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            self.refresh_formatting(rows=rows, columns=cols)

        elif new_vs["type"] == "delete ids":
            for obj in reversed(new_vs["rows"]):
                if obj.t == 1:
                    self.sheet.MT.data.insert(obj.rn, obj.row)
                else:
                    for h, par in zip(self.hiers, pickle.loads(zlib.decompress(obj.row))):
                        self.sheet.MT.data[obj.rn][h] = par
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            self.refresh_formatting(dehighlight=True)

        elif new_vs["type"] == "add col":
            c = new_vs["treecolsel"]
            for r in range(len(self.sheet.MT.data)):
                del self.sheet.MT.data[r][c]
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            self.refresh_formatting(dehighlight=True)

        elif new_vs["type"] == "del cols":
            for cn, rowdict in reversed(new_vs["cols"].items()):
                for rn, v in rowdict.items():
                    self.sheet.MT.data[rn].insert(cn, v)
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            self.refresh_formatting(dehighlight=True)

        elif new_vs["type"] == "edit validation":
            for rn, c in enumerate(pickle.loads(zlib.decompress(new_vs["col"]))):
                self.sheet.MT.data[rn][new_vs["col_num"]] = c
            self.refresh_formatting(dehighlight=True)

        elif new_vs["type"] == "rename col":
            ...

        elif new_vs["type"] == "col type text":
            self.refresh_formatting(columns=new_vs["col_num"])

        elif new_vs["type"] == "col type num date":
            for rn, c in enumerate(pickle.loads(zlib.decompress(new_vs["col"]))):
                self.sheet.MT.data[rn][new_vs["col_num"]] = c
            self.refresh_formatting(columns=new_vs["col_num"])

        elif new_vs["type"] == "sort":
            self.sheet.MT.data = [
                self.sheet.MT.data[self.rns[new_vs["ids"][oldrn]]] for oldrn in range(len(new_vs["ids"]))
            ]
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            self.refresh_formatting(dehighlight=True)

        elif new_vs["type"] == "prune changelog":
            self.changelog = new_vs["rows"] + self.changelog

        elif new_vs["type"] == "drag rows":
            self.sheet.mapping_move_rows(dict(zip(new_vs["row_mapping"].values(), new_vs["row_mapping"])), undo=False)
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}

        elif new_vs["type"] == "drag cols":
            self.sheet.mapping_move_columns(new_vs["column_mapping"], undo=False)
            self.tree.mapping_move_columns(new_vs["column_mapping"], undo=False)

        elif new_vs["type"] == "node sort":
            ...

        elif new_vs["type"] == "date form":
            self.DATE_FORM = new_vs["old_form"]
            xxform = new_vs["new_form"]
            date_cols = [i for i, h in enumerate(self.headers) if h.type_ == "Date"]
            for col in date_cols:
                for rn in range(len(self.sheet.MT.data)):
                    cell = self.sheet.MT.data[rn][col]
                    if "/" in cell or "-" in cell:
                        try:
                            cur_cell = datetime.datetime.strptime(cell, xxform)
                            cell = datetime.datetime.strftime(cur_cell, self.DATE_FORM)
                        except Exception:
                            pass
                    self.sheet.MT.data[rn][col] = cell
            self.refresh_formatting(columns=date_cols)

        elif new_vs["type"].startswith("full"):
            self.warnings_filepath = new_vs["og_file"]
            self.warnings_sheet = new_vs["og_sheet"]
            self.warnings = new_vs["build_warnings"]
            self.sheet.MT.data = pickle.loads(zlib.decompress(new_vs["sheet"]))
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            self.refresh_formatting(dehighlight=True)

        elif new_vs["type"] == "ctrl x, v, del key id par":
            self.sheet.MT.data = pickle.loads(zlib.decompress(new_vs["sheet"]))
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            self.refresh_formatting(dehighlight=True)

        elif new_vs["type"] == "ctrl x, v, del key":
            rows = set()
            cols = set()
            for k, v in new_vs["cells"].items():
                self.sheet.MT.data[k[0]][k[1]] = v
                rows.add(k[0])
                cols.add(k[1])
            self.refresh_formatting(rows=rows, columns=cols)

        self.sheet.row_index(newindex=self.ic)
        self.sheet.set_column_widths(new_vs["required_data"]["sheet_col_positions"], canvas_positions=True)
        self.sheet.set_safe_row_heights(new_vs["required_data"]["sheet_row_positions"])
        self.redo_tree_display()
        self.set_headers()
        self.refresh_hier_dropdown(self.hiers.index(self.pc))
        self.rehighlight_tagged_ids()
        self.edit_menu.entryconfig(0, label=f"Undo {self.vp}/30")
        if not self.vp:
            self.edit_menu.entryconfig(0, state="disabled")
        self.mirror_sels_disabler = True
        self.move_tree_pos()
        if new_vs["required_data"]["sheet_selections"] is not None:
            self.reselect_sheet_sel(new_vs["required_data"]["sheet_selections"])
            self.sheet_select_event()
        self.refresh_dropdowns()
        self.move_sheet_pos()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        del new_vs
        self.after_idle(self.undo_complete)
        self.stop_work(self.get_tree_editor_status_bar_text())

    def tree_gen_heights_from_saved(self) -> Generator[int]:
        heights_dict = self.saved_info[self.pc].theights
        default_row_height = self.tree.MT.get_default_row_height()
        return (
            heights_dict[iid] if (iid := self.tree.rowitem(r, data_index=True)) in heights_dict else default_row_height
            for r in self.tree.displayed_rows
        )

    def tree_gen_widths_from_saved(self) -> Generator[int]:
        widths_dict = self.saved_info[self.pc].twidths
        default_col_width = self.tree.ops.default_column_width
        return (widths_dict.get(h.name, default_col_width) for h in self.headers)

    def save_sheet_row_heights(self) -> None:
        quick_data = self.sheet.MT.data
        self.saved_sheet_row_heights = {
            quick_data[rn][self.ic].lower(): h for rn, h in enumerate(self.sheet.get_row_heights())
        }

    def set_undo_label(self, event=None):
        self.edit_menu.entryconfig(0, label=f"Undo {self.vp}/30")
        if not self.vp:
            self.edit_menu.entryconfig(0, state="disabled")

    def copy_headers(self):
        return [
            Header(
                f"{h.name}",
                f"{h.type_}",
                [tuple(t) for t in h.formatting],
                h.validation.copy(),
            )
            for h in self.headers
        ]

    def save_info_get_saved_info(self):
        self.saved_info[self.pc] = new_info_storage(
            scrolls=(
                float(self.tree.get_xview()[0]),
                float(self.tree.get_yview()[0]),
                float(self.sheet.get_xview()[0]),
                float(self.sheet.get_yview()[0]),
            ),
            opens=dict.fromkeys(self.tree.tree_get_open()),
            boxes=self.tree.boxes,
            selected=self.tree.selected,
            twidths={self.headers[i].name: width for i, width in enumerate(self.tree.get_column_widths())},
            theights={
                self.tree.rowitem(i, data_index=False): height
                for i, height in enumerate(self.tree.get_safe_row_heights())
            },
        )
        return self.saved_info

    def get_required_snapshot_data(self):
        return {
            "saved_info": pickle.dumps(self.save_info_get_saved_info()),
            "sheet_col_positions": list(self.sheet.get_column_widths(canvas_positions=True)),
            "sheet_row_positions": self.sheet.get_safe_row_heights(),
            "topnodes_order": {k: list(v) for k, v in self.topnodes_order.items()},
            "tv_label_col": int(self.tv_label_col),
            "tagged_ids": set(self.tagged_ids),
            "sheet_column_alignments": dict(self.sheet.get_column_alignments()),
            "headers": self.copy_headers(),
            "ic": int(self.ic),
            "pc": int(self.pc),
            "hiers": list(self.hiers),
            "row_len": int(self.row_len),
            "auto_sort_nodes_bool": bool(self.auto_sort_nodes_bool),
            "mirror_bool": bool(self.mirror_var),
            "nodes": pickle.dumps(self.nodes),
            "focus": self.tree.has_focus(),
            "sheet_selections": self.get_sheet_sel(),
        }

    def snapshot_ctrl_x_v_del_key_id_par(self):
        self.snapshot_chore()
        self.save_sheet_row_heights()
        self.vs.append(
            {
                "type": "ctrl x, v, del key id par",
                "sheet": zlib.compress(pickle.dumps(self.sheet.MT.data)),
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_ctrl_x_v_del_key(self):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "ctrl x, v, del key",
                "cells": {},
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_sheet(self, type_: str = "full sheet"):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": type_,
                "og_file": self.warnings_filepath,
                "og_sheet": self.warnings_sheet,
                "build_warnings": self.warnings,
                "sheet": zlib.compress(pickle.dumps(self.sheet.MT.data)),
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_add_id(self):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "add id",
                "row": {},
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_rename_id(self):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "rename id",
                "rows": [],
                "ikrow": (),
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_paste_id(self):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "paste id",
                "rows": [],
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_delete_ids(self):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "delete ids",
                "rows": [],
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_add_col(self, treecolsel):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "add col",
                "treecolsel": int(treecolsel),
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_del_cols(self):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "del cols",
                "cols": {},
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_rename_col(self):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "rename col",
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_edit_validation(self, col, validation):
        self.snapshot_chore()
        self.changelog_append(
            "Edit validation",
            f"Column #{col + 1} named: {self.headers[col].name} with type: {self.headers[col].type_}",
            f"{','.join(self.headers[col].validation)}",
            f"{','.join(validation)}",
        )
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        self.vs.append(
            {
                "type": "edit validation",
                "col_num": col,
                "col": zlib.compress(pickle.dumps([r[col] for r in self.sheet.MT.data])),
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_col_type_text(self, col):
        self.snapshot_chore()
        self.changelog_append(
            "Change detail column type",
            f"Column #{col + 1} named: {self.headers[col].name}",
            f"{self.headers[col].type_}",
            "Text",
        )
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        self.vs.append(
            {
                "type": "col type text",
                "col_num": col,
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_col_type_num_date(self, col, type_):
        self.snapshot_chore()
        self.changelog_append(
            "Change detail column type",
            f"Column #{col + 1} named: {self.headers[col].name}",
            f"{self.headers[col].type_}",
            f"{type_}",
        )
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        self.vs.append(
            {
                "type": "col type num date",
                "col_num": col,
                "col": zlib.compress(pickle.dumps([r[col] for r in self.sheet.MT.data])),
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_begin_drag_rows(self, event=None):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "drag rows",
                "row_mapping": {},
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_drag_rows(self, event_data):
        self.vs[-1]["row_mapping"] = self.sheet.full_move_rows_idxs(event_data["moved"]["rows"]["data"])
        old_locs = ",".join(f"{r}" for r in event_data["moved"]["rows"]["data"])
        new_locs = ",".join(f"{r}" for r in event_data["moved"]["rows"]["data"].values())
        self.changelog_append(
            "Move rows",
            f"{len(event_data['moved']['rows']['data'])} rows",
            f"Old locations: {old_locs}",
            f"New locations: {new_locs}",
        )
        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        self.disable_paste()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        self.redraw_sheets()

    def begin_tree_drag_drop_ids(self, event=None):
        self.save_info_get_saved_info()

    def get_ids_parent(self, iid) -> str:
        if self.nodes[iid.lower()].ps[self.pc]:
            return self.nodes[iid.lower()].ps[self.pc]
        return ""

    def tree_drag_drop_ids(self, event=None):
        if not event.moved.rows.data:
            return
        self.start_work("Moving IDs...")
        moved_rows = [self.tree.data_r(r) for r in event.moved.rows.displayed]
        as_sibling = []
        index_only = []
        if event.value > max(event.moved.rows.displayed):
            event.value -= 1
        move_to_iid = self.tree.rowitem(event.value)
        new_parent = self.tree.parent(move_to_iid)
        for r in moved_rows:
            iid = self.tree.rowitem(r, data_index=True)
            iid_parent = self.tree.parent(iid)
            if iid_parent == new_parent:
                index_only.append(iid)
            else:
                as_sibling.append(iid)
        successful = []
        if as_sibling:
            self.selected_ID = self.nodes[move_to_iid].name
            self.cut_ids(as_sibling, status_bar=False)
            if new_parent:
                self.selected_PAR = self.nodes[new_parent].name
            else:
                self.selected_PAR = new_parent
            successful = [dct["id"] for dct in self.paste_cut_sibling_all(redo_tree=False)]
        if successful and (not self.auto_sort_nodes_bool or index_only):
            index_only += successful
        if index_only:
            if self.auto_sort_nodes_bool:
                self.auto_sort_nodes_bool = False
                self.remake_topnodes_order()
            self.redo_tree_display(selections=False)
            move_to_index = self.tree.index(move_to_iid)
            if (
                not is_contiguous(event.moved.rows.displayed)
                and max(moved_rows) > self.tree.data_r(event.value)
                and min(moved_rows) < self.tree.data_r(event.value)
            ):
                move_to_index -= 1
            if parik := self.get_ids_parent(index_only[0]):
                self.nodes[parik].cn[self.pc].insert(
                    move_to_index,
                    self.nodes[parik].cn[self.pc].pop(self.tree.index(index_only[0])),
                )
            else:
                self.topnodes_order[self.pc].insert(
                    move_to_index,
                    self.topnodes_order[self.pc].pop(self.tree.index(index_only[0])),
                )
            if len(index_only) > 1:
                new_index = move_to_index
                for iid in islice(index_only, 1, None):
                    if parik := self.get_ids_parent(iid):
                        current_index = self.nodes[parik].cn[self.pc].index(iid)
                        if new_index < current_index:
                            new_index += 1
                        self.nodes[parik].cn[self.pc].insert(
                            new_index,
                            self.nodes[parik].cn[self.pc].pop(current_index),
                        )
                    else:
                        current_index = self.topnodes_order[self.pc].index(iid)
                        if new_index < current_index:
                            new_index += 1
                        self.topnodes_order[self.pc].insert(
                            new_index,
                            self.topnodes_order[self.pc].pop(current_index),
                        )
        self.redo_tree_display(selections=False)
        self.redraw_sheets()
        all_iids = index_only + successful
        self.tree.scroll_to_item(all_iids[0])
        self.tree.selection_set(all_iids)
        self.stop_work(self.get_tree_editor_status_bar_text())

    def snapshot_begin_drag_cols(self, event=None):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "drag cols",
                "column_mapping": {},
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_drag_cols(self, event_data):
        if self.tree.has_focus():
            full_new_idxs = self.tree.full_move_columns_idxs(event_data["moved"]["columns"]["data"])
            self.sheet.mapping_move_columns(
                event_data["moved"]["columns"]["data"],
                event_data["moved"]["columns"]["displayed"],
                undo=False,
            )
        else:
            full_new_idxs = self.sheet.full_move_columns_idxs(event_data["moved"]["columns"]["data"])
            self.tree.mapping_move_columns(
                event_data["moved"]["columns"]["data"],
                event_data["moved"]["columns"]["displayed"],
                undo=False,
            )
        old_locs = ",".join(f"{c}" for c in event_data["moved"]["columns"]["data"])
        new_locs = ",".join(f"{c}" for c in event_data["moved"]["columns"]["data"].values())
        self.changelog_append(
            "Move columns",
            f"{len(event_data['moved']['columns']['data'])} columns",
            f"Old locations: {old_locs}",
            f"New locations: {new_locs}",
        )
        self.ic = full_new_idxs[self.ic]
        self.pc = full_new_idxs[self.pc]
        self.headers = move_elements_by_mapping(
            self.headers,
            event_data["moved"]["columns"]["data"],
        )
        self.set_headers()
        self.hiers = sorted(full_new_idxs[c] for c in self.hiers)
        self.tv_label_col = full_new_idxs[self.tv_label_col]
        for node in self.nodes.values():
            node.cn = {full_new_idxs[k]: v for k, v in node.cn.items()}
            node.ps = {full_new_idxs[k]: v for k, v in node.ps.items()}
        self.saved_info = {full_new_idxs[k]: v for k, v in self.saved_info.items()}
        if not self.auto_sort_nodes_bool:
            self.topnodes_order = {full_new_idxs[k]: v for k, v in self.topnodes_order.items()}
        self.clear_copied_details()
        self.refresh_hier_dropdown(self.hiers.index(self.pc))
        self.sheet.row_index(newindex=self.ic)
        self.vs[-1]["column_mapping"] = dict(zip(full_new_idxs.values(), full_new_idxs))
        self.refresh_dropdowns()
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def set_headers(self, tree: bool = True, sheet: bool = True):
        headers = [
            "\n".join((f"{h.name}", f"{i + 1}/{_n2a(i)} {h.name}", f"{h.type_} {h.name}"))
            for i, h in enumerate(self.headers)
        ]
        if sheet:
            self.sheet.headers(
                headers,
                reset_col_positions=False,
                show_headers_if_not_sheet=False,
            )
        if tree:
            self.tree.headers(
                headers.copy(),
                reset_col_positions=False,
                show_headers_if_not_sheet=False,
            )

    def snapshot_sheet_sort(self):
        self.snapshot_chore()
        self.vs.append(
            {
                "type": "sort",
                "ids": {v: k for k, v in self.rns.items()},
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_prune_changelog(self, up_to):
        self.snapshot_chore()
        self.changelog_append(
            "Pruned changelog",
            f"From: {self.changelog[0][0]} To: {self.changelog[up_to][0]}",
            "",
            "",
        )
        self.vs.append(
            {
                "type": "prune changelog",
                "rows": self.changelog[: up_to + 1],
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_change_date_form(self, old_form, new_form):
        self.snapshot_chore()
        self.changelog_append(
            "Date format change",
            "",
            old_form.replace("%", ""),
            new_form.replace("%", ""),
        )
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        self.vs.append(
            {
                "type": "date form",
                "old_form": old_form,
                "new_form": new_form,
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_auto_sort_nodes(self):
        self.snapshot_chore()
        self.changelog_append(
            "Sort treeview",
            "Alphanumerically sorted order of Treeview IDs",
            "",
            "",
        )
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        self.vs.append(
            {
                "type": "node sort",
                "required_data": self.get_required_snapshot_data(),
            }
        )

    def snapshot_chore(self):
        self.save_info_get_saved_info()
        if self.vp < 30:
            self.vp += 1
        self.edit_menu.entryconfig(0, label=f"Undo {self.vp}/30", state="normal")

    def sort_sheet_choice(self):
        popup = Sort_Sheet_Popup(self, [h.name for h in self.headers], theme=self.C.theme)
        if popup.sort_decision["type"] is None:
            return
        if popup.sort_decision["type"] == "by column":
            self.sort_sheet(popup.sort_decision["col"], popup.sort_decision["order"])
        elif popup.sort_decision["type"] == "by tree":
            self.sort_sheet_walk()

    def sort_sheet_rc_asc(self):
        widget = self.tree if self.tree.has_focus() else self.sheet
        if widget.get_selected_columns():
            self.sort_sheet(header=self.headers[widget.selected.column].name, order="ASCENDING")

    def sort_sheet_rc_desc(self):
        widget = self.tree if self.tree.has_focus() else self.sheet
        if widget.get_selected_columns():
            self.sort_sheet(header=self.headers[widget.selected.column].name, order="DESCENDING")

    def sort_sheet(self, header, order, snapshot=True):
        col = next(i for i, h in enumerate(self.headers) if h.name == header)
        if snapshot:
            self.snapshot_sheet_sort()
            self.changelog_append(
                "Sort sheet",
                f"Sorted sheet by column #{col + 1} named: {header} in {order} order",
                "",
                "",
            )
        if self.headers[col].type_ == "Date":
            date_rows = []
            num_rows = []
            nothing_rows = []
            for row in self.sheet.MT.data:
                if "/" in row[col] or "-" in row[col]:
                    date_rows.append(row)
                elif row[col]:
                    num_rows.append(row)
                else:
                    nothing_rows.append(row)
            if order == "ASCENDING":
                try:
                    date_rows = sorted(
                        date_rows,
                        key=lambda row: datetime.datetime.strptime(row[col], self.DATE_FORM),
                    )
                except Exception:
                    date_rows = sorted(date_rows, key=lambda row: row[col])
                try:
                    num_rows = sorted(num_rows, key=lambda row: int(row[col]))
                except Exception:
                    num_rows = sorted(num_rows, key=lambda row: row[col])
            elif order == "DESCENDING":
                try:
                    date_rows = sorted(
                        date_rows,
                        key=lambda row: datetime.datetime.strptime(row[col], self.DATE_FORM),
                        reverse=True,
                    )
                except Exception:
                    date_rows = sorted(date_rows, key=lambda row: row[col], reverse=True)
                try:
                    num_rows = sorted(num_rows, key=lambda row: int(row[col]), reverse=True)
                except Exception:
                    num_rows = sorted(num_rows, key=lambda row: row[col], reverse=True)
            self.sheet.MT.data = date_rows + num_rows + nothing_rows
        else:
            ak = lambda row: tuple(  # noqa: E731
                int(c) if c.isdigit() else c.lower() for c in re.split("([0-9]+)", row[col])
            )
            if order == "ASCENDING":
                self.sheet.MT.data.sort(key=ak)
            elif order == "DESCENDING":
                self.sheet.MT.data.sort(key=ak, reverse=True)
        row_heights = self.sheet.get_row_heights()
        nrhs = []
        for i, r in enumerate(self.sheet.MT.data):
            ik = r[self.ic].lower()
            nrhs.append(row_heights[self.rns[ik]])
            self.rns[ik] = i
        self.sheet.set_row_heights(nrhs)
        if snapshot:
            self.disable_paste()
            self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
            self.refresh_formatting()
            self.reset_tagged_ids_dropdowns()
            self.rehighlight_tagged_ids()
            self.redraw_sheets()

    def sort_sheet_walk(self, snapshot=True):
        oldrns = self.rns.copy()
        oldpc = int(self.pc)
        if snapshot:
            self.snapshot_sheet_sort()
            self.changelog_append(
                "Sort sheet",
                "Sorted sheet in tree walk order",
                "",
                "",
            )
        for h in reversed(self.hiers):
            self.pc = int(h)
            self.sort_sheet_walk_pc_changer()
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        self.pc = int(oldpc)
        row_heights = self.sheet.get_row_heights()
        nrhs = []
        self.rns = {}
        for i, r in enumerate(self.sheet.MT.data):
            ik = r[self.ic].lower()
            nrhs.append(row_heights[oldrns[ik]])
            self.rns[ik] = i
        self.sheet.set_row_heights(nrhs)
        if snapshot:
            self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
            self.refresh_formatting()
            self.reset_tagged_ids_dropdowns()
            self.rehighlight_tagged_ids()
            self.disable_paste()
            self.redraw_sheets()

    def sort_sheet_walk_pc_changer(self):
        self.new_sheet = []
        self.visited = set()

        # Initialize stack with top nodes and their children
        stack = [(iid, self.nodes[iid].cn[self.pc]) for iid in self.top_iids()]
        stack.reverse()  # Reverse to maintain original processing order

        # Iterative depth-first traversal
        while stack:
            iid, children = stack.pop()
            rowno = self.rns[iid]
            if rowno not in self.visited:
                self.visited.add(rowno)
                self.new_sheet.append(self.sheet.MT.data[rowno])
                # Push children in reverse order to process them in original order
                child_stack = [(ciid, self.nodes[ciid].cn[self.pc]) for ciid in reversed(children)]
                stack.extend(child_stack)

        # Append any remaining unvisited rows
        for r in sorted(r for r in self.rns.values() if r not in self.visited):
            self.new_sheet.append(self.sheet.MT.data[r])

        # Update sheet data and reset temporary attributes
        self.sheet.MT.data = self.new_sheet
        self.new_sheet = []
        self.visited = set()

    def search_choice(self, event=None):
        choice = self.search_choice_displayed.get()
        if choice == "Non-exact":
            self.search_for_any(None)
        elif choice == "ID non-exact":
            self.search_for_ID(None, False)
        elif choice == "ID exact":
            self.search_for_ID(None, True)
        elif choice == "Detail non-exact":
            self.search_for_detail(None, False)
        elif choice == "Detail exact":
            self.search_for_detail(None, True)

    def sheet_search_choice(self, event=None):
        choice = self.sheet_search_choice_displayed.get()
        if choice == "Non-exact":
            self.sheet_search_for_any()
        elif choice == "ID non-exact":
            self.sheet_search_for_ID(None, False)
        elif choice == "ID exact":
            self.sheet_search_for_ID(None, True)
        elif choice == "Detail non-exact":
            self.sheet_search_for_detail(None, False)
        elif choice == "Detail exact":
            self.sheet_search_for_detail(None, True)

    def search_for_any(self, find=None):
        if not (search := self.search_entry.get() if find is None else find):
            return
        self.reset_tree_search_dropdown()
        search = search.lower()
        for iid, node in self.nodes.items():
            for i, e in enumerate(self.sheet.MT.data[self.rns[iid]]):
                if search in e.lower():
                    for h, par in node.ps.items():
                        if par is not None:
                            self.search_results.append(
                                SearchResult(
                                    hierarchy=h,
                                    text=(
                                        self.headers[h].name,
                                        node.name,
                                        self.headers[i].name,
                                        re.sub(remove_nrt, "", e),
                                    ),
                                    iid=iid,
                                    column=i,
                                    term=search,
                                    type_=2,
                                    exact=False,
                                )
                            )
        if self.search_results:
            col_chars = frame_w_to_nchars(
                frame_w=self.search_dropdown.winfo_width(),
                fixed_font_w=self.fixed_font_w,
                ncols=4,
            )
            process_search_results(
                self.search_results,
                search_results_max_column_chars(self.search_results, col_chars),
                col_chars,
            )
        self.display_search_results()

    def search_for_ID(self, find=None, exact=False):
        if not (search := self.search_entry.get() if find is None else find):
            return
        self.reset_tree_search_dropdown()
        search = search.lower()
        for iid, node in self.nodes.items():
            if (exact and search == iid) or (not exact and search in iid):
                for h, par in node.ps.items():
                    if par is not None:
                        self.search_results.append(
                            SearchResult(
                                hierarchy=h,
                                text=(
                                    self.headers[h].name,
                                    node.name,
                                ),
                                iid=iid,
                                column=self.ic,
                                term=search,
                                type_=0,
                                exact=exact,
                            )
                        )
        if self.search_results:
            col_chars = frame_w_to_nchars(
                frame_w=self.search_dropdown.winfo_width(),
                fixed_font_w=self.fixed_font_w,
                ncols=2,
            )
            process_search_results(
                self.search_results,
                search_results_max_column_chars(self.search_results, col_chars),
                col_chars,
            )
        self.display_search_results()

    def search_for_detail(self, find=None, exact=False):
        if not (search := self.search_entry.get() if find is None else find):
            return
        self.reset_tree_search_dropdown()
        search = search.lower()
        idcol_hiers = set(self.hiers) | {self.ic}
        for iid, node in self.nodes.items():
            for i, e in enumerate(self.sheet.MT.data[self.rns[iid]]):
                if i not in idcol_hiers and ((exact and search == e.lower()) or (not exact and search in e.lower())):
                    for h, par in node.ps.items():
                        if par is not None:
                            self.search_results.append(
                                SearchResult(
                                    hierarchy=h,
                                    text=(
                                        self.headers[h].name,
                                        node.name,
                                        self.headers[i].name,
                                        re.sub(remove_nrt, "", e),
                                    ),
                                    iid=iid,
                                    column=i,
                                    term=search,
                                    type_=1,
                                    exact=exact,
                                )
                            )
        if self.search_results:
            col_chars = frame_w_to_nchars(
                frame_w=self.search_dropdown.winfo_width(),
                fixed_font_w=self.fixed_font_w,
                ncols=4,
            )
            process_search_results(
                self.search_results,
                search_results_max_column_chars(self.search_results, col_chars),
                col_chars,
            )
        self.display_search_results()

    def display_search_results(self):
        if self.search_results:
            self.search_results.sort(key=attrgetter("hierarchy"))
            self.search_dropdown["values"] = [result.text for result in self.search_results]
            self.search_displayed.set(self.search_results[0].text)
            self.show_search_result()

    def sheet_search_for_any(self, find=None):
        if not (search := self.sheet_search_entry.get() if find is None else find):
            return
        self.reset_sheet_search_dropdown()
        search = search.lower()
        for r in self.sheet.MT.data:
            for i, e in enumerate(r):
                if search in e.lower():
                    self.sheet_search_results.append(
                        SearchResult(
                            hierarchy=self.pc,
                            text=(
                                r[self.ic],
                                self.headers[i].name,
                                re.sub(remove_nrt, "", e),
                            ),
                            iid=r[self.ic].lower(),
                            column=i,
                            term=search,
                            type_=2,
                            exact=False,
                        )
                    )
        if self.sheet_search_results:
            col_chars = frame_w_to_nchars(
                frame_w=self.sheet_search_dropdown.winfo_width(),
                fixed_font_w=self.fixed_font_w,
                ncols=3,
            )
            process_search_results(
                self.sheet_search_results,
                search_results_max_column_chars(self.sheet_search_results, col_chars),
                col_chars,
            )
        self.sheet_display_search_results()

    def sheet_search_for_ID(self, find=None, exact=False):
        if not (search := self.sheet_search_entry.get() if find is None else find):
            return
        self.reset_sheet_search_dropdown()
        search = search.lower()
        for r in self.sheet.MT.data:
            if (exact and search == r[self.ic].lower()) or (not exact and search in r[self.ic].lower()):
                self.sheet_search_results.append(
                    SearchResult(
                        hierarchy=self.pc,
                        text=(r[self.ic],),
                        iid=r[self.ic].lower(),
                        column=self.ic,
                        term=search,
                        type_=0,
                        exact=exact,
                    )
                )
        if self.sheet_search_results:
            col_chars = frame_w_to_nchars(
                frame_w=self.sheet_search_dropdown.winfo_width(),
                fixed_font_w=self.fixed_font_w,
                ncols=1,
            )
            process_search_results(
                self.sheet_search_results,
                search_results_max_column_chars(self.sheet_search_results, col_chars),
                col_chars,
            )
        self.sheet_display_search_results()

    def sheet_search_for_detail(self, find=None, exact=False):
        if not (search := self.sheet_search_entry.get() if find is None else find):
            return
        self.reset_sheet_search_dropdown()
        search = search.lower()
        idcol_hiers = set(self.hiers) | {self.ic}
        for r in self.sheet.MT.data:
            for i, e in enumerate(r):
                if i not in idcol_hiers and ((exact and search == e.lower()) or (not exact and search in e.lower())):
                    self.sheet_search_results.append(
                        SearchResult(
                            hierarchy=self.pc,
                            text=(
                                r[self.ic],
                                self.headers[i].name,
                                re.sub(remove_nrt, "", e),
                            ),
                            iid=r[self.ic].lower(),
                            column=i,
                            term=search,
                            type_=1,
                            exact=exact,
                        )
                    )
        if self.sheet_search_results:
            col_chars = frame_w_to_nchars(
                frame_w=self.sheet_search_dropdown.winfo_width(),
                fixed_font_w=self.fixed_font_w,
                ncols=3,
            )
            process_search_results(
                self.sheet_search_results,
                search_results_max_column_chars(self.sheet_search_results, col_chars),
                col_chars,
            )
        self.sheet_display_search_results()

    def sheet_display_search_results(self):
        if self.sheet_search_results:
            self.sheet_search_dropdown["values"] = [result.text for result in self.sheet_search_results]
            self.sheet_search_displayed.set(self.sheet_search_results[0].text)
            self.sheet_show_search_result()

    def enable_copy_paste(self):
        self.tree_rc_menu_single_row_paste.entryconfig(
            "Paste IDs as child",
            command=self.paste_copied_child,
            state="normal",
        )
        self.tree_rc_menu_single_row_paste.entryconfig(
            "Paste IDs as sibling",
            command=self.paste_copied_sibling,
            state="normal",
        )
        self.tree_rc_menu_single_row_paste.entryconfig(
            "Paste IDs and children as sibling",
            command=self.paste_copied_sibling_all,
            state="normal",
        )
        self.tree_rc_menu_single_row_paste.entryconfig(
            "Paste IDs and children as child",
            command=self.paste_copied_child_all,
            state="normal",
        )
        self.tree_rc_menu_empty.entryconfig(
            "Paste IDs",
            command=self.paste_copied_empty,
            state="normal",
        )
        self.tree_rc_menu_empty.entryconfig(
            "Paste IDs and children",
            command=self.paste_copied_empty_all,
            state="normal",
        )

    def enable_cut_paste(self):
        self.tree_rc_menu_single_row_paste.entryconfig(
            "Paste IDs as child",
            command=self.paste_cut_child,
            state="normal",
        )
        self.tree_rc_menu_single_row_paste.entryconfig(
            "Paste IDs as sibling",
            command=self.paste_cut_sibling,
            state="normal",
        )
        self.tree_rc_menu_single_row_paste.entryconfig(
            "Paste IDs and children as sibling",
            command=self.paste_cut_sibling_all,
            state="normal",
        )
        self.tree_rc_menu_single_row_paste.entryconfig(
            "Paste IDs and children as child",
            command=self.paste_cut_child_all,
            state="normal",
        )
        self.tree_rc_menu_empty.entryconfig(
            "Paste IDs",
            command=self.paste_cut_empty,
            state="normal",
        )
        self.tree_rc_menu_empty.entryconfig(
            "Paste IDs and children",
            command=self.paste_cut_empty_all,
            state="normal",
        )

    def enable_cut_paste_children(self):
        self.tree_rc_menu_single_row_paste.entryconfig(
            "Attach children",
            state="normal",
            command=self.paste_cut_children,
        )
        self.tree_rc_menu_empty.entryconfig(
            "Attach children",
            state="normal",
            command=self.paste_cut_children_empty,
        )

    def disable_paste(self):
        self.tree_rc_menu_single_row_paste.entryconfig("Paste IDs as child", state="disabled")
        self.tree_rc_menu_single_row_paste.entryconfig("Paste IDs as sibling", state="disabled")
        self.tree_rc_menu_single_row_paste.entryconfig("Paste IDs and children as sibling", state="disabled")
        self.tree_rc_menu_single_row_paste.entryconfig("Paste IDs and children as child", state="disabled")
        self.tree_rc_menu_single_row_paste.entryconfig("Attach children", state="disabled")
        self.tree_rc_menu_empty.entryconfig("Paste IDs", state="disabled")
        self.tree_rc_menu_empty.entryconfig("Paste IDs and children", state="disabled")
        self.tree_rc_menu_empty.entryconfig("Attach children", state="disabled")
        self.cut_columns = None
        self.cut = []
        self.copied = []
        self.cut_children_dct = {}
        return "break"

    def cut_children(self, event=None):
        if not self.selected_ID:
            return
        self.cut_children_dct["id"] = f"{self.selected_ID.lower()}"
        self.cut_children_dct["hier"] = int(self.pc)
        self.enable_cut_paste_children()

    def paste_cut_children(self):
        if not self.cut_children_dct:
            return
        self.snapshot_paste_id()
        if self.nodes[self.cut_children_dct["id"]].cn[self.cut_children_dct["hier"]]:
            see_iid = self.nodes[self.cut_children_dct["id"]].cn[self.cut_children_dct["hier"]][0]
            select_iids = tuple(self.nodes[self.cut_children_dct["id"]].cn[self.cut_children_dct["hier"]])
        else:
            see_iid = ""
            select_iids = ()
        success = self.cut_paste_children(
            self.cut_children_dct["id"], f"{self.selected_ID}", self.cut_children_dct["hier"]
        )
        if not success:
            self.vs.pop()
            self.vp -= 1
            self.set_undo_label()
            return
        iid = self.nodes[self.cut_children_dct["id"]].name
        np = self.nodes[self.selected_ID.lower()].name
        self.changelog_append(
            "Cut and paste children",
            "",
            f"Old parent: {iid} old column #{self.cut_children_dct['hier'] + 1} named: {self.headers[self.cut_children_dct['hier']].name}",
            f"New parent: {np} new column #{self.pc + 1} named: {self.headers[self.pc].name}",
        )
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.disable_paste()
        if see_iid:
            self.tree.scroll_to_item(see_iid)
            self.tree.selection_set(select_iids)
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def paste_cut_children_empty(self):
        if not self.cut_children_dct:
            return
        self.snapshot_paste_id()
        if self.nodes[self.cut_children_dct["id"]].cn[self.cut_children_dct["hier"]]:
            see_iid = self.nodes[self.cut_children_dct["id"]].cn[self.cut_children_dct["hier"]][0]
            select_iids = tuple(self.nodes[self.cut_children_dct["id"]].cn[self.cut_children_dct["hier"]])
        else:
            see_iid = ""
            select_iids = ()
        success = self.cut_paste_children(self.cut_children_dct["id"], "", self.cut_children_dct["hier"])
        if not success:
            self.vs.pop()
            self.vp -= 1
            self.set_undo_label()
            return
        iid = self.nodes[self.cut_children_dct["id"]].name
        self.changelog_append(
            "Cut and paste children",
            "",
            f"Old parent: {iid} old column #{self.cut_children_dct['hier'] + 1} named: {self.headers[self.cut_children_dct['hier']].name}",
            f"New parent: n/a - No parent new column #{self.pc + 1} named: {self.headers[self.pc].name}",
        )
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.disable_paste()
        if see_iid:
            self.tree.scroll_to_item(see_iid)
            self.tree.selection_set(select_iids)
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def paste_copied_child(self):
        if not self.copied or not self.selected_ID:
            return
        self.start_work(f"Pasting {len(self.copied)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.copied, 1):
            if self.copy_paste(dct["id"], dct["hier"], self.selected_ID, sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.copied)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            self.changelog_append_no_unsaved(
                "Copy and paste ID |",
                iid,
                f"From column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                f"New parent: {self.nodes[self.selected_ID.lower()].name} new column #{self.pc + 1} named: {self.headers[self.pc].name}",
            )
        if len(successful) > 1:
            self.changelog_append(
                f"Copy and paste {len(successful)} IDs",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Copy and paste ID")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def paste_copied_sibling(self):
        if not self.copied or not self.selected_ID:
            return
        self.start_work(f"Pasting {len(self.copied)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.copied, 1):
            if self.copy_paste(dct["id"], dct["hier"], "", sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.copied)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            if self.selected_PAR == "":
                self.changelog_append_no_unsaved(
                    "Copy and paste ID |",
                    iid,
                    f"From column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                    f"New parent: n/a - Top ID new column #{self.pc + 1} named: {self.headers[self.pc].name}",
                )
            else:
                self.changelog_append_no_unsaved(
                    "Copy and paste ID |",
                    iid,
                    f"From column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                    f"New parent: {self.nodes[self.selected_PAR.lower()].name} new column #{self.pc + 1} named: {self.headers[self.pc].name}",
                )
        if len(successful) > 1:
            self.changelog_append(
                f"Copy and paste {len(successful)} IDs",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Copy and paste ID")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def paste_copied_empty(self):
        if not self.copied:
            return
        self.start_work(f"Pasting {len(self.copied)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.copied, 1):
            if self.copy_paste(dct["id"], dct["hier"], "", sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.copied)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            self.changelog_append_no_unsaved(
                "Copy and paste ID |",
                iid,
                f"From column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                f"New parent: n/a - Top ID new column #{self.pc + 1} named: {self.headers[self.pc].name}",
            )
        if len(successful) > 1:
            self.changelog_append(
                f"Copy and paste {len(successful)} IDs",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Copy and paste ID")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def paste_copied_child_all(self):
        if not self.copied or not self.selected_ID:
            return
        self.start_work(f"Pasting {len(self.copied)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.copied, 1):
            if self.copy_paste_all(dct["id"], dct["hier"], self.selected_ID, sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.copied)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            self.changelog_append_no_unsaved(
                "Copy and paste ID + children |",
                iid,
                f"From column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                f"New parent: {self.nodes[self.selected_ID.lower()].name} new column #{self.pc + 1} named: {self.headers[self.pc].name}",
            )
        if len(successful) > 1:
            self.changelog_append(
                f"Copy and paste {len(successful)} IDs + children",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Copy and paste ID + children")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def paste_copied_sibling_all(self):
        if not self.copied or not self.selected_ID:
            return
        self.start_work(f"Pasting {len(self.copied)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.copied, 1):
            if self.copy_paste_all(dct["id"], dct["hier"], self.selected_PAR, sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.copied)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            if self.selected_PAR == "":
                self.changelog_append_no_unsaved(
                    "Copy and paste ID + children |",
                    iid,
                    f"From column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                    f"New parent: n/a - Top ID new column #{self.pc + 1} named: {self.headers[self.pc].name}",
                )
            else:
                self.changelog_append_no_unsaved(
                    "Copy and paste ID + children |",
                    iid,
                    f"From column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                    f"New parent: {self.nodes[self.selected_PAR.lower()].name} new column #{self.pc + 1} named: {self.headers[self.pc].name}",
                )
        if len(successful) > 1:
            self.changelog_append(
                f"Copy and paste {len(successful)} IDs + children",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Copy and paste ID + children")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def paste_copied_empty_all(self):
        if not self.copied:
            return
        self.start_work(f"Pasting {len(self.copied)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.copied, 1):
            if self.copy_paste_all(dct["id"], dct["hier"], "", sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.copied)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            self.changelog_append_no_unsaved(
                "Copy and paste ID + children |",
                iid,
                f"From column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                f"New parent: n/a - Top ID new column #{self.pc + 1} named: {self.headers[self.pc].name}",
            )
        if len(successful) > 1:
            self.changelog_append(
                f"Copy and paste {len(successful)} IDs + children",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Copy and paste ID + children")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def paste_cut_child(self):
        if not self.cut or not self.selected_ID:
            return
        self.start_work(f"Pasting {len(self.cut)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.cut, 1):
            if self.cut_paste(dct["id"], dct["parent"], dct["hier"], self.selected_ID, sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.cut)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, set):
                for idk in v:
                    self.nodes[idk].cn[self.sort_later_dct["old_hier"]] = self.sort_node_cn(
                        self.nodes[idk].cn[self.sort_later_dct["old_hier"]],
                        self.sort_later_dct["old_hier"],
                    )
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            self.changelog_append_no_unsaved(
                "Cut and paste ID |",
                iid,
                f"Old parent: {self.nodes[dct['parent']].name if dct['parent'] else 'n/a - Top ID'} old column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                f"New parent: {self.nodes[self.selected_ID.lower()].name} new column #{self.pc + 1} named: {self.headers[self.pc].name}",
            )
        if len(successful) > 1:
            self.changelog_append(
                f"Cut and paste {len(successful)} IDs",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Cut and paste ID")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def paste_cut_sibling(self):
        if not self.cut or not self.selected_ID:
            return
        self.start_work(f"Pasting {len(self.cut)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.cut, 1):
            if self.cut_paste(dct["id"], dct["parent"], dct["hier"], self.selected_PAR, sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.cut)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, set):
                for idk in v:
                    self.nodes[idk].cn[self.sort_later_dct["old_hier"]] = self.sort_node_cn(
                        self.nodes[idk].cn[self.sort_later_dct["old_hier"]],
                        self.sort_later_dct["old_hier"],
                    )
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            if self.selected_PAR == "":
                self.changelog_append_no_unsaved(
                    "Cut and paste ID |",
                    iid,
                    f"Old parent: {self.nodes[dct['parent']].name if dct['parent'] else 'n/a - Top ID'} old column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                    f"New parent: n/a - Top ID new column #{self.pc + 1} named: {self.headers[self.pc].name}",
                )
            else:
                self.changelog_append_no_unsaved(
                    "Cut and paste ID |",
                    iid,
                    f"Old parent: {self.nodes[dct['parent']].name if dct['parent'] else 'n/a - Top ID'} old column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                    f"New parent: {self.nodes[self.selected_PAR.lower()].name} new column #{self.pc + 1} named: {self.headers[self.pc].name}",
                )
        if len(successful) > 1:
            self.changelog_append(
                f"Cut and paste {len(successful)} IDs",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Cut and paste ID")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def paste_cut_empty(self):
        if not self.cut:
            return
        self.start_work(f"Pasting {len(self.cut)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.cut, 1):
            if self.cut_paste(dct["id"], dct["parent"], dct["hier"], "", sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.cut)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, set):
                for idk in v:
                    self.nodes[idk].cn[self.sort_later_dct["old_hier"]] = self.sort_node_cn(
                        self.nodes[idk].cn[self.sort_later_dct["old_hier"]],
                        self.sort_later_dct["old_hier"],
                    )
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            self.changelog_append_no_unsaved(
                "Cut and paste ID |",
                iid,
                f"Old parent: {self.nodes[dct['parent']].name if dct['parent'] else 'n/a - Top ID'} old column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                f"New parent: n/a - Top ID new column #{self.pc + 1} named: {self.headers[self.pc].name}",
            )
        if len(successful) > 1:
            self.changelog_append(
                f"Cut and paste {len(successful)} IDs",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Cut and paste ID")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def paste_cut_child_all(self):
        if not self.cut or not self.selected_ID:
            return
        self.start_work(f"Pasting {len(self.cut)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.cut, 1):
            if self.cut_paste_all(dct["id"], dct["parent"], dct["hier"], self.selected_ID, sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.cut)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, set):
                for idk in v:
                    self.nodes[idk].cn[self.sort_later_dct["old_hier"]] = self.sort_node_cn(
                        self.nodes[idk].cn[self.sort_later_dct["old_hier"]],
                        self.sort_later_dct["old_hier"],
                    )
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            self.changelog_append_no_unsaved(
                "Cut and paste ID + children |",
                iid,
                f"Old parent: {self.nodes[dct['parent']].name if dct['parent'] else 'n/a - Top ID'} old column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                f"New parent: {self.nodes[self.selected_ID.lower()].name} new column #{self.pc + 1} named: {self.headers[self.pc].name}",
            )
        if len(successful) > 1:
            self.changelog_append(
                f"Cut and paste {len(successful)} IDs + children",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Cut and paste ID + children")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def paste_cut_sibling_all(self, redo_tree=True) -> bool:
        successful = []
        if not self.cut or not self.selected_ID:
            return successful
        if redo_tree:
            self.start_work(f"Pasting {len(self.cut)} IDs...")
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.cut, 1):
            if self.cut_paste_all(dct["id"], dct["parent"], dct["hier"], self.selected_PAR, sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.cut)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return successful
        for v in self.sort_later_dct.values():
            if v and isinstance(v, set):
                for idk in v:
                    self.nodes[idk].cn[self.sort_later_dct["old_hier"]] = self.sort_node_cn(
                        self.nodes[idk].cn[self.sort_later_dct["old_hier"]],
                        self.sort_later_dct["old_hier"],
                    )
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            if self.selected_PAR == "":
                self.changelog_append_no_unsaved(
                    "Cut and paste ID + children |",
                    iid,
                    f"Old parent: {self.nodes[dct['parent']].name if dct['parent'] else 'n/a - Top ID'} old column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                    f"New parent: n/a - Top ID new column #{self.pc + 1} named: {self.headers[self.pc].name}",
                )
            else:
                self.changelog_append_no_unsaved(
                    "Cut and paste ID + children |",
                    iid,
                    f"Old parent: {self.nodes[dct['parent']].name if dct['parent'] else 'n/a - Top ID'} old column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                    f"New parent: {self.nodes[self.selected_PAR.lower()].name} new column #{self.pc + 1} named: {self.headers[self.pc].name}",
                )
        if len(successful) > 1:
            self.changelog_append(
                f"Cut and paste {len(successful)} IDs + children",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Cut and paste ID + children")
        self.refresh_formatting(rows=self.refresh_rows)
        self.refresh_rows = set()
        if redo_tree:
            self.redo_tree_display()
            self.redraw_sheets()
            self.tree.scroll_to_item(iid.lower())
            self.tree.selection_set(tuple(dct["id"] for dct in successful))
            self.disable_paste()
            self.stop_work(self.get_tree_editor_status_bar_text())
        return successful

    def paste_cut_empty_all(self):
        if not self.cut:
            return
        self.start_work(f"Pasting {len(self.cut)} IDs...")
        successful = []
        self.sort_later_dct = None
        self.snapshot_paste_id()
        for i, dct in enumerate(self.cut, 1):
            if self.cut_paste_all(dct["id"], dct["parent"], dct["hier"], "", sort_later=True):
                successful.append(dct)
                if not i % 50:
                    self.C.status_bar.change_text(
                        f"Pasting {len(self.cut)} IDs... attempted: {i} | successful: {len(successful)} "
                    )
                    self.C.update()
        if not successful:
            self.unsuccessful_paste()
            return
        for v in self.sort_later_dct.values():
            if v and isinstance(v, set):
                for idk in v:
                    self.nodes[idk].cn[self.sort_later_dct["old_hier"]] = self.sort_node_cn(
                        self.nodes[idk].cn[self.sort_later_dct["old_hier"]],
                        self.sort_later_dct["old_hier"],
                    )
            if v and isinstance(v, tuple):  # v[0] is node name.lower() v[1] is hier, always sorting .cn[hier int]
                self.nodes[v[0]].cn[v[1]] = self.sort_node_cn(self.nodes[v[0]].cn[v[1]], v[1])
        for dct in successful:
            iid = self.nodes[dct["id"]].name
            self.changelog_append_no_unsaved(
                "Cut and paste ID + children |",
                iid,
                f"Old parent: {self.nodes[dct['parent']].name if dct['parent'] else 'n/a - Top ID'} old column #{dct['hier'] + 1} named: {self.headers[dct['hier']].name}",
                f"New parent: n/a - Top ID new column #{self.pc + 1} named: {self.headers[self.pc].name}",
            )
        if len(successful) > 1:
            self.changelog_append(
                f"Cut and paste {len(successful)} IDs + children",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Cut and paste ID + children")
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(iid.lower())
        self.tree.selection_set(tuple(dct["id"] for dct in successful))
        self.disable_paste()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def unsuccessful_paste(self):
        self.vs.pop()
        self.vp -= 1
        self.set_undo_label()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def copy_ID(self, iids: Sequence[str]) -> None | Literal["break"]:
        if not self.selected_ID:
            return
        if self.cut:
            self.cut = []
        self.copied = []
        first_iid = iids[0]
        first_iid_par = self.tree.parent(first_iid)
        h = int(self.pc)
        self.copied.append({"id": first_iid.lower(), "parent": first_iid_par.lower(), "hier": h})
        self.levels = defaultdict(list)
        self.get_par_lvls(h, first_iid.lower())
        first_iid_level = max(self.levels, default=0)
        tr = []
        for iid in islice(iids, 1, None):
            self.levels = defaultdict(list)
            self.get_par_lvls(h, iid.lower())
            iid_level = max(self.levels, default=0)
            if self.tree.parent(iid) == first_iid_par or iid_level == first_iid_level:
                self.copied.append(
                    {
                        "id": iid.lower(),
                        "parent": self.tree.parent(iid).lower(),
                        "hier": h,
                    }
                )
            else:
                tr.append(iid)
        if tr:
            self.tree.selection_remove(tr)
        self.enable_copy_paste()
        self.levels = defaultdict(list)
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        return "break"

    def cut_ids(self, iids: list[str] | None = None, status_bar=True):
        if not self.selected_ID:
            return
        if self.copied:
            self.copied = []
        self.cut = []
        if iids is None:
            iids = self.tree.selection()
        first_iid = iids[0]
        first_iid_par = self.tree.parent(first_iid)
        h = int(self.pc)
        self.cut.append({"id": first_iid.lower(), "parent": first_iid_par.lower(), "hier": h})
        self.levels = defaultdict(list)
        self.get_par_lvls(h, first_iid.lower())
        first_iid_level = max(self.levels, default=0)
        tr = []
        for iid in islice(iids, 1, None):
            self.levels = defaultdict(list)
            self.get_par_lvls(h, iid.lower())
            iid_level = max(self.levels, default=0)
            if self.tree.parent(iid) == first_iid_par or iid_level == first_iid_level:
                self.cut.append(
                    {
                        "id": iid.lower(),
                        "parent": self.tree.parent(iid).lower(),
                        "hier": h,
                    }
                )
            else:
                tr.append(iid)
        if tr:
            self.tree.selection_remove(tr)
        self.enable_cut_paste()
        self.levels = defaultdict(list)
        if status_bar:
            self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        return "break"

    def add_child_node(self):
        if not self.selected_ID:
            return
        sel = self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        if sel:
            popup = Add_Child_Or_Sibling_Id_Popup(
                self,
                "child",
                self.selected_ID,
                self.sheet.MT.data[sel[0]][self.ic],
                theme=self.C.theme,
            )
        else:
            popup = Add_Child_Or_Sibling_Id_Popup(self, "child", self.selected_ID, None, theme=self.C.theme)
        if not popup.result:
            return
        new_id = popup.result
        new_ik = new_id.lower()
        success = self.add(new_id, self.selected_ID)
        if not success:
            return
        self.changelog_append(
            "Add ID",
            f"Name: {new_id} Parent: {self.selected_ID} column #{self.pc + 1} named: {self.headers[self.pc].name}",
            "",
            "",
        )
        if self.tv_label_col != self.ic:
            new_label = popup.id_label
            if not new_label:
                new_label = new_id
            self.changelog_append(
                "Edit cell",
                f"ID: {new_id} column #{self.tv_label_col + 1} named: {self.headers[self.tv_label_col].name} with type: {self.headers[self.tv_label_col].type_}"
                f"{self.sheet.MT.data[self.rns[new_ik]][self.tv_label_col]}",
                f"{new_label}",
            )
            self.sheet.MT.data[self.rns[new_ik]][self.tv_label_col] = new_label
        self.disable_paste()
        self.redo_tree_display()
        self.refresh_dropdowns()
        self.tree.scroll_to_item(new_ik)
        self.tree.selection_set(new_ik)
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def add_sibling_node(self):
        if not self.selected_ID:
            return
        sel = self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        if sel:
            popup = Add_Child_Or_Sibling_Id_Popup(
                self,
                "sibling",
                self.selected_PAR,
                self.sheet.MT.data[sel[0]][self.ic],
                theme=self.C.theme,
            )
        else:
            popup = Add_Child_Or_Sibling_Id_Popup(self, "sibling", self.selected_PAR, None, theme=self.C.theme)
        if not popup.result:
            return
        new_id = popup.result
        new_ik = new_id.lower()
        success = self.add(new_id, self.selected_PAR)
        if not success:
            return
        if self.selected_PAR == "":
            self.changelog_append(
                "Add ID",
                f"Name: {new_id} Parent: n/a - Top ID column #{self.pc + 1} named: {self.headers[self.pc].name}",
                "",
                "",
            )
        else:
            self.changelog_append(
                "Add ID",
                f"Name: {new_id} Parent: {self.selected_PAR} column #{self.pc + 1} named: {self.headers[self.pc].name}",
                "",
                "",
            )
        if self.tv_label_col != self.ic:
            new_label = popup.id_label
            if not new_label:
                new_label = new_id
            self.changelog_append(
                "Edit cell",
                f"ID: {new_id} column #{self.tv_label_col + 1} named: {self.headers[self.tv_label_col].name} with type: {self.headers[self.tv_label_col].type_}",
                f"{self.sheet.MT.data[self.rns[new_ik]][self.tv_label_col]}",
                f"{new_label}",
            )
            self.sheet.MT.data[self.rns[new_ik]][self.tv_label_col] = new_label
        self.disable_paste()
        self.redo_tree_display()
        self.refresh_dropdowns()
        self.tree.scroll_to_item(new_ik)
        self.tree.selection_set(new_ik)
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def sheet_add_top_node(self):
        popup = Add_Top_Id_Popup(self, None, theme=self.C.theme)
        if not popup.result:
            return
        new_id = popup.result
        new_ik = new_id.lower()
        if not self.sheet.anything_selected(exclude_columns=True):
            insert_row = len(self.sheet.MT.data)
        else:
            insert_row = self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)[0]
        success = self.add(new_id, "", insert_row)
        if not success:
            return
        self.changelog_append(
            "Add ID",
            f"Name: {new_id} Parent: n/a - Top ID column #{self.pc + 1} named: {self.headers[self.pc].name}",
            "",
            "",
        )
        if self.tv_label_col != self.ic:
            new_label = popup.id_label
            if not new_label:
                new_label = new_id
            self.changelog_append(
                "Edit cell",
                f"ID: {new_id} column #{self.tv_label_col + 1} named: {self.headers[self.tv_label_col].name} with type: {self.headers[self.tv_label_col].type_}",
                f"{self.sheet.MT.data[self.rns[new_ik]][self.tv_label_col]}",
                f"{new_label}",
            )
            self.sheet.MT.data[self.rns[new_ik]][self.tv_label_col] = new_label
        self.disable_paste()
        self.redo_tree_display()
        self.refresh_dropdowns()
        if self.tree.exists(new_ik):
            self.tree.scroll_to_item(new_ik)
            self.tree.selection_set(new_ik)
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def add_top_node(self):
        sel = self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        if sel:
            popup = Add_Top_Id_Popup(self, self.sheet.MT.data[sel[0]][self.ic], theme=self.C.theme)
        else:
            popup = Add_Top_Id_Popup(self, None, theme=self.C.theme)
        if not popup.result:
            return
        new_id = popup.result
        new_ik = new_id.lower()
        success = self.add(new_id, "")
        if not success:
            return
        self.changelog_append(
            "Add ID",
            f"Name: {new_id} Parent: n/a - Top ID column #{self.pc + 1} named: {self.headers[self.pc].name}",
            "",
            "",
        )
        if self.tv_label_col != self.ic:
            new_label = popup.id_label
            if not new_label:
                new_label = new_id
            self.changelog_append(
                "Edit cell",
                f"ID: {new_id} column #{self.tv_label_col + 1} named: {self.headers[self.tv_label_col].name} with type: {self.headers[self.tv_label_col].type_}",
                f"{self.sheet.MT.data[self.rns[new_ik]][self.tv_label_col]}",
                f"{new_label}",
            )
            self.sheet.MT.data[self.rns[new_ik]][self.tv_label_col] = new_label
        self.disable_paste()
        self.redo_tree_display()
        self.refresh_dropdowns()
        self.tree.scroll_to_item(new_ik)
        self.tree.selection_set(new_ik)
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def sheet_rename_node(self):
        rn = self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)[0]
        id_ = self.sheet.MT.data[rn][self.ic]
        ik = id_.lower()
        popup = Rename_Id_Popup(self, id_, theme=self.C.theme)
        if not popup.result:
            return
        tree_sel = self.tree.selection()[0] if self.tree.selection() else False
        success = self.change_ID_name(id_, popup.result)
        if not success:
            return
        self.changelog_append(
            "Rename ID",
            id_,
            id_,
            str(popup.result),
        )
        new_ik = popup.result.lower()
        if ik in self.tagged_ids:
            self.tagged_ids.discard(ik)
            self.tagged_ids.add(new_ik)
            self.reset_tagged_ids_dropdowns()
        self.disable_paste()
        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        if tree_sel:
            try:
                self.tree.scroll_to_item(tree_sel)
                self.tree.selection_set(tree_sel)
            except Exception:
                self.tree.scroll_to_item(popup.result.lower())
                self.tree.selection_set(popup.result.lower())
        else:
            self.move_tree_pos()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def rename_node(self):
        if not self.selected_ID:
            return
        ik = self.selected_ID.lower()
        popup = Rename_Id_Popup(self, self.selected_ID, theme=self.C.theme)
        if not popup.result:
            return
        success = self.change_ID_name(self.selected_ID, popup.result)
        if not success:
            return
        self.changelog_append(
            "Rename ID",
            self.selected_ID,
            self.selected_ID,
            f"{popup.result}",
        )
        new_ik = popup.result.lower()
        if ik in self.tagged_ids:
            self.tagged_ids.discard(ik)
            self.tagged_ids.add(new_ik)
            self.reset_tagged_ids_dropdowns()
        self.disable_paste()
        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        self.refresh_formatting(rows=self.refresh_rows)
        self.redo_tree_display()
        self.refresh_rows = set()
        self.redraw_sheets()
        self.tree.scroll_to_item(popup.result.lower())
        self.tree.selection_set(popup.result.lower())
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def del_id(self, iids: Sequence[str]):
        if not iids:
            return
        self.start_work(f"Deleting {len(iids)} IDs")
        self.snapshot_delete_ids()
        to_del = []
        self.refresh_rows = set()
        for iid in iids:
            par = self.nodes[self.nodes[iid].ps[self.pc]].name if self.nodes[iid].ps[self.pc] else ""
            to_del = self._del_id_core(iid, to_del, snapshot=True)
            self.changelog_append_no_unsaved(
                "Delete ID |",
                f"ID: {self.sheet.data[self.rns[iid]][self.ic]} parent: {par if par else 'n/a - Top ID'} column #{self.pc + 1} named: {self.headers[self.pc].name}",
                "",
                "",
            )
        if len(iids) > 1:
            self.changelog_append(
                f"Delete {len(iids)} IDs",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Delete ID")
        if to_del:
            self.sheet.del_rows(map(self.rns.get, to_del), redraw=False)
            self.sheet.deselect("all", redraw=False)
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            self.refresh_formatting(rows=map(self.rns.__getitem__, self.refresh_rows))
        self.disable_paste()
        self.move_tree_pos()
        self.reset_tagged_ids_dropdowns()
        self.rehighlight_tagged_ids()
        self.redo_tree_display()
        self.redraw_sheets()
        self.focus_tree()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def del_id_all(self, iids: Iterator[str] | None = None) -> None:
        if not iids:
            iids = self.tree.selection()
        self.start_work(f"Deleting {len(iids)} IDs")
        self.snapshot_delete_ids()
        to_del = []
        self.refresh_rows = set()
        for iid in iids:
            to_del = self._del_id_all_core(iid, to_del, snapshot=True)
            self.changelog_append_no_unsaved(
                "Delete ID from all hierarchies |",
                f"{self.sheet.data[self.rns[iid]][self.ic]}",
                "",
                "",
            )
        if len(to_del) > 1:
            self.changelog_append(
                f"Deleted {len(to_del)} IDs from all hierarchies",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Delete ID from all hierarchies")
        if to_del:
            self.sheet.del_rows(map(self.rns.get, to_del), redraw=False)
            self.sheet.deselect("all", redraw=False)
            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
            self.refresh_formatting(rows=map(self.rns.__getitem__, self.refresh_rows))
        self.disable_paste()
        self.move_tree_pos()
        self.reset_tagged_ids_dropdowns()
        self.rehighlight_tagged_ids()
        self.redo_tree_display()
        self.redraw_sheets()
        self.focus_tree()
        self.stop_work(self.get_tree_editor_status_bar_text())

    def del_id_orphan(self, event=None):
        if not self.selected_ID:
            return
        self.changelog_append(
            "Delete ID, orphan children",
            f"ID: {self.selected_ID} parent: {self.selected_PAR if self.selected_PAR else 'n/a - Top ID'} column #{self.pc + 1} named: {self.headers[self.pc].name}",
            "",
            "",
        )
        self.snapshot_delete_ids()
        self.sheet.deselect("all", redraw=False)
        self.disable_paste()
        self._del_id_orphan_core(self.selected_ID, self.selected_PAR if self.selected_PAR else "")
        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        self.refresh_formatting(rows=map(self.rns.__getitem__, self.refresh_rows))
        self.move_tree_pos()
        self.reset_tagged_ids_dropdowns()
        self.rehighlight_tagged_ids()
        self.redo_tree_display()
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        self.focus_tree()

    def del_id_children(self):
        if not self.selected_ID:
            return
        ik = self.selected_ID.lower()
        self.start_work(f"Deleting {ik} and all children...")
        self.changelog_append(
            "Delete ID + all children",
            f"ID: {self.selected_ID} parent: {self.selected_PAR if self.selected_PAR else 'n/a - Top ID'} column #{self.pc + 1} named: {self.headers[self.pc].name}",
            "",
            "",
        )
        self.snapshot_delete_ids()
        self.sheet.deselect("all", redraw=False)
        self.disable_paste()
        self._del_id_children_core(self.selected_ID, self.selected_PAR if self.selected_PAR else "")
        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        self.refresh_formatting(rows=map(self.rns.__getitem__, self.refresh_rows))
        self.redo_tree_display()
        self.move_tree_pos()
        self.reset_tagged_ids_dropdowns()
        self.rehighlight_tagged_ids()
        self.redraw_sheets()
        self.stop_work(self.get_tree_editor_status_bar_text())
        self.focus_tree()

    def del_id_children_all(self):
        if not self.selected_ID:
            return
        ik = self.selected_ID.lower()
        self.start_work(f"Deleting {ik} and all children...")
        self.changelog_append(
            "Delete ID + all children from all hierarchies",
            f"ID: {self.selected_ID} parent: {self.selected_PAR if self.selected_PAR else 'n/a - Top ID'} column #{self.pc + 1} named: {self.headers[self.pc].name}",
            "",
            "",
        )
        self.snapshot_delete_ids()
        self.sheet.deselect("all", redraw=False)
        self.disable_paste()
        self._del_id_children_all_core(self.selected_ID, self.selected_PAR if self.selected_PAR else "")
        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        self.redo_tree_display()
        self.move_tree_pos()
        self.reset_tagged_ids_dropdowns()
        self.rehighlight_tagged_ids()
        self.redraw_sheets()
        self.stop_work(self.get_tree_editor_status_bar_text())
        self.focus_tree()

    def del_id_all_orphan(self):
        if not self.selected_ID:
            return
        self.changelog_append(
            "Delete ID from all hierarchies, orphan children",
            self.selected_ID,
            "",
            "",
        )
        self.snapshot_delete_ids()
        self.sheet.deselect("all", redraw=False)
        self.disable_paste()
        self._del_id_all_orphan_core(self.selected_ID)
        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        self.refresh_formatting(rows=map(self.rns.__getitem__, self.refresh_rows))
        self.redo_tree_display()
        self.move_tree_pos()
        self.reset_tagged_ids_dropdowns()
        self.rehighlight_tagged_ids()
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        self.focus_tree()

    def reselect_sheet_sel(self, boxes):
        self.sheet.deselect("all")
        if boxes[0]:
            self.sheet.set_currently_selected(boxes[0][0], boxes[0][1])
        for box in boxes[1]:
            r1, c1, r2, c2 = box[0]
            self.sheet.create_selection_box(r1, c1, r2, c2, box[1])

    def get_sheet_sel(self):
        return (
            self.sheet.get_currently_selected(),
            self.sheet.get_all_selection_boxes_with_types(),
        )

    def clear_copied_details(self):
        self.disable_paste()
        self.tree_rc_menu_single_row_paste.entryconfig("Paste details", state="disabled")
        self.sheet_rc_menu_single_row.entryconfig("Paste details", state="disabled")
        self.sheet_rc_menu_multi_row.entryconfig("Paste details", state="disabled")
        self.tree_rc_menu_multi_row.entryconfig("Paste details", state="disabled")
        self.copied_details = {"copied": [], "id": ""}
        return "break"

    def tree_sheet_edit_detail(self):
        if self.tree.has_focus():
            selected = self.tree.selected
            if not selected:
                return
            rn = self.rns[self.tree.rowitem(selected.row)]
        else:
            selected = self.sheet.selected
            if not selected:
                return
            rn = selected.row
        col = selected.column
        currentdetail = self.sheet.MT.data[rn][col]
        heading = self.headers[col].name
        ID = self.sheet.MT.data[rn][self.ic]
        if self.headers[col].type_ in ("ID", "Parent"):
            popup = Edit_Detail_Text_Popup(self, ID, heading, currentdetail, theme=self.C.theme)
        else:
            validation = self.headers[col].validation
            if validation:
                set_value = currentdetail if currentdetail in set(validation) else validation[0]
                if self.headers[col].type_ == "Text":
                    popup = Edit_Detail_Text_Popup(
                        self,
                        ID,
                        heading,
                        currentdetail,
                        validation_values=validation,
                        set_value=set_value,
                        theme=self.C.theme,
                    )
                elif self.headers[col].type_ == "Number":
                    popup = Edit_Detail_Number_Popup(
                        self,
                        ID,
                        heading,
                        currentdetail,
                        validation_values=validation,
                        set_value=set_value,
                        theme=self.C.theme,
                    )
                elif self.headers[col].type_ == "Date":
                    popup = Edit_Detail_Date_Popup(
                        self,
                        ID,
                        heading,
                        currentdetail,
                        self.DATE_FORM,
                        validation_values=validation,
                        set_value=set_value,
                        theme=self.C.theme,
                    )
            else:
                if self.headers[col].type_ == "Text":
                    popup = Edit_Detail_Text_Popup(self, ID, heading, currentdetail, theme=self.C.theme)
                elif self.headers[col].type_ == "Number":
                    popup = Edit_Detail_Number_Popup(self, ID, heading, currentdetail, theme=self.C.theme)
                elif self.headers[col].type_ == "Date":
                    popup = Edit_Detail_Date_Popup(
                        self,
                        ID,
                        heading,
                        currentdetail,
                        self.DATE_FORM,
                        theme=self.C.theme,
                    )
        if popup.result:
            self.tree_sheet_edit_table(
                event=DotDict(
                    sheetname="sheet", value=popup.saved_string, loc=(rn, col), data={(rn, col): popup.saved_string}
                )
            )

    def sheet_copy_details(self):
        rn = self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)[0]
        ik = self.sheet.MT.data[rn][self.ic].lower()
        self.copied_details["copied"] = self.sheet.MT.data[rn].copy()
        self.copied_details["id"] = ik
        self.tree_rc_menu_single_row_paste.entryconfig("Paste details", state="normal")
        self.sheet_rc_menu_single_row.entryconfig("Paste details", state="normal")
        self.sheet_rc_menu_multi_row.entryconfig("Paste details", state="normal")
        self.tree_rc_menu_multi_row.entryconfig("Paste details", state="normal")
        s, writer = str_io_csv_writer(dialect=csv.excel_tab)
        writer.writerow(self.sheet.MT.data[rn])
        to_clipboard(widget=self, s=s.getvalue().rstrip())

    def copy_details(self):
        if not self.selected_ID:
            return
        rn = self.rns[self.selected_ID.lower()]
        self.copied_details["copied"] = self.sheet.MT.data[rn].copy()
        self.copied_details["id"] = self.selected_ID.lower()
        self.tree_rc_menu_single_row_paste.entryconfig("Paste details", state="normal")
        self.sheet_rc_menu_single_row.entryconfig("Paste details", state="normal")
        self.sheet_rc_menu_multi_row.entryconfig("Paste details", state="normal")
        self.tree_rc_menu_multi_row.entryconfig("Paste details", state="normal")
        s, writer = str_io_csv_writer(dialect=csv.excel_tab)
        writer.writerow(self.sheet.MT.data[rn])
        to_clipboard(widget=self, s=s.getvalue().rstrip())

    def sheet_paste_details(self):
        self.snapshot_ctrl_x_v_del_key()
        cells_changed, idcol_hiers = 0, set(self.hiers) | {self.ic}
        selected_rows = self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        rows, cols = set(), set()
        for rn in selected_rows:
            ID = self.sheet.MT.data[rn][self.ic]
            for c, e in enumerate(self.copied_details["copied"]):
                if c not in idcol_hiers and self.sheet.MT.data[rn][c] != e:
                    self.changelog_append_no_unsaved(
                        "Edit cell |",
                        f"ID: {ID} column #{c + 1} named: {self.headers[c].name} with type: {self.headers[c].type_}",
                        f"{self.sheet.MT.data[rn][c]}",
                        f"{e}",
                    )
                    self.vs[-1]["cells"][(rn, c)] = f"{self.sheet.MT.data[rn][c]}"
                    self.sheet.MT.data[rn][c] = e
                    rows.add(rn)
                    cols.add(c)
                    cells_changed += 1
        self.refresh_formatting(rows=rows, columns=cols)
        for rn in rows:
            self.refresh_tree_item(self.sheet.MT.data[rn][self.ic])
        if not cells_changed:
            self.vp -= 1
            self.set_undo_label()
            self.vs.pop()
            self.disable_paste()
            self.redraw_sheets()
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        if cells_changed > 1:
            self.changelog_append(
                f"Edit {cells_changed} cells",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Edit cell")
        self.disable_paste()
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def paste_details(self):
        self.snapshot_ctrl_x_v_del_key()
        cells_changed, idcol_hiers = 0, set(self.hiers) | {self.ic}
        rows, cols = set(), set()
        for iid in self.tree.selection():
            ID = self.nodes[iid].name
            rn = self.rns[iid]
            for c, e in enumerate(self.copied_details["copied"]):
                if c not in idcol_hiers and self.sheet.MT.data[rn][c] != e:
                    self.changelog_append_no_unsaved(
                        "Edit cell |",
                        f"ID: {ID} column #{c + 1} named: {self.headers[c].name} with type: {self.headers[c].type_}",
                        f"{self.sheet.MT.data[rn][c]}",
                        f"{e}",
                    )
                    self.vs[-1]["cells"][(rn, c)] = f"{self.sheet.MT.data[rn][c]}"
                    self.sheet.MT.data[rn][c] = e
                    rows.add(rn)
                    cols.add(c)
                    cells_changed += 1
        self.refresh_formatting(rows=rows, columns=cols)
        for rn in rows:
            self.refresh_tree_item(self.sheet.MT.data[rn][self.ic])
        if not cells_changed:
            self.vp -= 1
            self.set_undo_label()
            self.vs.pop()
            self.disable_paste()
            self.redraw_sheets()
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        if cells_changed > 1:
            self.changelog_append(
                f"Edit {cells_changed} cells",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Edit cell")
        self.disable_paste()
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def sheet_del_all_details(self):
        self.snapshot_ctrl_x_v_del_key()
        cells_changed, idcol_hiers = 0, set(self.hiers) | {self.ic}
        rows = set()
        cols = set()
        for rn in self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True):
            ID = self.sheet.MT.data[rn][self.ic]
            for c in range(len(self.sheet.MT.data[rn])):
                if c not in idcol_hiers and self.sheet.MT.data[rn][c] != "":
                    self.changelog_append_no_unsaved(
                        "Edit cell |",
                        f"ID: {ID} column #{c + 1} named: {self.headers[c].name} with type: {self.headers[c].type_}",
                        f"{self.sheet.MT.data[rn][c]}",
                        "",
                    )
                    self.vs[-1]["cells"][(rn, c)] = f"{self.sheet.MT.data[rn][c]}"
                    self.sheet.MT.data[rn][c] = ""
                    rows.add(rn)
                    cols.add(c)
                    cells_changed += 1
        self.refresh_formatting(rows=rows, columns=cols)
        for rn in rows:
            self.refresh_tree_item(self.sheet.MT.data[rn][self.ic])
        if not cells_changed:
            self.vp -= 1
            self.set_undo_label()
            self.vs.pop()
            self.disable_paste()
            self.redraw_sheets()
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        if cells_changed > 1:
            self.changelog_append(
                f"Edit {cells_changed} cells",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Edit cell")
        self.disable_paste()
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def del_all_details(self):
        self.snapshot_ctrl_x_v_del_key()
        cells_changed, idcol_hiers = 0, set(self.hiers) | {self.ic}
        rows = set()
        cols = set()
        for iid in self.tree.selection():
            ID = self.nodes[iid].name
            rn = self.rns[iid]
            for c in range(len(self.sheet.MT.data[rn])):
                if c not in idcol_hiers and self.sheet.MT.data[rn][c] != "":
                    self.changelog_append_no_unsaved(
                        "Edit cell |",
                        f"ID: {ID} column #{c + 1} named: {self.headers[c].name} with type: {self.headers[c].type_}",
                        f"{self.sheet.MT.data[rn][c]}",
                        "",
                    )
                    self.vs[-1]["cells"][(rn, c)] = f"{self.sheet.MT.data[rn][c]}"
                    self.sheet.MT.data[rn][c] = ""
                    rows.add(rn)
                    cols.add(c)
                    cells_changed += 1
        self.refresh_formatting(rows=rows, columns=cols)
        for rn in rows:
            self.refresh_tree_item(self.sheet.MT.data[rn][self.ic])
        if not cells_changed:
            self.vp -= 1
            self.set_undo_label()
            self.vs.pop()
            self.disable_paste()
            self.redraw_sheets()
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        if cells_changed > 1:
            self.changelog_append(
                f"Edit {cells_changed} cells",
                "",
                "",
                "",
            )
        else:
            self.changelog_singular("Edit cell")
        self.disable_paste()
        self.redraw_sheets()
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def tag_ids(
        self,
        event=None,
        selection: Iterator[str] | None = None,
        toggle: bool = True,
        do_tree: bool = True,
    ):
        if selection is None:
            if self.tree_has_focus:
                selection = self.tree.selection(cells=True)
            elif self.sheet_has_focus:
                selection = (
                    self.sheet.data[r][self.ic].lower() for r in self.sheet.get_selected_rows(get_cells_as_rows=True)
                )
            if not selection:
                return
        for iid in selection:
            if (ik := iid.lower()) not in self.rns:
                self.tagged_ids.discard(ik)
                continue
            rn = self.rns[ik]
            if toggle and ik in self.tagged_ids:
                self.tagged_ids.discard(ik)
                self.sheet.dehighlight_cells(
                    row=rn,
                    canvas="row_index",
                    redraw=False,
                )
                if do_tree and self.tree.exists(ik):
                    self.tree.dehighlight_cells(
                        row=self.tree.itemrow(ik),
                        canvas="row_index",
                        redraw=False,
                    )
            else:
                self.tagged_ids.add(ik)
                self.sheet.highlight_cells(
                    row=rn,
                    bg="orange",
                    fg="black",
                    canvas="row_index",
                    redraw=False,
                )
                if do_tree and self.tree.exists(ik):
                    self.tree.highlight_cells(
                        row=self.tree.itemrow(ik),
                        bg="orange",
                        fg="black",
                        canvas="row_index",
                        redraw=False,
                    )
        self.reset_tagged_ids_dropdowns()
        self.redraw_sheets()

    def tree_sheet_align(self, align):
        boxes = self.sheet.boxes if self.sheet.has_focus() else self.tree.boxes
        for box in boxes:
            if box.type_ == "columns":
                self.sheet.align(
                    f"{_n2a(box.coords.from_c)}:{_n2a(box.coords.upto_c - 1)}",
                    align=align,
                )
                self.tree.align(
                    f"{_n2a(box.coords.from_c)}:{_n2a(box.coords.upto_c - 1)}",
                    align=align,
                )

    def untag_id(self, ik):
        if ik in self.tagged_ids:
            self.tagged_ids.discard(ik)
            self.sheet.dehighlight_cells(row=self.rns[ik], canvas="row_index")
            if self.tree.exists(ik):
                self.tree.dehighlight_rows(self.tree.itemrow(ik))

    def clear_tagged_ids(self, event=None):
        self.tagged_ids = set()
        self.reset_tagged_ids_dropdowns()
        self.sheet.dehighlight_cells(canvas="row_index", all_=True, redraw=True)
        self.redo_tree_display()

    def reset_tagged_ids_dropdowns(self, event=None):
        res = []
        for ik in tuple(self.tagged_ids):
            if ik in self.nodes:
                res.append(self.nodes[ik].name)
            else:
                self.tagged_ids.discard(ik)
        self.tree_tagged_ids_dropdown["values"] = res
        self.sheet_tagged_ids_dropdown["values"] = res
        if self.tagged_ids:
            self.sheet_tagged_ids_dropdown.set_my_value(res[0])
            self.tree_tagged_ids_dropdown.set_my_value(res[0])
        else:
            self.sheet_tagged_ids_dropdown.set_my_value("")
            self.tree_tagged_ids_dropdown.set_my_value("")
        return "break"

    def rehighlight_tagged_ids(self, event=None):
        self.sheet.dehighlight_cells(canvas="row_index", all_=True, redraw=False)
        self.tree.dehighlight_cells(canvas="row_index", all_=True, redraw=False)
        for ik in tuple(self.tagged_ids):
            try:
                self.sheet.highlight_cells(
                    row=self.rns[ik],
                    bg="orange",
                    fg="black",
                    canvas="row_index",
                    redraw=False,
                )
            except Exception:
                self.tagged_ids.discard(ik)
        for ik in filter(self.tree.RI.rns.__contains__, self.tagged_ids):
            self.tree.highlight_cells(
                row=self.tree.itemrow(ik),
                bg="orange",
                fg="black",
                canvas="row_index",
                redraw=False,
            )
        return "break"

    def go_to_treeview_id_finder(self, ik: str):
        hs = [self.headers[h].name for h, p in self.nodes[ik].ps.items() if p is not None]
        if len(hs) > 1:
            popup = Treeview_Id_Finder(self, hs, theme=self.C.theme)
            if not popup.GO:
                return
            selected = popup.selected
        else:
            selected = hs[0]
        self.switch_displayed.set(f"{selected}")
        self.switch_hier()
        self.tree.scroll_to_item(ik)
        self.tree.selection_set(ik)

    def tree_go_to_tagged_id(self, event=None):
        if not (ik := self.tree_tagged_ids_dropdown.get_my_value().lower()):
            return
        if ik in self.nodes:
            if self.tree.exists(ik):
                self.tree.scroll_to_item(ik)
                self.tree.selection_set(ik)
            else:
                self.go_to_treeview_id_finder(ik)
        else:
            Error(self, f"{ik} no longer exists and has been removed from tagged ids.", theme=self.C.theme)
            self.discard_tagged_id(ik)

    def discard_tagged_id(self, ik):
        self.tagged_ids.discard(ik)
        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        self.reset_tagged_ids_dropdowns()
        self.rehighlight_tagged_ids()
        self.redraw_sheets()

    def sheet_go_to_tagged_id(self, event=None):
        if not (ik := self.sheet_tagged_ids_dropdown.get_my_value().lower()):
            return
        if ik in self.rns:
            self.sheet.select_row(self.rns[ik])
            self.sheet.see(row=self.rns[ik], keep_xscroll=True)
            if self.tree.exists(ik):
                self.tree.scroll_to_item(ik)
                self.tree.selection_set(ik)
        else:
            self.discard_tagged_id(ik)

    def find_next_main(self, event=None):
        if self.find_popup:
            self.find_popup.find_next()
        elif self.search_results and not self.sheet_search_results:
            self.find_next()
        elif self.sheet_search_results and not self.search_results:
            self.sheet_find_next()
        elif self.sheet_search_results and self.search_results and self.tree_has_focus:
            self.find_next()
        elif self.sheet_search_results and self.search_results and self.sheet_has_focus:
            self.sheet_find_next()

    def find_next(self, event=None):
        if self.search_dropdown["values"]:
            idx = self.search_dropdown.current()
            if idx + 1 == len(self.search_dropdown["values"]):
                idx = 0
            else:
                idx += 1
            self.search_dropdown.current(idx)
            self.show_search_result()

    def sheet_find_next(self, event=None):
        if self.sheet_search_dropdown["values"]:
            idx = self.sheet_search_dropdown.current()
            if idx + 1 == len(self.sheet_search_dropdown["values"]):
                idx = 0
            else:
                idx += 1
            self.sheet_search_dropdown.current(idx)
            self.sheet_show_search_result()

    def show_search_result(self, event=None):
        result = self.search_results[self.search_dropdown.current()]
        if not self.check_search_result(result):
            return
        if self.pc != result.hierarchy:
            self.switch_hier(hier=result.hierarchy)
        self.tree.scroll_to_item(result.iid, redraw=True)
        self.tree.selection_set(result.iid)
        self.tree.set_currently_selected(column=result.column, item=self.tree.selected.fill_iid)
        self.tree.see(column=result.column, keep_yscroll=True)
        self.focus_tree()

    def sheet_show_search_result(self, event=None):
        result = self.sheet_search_results[self.sheet_search_dropdown.current()]
        if not self.check_search_result(result):
            return
        self.sheet.select_row(row=self.rns[result.iid])
        self.sheet.see(row=self.rns[result.iid], column=result.column, redraw=True)
        self.sheet.set_currently_selected(column=result.column, item=self.sheet.selected.fill_iid)
        self.focus_sheet()

    def check_search_result(self, result: SearchResult) -> bool:
        result_ok = True
        if result.iid not in self.rns:
            result_ok = False
        else:
            sheet_rn = self.rns[result.iid]
            try:
                sheet_cell = self.sheet.data[sheet_rn][result.column].lower()
                if (
                    (result.hierarchy not in self.hiers)
                    or (result.exact and sheet_cell != result.term)
                    or (not result.exact and result.term not in sheet_cell)
                ):
                    result_ok = False
            except Exception:
                result_ok = False
        if not result_ok:
            Error(
                self,
                "Search result not found, refresh the search. Data may have been modified after searching.",
                theme=self.C.theme,
            )
        return result_ok

    def sheet_rc_tv_label(self, event=None):
        if (
            self.tree.has_focus()
            and self.tree.selected
            and self.tree.selected.type_ == "columns"
            or self.sheet.has_focus()
            and self.sheet.selected
            and self.sheet.selected.type_ == "columns"
        ):
            selected_col = self.tree.selected.column
        if self.headers[selected_col].type_ == "Parent":
            Error(
                self,
                "Cannot select Parent column as Treeview label",
                theme=self.C.theme,
            )
            return
        self.tv_label_col = selected_col
        self.save_info_get_saved_info()
        self.redo_tree_display()

    def set_all_col_widths(self, event=None):
        self.tree.set_all_cell_sizes_to_text()
        self.sheet.set_all_cell_sizes_to_text()

    def toggle_auto_resize_index(self, enabled):
        self.tree.set_options(auto_resize_row_index=enabled)
        self.sheet.set_options(auto_resize_row_index=enabled)
        self.auto_resize_indexes = enabled

    def toggle_mirror(self, enabled, select_row=True):
        self.mirror_var = enabled
        if select_row and self.mirror_var and self.tree.selection():
            self.go_to_row()

    def focus_tree(self):
        self.sheet.focus_set()
        self.sheet_focus_leave()
        self.tree_focus_enter()
        self.tree.focus_set()

    def focus_sheet(self):
        self.tree.focus_set()
        self.sheet_focus_enter()
        self.tree_focus_leave()
        self.sheet.focus_set()

    def tree_focus_leave(self, event=None):
        self.l_frame.config(
            highlightbackground=themes[self.C.theme].table_bg,
            highlightcolor=themes[self.C.theme].table_bg,
        )
        self.l_frame.update_idletasks()

    def tree_focus_enter(self, event=None):
        if self.get_display_option() in ("50/50", "adjustable"):
            self.l_frame.config(
                highlightbackground=themes[self.C.theme].table_selected_box_cells_fg,
                highlightcolor=themes[self.C.theme].table_selected_box_cells_fg,
                highlightthickness=2,
            )
        else:
            self.tree_focus_leave()
            self.l_frame.config(highlightthickness=0)
        self.l_frame.update_idletasks()
        self.tree_has_focus = True
        self.sheet_has_focus = False

    def sheet_focus_leave(self, event=None):
        self.r_frame.config(
            highlightbackground=themes[self.C.theme].table_bg,
            highlightcolor=themes[self.C.theme].table_bg,
        )
        self.r_frame.update_idletasks()

    def sheet_focus_enter(self, event=None):
        if self.get_display_option() in ("50/50", "adjustable"):
            self.r_frame.config(
                highlightbackground=themes[self.C.theme].table_selected_box_cells_fg,
                highlightcolor=themes[self.C.theme].table_selected_box_cells_fg,
                highlightthickness=2,
            )
        else:
            self.sheet_focus_leave()
            self.r_frame.config(highlightthickness=0)
        self.r_frame.update_idletasks()
        self.tree_has_focus = False
        self.sheet_has_focus = True

    def details_focus_set(self, event=None):
        if self.show_ids_details_dropdown.get_my_value() == "Treeview selection":
            self.focus_tree()
        else:
            self.focus_sheet()

    def show_ids_details_sheet(self, event=None):
        sel = self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        if sel:
            View_Id_Popup(
                self,
                ids_row={"row": self.sheet.MT.data[sel[0]], "rn": sel[0]},
                theme=self.C.theme,
            )

    def show_ids_details_tree(self, event=None):
        if self.selected_ID:
            rn = self.rns[self.selected_ID.lower()]
            View_Id_Popup(
                self,
                ids_row={"row": self.sheet.MT.data[rn], "rn": rn},
                theme=self.C.theme,
            )

    def show_ids_full_info_sheet(self, event=None):
        sel = self.sheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        if sel:
            ik = self.sheet.MT.data[sel[0]][self.ic].lower()
            Text_Popup(self, self.details(ik), theme=self.C.theme)
        else:
            Error(self, "Select an ID in the sheet\nTo display the sheet go to View > Layout", theme=self.C.theme)

    def show_ids_full_info_tree(self, event=None):
        if self.selected_ID:
            Text_Popup(self, self.details(self.selected_ID.lower()), theme=self.C.theme)
        else:
            Error(self, "Select an ID in the tree", theme=self.C.theme)

    def show_warnings(self, filepath=None, sheetname=None, show_regardless=False):
        if filepath and sheetname:
            self.warnings_filepath = filepath
            self.warnings_sheet = sheetname
        top = "".join(
            (
                "File opened: ",
                self.warnings_filepath,
                "\nSheet opened: ",
                self.warnings_sheet,
                "\n\n",
            )
        )
        if show_regardless:
            if self.warnings:
                Text_Popup(
                    self,
                    "".join((top, warnings_header, "\n", "\n".join(self.warnings))),
                    theme=self.C.theme,
                )
            else:
                Text_Popup(
                    self,
                    "".join((top, warnings_header, "\n - NO WARNINGS TO DISPLAY - ")),
                    theme=self.C.theme,
                )
        else:
            if self.warnings:
                Text_Popup(
                    self,
                    "".join((top, warnings_header, "\n", "\n".join(self.warnings))),
                    theme=self.C.theme,
                )

    def show_changelog(self, event=None):
        if not isinstance(event, str) or event == "specific":
            Changelog_Popup(self, theme=self.C.theme)
        else:
            self.start_work("Opened save dialog")
            newfile = filedialog.asksaveasfilename(
                parent=self,
                title="Save changes as",
                filetypes=[
                    ("CSV File", ".csv"),
                    ("TSV File", ".tsv"),
                    ("Excel file", ".xlsx"),
                    ("JSON File", ".json"),
                ],
                defaultextension=".csv",
                confirmoverwrite=True,
            )
            if not newfile:
                self.stop_work()
                return
            newfile = os.path.normpath(newfile)
            if not newfile.lower().endswith((".csv", ".xlsx", ".json", ".tsv")):
                self.stop_work("Can only save .csv/.xlsx/.json file types")
                return
            self.C.status_bar.change_text("Saving changelog...")
            if event == "all":
                try:
                    if newfile.lower().endswith(".xlsx"):
                        self.C.wb = Workbook(write_only=True)
                        ws = self.C.wb.create_sheet(title="Changelog")
                        ws.append(xlsx_changelog_header(ws))
                        for row in self.changelog:
                            ws.append((e if e else None for e in row))
                        self.C.wb.save(newfile)
                        self.C.try_to_close_workbook()
                    elif newfile.lower().endswith((".csv", ".tsv")):
                        with open(newfile, "w", newline="", encoding="utf-8") as fh:
                            writer = csv.writer(
                                fh,
                                dialect=csv.excel_tab if newfile.lower().endswith(".tsv") else csv.excel,
                                lineterminator="\n",
                            )
                            writer.writerow(changelog_header)
                            writer.writerows(self.changelog)
                    elif newfile.lower().endswith(".json"):
                        with open(newfile, "w", newline="") as fh:
                            fh.write(
                                json.dumps(
                                    full_sheet_to_dict(
                                        changelog_header,
                                        self.changelog,
                                        include_headers=True,
                                        format_=self.json_format,
                                    ),
                                    indent=4,
                                )
                            )
                except Exception as error_msg:
                    self.C.try_to_close_workbook()
                    self.stop_work(f"Error saving file: {error_msg}")
                    return
                self.stop_work("Success! Changelog saved")
            elif event == "sheet":
                from_row = len(self.changelog) - self.sheet_changes
                to_row = len(self.changelog)
                try:
                    if newfile.lower().endswith(".xlsx"):
                        self.C.wb = Workbook(write_only=True)
                        ws = self.C.wb.create_sheet(title="Changelog")
                        ws.append(xlsx_changelog_header(ws))
                        if self.sheet_changes:
                            for row in islice(self.changelog, from_row, to_row):
                                ws.append((e if e else None for e in row))
                        self.C.wb.save(newfile)
                        self.C.try_to_close_workbook()
                    elif newfile.lower().endswith((".csv", ".tsv")):
                        with open(newfile, "w", newline="", encoding="utf-8") as fh:
                            writer = csv.writer(
                                fh,
                                dialect=csv.excel_tab if newfile.lower().endswith(".tsv") else csv.excel,
                                lineterminator="\n",
                            )
                            writer.writerow(changelog_header)
                            if self.sheet_changes:
                                writer.writerows(islice(self.changelog, from_row, to_row))
                    elif newfile.lower().endswith(".json"):
                        with open(newfile, "w", newline="") as fh:
                            fh.write(
                                json.dumps(
                                    full_sheet_to_dict(
                                        changelog_header,
                                        self.changelog[from_row:to_row] if self.sheet_changes else [],
                                        include_headers=True,
                                        format_=self.json_format,
                                    ),
                                    indent=4,
                                )
                            )
                except Exception as error_msg:
                    self.C.try_to_close_workbook()
                    self.stop_work(f"Error saving file: {error_msg}")
                    return
                self.stop_work("Success! Changelog saved")

    def go_to_row(self):
        if not self.selected_ID:
            return
        rn, sheet_rn = self.rns[self.selected_ID.lower()], self.sheet.selected
        if not sheet_rn or rn != sheet_rn.row or not self.sheet.cell_visible(sheet_rn.row, sheet_rn.column):
            self.sheet.select_row(rn, redraw=False)
            self.sheet.see(row=rn, keep_xscroll=True, redraw=True)
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())

    def zoom_in(self, event=None):
        self.tree.zoom_in()
        self.sheet.zoom_in()

    def zoom_out(self, event=None):
        self.tree.zoom_out()
        self.sheet.zoom_out()

    def expand_id(self, event=None):
        if current := self.tree.selected:
            selections = self.tree.selection()
            self.tree.tree_open({self.selected_ID.lower()} | set(self.tree.descendants(self.selected_ID.lower())))
            self.tree.selection_set(selections)
            self.tree.selected = current
        else:
            self.tree.tree_open()

    def collapse_id(self, event=None):
        if current := self.tree.selected:
            selections = self.tree.selection()
            current_iid = self.tree.tree_selected
            self.tree.tree_close({self.selected_ID.lower()} | set(self.tree.descendants(self.selected_ID.lower())))
            self.tree.selection_set(set(filter(self.tree.item_displayed, selections)))
            if self.tree.item_displayed(current_iid):
                self.tree.selected = current
        else:
            self.tree.tree_close()

    def apply_date_format_change(self, new_format, snapshot=True):
        if snapshot:
            self.snapshot_change_date_form(old_form=str(self.DATE_FORM), new_form=str(new_format))
        date_cols = [i for i, h in enumerate(self.headers) if h.type_ == "Date"]
        old_formula_and_condition_dateform = self.convert_hyphen_to_slash_date_form(self.DATE_FORM)
        new_formula_and_condition_dateform = self.convert_hyphen_to_slash_date_form(new_format)
        for col in date_cols:
            for i in range(len(self.headers[col].formatting)):
                new = []
                cond = self.headers[col].formatting[i]
                for e in re.split("([0-9/]+)", cond[0]):
                    if "/" in e:
                        try:
                            z = datetime.datetime.strptime(e, old_formula_and_condition_dateform)
                            e = datetime.datetime.strftime(z, new_formula_and_condition_dateform)
                        except Exception:
                            continue
                    new.append(e)
                self.headers[col].formatting[i] = ("".join(new), cond[1])
            for i in range(len(self.headers[col].validation)):
                if not isint(self.headers[col].validation[i]):
                    try:
                        self.headers[col].validation[i] = datetime.datetime.strftime(
                            datetime.datetime.strptime(self.headers[col].validation[i], self.DATE_FORM),
                            new_format,
                        )
                    except Exception:
                        continue
        for rn in range(len(self.sheet.MT.data)):
            for col in date_cols:
                cell = self.sheet.MT.data[rn][col]
                if "/" in cell or "-" in cell:
                    try:
                        self.sheet.MT.data[rn][col] = datetime.datetime.strftime(
                            datetime.datetime.strptime(cell, self.DATE_FORM), new_format
                        )
                    except Exception:
                        continue
        if snapshot:
            self.disable_paste()
            self.redo_tree_display()
            self.redraw_sheets()
        else:
            self.DATE_FORM = new_format

    def change_date_format(self, new_form):
        if new_form != self.DATE_FORM:
            self.apply_date_format_change(new_form)
            self.DATE_FORM = new_form

    def move_tree_pos(self):
        self.tree.set_yview(self.saved_info[self.pc].scrolls.treey)
        self.tree.set_xview(self.saved_info[self.pc].scrolls.treex)

    def move_sheet_pos(self):
        self.sheet.set_yview(self.saved_info[self.pc].scrolls.sheety)
        self.sheet.set_xview(self.saved_info[self.pc].scrolls.sheetx)

    def refresh_tree_item(self, ID):
        iid = ID.lower()
        if self.tree.exists(iid):
            rn = self.rns[iid]
            highlights = {
                (rn, c): self.sheet.MT.cell_options[(rn, c)]["highlight"]
                for c in range(self.row_len)
                if (rn, c) in self.sheet.MT.cell_options and "highlight" in self.sheet.MT.cell_options[(rn, c)]
            }
            tree_row = self.tree.itemrow(iid)
            self.tree.dehighlight_cells(cells=[(tree_row, c) for c in range(self.row_len)])
            if highlights:
                for cell, highlight in highlights.items():
                    self.tree.highlight_cells(tree_row, cell[1], bg=highlight.bg, fg="black")
            r = self.sheet.MT.data[rn]
            if self.tv_lvls_bool:
                self.tree.item(
                    iid,
                    text=f"{self.get_node_level(self.nodes[iid])}. {r[self.tv_label_col]}",
                    values=r,
                )
            else:
                self.tree.item(
                    iid,
                    text=f"{r[self.tv_label_col]}",
                    values=r,
                )

    def redraw_sheets(self):
        self.sheet.set_refresh_timer()
        self.tree.set_refresh_timer()

    def reset_tree_search_dropdown(self):
        self.search_dropdown["values"] = []
        self.search_displayed.set("")
        self.search_results = []

    def reset_sheet_search_dropdown(self):
        self.sheet_search_dropdown["values"] = []
        self.sheet_search_displayed.set("")
        self.sheet_search_results = []

    def get_node_level(self, node, level=1):
        current_node = node
        current_level = level
        while True:
            if not current_node.ps[self.pc]:
                break
            current_node = self.nodes[current_node.ps[self.pc]]
            current_level += 1
        return current_level

    def redo_tree_display(self, selections=True):
        if self.saved_info[self.pc].twidths:
            self.tree.set_column_widths(self.tree_gen_widths_from_saved())
        else:
            self.tree.set_column_widths()
        self.selected_ID = ""
        self.selected_PAR = ""
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        if self.sheet.data:
            open_ids = self.saved_info[self.pc].opens if self.saved_info[self.pc].opens else None
            if self.tv_lvls_bool:
                data = []
                labels = []
                for iid in self.pc_iids():
                    data.append(self.sheet.data[self.rns[iid]])
                    labels.append(
                        f"{self.get_node_level(self.nodes[iid])}. {self.sheet.data[self.rns[iid]][self.tv_label_col]}"
                    )
                self.tree.tree_build(
                    data=data,
                    iid_column=self.ic,
                    parent_column=self.pc,
                    text_column=labels,
                    row_heights=False,
                    open_ids=open_ids,
                    safety=False,
                    ncols=self.row_len,
                    lower=True,
                ).dehighlight_all()
            else:
                self.tree.tree_build(
                    data=[self.sheet.data[self.rns[iid]] for iid in self.pc_iids()],
                    iid_column=self.ic,
                    parent_column=self.pc,
                    text_column=self.tv_label_col,
                    row_heights=False,
                    open_ids=open_ids,
                    safety=False,
                    ncols=self.row_len,
                    lower=True,
                ).dehighlight_all()
        else:
            self.tree.reset(cell_options=False, column_widths=False, header=False, redraw=False)

        if self.saved_info[self.pc].theights:
            self.tree.set_safe_row_heights(self.tree_gen_heights_from_saved())
        else:
            self.tree.set_row_heights()
        if selections:
            try:
                self.tree.boxes = self.saved_info[self.pc].boxes
                self.tree.selected = self.saved_info[self.pc].selected
            except Exception:
                self.saved_info[self.pc].boxes = ()
                self.saved_info[self.pc].selected = ()
        tree_rns = self.tree.RI.rns
        if self.tagged_ids:
            options = self.tree.RI.cell_options
            highlight = Highlight(
                bg="orange",
                fg="black",
                end=False,
            )
            for ik in filter(tree_rns.__contains__, self.tagged_ids):
                options[tree_rns[ik]] = {}
                options[tree_rns[ik]]["highlight"] = highlight
        options = self.tree.MT.cell_options
        sheet = self.sheet.MT.data
        numrows = len(sheet)
        out_of_bounds = []
        for cell, dct in self.sheet.MT.cell_options.items():
            if cell[0] >= numrows or cell[1] >= self.row_len:
                out_of_bounds.append(cell)
            elif "highlight" in dct and (iid := sheet[cell[0]][self.ic].lower()) in tree_rns:
                options[key := (tree_rns[iid], cell[1])] = {}
                options[key]["highlight"] = dct["highlight"]
        for cell in out_of_bounds:
            del self.sheet.MT.cell_options[cell]
        return "break"

    def get_clipboard_data(self, event=None):
        self.start_work("Loading data from clipboard...")
        self.new_sheet = []
        try:
            data = self.C.clipboard_get()
        except Exception as error_msg:
            Error(self, f"Error: {error_msg}", theme=self.C.theme)
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        try:
            if data.startswith("{") and data.endswith("}"):
                self.new_sheet = json_to_sheet(json.loads(data))
            else:
                self.new_sheet = csv_str_x_data(data)
        except Exception as error_msg:
            self.new_sheet = []
            self.stop_work(self.get_tree_editor_status_bar_text())
            Error(self, f"Error parsing clipboard data: {error_msg}", theme=self.C.theme)
            return
        if not self.new_sheet:
            self.new_sheet = []
            self.stop_work(self.get_tree_editor_status_bar_text())
            Error(self, "No data found on clipboard", theme=self.C.theme)
            return
        new_row_len = equalize_sublist_lens(self.new_sheet)
        self.C.status_bar.change_text(self.get_tree_editor_status_bar_text())
        popup = Get_Clipboard_Data_Popup(
            self,
            cols=self.new_sheet[0],
            row_len=new_row_len,
            theme=self.C.theme,
        )
        if not popup.result:
            self.new_sheet = []
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        new_row_len = equalize_sublist_lens(self.new_sheet)
        flattened = popup.flattened
        order = popup.order
        if flattened:
            hier_cols = popup.flattened_pcols
            if not hier_cols:
                return
        self.C.status_bar.change_text("Building tree...")
        self.snapshot_sheet("full overwrite")
        self.C.status_bar.change_text("Loading...   ")
        self.C.disable_at_start()
        self.warnings = []
        if flattened:
            self.new_sheet, self.row_len, self.ic, self.hiers = TreeBuilder().convert_flattened_to_normal(
                data=self.new_sheet,
                hier_cols=hier_cols,
                rowlen=new_row_len,
                order=order,
                warnings=self.warnings,
            )
        else:
            self.row_len = new_row_len
            self.ic = popup.ic
            self.hiers = popup.pcols
        self.selected_ID = ""
        self.selected_PAR = ""
        self.pc = int(self.hiers[0])
        self.tv_label_col = self.ic

        new_headers = self.fix_headers(self.new_sheet.pop(0), self.row_len)
        new_headers = [Header(name) for name in new_headers]
        new_headers[self.ic].type_ = "ID"
        for h in self.hiers:
            new_headers[h].type_ = "Parent"
        existing_headers = {h.name: i for i, h in enumerate(self.headers)}
        existing_col_alignments = {
            self.headers[c].name: align for c, align in self.sheet.get_column_alignments().items()
        }

        self.tree.reset()
        self.sheet.reset()

        for h in new_headers:
            if h.name in existing_headers and h.type_ == self.headers[existing_headers[h.name]].type_:
                h.formatting = self.headers[existing_headers[h.name]].formatting
            if h.name in existing_col_alignments:
                self.tree.align_columns(existing_headers[h.name], existing_col_alignments[h.name])
                self.sheet.align_columns(existing_headers[h.name], existing_col_alignments[h.name])

        self.headers = new_headers
        self.sheet.MT.data = self.new_sheet
        self.new_sheet = []
        self.saved_info = new_saved_info(self.hiers)
        self.clear_copied_details()
        self.cut = []
        self.copied = []
        self.cut_children_dct = {}
        self.sheet.set_xview(0.0)
        self.sheet.set_yview(0.0)
        self.auto_sort_nodes_bool = True
        self.sheet.MT.data, self.nodes, self.warnings = TreeBuilder().build(
            input_sheet=self.sheet.MT.data,
            output_sheet=self.new_sheet,
            row_len=self.row_len,
            ic=self.ic,
            hiers=self.hiers,
            nodes={},
            warnings=self.warnings,
            add_warnings=True,
            strip=not self.allow_spaces_ids_var,
        )
        self.new_sheet = []
        self.fix_associate_sort(startup=True)
        self.set_headers()
        self.refresh_hier_dropdown(self.hiers.index(self.pc))
        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
        self.sheet.set_row_heights().set_column_widths().row_index(newindex=self.ic)
        self.tagged_ids = set(filter(self.rns.__contains__, self.tagged_ids))
        self.reset_tagged_ids_dropdowns()
        self.rehighlight_tagged_ids()
        self.refresh_formatting(dehighlight=True)
        self.redo_tree_display()
        self.refresh_dropdowns()
        self.changelog_append(
            "Overwrite sheet with clipboard data",
            "",
            "",
            "",
        )
        self.stop_work(self.get_tree_editor_status_bar_text())
        self.show_warnings("n/a - Data obtained from clipboard", "n/a")

    def import_changes(self):
        fp = filedialog.askopenfilename(parent=self, title="Select file")
        if not fp:
            return
        self.start_work("Import changes...")
        try:
            fp = os.path.normpath(fp)
        except Exception:
            Error(self, "Filepath invalid   ", theme=self.C.theme)
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        if not fp.lower().endswith((".tsv", ".csv", ".xls", ".xlsx", ".xlsm", ".json")):
            Error(self, "Invalid file format   ", theme=self.C.theme)
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        if not os.path.isfile(fp):
            Error(self, "Filepath invalid   ", theme=self.C.theme)
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        changes = []
        row_len = 0
        if fp.lower().endswith((".csv", ".tsv")):
            try:
                with open(fp, "r") as fh:
                    temp_data = fh.read()
                if not temp_data:
                    Error(
                        self,
                        "No data found in file",
                        theme=self.C.theme,
                    )
                    self.stop_work(self.get_tree_editor_status_bar_text())
                    return
                changes = csv_str_x_data(temp_data)
            except Exception as error_msg:
                Error(self, f"Error: {error_msg}", theme=self.C.theme)
                self.stop_work(self.get_tree_editor_status_bar_text())
                return
        elif fp.lower().endswith(".json"):
            try:
                j = get_json_from_file(fp)
                json_format = get_json_format(j)
                if not json_format:
                    Error(
                        self,
                        "Error opening file, could not find data of correct format   ",
                        theme=self.C.theme,
                    )
                    self.stop_work(self.get_tree_editor_status_bar_text())
                    return
                changes, row_len = json_to_sheet(
                    j,
                    format_=json_format[0],
                    key=json_format[1],
                    get_format=False,
                    return_rowlen=True,
                )
            except Exception as error_msg:
                Error(self, f"Error: {error_msg}", theme=self.C.theme)
                self.stop_work(self.get_tree_editor_status_bar_text())
                return
        elif fp.lower().endswith((".xls", ".xlsx", ".xlsm")):
            try:
                wb = load_workbook(bytes_io_wb(fp), read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                ws.reset_dimensions()
                changes = ws_x_data(ws)
                wb.close()
            except Exception as error_msg:
                Error(self, f"Error: {error_msg}", theme=self.C.theme)
                self.stop_work(self.get_tree_editor_status_bar_text())
                return
        if not changes:
            Error(self, "File contains no data   ", theme=self.C.theme)
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        row_len = max(map(len, changes), default=0)
        if row_len != 5:
            Error(self, "Invalid changelog format   ", theme=self.C.theme)
            self.stop_work(self.get_tree_editor_status_bar_text())
            return
        equalize_sublist_lens(seq=changes, len_=row_len)
        successful = []
        excluded = 0
        self.snapshot_sheet()
        changes_len = len(changes)
        for changenum, change in enumerate(changes):
            if not changenum % 10:
                self.C.update()
                self.C.status_bar.change_text(f"Imported {changenum} / {changes_len} changes")
            ctyp = change[1]
            if ctyp.startswith("Imported change |"):
                ctyp = ctyp.split("Imported change | ")[1]
            elif ctyp.startswith("Merge |"):
                ctyp = ctyp.split("Merge | ")[1]
            try:
                #  "Edit cell"
                if ctyp == "Edit cell |" or ctyp == "Edit cell":
                    c3s = change[2].split(" ")
                    cik = c3s[1].lower()
                    name = c3s[5]  # col name in change
                    col = next(i for i, h in enumerate(self.headers) if h.name.lower() == name.lower())
                    type_ = c3s[-1]  # col type in change
                    if type_ == "Detail":
                        type_ = f"{c3s[-2]} {type_}"
                    if self.headers[col].validation:
                        validation_check = self.is_in_validation(self.headers[col].validation, change[4])
                    else:
                        validation_check = True
                    if (
                        self.headers[col].type_ == type_
                        and cik in self.rns
                        and self.sheet.MT.data[self.rns[cik]][col] == change[3]
                        and validation_check
                    ):
                        oldv = f"{self.sheet.MT.data[self.rns[cik]][col]}"
                        newv = f"{change[4]}"
                        if self.sheet.MT.data[self.rns[cik]][col] != change[4]:
                            self.changelog_append_no_unsaved(
                                "Imported change | Edit cell",
                                change[2],
                                change[3],
                                change[4],
                            )
                            self.sheet.MT.data[self.rns[cik]][col] = change[4]
                            if oldv != newv and type_ == "ID" or type_ == "Parent":
                                self.nodes = {}
                                self.auto_sort_nodes_bool = True
                                self.sheet.MT.data, self.nodes = TreeBuilder().build(
                                    self.sheet.MT.data,
                                    self.new_sheet,
                                    self.row_len,
                                    self.ic,
                                    self.hiers,
                                    self.nodes,
                                    add_warnings=False,
                                    strip=not self.allow_spaces_ids_var,
                                )
                                self.new_sheet = []
                                self.fix_associate_sort_edit_cells()
                                self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                            successful.append(True)
                        else:
                            excluded += 1
                    else:
                        successful.append(False)

                #  "Move rows"

                elif ctyp == "Move rows":
                    old_locs = change[3].split(",")
                    new_locs = change[4].split(",")
                    if len(old_locs) != len(new_locs):
                        successful.append(False)
                        continue
                    if len(old_locs) == 1:
                        old_locs = [old_locs[0].split("Old locations: ")[1]]
                        new_locs = [new_locs[0].split("New locations: ")[1]]
                    new_idxs = dict(zip(map(int, old_locs), map(int, new_locs)))
                    if all(i <= len(self.sheet.data) and i >= 0 for i in new_idxs) and all(
                        i <= len(self.sheet.data) and i >= 0 for i in new_idxs.values()
                    ):
                        self.sheet.mapping_move_rows(
                            data_new_idxs=new_idxs,
                            disp_new_idxs=new_idxs,
                            undo=False,
                            create_selections=False,
                            redraw=False,
                        )
                        self.changelog_append_no_unsaved(
                            "Imported change | Move rows",
                            change[2],
                            change[3],
                            change[4],
                        )
                        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Move columns"

                elif ctyp == "Move columns":
                    old_locs = change[3].split(",")
                    new_locs = change[4].split(",")
                    if len(old_locs) != len(new_locs):
                        successful.append(False)
                        continue
                    if len(old_locs) == 1:
                        old_locs = [old_locs[0].split("Old locations: ")[1]]
                        new_locs = [new_locs[0].split("New locations: ")[1]]
                    new_idxs = dict(zip(map(int, old_locs), map(int, new_locs)))
                    event_data = {
                        "moved": {
                            "columns": {
                                "data": new_idxs,
                                "displayed": new_idxs,
                            }
                        }
                    }
                    if max(new_idxs.values()) < self.row_len:
                        self.snapshot_drag_cols(event_data=event_data)
                        self.changelog_append_no_unsaved(
                            "Imported change | Move columns",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Add new hierarchy column"

                elif ctyp == "Add new hierarchy column":
                    c3s = change[2].split(" ")
                    colname = "".join(c3s[-1].split(" ")).strip()
                    colnum = int(c3s[1][1:]) - 1
                    if (
                        colname.lower() not in (h.name.lower() for h in self.headers)
                        and colnum >= 0
                        and colnum <= len(self.headers)
                    ):
                        self.add_hier_col(colnum, colname, snapshot=False)
                        self.changelog_append_no_unsaved(
                            "Imported change | Add new hierarchy column",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Add new detail column"

                elif ctyp == "Add new detail column":
                    c3s = change[2].split(" ")
                    colname = "".join(c3s[4].split(" ")).strip()
                    colnum = int(c3s[1][1:]) - 1
                    coltype = f"{c3s[-2]} {c3s[-1]}"
                    if (
                        colname.lower() not in (h.name.lower() for h in self.headers)
                        and colnum >= 0
                        and colnum <= len(self.headers)
                    ):
                        self.add_col(colnum, colname, coltype, snapshot=False)
                        self.changelog_append_no_unsaved(
                            "Imported change | Add new detail column",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Delete hierarchy column"

                elif ctyp == "Delete hierarchy column":
                    c3s = change[2].split(" ")
                    colname = c3s[-1]
                    colnum = next(i for i, h in enumerate(self.headers) if h.name.lower() == colname.lower())
                    if self.headers[colnum].type_ == "Parent" and len(self.hiers) > 1:
                        if self.pc == colnum:
                            self.pc = int(next(i for i in self.hiers if i != colnum))
                        self.del_cols(cols=[colnum], snapshot=False)
                        self.changelog_append_no_unsaved(
                            "Imported change | Delete hierarchy column",
                            change[2],
                            "",
                            "",
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Delete detail column"

                elif ctyp == "Delete detail column":
                    c3s = change[2].split(" ")
                    colname = c3s[4]
                    colnum = next(i for i, h in enumerate(self.headers) if h.name.lower() == colname.lower())
                    coltype = f"{c3s[-2]} {c3s[-1]}"
                    if self.headers[colnum].type_ == coltype and self.headers[colnum].type_ in (
                        "Number",
                        "Text",
                        "Date",
                    ):
                        self.del_cols(cols=[colnum], snapshot=False)
                        self.changelog_append_no_unsaved(
                            "Imported change | Delete detail column",
                            change[2],
                            "",
                            "",
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Column rename"

                elif ctyp == "Column rename":
                    c3s = change[2].split(" ")
                    coltype = f"{c3s[-2]} {c3s[-1]}"
                    colname = "".join(change[4].split(" ")).strip()
                    colnum = next(i for i, h in enumerate(self.headers) if h.name.lower() == colname.lower())
                    if (
                        self.headers[colnum].name.lower() == change[3].lower()
                        and self.headers[colnum].type_ == coltype
                        and colname.lower() not in (h.name.lower() for h in self.headers)
                    ):
                        self.rename_col(colnum, colname, snapshot=False)
                        self.changelog_append_no_unsaved(
                            "Imported change | Column rename",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Edit validation"

                elif ctyp == "Edit validation":
                    c3s = change[2].split(" ")
                    colname = c3s[3]
                    colnum = next(i for i, h in enumerate(self.headers) if h.name.lower() == colname.lower())
                    coltype = f"{c3s[-2]} {c3s[-1]}"
                    validation = change[4]
                    if (
                        self.headers[colnum].type_ == coltype
                        and coltype in ("Text", "Number", "Date")
                        and change[3] == ",".join(self.headers[colnum].validation)
                    ):
                        if validation:
                            validation = self.check_validation_validity(colnum, validation.split(","))
                            if isinstance(validation, str):
                                successful.append(False)
                                continue
                        else:
                            validation = []
                        self.headers[colnum].validation = validation
                        if validation:
                            self.apply_validation_to_col(colnum)
                        self.changelog_append_no_unsaved(
                            "Imported change | Edit validation",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Change detail column type"

                elif ctyp == "Change detail column type":
                    c3s = change[2].split(" ")
                    colname = c3s[-1]
                    colnum = next(i for i, h in enumerate(self.headers) if h.name.lower() == colname.lower())
                    oldtype = change[3]
                    newtype = change[4]
                    if self.headers[colnum].type_ == oldtype and newtype in (
                        "Text",
                        "Number",
                        "Date",
                    ):
                        if newtype == "Text":
                            self.change_coltype_text(colnum)
                        elif newtype == "Number":
                            self.headers[colnum].type_ = "Number"
                            self.change_coltype_number(colnum)
                            if isinstance(self.check_validation_validity(colnum, self.headers[colnum].validation), str):
                                self.headers[colnum].validation = []
                            self.headers[colnum].formatting = [
                                tup
                                for tup in self.headers[colnum].formatting
                                if not self.check_condition_validity(colnum, tup[0]).startswith("Error:")
                            ]
                        else:
                            self.headers[colnum].type_ = "Date"
                            self.change_coltype_date(colnum, detect_date_form=True)
                            if isinstance(self.check_validation_validity(colnum, self.headers[colnum].validation), str):
                                self.headers[colnum].validation = []
                            self.headers[colnum].formatting = [
                                tup
                                for tup in self.headers[colnum].formatting
                                if not self.check_condition_validity(colnum, tup[0]).startswith("Error:")
                            ]
                        self.changelog_append_no_unsaved(
                            "Imported change | Change detail column type",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Date format change"

                elif ctyp == "Date format change":
                    old_form = "%" + change[3][:2] + "%" + change[3][2:4] + "%" + change[3][4:]
                    new_form = "%" + change[4][:2] + "%" + change[4][2:4] + "%" + change[4][4:]
                    if old_form in date_formats_usable and new_form in date_formats_usable:
                        self.apply_date_format_change(new_form, snapshot=False)
                        self.changelog_append_no_unsaved(
                            "Imported change | Date format change",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Cut and paste ID"

                elif ctyp == "Cut and paste ID" or ctyp == "Cut and paste ID |":
                    cik = change[2].lower()
                    old = change[3].split(" ")
                    oldcolname = old[-1]
                    oldcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == oldcolname.lower())
                    if "n/a - Top ID" in change[3]:
                        oldpar = ""
                        oldpar_check = True
                    else:
                        oldpar = old[2]
                        if oldpar.lower() not in self.nodes or oldpar != self.nodes[self.nodes[cik].ps[oldcol]].name:
                            oldpar_check = False
                        else:
                            oldpar_check = True

                    new = change[4].split(" ")
                    newcolname = new[-1]
                    newcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == newcolname.lower())
                    if "n/a - Top ID" in change[4]:
                        newpar = ""
                        newpar_check = True
                    else:
                        newpar = new[2]
                        if newpar.lower() not in self.nodes or self.nodes[newpar.lower()].ps[newcol] is None:
                            newpar_check = False
                        else:
                            newpar_check = True

                    if (
                        self.headers[oldcol].type_ == "Parent"
                        and self.headers[newcol].type_ == "Parent"
                        and cik in self.rns
                        and oldpar_check
                        and newpar_check
                    ):
                        oldpc = int(self.pc)
                        self.pc = newcol
                        if self.cut_paste(
                            f"{change[2]}",
                            oldpar,
                            oldcol,
                            newpar,
                            snapshot=False,
                            errors=False,
                        ):
                            self.changelog_append_no_unsaved(
                                "Imported change | Cut and paste ID",
                                change[2],
                                change[3],
                                change[4],
                            )
                            successful.append(True)
                        else:
                            successful.append(False)
                        self.pc = int(oldpc)
                    else:
                        successful.append(False)

                #  "Cut and paste ID + children |"

                elif ctyp == "Cut and paste ID + children" or ctyp == "Cut and paste ID + children |":
                    cik = change[2].lower()
                    old = change[3].split(" ")
                    oldcolname = old[-1]
                    oldcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == oldcolname.lower())
                    if "n/a - Top ID" in change[3]:
                        oldpar = ""
                        oldpar_check = True
                    else:
                        oldpar = old[2]
                        if oldpar.lower() not in self.nodes or oldpar != self.nodes[self.nodes[cik].ps[oldcol]].name:
                            oldpar_check = False
                        else:
                            oldpar_check = True

                    new = change[4].split(" ")
                    newcolname = new[-1]
                    newcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == newcolname.lower())
                    if "n/a - Top ID" in change[4]:
                        newpar = ""
                        newpar_check = True
                    else:
                        newpar = new[2]
                        if newpar.lower() not in self.nodes or self.nodes[newpar.lower()].ps[newcol] is None:
                            newpar_check = False
                        else:
                            newpar_check = True

                    if (
                        self.headers[oldcol].type_ == "Parent"
                        and self.headers[newcol].type_ == "Parent"
                        and cik in self.rns
                        and oldpar_check
                        and newpar_check
                    ):
                        oldpc = int(self.pc)
                        self.pc = newcol
                        if self.cut_paste_all(
                            f"{change[2]}",
                            oldpar,
                            oldcol,
                            newpar,
                            snapshot=False,
                            errors=False,
                        ):
                            self.changelog_append_no_unsaved(
                                "Imported change | Cut and paste ID + children",
                                change[2],
                                change[3],
                                change[4],
                            )
                            successful.append(True)
                        else:
                            successful.append(False)
                        self.pc = int(oldpc)
                    else:
                        successful.append(False)

                #  "Cut and paste children"

                elif ctyp == "Cut and paste children":
                    old = change[3].split(" ")
                    oldcolname = old[-1]
                    oldcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == oldcolname.lower())
                    if "n/a - Top ID" in change[3]:
                        oldpar = ""
                        oldpar_check = True
                    else:
                        oldpar = old[2]
                        if oldpar.lower() not in self.nodes or self.nodes[oldpar.lower()].ps[oldcol] is None:
                            oldpar_check = False
                        else:
                            oldpar_check = True

                    new = change[4].split(" ")
                    newcolname = new[-1]
                    newcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == newcolname.lower())
                    if "n/a - Top ID" in change[4]:
                        newpar = ""
                        newpar_check = True
                    else:
                        newpar = new[2]
                        if newpar.lower() not in self.nodes or self.nodes[newpar.lower()].ps[newcol] is None:
                            newpar_check = False
                        else:
                            newpar_check = True

                    if (
                        self.headers[oldcol].type_ == "Parent"
                        and self.headers[newcol].type_ == "Parent"
                        and oldpar_check
                        and newpar_check
                    ):
                        oldpc = int(self.pc)
                        self.pc = newcol
                        if self.cut_paste_children(oldpar, newpar, oldcol, snapshot=False, errors=False):
                            self.changelog_append_no_unsaved(
                                "Imported change | Cut and paste children",
                                change[2],
                                change[3],
                                change[4],
                            )
                            successful.append(True)
                        else:
                            successful.append(False)
                        self.pc = int(oldpc)
                    else:
                        successful.append(False)

                #  "Copy and paste ID |"

                elif ctyp == "Copy and paste ID |" or ctyp == "Copy and paste ID":
                    cik = change[2].lower()
                    old = change[3].split(" ")
                    oldcolname = old[-1]
                    oldcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == oldcolname.lower())
                    new = change[4].split(" ")
                    newcolname = new[-1]
                    newcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == newcolname.lower())
                    if "n/a - Top ID" in change[4]:
                        newpar = ""
                        newpar_check = True
                    else:
                        newpar = new[2]
                        if newpar.lower() not in self.nodes or self.nodes[newpar.lower()].ps[newcol] is None:
                            newpar_check = False
                        else:
                            newpar_check = True

                    if (
                        self.headers[newcol].type_ == "Parent"
                        and self.headers[oldcol].type_ == "Parent"
                        and cik in self.rns
                        and newpar_check
                    ):
                        oldpc = int(self.pc)
                        self.pc = newcol
                        if self.copy_paste(change[2], oldcol, newpar, snapshot=False, errors=False):
                            self.changelog_append_no_unsaved(
                                "Imported change | Copy and paste ID",
                                change[2],
                                change[3],
                                change[4],
                            )
                            successful.append(True)
                        else:
                            successful.append(False)
                        self.pc = int(oldpc)
                    else:
                        successful.append(False)

                #  "Copy and paste ID + children |"

                elif ctyp == "Copy and paste ID + children |" or ctyp == "Copy and paste ID + children":
                    cik = change[2].lower()
                    old = change[3].split(" ")
                    oldcolname = old[-1]
                    oldcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == oldcolname.lower())
                    new = change[4].split(" ")
                    newcolname = new[-1]
                    newcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == newcolname.lower())
                    if "n/a - Top ID" in change[4]:
                        newpar = ""
                        newpar_check = True
                    else:
                        newpar = new[2]
                        if newpar.lower() not in self.nodes or self.nodes[newpar.lower()].ps[newcol] is None:
                            newpar_check = False
                        else:
                            newpar_check = True

                    if (
                        self.headers[newcol].type_ == "Parent"
                        and self.headers[oldcol].type_ == "Parent"
                        and cik in self.rns
                        and newpar_check
                    ):
                        oldpc = int(self.pc)
                        self.pc = newcol
                        if self.copy_paste_all(change[2], oldcol, newpar, snapshot=False, errors=False):
                            self.changelog_append_no_unsaved(
                                "Imported change | Copy and paste ID + children",
                                change[2],
                                change[3],
                                change[4],
                            )
                            successful.append(True)
                        else:
                            successful.append(False)
                        self.pc = int(oldpc)
                    else:
                        successful.append(False)

                #  "Add ID"

                elif ctyp == "Add ID":
                    new = change[2].split(" ")
                    newcolname = new[-1]
                    newcol = next(i for i, h in enumerate(self.headers) if h.name.lower() == newcolname.lower())
                    cid = new[1]
                    cik = cid.lower()
                    if "n/a - Top ID" in change[2]:
                        newpar = ""
                        newpk = ""
                    else:
                        newpar = new[3]
                        newpk = newpar.lower()
                    newpar_check = bool(not newpk or newpk in self.rns)

                    if self.headers[newcol].type_ == "Parent" and newpar_check:
                        oldpc = int(self.pc)
                        self.pc = newcol
                        if self.add(cid, newpar, snapshot=False, errors=False):
                            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                            self.changelog_append_no_unsaved(
                                "Imported change | Add ID",
                                change[2],
                                change[3],
                                change[4],
                            )
                            successful.append(True)
                        else:
                            successful.append(False)
                        self.pc = int(oldpc)
                    else:
                        successful.append(False)

                #  "Rename ID"

                elif ctyp == "Rename ID":
                    oldname = change[3]
                    newname = change[4]
                    if oldname.lower() in self.rns and newname.lower() not in self.rns:
                        if self.change_ID_name(oldname, newname, snapshot=False, errors=False):
                            self.changelog_append_no_unsaved(
                                "Imported change | Rename ID",
                                change[2],
                                change[3],
                                change[4],
                            )
                            self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                            if oldname.lower() in self.tagged_ids:
                                self.tagged_ids.discard(oldname.lower())
                                self.tagged_ids.add(newname.lower())
                            successful.append(True)
                        else:
                            successful.append(False)
                    else:
                        successful.append(False)

                #  "Delete ID |"

                elif ctyp == "Delete ID |" or ctyp == "Delete ID":
                    info = change[2].split(" ")
                    colname = info[-1]
                    colnum = next(i for i, h in enumerate(self.headers) if h.name.lower() == colname.lower())
                    cid = info[1]
                    cpar = "" if "n/a - Top ID" in change[2] else info[3]
                    if cpar:
                        if (
                            cpar.lower() not in self.nodes
                            or self.nodes[self.nodes[cid.lower()].ps[colnum]].name != cpar
                        ):
                            cpar_check = False
                        else:
                            cpar_check = True
                    else:
                        cpar_check = True
                    if cid.lower() in self.rns and cpar_check and self.headers[colnum].type_ == "Parent":
                        oldpc = int(self.pc)
                        self.pc = colnum
                        self._del_id_core(cid.lower(), cpar.lower(), snapshot=False)
                        self.pc = int(oldpc)
                        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                        self.changelog_append_no_unsaved(
                            "Imported change | Delete ID",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Delete ID, orphan children"

                elif ctyp == "Delete ID, orphan children":
                    info = change[2].split(" ")
                    colname = info[-1]
                    colnum = next(i for i, h in enumerate(self.headers) if h.name.lower() == colname.lower())
                    cid = info[1]
                    cpar = "" if "n/a - Top ID" in change[2] else info[3]
                    if cpar:
                        if (
                            cpar.lower() not in self.nodes
                            or self.nodes[self.nodes[cid.lower()].ps[colnum]].name != cpar
                        ):
                            cpar_check = False
                        else:
                            cpar_check = True
                    else:
                        cpar_check = True
                    if cid.lower() in self.rns and cpar_check and self.headers[colnum].type_ == "Parent":
                        oldpc = int(self.pc)
                        self.pc = colnum
                        self._del_id_orphan_core(cid.lower(), cpar.lower(), snapshot=False)
                        self.pc = int(oldpc)
                        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                        self.changelog_append_no_unsaved(
                            "Imported change | Delete ID",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Delete ID + all children"

                elif ctyp == "Delete ID + all children":
                    info = change[2].split(" ")
                    colname = info[-1]
                    colnum = next(i for i, h in enumerate(self.headers) if h.name.lower() == colname.lower())
                    cid = info[1]
                    cpar = "" if "n/a - Top ID" in change[2] else info[3]
                    if cpar:
                        if (
                            cpar.lower() not in self.nodes
                            or self.nodes[self.nodes[cid.lower()].ps[colnum]].name != cpar
                        ):
                            cpar_check = False
                        else:
                            cpar_check = True
                    else:
                        cpar_check = True
                    if cid.lower() in self.rns and cpar_check and self.headers[colnum].type_ == "Parent":
                        oldpc = int(self.pc)
                        self.pc = colnum
                        self._del_id_children_core(cid.lower(), cpar.lower(), snapshot=False)
                        self.pc = int(oldpc)
                        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                        self.changelog_append_no_unsaved(
                            "Imported change | Delete ID + all children",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Delete ID + all children from all hierarchies"

                elif ctyp == "Delete ID + all children from all hierarchies":
                    info = change[2].split(" ")
                    colname = info[-1]
                    colnum = next(i for i, h in enumerate(self.headers) if h.name.lower() == colname.lower())
                    cid = info[1]
                    cpar = "" if "n/a - Top ID" in change[2] else info[3]
                    if cpar:
                        if (
                            cpar.lower() not in self.nodes
                            or self.nodes[self.nodes[cid.lower()].ps[colnum]].name != cpar
                        ):
                            cpar_check = False
                        else:
                            cpar_check = True
                    else:
                        cpar_check = True
                    if cid.lower() in self.rns and cpar_check and self.headers[colnum].type_ == "Parent":
                        oldpc = int(self.pc)
                        self.pc = colnum
                        self._del_id_children_all_core(cid.lower(), cpar.lower(), snapshot=False)
                        self.pc = int(oldpc)
                        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                        self.changelog_append_no_unsaved(
                            "Imported change | Delete ID + all children from all hierarchies",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Delete ID from all hierarchies |"

                elif ctyp == "Delete ID from all hierarchies |" or ctyp == "Delete ID from all hierarchies":
                    cid = change[2]
                    if cid.lower() in self.rns:
                        to_del = self._del_id_all_core(cid.lower(), snapshot=False)
                        self.sheet.del_rows(map(self.rns.get, to_del), redraw=False)
                        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                        self.changelog_append_no_unsaved(
                            "Imported change | Delete ID from all hierarchies",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Delete ID from all hierarchies, orphan children"

                elif ctyp == "Delete ID from all hierarchies, orphan children":
                    cid = change[2]
                    if cid.lower() in self.rns:
                        self._del_id_all_orphan_core(cid.lower(), snapshot=False)
                        self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                        self.changelog_append_no_unsaved(
                            "Imported change | Delete ID from all hierarchies, orphan children",
                            change[2],
                            change[3],
                            change[4],
                        )
                        successful.append(True)
                    else:
                        successful.append(False)

                #  "Sort sheet"

                elif ctyp == "Sort sheet":
                    if change[2] == "Sorted sheet in tree walk order":
                        if self.sheet.MT.data:
                            self.sort_sheet_walk(snapshot=False)
                            self.changelog_append_no_unsaved(
                                f"Imported change | {change[1]}",
                                change[2],
                                change[3],
                                change[4],
                            )
                            successful.append(True)
                        else:
                            successful.append(False)
                    else:
                        c3s = change[2].split(" ")
                        colname = info[6]
                        colnum = next(
                            i for i, h in enumerate(self.headers) if h.name.lower() == colname.lower()
                        )  # checks if column name exists
                        order = info[8]
                        if order in ("ASCENDING", "DESCENDING"):
                            self.sort_sheet(colname, order, snapshot=False)
                            self.changelog_append_no_unsaved(
                                f"Imported change | {change[1]}",
                                change[2],
                                change[3],
                                change[4],
                            )
                            successful.append(True)
                        else:
                            successful.append(False)
            except Exception:
                successful.append(False)
                continue
        num_successful = sum(successful)
        if num_successful:
            self.changelog_append(
                f"Imported {num_successful} changes from: {os.path.basename(fp)}",
                f"Unsuccessful: {len(successful) - num_successful} Unnecessary: {excluded}",
                "",
                "",
            )
        else:
            self.vs.pop()
            self.vp -= 1
            self.set_undo_label()
        self.pc = int(self.hiers[0])
        self.clear_copied_details()
        self.refresh_hier_dropdown(0)
        self.set_headers()
        self.sheet.deselect().set_column_widths().row_index(newindex=self.ic)
        self.reset_tagged_ids_dropdowns()
        self.rehighlight_tagged_ids()
        self.refresh_rows = set()
        self.refresh_formatting()
        self.redo_tree_display()
        self.refresh_dropdowns()
        self.stop_work(self.get_tree_editor_status_bar_text())
        applicable_changes = {
            "Edit cell",
            "Edit cell |",
            "Move rows",
            "Move columns",
            "Add new hierarchy column",
            "Add new detail column",
            "Delete hierarchy column",
            "Delete detail column",
            "Column rename",
            "Edit validation",
            "Change detail column type",
            "Date format change",
            "Cut and paste ID",
            "Cut and paste ID |",
            "Cut and paste ID + children",
            "Cut and paste ID + children |",
            "Cut and paste children",
            "Copy and paste ID",
            "Copy and paste ID |",
            "Copy and paste ID + children",
            "Copy and paste ID + children |",
            "Add ID",
            "Rename ID",
            "Delete ID",
            "Delete ID |",
            "Delete ID, orphan children",
            "Delete ID + all children",
            "Delete ID + all children from all hierarchies",
            "Delete ID from all hierarchies",
            "Delete ID from all hierarchies |",
            "Delete ID from all hierarchies, orphan children",
            "Sort sheet",
        }
        applicable_changes = applicable_changes | {f"Imported change | {change}" for change in applicable_changes}
        Post_Import_Changes_Popup(
            self,
            [change for change in changes if change[1] in applicable_changes or change[1].startswith("Merge | ")],
            successful,
            theme=self.C.theme,
        )
        self.focus_tree()

    def add_rows_rc(self, insert=False):
        self.new_sheet = []
        popup = Merge_Sheets_Popup(self, theme=self.C.theme, add_rows=True)
        if not popup.result:
            self.new_sheet = []
            return
        self.merge_sheets(
            insert_row=sorted(self.sheet.get_selected_rows())[0] if insert else len(self.sheet.MT.data),
            popup_=popup,
        )

    def merge_sheets(self, insert_row=None, popup_=None):
        try:
            if insert_row is None:
                self.new_sheet = []
                popup = Merge_Sheets_Popup(self, theme=self.C.theme)
                if not popup.result:
                    self.new_sheet = []
                    return
            else:
                popup = popup_
            self.start_work("Merging sheets...")
            self.snapshot_sheet()
            self.warnings = []
            ns_ic = popup.ic
            ns_hiers = popup.pcols
            ns_hiers_set = set(ns_hiers)
            ns_row_len = popup.row_len
            ns_headers = self.new_sheet.pop(0)
            ns_headers = self.fix_headers(ns_headers, ns_row_len)
            ns_num_hdrs = len(ns_headers)
            equalize_sublist_lens(seq=self.new_sheet, len_=ns_num_hdrs)
            ns_pcol_names = {cell.lower(): i for i, cell in enumerate(ns_headers) if i in ns_hiers_set}
            ns_dcol_names = {
                cell.lower(): i for i, cell in enumerate(ns_headers) if i not in ns_hiers_set and i != ns_ic
            }
            os_header_names = {h.name.lower() for h in self.headers}
            ns_rns = {row[ns_ic].lower(): i for i, row in enumerate(self.new_sheet)}
            shared_ids = {i: ik for ik, i in self.rns.items() if ik in ns_rns}
            os_pcol_names = {h.name.lower(): i for i, h in enumerate(self.headers) if h.type_ == "Parent"}
            os_dcol_names = {h.name.lower(): i for i, h in enumerate(self.headers) if h.type_ in detail_column_types}
            changes_made = 0
            rows_to_insert = []

            # add new details columns option
            if popup.add_new_dcols:
                new_dcols = [idx for colname, idx in ns_dcol_names.items() if colname not in os_header_names]
                num_new_dcols = len(new_dcols)
                self.headers.extend([Header(ns_headers[idx], "Text") for idx in new_dcols])
                if num_new_dcols:
                    self.tree.insert_columns(num_new_dcols)
                    self.sheet.insert_columns(num_new_dcols)
                for num, idx in enumerate(new_dcols, 1):
                    self.changelog_append_no_unsaved(
                        "Merge | Add new detail column",
                        f"Column #{self.row_len + num} with name: {ns_headers[idx]} and type: Text",
                        "",
                        "",
                    )
                    changes_made += 1
                for rn in range(len(self.sheet.MT.data)):
                    row = self.sheet.MT.data[rn]
                    if rn in shared_ids:
                        ns_rn = ns_rns[shared_ids[rn]]
                        for num, idx in enumerate(new_dcols):
                            row[self.row_len + num] = self.new_sheet[ns_rn][idx]
                            if row[self.row_len + num] != "":
                                self.changelog_append_no_unsaved(
                                    "Merge | Edit cell",
                                    f"ID: {row[self.ic]} column #{self.row_len + num + 1} named: {self.headers[self.row_len + num].name} with type: {self.headers[self.row_len + num].type_}",
                                    "",
                                    f"{row[self.row_len + num]}",
                                )
                                changes_made += 1
                    self.sheet.MT.data[rn] = row
                self.row_len += num_new_dcols

            # add new parent columns option
            if popup.add_new_pcols:
                new_pcols = [idx for colname, idx in ns_pcol_names.items() if colname not in os_header_names]
                num_new_pcols = len(new_pcols)
                self.headers.extend([Header(ns_headers[idx], "Parent") for idx in new_pcols])
                if num_new_pcols:
                    self.tree.insert_columns(num_new_pcols)
                    self.sheet.insert_columns(num_new_pcols)
                for num, idx in enumerate(new_pcols, 1):
                    self.changelog_append_no_unsaved(
                        "Merge | Add new hierarchy column",
                        f"Column #{self.row_len + num} with name: {ns_headers[idx]}",
                        "",
                        "",
                    )
                    changes_made += 1
                range_end = self.row_len + num_new_pcols
                self.hiers.extend(list(range(self.row_len, range_end)))
                for node in self.nodes.values():
                    for i in range(self.row_len, range_end):
                        node.ps[i] = None
                        node.cn[i] = []
                for i in range(self.row_len, range_end):
                    self.saved_info[i] = new_info_storage()
                for rn in range(len(self.sheet.MT.data)):
                    row = self.sheet.MT.data[rn]
                    if rn in shared_ids:
                        ns_rn = ns_rns[shared_ids[rn]]
                        for num, idx in enumerate(new_pcols):
                            row[self.row_len + num] = self.new_sheet[ns_rn][idx]
                            if row[self.row_len + num] != "":
                                self.changelog_append_no_unsaved(
                                    "Merge | Edit cell",
                                    f"ID: {row[self.ic]} column #{self.row_len + num + 1} named: {self.headers[self.row_len + num].name} with type: {self.headers[self.row_len + num].type_}",
                                    "",
                                    f"{row[self.row_len + num]}",
                                )
                                changes_made += 1
                    self.sheet.MT.data[rn] = row
                self.row_len += num_new_pcols

            # add any new ids
            # AND if add new detail columns then add the details for those ids
            # AND if add new parent columns then add the parents for those ids
            if popup.add_new_ids:
                new_ids = {ik for ik in ns_rns if ik not in self.rns and ik}
                shared_dcols = tuple(name for name in os_dcol_names if name in ns_dcol_names)
                shared_pcols = tuple(name for name in os_pcol_names if name in ns_pcol_names)
                if not popup.add_new_dcols and not popup.add_new_pcols:
                    for row in self.new_sheet:
                        if row[ns_ic].lower() in new_ids:
                            newrow = list(repeat("", self.row_len))
                            newrow[self.ic] = row[ns_ic]
                            self.changelog_append_no_unsaved(
                                "Merge | Add ID",
                                f"Name: {newrow[self.ic]} Parent: n/a - Top ID column #{self.hiers[0] + 1} named: {self.headers[self.hiers[0]].name}",
                                "",
                                "",
                            )
                            changes_made += 1
                            for name in shared_dcols:
                                if self.detail_is_valid_for_col(os_dcol_names[name], row[ns_dcol_names[name]]):
                                    newrow[os_dcol_names[name]] = row[ns_dcol_names[name]]
                                    hdr_idx = os_dcol_names[name]
                                    if newrow[hdr_idx] != "":
                                        self.changelog_append_no_unsaved(
                                            "Merge | Edit cell",
                                            f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                            "",
                                            f"{newrow[hdr_idx]}",
                                        )
                                        changes_made += 1
                            for name in shared_pcols:
                                newrow[os_pcol_names[name]] = row[ns_pcol_names[name]]
                                hdr_idx = os_pcol_names[name]
                                if newrow[hdr_idx] != "":
                                    self.changelog_append_no_unsaved(
                                        "Merge | Edit cell",
                                        f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                        "",
                                        f"{newrow[hdr_idx]}",
                                    )
                                    changes_made += 1
                            rows_to_insert.append(newrow)
                elif popup.add_new_dcols and not popup.add_new_pcols:
                    new_dcol_indexes = {
                        i: h.name.lower()
                        for i, h in enumerate(self.headers)
                        if h.name.lower() in ns_dcol_names and h.name.lower() not in os_dcol_names
                    }
                    for row in self.new_sheet:
                        if row[ns_ic].lower() in new_ids:
                            newrow = list(repeat("", self.row_len))
                            newrow[self.ic] = row[ns_ic]
                            self.changelog_append_no_unsaved(
                                "Merge | Add ID",
                                f"Name: {newrow[self.ic]} Parent: n/a - Top ID column #{self.hiers[0] + 1} named: {self.headers[self.hiers[0]].name}",
                                "",
                                "",
                            )
                            changes_made += 1
                            for idx, colname in new_dcol_indexes.items():
                                newrow[idx] = row[ns_dcol_names[colname]]
                                hdr_idx = idx
                                if newrow[hdr_idx] != "":
                                    self.changelog_append_no_unsaved(
                                        "Merge | Edit cell",
                                        f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                        "",
                                        f"{newrow[hdr_idx]}",
                                    )
                                    changes_made += 1
                            for name in shared_dcols:
                                if self.detail_is_valid_for_col(os_dcol_names[name], row[ns_dcol_names[name]]):
                                    newrow[os_dcol_names[name]] = row[ns_dcol_names[name]]
                                    hdr_idx = os_dcol_names[name]
                                    if newrow[hdr_idx] != "":
                                        self.changelog_append_no_unsaved(
                                            "Merge | Edit cell",
                                            f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                            "",
                                            f"{newrow[hdr_idx]}",
                                        )
                                        changes_made += 1
                            for name in shared_pcols:
                                newrow[os_pcol_names[name]] = row[ns_pcol_names[name]]
                                hdr_idx = os_pcol_names[name]
                                if newrow[hdr_idx] != "":
                                    self.changelog_append_no_unsaved(
                                        "Merge | Edit cell",
                                        f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                        "",
                                        f"{newrow[hdr_idx]}",
                                    )
                                    changes_made += 1
                            rows_to_insert.append(newrow)
                elif popup.add_new_pcols and not popup.add_new_dcols:
                    new_pcol_indexes = {
                        i: h.name.lower()
                        for i, h in enumerate(self.headers)
                        if h.name.lower() in ns_pcol_names and h.name.lower() not in os_pcol_names
                    }
                    for row in self.new_sheet:
                        if row[ns_ic].lower() in new_ids:
                            newrow = list(repeat("", self.row_len))
                            newrow[self.ic] = row[ns_ic]
                            self.changelog_append_no_unsaved(
                                "Merge | Add ID",
                                f"Name: {newrow[self.ic]} Parent: n/a - Top ID column #{self.hiers[0] + 1} named: {self.headers[self.hiers[0]].name}",
                                "",
                                "",
                            )
                            changes_made += 1
                            for idx, colname in new_pcol_indexes.items():
                                newrow[idx] = row[ns_pcol_names[colname]]
                                hdr_idx = idx
                                if newrow[hdr_idx] != "":
                                    self.changelog_append_no_unsaved(
                                        "Merge | Edit cell",
                                        f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                        "",
                                        f"{newrow[hdr_idx]}",
                                    )
                                    changes_made += 1
                            for name in shared_dcols:
                                if self.detail_is_valid_for_col(os_dcol_names[name], row[ns_dcol_names[name]]):
                                    newrow[os_dcol_names[name]] = row[ns_dcol_names[name]]
                                    hdr_idx = os_dcol_names[name]
                                    if newrow[hdr_idx] != "":
                                        self.changelog_append_no_unsaved(
                                            "Merge | Edit cell",
                                            f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                            "",
                                            f"{newrow[hdr_idx]}",
                                        )
                                        changes_made += 1
                            for name in shared_pcols:
                                newrow[os_pcol_names[name]] = row[ns_pcol_names[name]]
                                hdr_idx = os_pcol_names[name]
                                if newrow[hdr_idx] != "":
                                    self.changelog_append_no_unsaved(
                                        "Merge | Edit cell",
                                        f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                        "",
                                        f"{newrow[hdr_idx]}",
                                    )
                                    changes_made += 1
                            rows_to_insert.append(newrow)
                elif popup.add_new_pcols and popup.add_new_dcols:
                    new_dcol_indexes = {
                        i: h.name.lower()
                        for i, h in enumerate(self.headers)
                        if h.name.lower() in ns_dcol_names and h.name.lower() not in os_dcol_names
                    }
                    new_pcol_indexes = {
                        i: h.name.lower()
                        for i, h in enumerate(self.headers)
                        if h.name.lower() in ns_pcol_names and h.name.lower() not in os_pcol_names
                    }
                    for row in self.new_sheet:
                        if row[ns_ic].lower() in new_ids:
                            newrow = list(repeat("", self.row_len))
                            newrow[self.ic] = row[ns_ic]
                            self.changelog_append_no_unsaved(
                                "Merge | Add ID",
                                f"Name: {newrow[self.ic]} Parent: n/a - Top ID column #{self.hiers[0] + 1} named: {self.headers[self.hiers[0]].name}",
                                "",
                                "",
                            )
                            changes_made += 1
                            for idx, colname in new_dcol_indexes.items():
                                newrow[idx] = row[ns_dcol_names[colname]]
                                hdr_idx = idx
                                if newrow[hdr_idx] != "":
                                    self.changelog_append_no_unsaved(
                                        "Merge | Edit cell",
                                        f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                        "",
                                        f"{newrow[hdr_idx]}",
                                    )
                                    changes_made += 1
                            for idx, colname in new_pcol_indexes.items():
                                newrow[idx] = row[ns_pcol_names[colname]]
                                hdr_idx = idx
                                if newrow[hdr_idx] != "":
                                    self.changelog_append_no_unsaved(
                                        "Merge | Edit cell",
                                        f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                        "",
                                        f"{newrow[hdr_idx]}",
                                    )
                                    changes_made += 1
                            for name in shared_dcols:
                                if self.detail_is_valid_for_col(os_dcol_names[name], row[ns_dcol_names[name]]):
                                    newrow[os_dcol_names[name]] = row[ns_dcol_names[name]]
                                    hdr_idx = os_dcol_names[name]
                                    if newrow[hdr_idx] != "":
                                        self.changelog_append_no_unsaved(
                                            "Merge | Edit cell",
                                            f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                            "",
                                            f"{newrow[hdr_idx]}",
                                        )
                                        changes_made += 1
                            for name in shared_pcols:
                                newrow[os_pcol_names[name]] = row[ns_pcol_names[name]]
                                hdr_idx = os_pcol_names[name]
                                if newrow[hdr_idx] != "":
                                    self.changelog_append_no_unsaved(
                                        "Merge | Edit cell",
                                        f"ID: {newrow[self.ic]} column #{hdr_idx + 1} named: {self.headers[hdr_idx].name} with type: {self.headers[hdr_idx].type_}",
                                        "",
                                        f"{newrow[hdr_idx]}",
                                    )
                                    changes_made += 1
                            rows_to_insert.append(newrow)

            # overwrite details for same ids with shared detail columns
            if popup.overwrite_details:
                shared_dcols = {name: idx for name, idx in os_dcol_names.items() if name in ns_dcol_names}
                for rn in range(len(self.sheet.MT.data)):
                    row = self.sheet.MT.data[rn]
                    if rn in shared_ids:
                        ns_rn = ns_rns[shared_ids[rn]]
                        for name, idx in shared_dcols.items():
                            ns_dcol_idx = ns_dcol_names[name]
                            if (
                                self.detail_is_valid_for_col(idx, self.new_sheet[ns_rn][ns_dcol_idx])
                                and row[idx] != self.new_sheet[ns_rn][ns_dcol_idx]
                            ):
                                self.changelog_append_no_unsaved(
                                    "Merge | Edit cell",
                                    f"ID: {row[self.ic]} column #{idx + 1} named: {self.headers[idx].name} with type: {self.headers[idx].type_}",
                                    f"{row[idx]}",
                                    self.new_sheet[ns_rn][ns_dcol_idx],
                                )
                                changes_made += 1
                                row[idx] = self.new_sheet[ns_rn][ns_dcol_idx]
                    self.sheet.MT.data[rn] = row

            # overwrite parents for same ids with shared parent columns
            if popup.overwrite_parents:
                shared_pcols = {name: idx for name, idx in os_pcol_names.items() if name in ns_pcol_names}
                for rn in range(len(self.sheet.MT.data)):
                    row = self.sheet.MT.data[rn]
                    if rn in shared_ids:
                        ns_rn = ns_rns[shared_ids[rn]]
                        for name, idx in shared_pcols.items():
                            ns_pcol_idx = ns_pcol_names[name]
                            if row[idx] != self.new_sheet[ns_rn][ns_pcol_idx]:
                                self.changelog_append_no_unsaved(
                                    "Merge | Edit cell",
                                    f"ID: {row[self.ic]} column #{idx + 1} named: {self.headers[idx].name} with type: {self.headers[idx].type_}",
                                    f"{row[idx]}",
                                    self.new_sheet[ns_rn][ns_pcol_idx],
                                )
                                changes_made += 1
                                row[idx] = self.new_sheet[ns_rn][ns_pcol_idx]
                    self.sheet.MT.data[rn] = row

            if rows_to_insert:
                self.sheet.insert_rows(rows_to_insert, insert_row)
            if changes_made:
                self.changelog_append(
                    f"Merged sheets making {changes_made} {'changes' if changes_made > 1 else 'change'}",
                    f"{'With file:' if popup.file_opened else ''} {popup.file_opened}",
                    "",
                    "",
                )
                self.new_sheet = []
                self.nodes = {}
                self.clear_copied_details()
                self.auto_sort_nodes_bool = True
                self.sheet.MT.data, self.nodes, self.warnings = TreeBuilder().build(
                    self.sheet.MT.data,
                    self.new_sheet,
                    self.row_len,
                    self.ic,
                    self.hiers,
                    self.nodes,
                    warnings=self.warnings,
                    add_warnings=True,
                    strip=not self.allow_spaces_ids_var,
                )
                self.new_sheet = []
                self.fix_associate_sort(startup=False)
                self.refresh_hier_dropdown(self.hiers.index(self.pc))
                self.rns = {r[self.ic].lower(): i for i, r in enumerate(self.sheet.data)}
                self.sheet.deselect()
                self.set_headers()
                self.refresh_formatting()
                self.reset_tagged_ids_dropdowns()
                self.rehighlight_tagged_ids()
                self.redo_tree_display()
                self.refresh_dropdowns()
                self.show_warnings("n/a - Data imported from: " + popup.file_opened, popup.sheet_opened)
            else:
                self.vs.pop()
                self.vp -= 1
                self.set_undo_label()
                Error(self, "No applicable changes were made", theme=self.C.theme)
            self.stop_work(self.get_tree_editor_status_bar_text())
            self.focus_sheet()
        except Exception as error_msg:
            Error(self, f"Error: {error_msg}", theme=self.C.theme)

    def get_par_lvls(self, h: int, iid: str, lvl=1):
        current_iid = iid
        current_level = lvl
        while self.nodes[current_iid].ps[h]:
            pid = self.nodes[current_iid].ps[h]
            self.levels[current_level] = self.nodes[pid].name
            current_iid = pid
            current_level += 1

    def export_flattened(self, event=None):
        self.start_work("Flattening sheet...")
        self.new_sheet = []
        Export_Flattened_Popup(self, theme=self.C.theme)
        self.stop_work(self.get_tree_editor_status_bar_text())
        self.new_sheet = []

    def get_save_json(self, program_data=False):
        if not program_data:
            d = full_sheet_to_dict(
                [h.name for h in self.headers],
                self.sheet.MT.data,
                format_=self.json_format,
            )
        else:
            d = {}
        if self.save_with_program_data:
            d["version"] = software_version_number
            if not program_data:
                d["changelog"] = self.changelog
            d["program_data"] = dict_x_b32(self.get_program_data_dict())
        return d

    def get_program_data_dict(self, sheetname="n/a"):
        d = {}
        d["records"] = self.sheet.data
        d["ic"] = self.ic
        d["pc"] = self.pc
        d["hiers"] = self.hiers
        d["headers"] = [
            {
                "name": h.name,
                "type": h.type_,
                "formatting": h.formatting,
                "validation": h.validation,
            }
            for h in self.headers
        ]
        d["nodes"] = self.jsonify_nodes()
        d["changelog"] = self.changelog
        d["row_heights"] = self.sheet.get_safe_row_heights()
        d["column_widths"] = self.sheet.get_column_widths()
        d["sheet_column_alignments"] = self.sheet.get_column_alignments()
        d["sheet_table_align"] = self.sheet.table_align()
        d["sheet_header_align"] = self.sheet.header_align()
        d["sheet_index_align"] = self.sheet.index_align()
        d["saved_info"] = self.save_info_get_saved_info()
        d["tv_label_col"] = self.tv_label_col
        d["topnodes_order"] = self.topnodes_order
        d["tagged_ids"] = list(self.tagged_ids)
        d["auto_sort_nodes_bool"] = self.auto_sort_nodes_bool
        d["sheetname"] = sheetname
        d["allow_spaces_ids"] = self.allow_spaces_ids_var
        d["allow_spaces_columns"] = self.allow_spaces_columns_var
        d["date_format"] = self.DATE_FORM
        return d

    def jsonify_nodes(self):
        return {
            n.name: {
                "cn": n.cn,
                "ps": n.ps,
            }
            for n in self.nodes.values()
        }

    def nodes_json_x_dict(self, njson: dict, hiers: Sequence[int]) -> dict:
        nodes = {}
        for name, nodedict in njson.items():
            node = Node(
                name=name,
                hrs=hiers,
            )
            node.ps = {int(h): pk for h, pk in nodedict["ps"].items()}
            node.cn = {int(h): cnl for h, cnl in nodedict["cn"].items()}
            nodes[name.lower()] = node
        return nodes

    def xlsx_chunker(self, seq):
        size = len(seq) if len(seq) <= 32000 else 32000
        return (seq[pos : pos + size] for pos in range(0, len(seq), size))

    def write_program_data_to_workbook(self, wb, sheetnames_):
        with suppress(Exception):
            wb.remove(wb["program_data"])
        ws = wb.create_sheet(title="program_data")
        ws.append([f"{software_version_number}"])
        for chunk in self.xlsx_chunker(dict_x_b32(self.get_program_data_dict(sheetnames_[1]))):
            ws.append([chunk])
        ws.sheet_state = "hidden"

    def write_changelog_to_workbook(self, wb, sheetnames_):
        sheetname = sheetnames_[1]
        new_title1 = sheetname + " Changelog"
        for sname in (i for i in wb.sheetnames if "Changelog" in i):
            try:
                wb.remove(wb[sname])
            except Exception:
                continue
        ws = wb.create_sheet(title=new_title1)
        ws.append(xlsx_changelog_header(ws))
        for r in reversed(self.changelog):
            ws.append((e if e else None for e in r))

    def write_flattened_to_workbook(self, wb, sheetnames_):
        sheetname = sheetnames_[1]
        new_title1 = sheetname + " Flattened"
        for sname in (i for i in wb.sheetnames if "flattened" in i or "Flattened" in i):
            try:
                wb.remove(wb[sname])
            except Exception:
                continue
        ws = wb.create_sheet(title=new_title1)
        ws.freeze_panes = "A2"
        self.new_sheet = []
        for r in TreeBuilder().build_flattened(
            input_sheet=self.sheet.MT.data,
            output_sheet=self.new_sheet,
            nodes=self.nodes,
            headers=[f"{hdr.name}" for hdr in self.headers],
            ic=int(self.ic),
            pc=int(self.pc),
            hiers=list(self.hiers),
            detail_columns=self.xlsx_flattened_detail_columns,
            justify_left=self.xlsx_flattened_justify,
            reverse=self.xlsx_flattened_reverse_order,
            add_index=self.xlsx_flattened_add_index,
        ):
            ws.append((e if e else None for e in r))
        self.new_sheet = []

    def write_treeview_to_workbook(self, wb: Workbook, sheetnames_):
        sheetname = sheetnames_[1]
        new_title1 = sheetname + " Treeview"
        for sname in (i for i in wb.sheetnames if "Treeview" in i):
            try:
                wb.remove(wb[sname])
            except Exception:
                continue
        ws = wb.create_sheet(title=new_title1)
        ws.freeze_panes = "A2"
        oldpc = int(self.pc)
        self.levels = defaultdict(list)
        for h in self.hiers:
            for iid, node in self.nodes.items():
                if node.ps[h] and not node.cn[h]:
                    self.get_par_lvls(h, iid)
        maxlvls = max(self.levels, default=0) + 1
        self.xl_tv_detail_cols = tuple(i for i, h in enumerate(self.headers) if h.type_ not in ("ID", "Parent"))
        self.level_colors = tuple(tv_lvls_colors[level_to_color(i)] for i in range(maxlvls + 1))
        cycle_colors = cycle(tv_lvls_colors)

        # Write header row
        row = []
        for lvl in range(1, maxlvls + 1):
            cell = WriteOnlyCell(ws, value=lvl)
            cell.fill = next(cycle_colors)
            row.append(cell)
        for _, hdr in enumerate((h for h in self.headers if h.type_ not in ("ID", "Parent")), start=maxlvls + 1):
            cell = WriteOnlyCell(ws, value=hdr.name)
            row.append(cell)
        ws.append(row)

        # Process hierarchies iteratively
        for h in self.hiers:
            self.pc = int(h)  # Ensure self.pc is an integer as expected elsewhere
            self.hier_disp = f"{self.headers[self.pc].name} - "

            # Initialize stack with top nodes at level 1
            stack = [(iid, 1) for iid in self.top_iids()]

            # Process nodes using the stack
            while stack:
                iid, level = stack.pop()
                # Construct row with indentation
                row = list(repeat(None, level - 1))
                cell = WriteOnlyCell(
                    ws,
                    value=f"{self.hier_disp}{self.sheet.MT.data[self.rns[iid]][self.tv_label_col]}",
                )
                cell.fill = self.level_colors[level - 1]
                row.append(cell)
                # Add None values up to detail columns
                num_nones = maxlvls - level
                row.extend(list(repeat(None, num_nones)))
                # Add detail columns
                for col in self.xl_tv_detail_cols:
                    cell = WriteOnlyCell(ws, value=self.sheet.MT.data[self.rns[iid]][col])
                    row.append(cell)
                ws.append(row)
                # Push children in reverse order to maintain original order
                for ciid in reversed(self.nodes[iid].cn[self.pc]):
                    stack.append((ciid, level + 1))

        # Restore state
        self.pc = int(oldpc)
        self.levels = defaultdict(list)

    def write_additional_sheets_to_workbook(self, new_sheet_name=None):
        if self.save_xlsx_with_flattened:
            self.C.status_bar.change_text("Saving flattened sheet...")
            self.write_flattened_to_workbook(
                self.C.wb,
                (
                    self.C.open_dict["sheet"],
                    self.C.open_dict["sheet"] if new_sheet_name is None else new_sheet_name,
                ),
            )
        if self.save_xlsx_with_changelog:
            self.C.status_bar.change_text("Saving changelog...")
            self.write_changelog_to_workbook(
                self.C.wb,
                (
                    self.C.open_dict["sheet"],
                    self.C.open_dict["sheet"] if new_sheet_name is None else new_sheet_name,
                ),
            )
        if self.save_xlsx_with_treeview:
            self.C.status_bar.change_text("Saving treeview...")
            self.write_treeview_to_workbook(
                self.C.wb,
                (
                    self.C.open_dict["sheet"],
                    self.C.open_dict["sheet"] if new_sheet_name is None else new_sheet_name,
                ),
            )
        if self.save_xlsx_with_program_data:
            self.C.status_bar.change_text("Saving program data...")
            self.write_program_data_to_workbook(
                self.C.wb,
                (
                    self.C.open_dict["sheet"],
                    self.C.open_dict["sheet"] if new_sheet_name is None else new_sheet_name,
                ),
            )
        self.C.status_bar.change_text("Writing file...")

    def save_workbook(self, filepath, sheetname):
        self.C.wb = Workbook(write_only=True)
        ws = self.C.wb.create_sheet(title=sheetname)
        if not self.ic:
            ws.freeze_panes = "B2"
        else:
            ws.freeze_panes = "A2"
        for row in self.gen_sheet_w_headers():
            ws.append(row)
        self.write_additional_sheets_to_workbook(sheetname)
        self.C.wb.active = self.C.wb[sheetname]
        filepath = convert_old_xl_to_xlsx(filepath)
        self.C.wb.save(filepath)
        self.C.open_dict["filepath"] = filepath
        self.C.change_app_title(title=os.path.basename(filepath))
        return True

    def save_csv(self, filepath):
        with open(filepath, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(
                fh,
                dialect=csv.excel_tab if filepath.lower().endswith(".tsv") else csv.excel,
                lineterminator="\n",
            )
            writer.writerows(self.gen_sheet_w_headers())
        self.C.open_dict["filepath"] = filepath
        self.C.change_app_title(title=os.path.basename(filepath))
        self.C.open_dict["sheet"] = "Sheet1"
        return True

    def save_json(self, filepath):
        with open(filepath, "w") as fh:
            fh.write(json.dumps(self.get_save_json(), indent=4))
        self.C.open_dict["filepath"] = filepath
        self.C.change_app_title(title=os.path.basename(filepath))
        self.C.open_dict["sheet"] = "Sheet1"
        return True

    def save_(self, event=None, quitting=False):
        if self.C.current_frame != "tree_edit":
            return
        newfile = os.path.normpath(self.C.open_dict["filepath"])
        self.start_work("Saving... ")
        successful = False
        try:
            if newfile.lower().endswith((".csv", ".tsv")):
                successful = self.save_csv(newfile)
            elif newfile.lower().endswith(".json"):
                successful = self.save_json(newfile)
            elif newfile.lower().endswith((".xlsx", ".xls", ".xlsm")):
                successful = self.save_workbook(newfile, self.C.open_dict["sheet"])
                self.C.try_to_close_workbook()
        except Exception as error_msg:
            Error(self, f"Error: {error_msg}", theme=self.C.theme)
        if successful:
            self.C.created_new = False
            self.bind_or_unbind_save("normal")
            self.C.unsaved_changes = False
        if not quitting:
            self.stop_work(self.get_tree_editor_status_bar_text())
        return successful

    def save_as(self, event=None, quitting=False):
        if self.C.current_frame != "tree_edit":
            return
        newfile = filedialog.asksaveasfilename(
            parent=self.C,
            title="Save as",
            filetypes=[
                ("Excel file", ".xlsx"),
                ("JSON file", ".json"),
                ("CSV File (Comma separated values)", ".csv"),
                ("TSV File (Tab separated values)", ".tsv"),
            ],
            defaultextension=".xlsx",
            confirmoverwrite=True,
        )
        if not newfile:
            return False
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".json", ".csv", ".xlsx", ".tsv")):
            Error(self, "Can only write .json, .xlsx or .csv    ", theme=self.C.theme)
            return False
        self.start_work("Saving... ")
        successful = False
        try:
            if newfile.lower().endswith((".csv", ".tsv")):
                successful = self.save_csv(newfile)
            elif newfile.lower().endswith(".json"):
                successful = self.save_json(newfile)
            elif newfile.lower().endswith(".xlsx"):
                popup = Enter_Sheet_Name_Popup(self, theme=self.C.theme)
                if popup.result and all(
                    reserved not in popup.result.lower()
                    for reserved in (
                        "program_data",
                        " treeview",
                        " changelog",
                        " flattened",
                    )
                ):
                    successful = self.save_workbook(newfile, popup.result)
                else:
                    error_msg = "Enter a sheet name / sheet name must not be equal to 'program data'   "
                    Error(self, error_msg, theme=self.C.theme)
        except Exception as error_msg:
            Error(self, f"Error: {error_msg}", theme=self.C.theme)
        if successful:
            self.C.created_new = False
            self.bind_or_unbind_save("normal")
            self.C.unsaved_changes = False
        if not quitting:
            self.stop_work(self.get_tree_editor_status_bar_text())
        return successful

    def save_new_vrsn(self, event=None):
        newfile = self.C.open_dict["filepath"]
        folder = os.path.dirname(newfile)
        if not folder:
            folder = os.path.dirname(os.path.abspath(__file__))
        popup = Save_New_Version_Presave_Popup(self, folder, theme=self.C.theme)
        if not popup.result:
            return False
        self.start_work("Saving... ")
        folder = popup.result
        newfile = os.path.join(folder, os.path.basename(newfile))
        if newfile.lower().endswith((".csv", ".xls", ".tsv")):
            ext = newfile[-4:]
            path = newfile[:-4]
        elif newfile.lower().endswith((".xlsx", ".json", ".xlsm")):
            ext = newfile[-5:]
            path = newfile[:-5]
        else:
            Error(
                self,
                "Error saving file, file extension must be .csv/.xlsx/.json   ",
                theme=self.C.theme,
            )
            self.stop_work(self.get_tree_editor_status_bar_text())
            return False
        last_index = 0
        for i, c in enumerate(reversed(path), 1):
            if c.isdigit():
                last_index = i
            else:
                break
        newfile_without_numbers = path[:-last_index] + ext
        matches = {}
        found_suitable_folder = False
        while not found_suitable_folder:
            try:
                for file in os.listdir(folder):
                    if (
                        file.lower().endswith((".json", ".xlsx", ".csv", ".xls", ".xlsm", ".tsv"))
                        and path_without_numbers(file) == newfile_without_numbers
                    ):
                        matches[file] = path_numbers(file)
                found_suitable_folder = True
            except Exception:
                popup = Save_New_Version_Error_Popup(self, theme=self.C.theme)
                if popup.result:
                    folder = os.path.normpath(
                        filedialog.askdirectory(
                            parent=self.C,
                            title="Select a folder to save new version in",
                        )
                    )
                    if folder == ".":
                        self.stop_work(self.get_tree_editor_status_bar_text())
                        return False
                    newfile = os.path.join(folder, os.path.basename(newfile))
                else:
                    self.stop_work(self.get_tree_editor_status_bar_text())
                    return False
        if matches:
            latest_num = float("-inf")
            latest_name = None
            for k, v in matches.items():
                if v > latest_num:
                    latest_num = v
                    latest_name = k
            newfile = os.path.join(folder, increment_file_version(latest_name))
        else:
            newfile = increment_file_version(newfile)
        x = 0
        while os.path.isfile(newfile):
            newfile = increment_file_version(newfile)
            x += 1
            if x > 200:
                Error(
                    self,
                    "Error saving file, could not get name for new version   ",
                    theme=self.C.theme,
                )
                self.stop_work(self.get_tree_editor_status_bar_text())
                return False
        successful = False
        try:
            if newfile.lower().endswith((".csv", ".tsv")):
                successful = self.save_csv(newfile)
            elif newfile.lower().endswith(".json"):
                successful = self.save_json(newfile)
            elif newfile.lower().endswith((".xlsx", ".xls", ".xlsm")):
                successful = self.save_workbook(newfile, self.C.open_dict["sheet"])
                self.C.try_to_close_workbook()
        except Exception as error_msg:
            Error(self, f"Error: {error_msg}", theme=self.C.theme)
        if successful:
            self.C.unsaved_changes = False
            popup = Save_New_Version_Postsave_Popup(self, folder, os.path.basename(newfile), theme=self.C.theme)
        self.stop_work(self.get_tree_editor_status_bar_text())
        return successful
