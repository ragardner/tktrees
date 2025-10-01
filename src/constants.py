# SPDX-License-Identifier: AGPL-3.0-only
# Copyright (c) 2025 R. A. Gardner

import datetime
import os
import re
import tkinter as tk
from platform import (
    release as get_os_version,
)
from platform import (
    system as get_os,
)
from sys import (
    version as get_python_version,
)

from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import Color
from tksheet import (
    DotDict,
    theme_black,
    theme_dark,
    theme_dark_blue,
    theme_light_blue,
    theme_light_green,
)

_ts_path = os.path.realpath(__file__)
current_dir = os.path.join(os.path.normpath(os.path.dirname(_ts_path)), "")
upone_dir = os.path.join(os.path.normpath(os.path.dirname(os.path.dirname(_ts_path))), "")

# ________________________ OS BINDINGS ________________________

USER_OS = f"{get_os()}".lower()
USER_OS_VERSION = f"{get_os_version()}"
USER_PYTHON_VERSION = f"{get_python_version}"
USER_TK_VERSION = f"{tk.TkVersion}"
USER_TCL_VERSION = f"{tk.TclVersion}"

rc_button = "<2>" if USER_OS == "darwin" else "<3>"
rc_press = "<ButtonPress-2>" if USER_OS == "darwin" else "<ButtonPress-3>"
rc_motion = "<B2-Motion>" if USER_OS == "darwin" else "<B3-Motion>"
rc_release = "<ButtonRelease-2>" if USER_OS == "darwin" else "<ButtonRelease-3>"
ctrl_button = "Command" if USER_OS == "darwin" else "Control"
from_clipboard_delimiters = "\t,|"

software_version_number = "1.14.6"
software_version_full = "Version: " + software_version_number
app_title = "tktrees"
contact_email = "github@ragardner.simplelogin.com"
website1 = "github.com/ragardner"
current_year = f"{datetime.datetime.now().year}"
app_copyright = f"Copyright © 2019-{current_year} R. A. Gardner."
contact_info = f" {software_version_full}\n {app_copyright}\n {contact_email}\n {website1}"
about_system = "\n".join(
    (
        f"TkTrees: {software_version_number}",
        f"OS: {USER_OS}",
        f"OS Version: {USER_OS_VERSION}",
        f"Python: {USER_PYTHON_VERSION}",
        f"Tk: {USER_TK_VERSION}",
        f"Tcl: {USER_TCL_VERSION}",
    )
)
config_name = ".tktrees.json"
default_app_window_size = (1000, 760)

if USER_OS == "darwin":
    TF = ("Calibri", 16, "bold")
    BF = ("Calibri", 13, "normal")
    BFB = ("Calibri", 13, "bold")
    STSF = ("Calibri", 13, "bold")
    EF = ("Calibri", 13, "normal")
    EFB = ("Calibri", 14, "bold")
    ERR_ASK_FNT = ("Calibri", 13, "bold")
    std_font_size = 13
    lge_font_size = 15
    sheet_header_font = ("Calibri", std_font_size, "normal")
else:
    TF = ("Calibri", 15, "bold")
    BF = ("Calibri", 11, "normal")
    BFB = ("Calibri", 11, "bold")
    STSF = ("Calibri", 11, "bold")
    EF = ("Calibri", 11, "normal")
    EFB = ("Calibri", 13, "bold")
    ERR_ASK_FNT = ("Calibri", 12, "bold")
    std_font_size = 11
    lge_font_size = 14
    sheet_header_font = ("Calibri", std_font_size, "normal")
dropdown_font = "TkFixedFont"
lge_font = ("Calibri", lge_font_size, "normal")

tree_bindings = (
    "single",
    "ctrl_select",
    "drag_select",
    "rc_select",
    "select_all",
    "row_select",
    "column_select",
    "column_height_resize",
    "arrowkeys",
    "column_width_resize",
    "double_click_column_resize",
    "row_height_resize",
    "double_click_row_resize",
    "row_width_resize",
    "edit_cell",
    "column_drag_and_drop",
    "row_drag_and_drop",
    "find",
    "replace",
)
sheet_bindings = tree_bindings

detail_column_types = {"Text", "Number", "Date"}

validation_allowed_num_chars = {"0", "1", "2", "3", "4", "5", "6", "7", "8", "9", ",", "-", ".", "e"}
validation_allowed_date_chars = {"0", "1", "2", "3", "4", "5", "6", "7", "8", "9", ",", "/", "-", " "}

# dict to maintain order
date_formats_usable = {
    "%d/%m/%Y": None,
    "%m/%d/%Y": None,
    "%Y/%m/%d": None,
    "%d-%m-%Y": None,
    "%m-%d-%Y": None,
    "%Y-%m-%d": None,
}
# date_formats_entry = date_formats_usable.copy()
# date_formats_entry["%B %d, %Y"] = None, # Full month name, e.g., January 01, 2023
# date_formats_entry["%b %d, %Y"] = None, # Abbreviated month name, e.g., Jan 01, 2023

themes = DotDict(
    {
        "light_blue": theme_light_blue,
        "light_green": theme_light_green,
        "dark": theme_dark,
        "black": theme_black,
        "dark_blue": theme_dark_blue,
    }
)

# BUILD START WARNINGS HEADER
warnings_header = """## TREE BUILD WARNINGS"""

colors = (
    "white",
    "gray93",
    "LightSkyBlue1",
    "light grey",
    "antique white",
    "papaya whip",
    "bisque",
    "peach puff",
    "navajo white",
    "NavajoWhite2",
    "wheat1",
    "wheat2",
    "khaki",
    "pale goldenrod",
    "gold",
    "LightGoldenrod1",
    "LightGoldenrod2",
    "goldenrod1",
    "goldenrod2",
    "LightYellow2",
    "LightYellow3",
    "RosyBrown1",
    "burlywood1",
    "LemonChiffon2",
    "LemonChiffon3",
    "cornsilk2",
    "ivory2",
    "sky blue",
    "light sky blue",
    "light blue",
    "LightBlue1",
    "powder blue",
    "CadetBlue1",
    "pale turquoise",
    "medium turquoise",
    "turquoise",
    "medium aquamarine",
    "aquamarine2",
    "medium sea green",
    "dark sea green",
    "DarkSeaGreen2",
    "DarkOliveGreen2",
    "salmon1",
    "coral1",
    "pink",
    "pink1",
    "orchid1",
    "plum1",
    "LavenderBlush2",
    "MistyRose2",
    "thistle",
    "thistle1",
    "thistle2",
    "thistle3",
)

menu_kwargs = {
    "font": ("Calibri", std_font_size, "normal"),
    "background": "#f2f2f2",
    "foreground": "gray2",
    "activebackground": "#91c9f7",
    "activeforeground": "black",
}

openpyxl_thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

isrealre = re.compile(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?$")
isfloatre = re.compile(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?$")
isintre = re.compile(r"[-+]?\d+$")
isintlikere = re.compile(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?$")
remove_whitespace = re.compile(r"\s+")
remove_nrt = re.compile(r"[\n\r\t]")

blue_fill = PatternFill(start_color=Color("0078d7"), end_color=Color("0078d7"), fill_type="solid")
green_fill = PatternFill(start_color=Color("648748"), end_color=Color("648748"), fill_type="solid")
orange_fill = PatternFill(start_color=Color("FFA51E"), end_color=Color("FFA51E"), fill_type="solid")
slate_fill = PatternFill(start_color=Color("E1E1E1"), end_color=Color("E1E1E1"), fill_type="solid")
tan_fill = PatternFill(start_color=Color("EDEBE1"), end_color=Color("EDEBE1"), fill_type="solid")
green_add_fill = PatternFill(start_color=Color("E6FFED"), end_color=Color("E6FFED"), fill_type="solid")
red_remove_fill = PatternFill(start_color=Color("FFEEF0"), end_color=Color("FFEEF0"), fill_type="solid")
openpyxl_left_align = Alignment(horizontal="left")
openpyxl_center_align = Alignment(horizontal="center")
tv_lvls_colors = [
    PatternFill(start_color=Color("d2abff"), end_color=Color("d2abff"), fill_type="solid"),
    PatternFill(start_color=Color("88d2fc"), end_color=Color("88d2fc"), fill_type="solid"),
    PatternFill(start_color=Color("A0C36C"), end_color=Color("A0C36C"), fill_type="solid"),
    PatternFill(start_color=Color("FFEC87"), end_color=Color("FFEC87"), fill_type="solid"),
    PatternFill(start_color=Color("fea05f"), end_color=Color("fea05f"), fill_type="solid"),
]

changelog_header = [
    "Date YYYY/MM/DD",
    "Type",
    "ID/Name/Number",
    "Old Value",
    "New Value",
]

align_w_icon = (
    """iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAAAXNSR0IArs4c6QAAAWFJREFUSEuVVdtx"""
    """wzAMIztZOkrHSM9xEjdp18gm8WbuRQ+LtADJ0Y99ehAiCEIqcKiILGkl/YeP/1dZtrs20VTUhgqrNhAA"""
    """r6DTfgtW4kgCAEFZoJIZzt3Pogw8MQ0GLV0pcXD2tYuvNiny5FZUG7DwO93+niLLIc6rBE7ZqQq4Kv58"""
    """Gr4/891DBtfpfvlQPWftuEIZTTWZXy+kj3E4fhUAdqqTARDzhvEXE93BwpCKul6J5e3qBvRY1YetsrCr"""
    """eHW9K7cslkzAdPt9qsghZrhHPmHPPAa18P0rRdef+0VVz44w3M4mXlELo9p5UbOcnVozBtsq8m5QGnBf"""
    """M0RM7KY+QqUiIJnS+bWFeBUxZ93TDuRs8iKrIp4BLqbO43AM3rMdq5tCL7KVayoXqSlC1SoCVJQpgNIG"""
    """TlbBNr1hdpa68nxms+vLBDxK6c1gkk0xaR8wTNbcGMfZtb0RLwRmjXP5D6Z1tiGagIPwAAAAAElFTkSu"""
    """QmCC"""
)

align_c_icon = (
    """iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAAAXNSR0IArs4c6QAAAXFJREFUSEuNlVl6"""
    """gzAMhCVOlhylx2jCZxJKeg5uktyMFq+yNTLkJSzGWubXmCn/mIi2ciuu5JtyHa/8X/9bIvUeb9lGt7dN"""
    """K9lvHQPsNxtMphtfRlXVMHE3i9MvZbvqOkMF6herKeWBNqrUjX3i43n5fTPRBcvcfHsgLBF93Hi77lll"""
    """DZ7z8mAeJi24LjkkAUDIgWl14/dXWpXXWuxIND0IIIAJCNbRUhdoEwssXau/FSKfH5wwVzIYRsVrIPVS"""
    """rTjAVLULDIyYg1qBeXm9/2m4BBJS361MEzUIeMOLnj+vBw88QXvSla1uvHlqFF9gurv9xaQB/dIj2Xfp"""
    """pmfoOnaSNGiotOzCBxocEKUxtSzG9jNpyL4JFZmtBrsnBXoaGwdVgjUfN96vkjrlpp4eHiaLcTmx7UYb"""
    """8+ruwYNSGfo8sHoKFO2CoCmC/ghOisLaqQBY0/r4VLOi9LAJ6FOknLJ+YFl0Sbo69Gu8rMq6wwgG+g8H"""
    """yssioq1ZgAAAAABJRU5ErkJggg=="""
)

align_e_icon = (
    """iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAAAXNSR0IArs4c6QAAAVtJREFUSEuVVVGW"""
    """wiAMTDyZHsVjuM9atXb3Gt5Eb4avCJQ0mcD2qwWaTDKTgcl8mIiCvbVdbRxdtoniofqk8d4RSEPiGFo+"""
    """KBDKnxAyhYRxCRpK/TpBSqeKqtddVBIJu41Jm24HjcT1+cLBffp7EYW9APffyGu73sP555DpjSTf7vN1"""
    """xzwKrlFbmoTzczifjkvgb4tUn+reILXKLOvXhoNVpkmtiMAm6jw5OrEBsU+rtRztefqCNhO0f056V0qz"""
    """KjA4yINT1LXVnhhbVfH7slURsp7bNF+ZedR2giYyNuV5GU7HRT2iRbYSJPmuYwA2y7KrVqXcmgMkWUEy"""
    """1r7qsJEM2TtQUY/ByixqYCuKTKFOj99XYNp79449LXF1VREaqaieHY/mRYTgFtTZiyh5Ue0XdT1G9m4V"""
    """JRAdkyxvKITFtorspu5t1dhEivIrQM6K1z1TR3voMjRLQlb1ATWvtSEFysTAAAAAAElFTkSuQmCC"""
)

nums_icon = (
    """iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAAAXNSR0IArs4c6QAAAjhJREFUSEuVVu11"""
    """4jAQ3EVJH6QTU0lCGWAhY/FVxnGVhE7OhYB1WVmyVlr53jv/yFMWr2Y/ZgYQ2IMA4PwfHgUQoRDg76d3"""
    """EBDcfA/Flx9+kX9LBEJuXkIOVr0+VFFNryX8D0B4tz9djoj46a8bXW9Me49X2/P1GwCa8P9gdPvB58rh"""
    """2IhS2NrLF6zwV1br6LYRJIB38XOjW0wARTclGu2XKkSA5uleG3iqQb3DHwAYjN5/cFB7vno6GL2vFhq3"""
    """JmhizzdHFcXE0/nq6Ka50lBVHUAycI7EbsrEBMAr9Z06IpZp+YhCvYxwc8TzN4zIL3F02xeOg0JFS30Y"""
    """3W78rHkHASDmlhzLdDB3EJfMac+WXLAoKuRx0PsNoScwOleerrutlRobUPgJDtbOud+daY9R4ZFFuZLx"""
    """ftC7bTlvrLFILKXWd5oU13g6B/TqiBbELajOLYHvpnQVwSt7mnQQPE/wnwbb2yWl5zOZRsSH+WNpuVIR"""
    """jN7NndKhDyTI0hgJsg5Em8zg+qpSJ6UDYPNyzw083wb17pjSC7suis/221+uXhjJCqbkCRih0zsMumGW"""
    """UTjxEgAhWQEw4WdxBLAn7kmLZpd/wC/q9N5XKuyalA7joFZR6SS0RatIwycWeavI23sYr1SAzM5jXbTk"""
    """rr3Xvm6Fm5Z+P92Bd6N3Wzobc1urt7GBlf9C8ko/HtpjZCP3JSk0IW3OivwsxMUVJZUsfw3EXxj/IsJ0"""
    """J3fFQmjCFsSu5fKrVsKCPOMvPi9aL6aCzOMAAAAASUVORK5CYII="""
)

letters_icon = (
    """iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAAAXNSR0IArs4c6QAAAo9JREFUSEuNVs1u"""
    """2kAQ/sZRlajXpkTwGhXKuXfad+BYCR7A4GwaY8wDGIkj7xC491zxHqSQ9FSpCod4qv2xvYvXFJ/Mena+"""
    """mW++mYFQPgSAzS8CgctflY3nTV1rtpefAdt3zYf5aBxVQTSBue40gEE5O+ryhgr+jAzsC0182I7sbFxm"""
    """VYLHRycp9jJYo8um0XVnUVRH1ul7zmuFM0X21NKpga2iE54RJ+kPiSGi0Wdf+lWCBGoSEAFkCsjHxY+n"""
    """qdKwiEbyujSVqRYadzBrMo3TWRd5fkegngbgNUAdAH+KiAsAgFcg6umG4TWCYCJGo40tZYeieJZ28cY/"""
    """oRLj3IQSFJSLsYoYBYDJ3rZjXNCtCMNNKX1bEJNk+sigL8xYvuyfvsnIrm86CyLuG841QJKq0ufMy9/S"""
    """DsCHVnsRBOiD85WIoq9Vd1mMTZI0Z4Cfd9v3WTY/SGqHg+Hlx5v2qwFQ2cRJqqJ+3j1dZVl2kGDD4UDa"""
    """/ZXZ30dhUA0dC8BcVADzbH6QRvpi51WCifFIA0zTXLbwfre9ms+zg8x0MBxethQASETaTgI7fRAn00co"""
    """imj5st+q1K9bnUVA3JdgIhrL2pgaEJhZU6ns2gtSFPFKRGMfRYSqyAo41+qjoJCplKU8elA1KJ+qyNIw"""
    """oNv7MNyUFB33QTybdfGW3wHU0y7yNSP4RaB3IgpVsR/i6XcK6JPSvpEpgdd8EUxE6JNprdvc1q8+u2PM"""
    """bkCPi3IYmEZ0B4/T7nL5eCZo7cg3i/zInkhPzBR3IB7d9UfRMBZ9q9F2UEqzWrfOqDhno/0nkUpbBtjp"""
    """A7uYFZi90N332ngvhpb1d8EB8M123Y8nVn2TfJxObqK8ISJ7zhRbzw7OdvcP9QdMLotFpAkAAAAASUVO"""
    """RK5CYII="""
)

date_icon = (
    """iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAAAXNSR0IArs4c6QAAAktJREFUSEuFVeuZ"""
    """2yAQXNCXOnzpRK4k5zKsB5IQksuIU4muk7iRM7nluRi48EcSIGZ3dnZgEAYDAF358tO1PW4+fQAAA5xy"""
    """I1sFue0HYk5Dd9ZmOQVY1X5gSGLozvGc/DgEij+Td6l2k5IYOhKICczsl2rT+C76l3X8ycdLz17WbWaM"""
    """/QKANxNHnpQNL02EBv+Ap14m0d8t2YQiKdU7cP47SdWhM9ChOvWzCfBTX4To7ymA2g8AaD/153kZxw88"""
    """aNl2jVRMw5X5iKgQVrVrW4Oe4fw03d6aH/ovADzE0P00AD4iqW5fR2kYDd921lF2mobu4vVFM5jXbeaM"""
    """n8RwvXiq5ZrWjQDsRiMWoD5izQrSZAAIYM4ZO4w3yhQVEwFq1U2BYzZxv1xvX8LWTnm0yE6SVYV8W908"""
    """Y1HLIGkomkgGUFvEqHWkKBY5586DRZlSlMK7qYGjKMnAFYd2LW1ASkCJd2ohr91vVUSRiUy9F3mvkbZX"""
    """gveUvCgD8NGlKrLyQp0zxk/TcDV9sKxqZqwxusf/TJ9wfhK9/cYRAKoUmYW0omUBFWZpH7h+Ck1VdE5a"""
    """iFp9U3c2GdCGJQC0QSxFywvnUt0OZydn3LGq7cDTRB/vgwyAesgrsq9BwrnznlAD891djAmWKKr1QdmN"""
    """/t/OaJrRKsC5qZXpAaBbCF7+neW93jsWeJJr2/AGpfzhpR3dVG7vwFl64YRbKV442U1X6kgMcurvqMYA"""
    """YC6L5tkCN1dmG+MvOGvdhh76qf9MYz/7HANAkZAa5RlmvTb/ALlVeC92xBz2AAAAAElFTkSuQmCC"""
)

tag_icon = (
    """iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAAAXNSR0IArs4c6QAAAm1JREFUSEuNVe1h"""
    """4jAMlSh0jrAJTNLeGECI4xzQMSiTpJuc57gS1EqWEztxrpcfkPjrSe89yQj8IACQ/Mhn+qTj8Vd+BwIC"""
    """9SfxGnns71ONuHgBgCKDEobcg+hWV4d6HFAemMEAoGnOr7DAa3zwJLp0wHWfuLV25/wenUz//AyP2fOl"""
    """BcJNR/etPR4/+g3x5m8CjXkrnlbUAkIBBCOQfN4+g9OF+NCq3PmcEinisBAqcymWz9QCCZUpSEaUCACg"""
    """KvfzABogb6gkE2gBqAAE1/1N6ZqInGSgB6kuhuhxM8eyDlyHzUzXckUtCV3ouk/Y2mbncuIrRZqB8t6c"""
    """zkoUOlPu1v4j5aDXxDuvpytxlIpMjCwa6EHetvgCD7KmOrxPZfHHTEFga+1e3TW4SCpDNPjhyRVXCqJ0"""
    """qYVnRP6hXjNOE3exhZWuqtyvmdIAwBObO3XbZq4O5jpJpM0ok/eq3P/yhaaVjADurhU6nBf3Fn2fJDis"""
    """UZA/LDpnIRmY6q14eqYWAQpSNzR254Jzxr4OggvOqD1EAKKpZMAbjLlo8bDlUqEyDh2Gomz0cO5pGyKy"""
    """5niovWt0kSyI2wBXaMOWy7XxtDX6AOkKgBsEcscg8pBm7GtoEShDV97Dmr1EHrgPKxPfD3Rp14wrVNpA"""
    """1JD0PaYlFrZfOtc8uaHFvvb9X+mKKVVaAEhc0+cYxI+dMFYu12vEXeidt1zRlZQWU+7X01sX/YWTybwP"""
    """JNcGeJIFRXYLgqsOUeRSveFO/g8AqZNwk4X7eoiqpyU2ywzAvyuWQRbLx+sC0XAxEtGNfT69AVMjzHTP"""
    """FExInHPDXBUqzheoS60vHv02dQAAAABJRU5ErkJggg=="""
)

search_icon = (
    """iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAAAXNSR0IArs4c6QAAAs9JREFUSEuVVj1z"""
    """00AQfSv7hzjpSB26TJj8AAhFCgpSUJBJz0ws2bIcK7JsMdBQZUgJMzQUYfIDGDJ0UCcd1g+Jb2FPXyfp"""
    """FOAKyz7d7dt9++6dCTIIAOuP+nf90hwEAsuqauTbWlN5THn8dZj4VUKWbZYcM4ASjctCZHoWLfeJcADi"""
    """HTAGUiERpcz0nRmfp777pQ1jliQVNykCEEXJ9prUW2Ls5ivyOEIOaSbloYDrPpxX/nj4s4uzFkVZ1nyp"""
    """NzCt4NA59XA1GQ5vZSpMkgdQ/BiKjgG1oZcxPTWrqddgVBBFy21F6kce/CLw3aPytYXys2j5nolfyiuH"""
    """nYe+L5XYKMo3z+aLb0ILgS4mefD7ACRUmIMwcD31vUelxIyG6/AVNbQKxt5mS45dWiMBSX4JXU2qpBoq"""
    """JBhGyQdAPQfRMBi5r1vyNiXYEH04X56AOQHoYzD2DjMVZKNschgvViJF6jlbE9e9NRfZk69SOJPG36kb"""
    """JqTByNONbwPMFxo28D2Sw2BSZDmslbrzSGGxf+yVrGRxygWx9olg7Ok5e9CmAAtrATQAS4IjaianIcI4"""
    """XoFpQH1nq9B8jZqWX1RJCEV8p27gIA280Ya9B/PuJlvyrlEUxskJWOVNdg9r64sfsyjZd0hdMkSm7maX"""
    """KlvkS/XzLpkaKtJnwXrQLBZtSHA2X34C+Jn40qntoFXUElpWMXGPmooyKwvz4DLnMD3xJ+5V/bJoVCBg"""
    """p9rs1GXmmmJ2OKcelWanG6rNDscE3uDcWQGka8LebOSl5ikwNFtNl3YN7NYla7FrpjeK1DuABmCk6x72"""
    """Zl4FkgF0dDTzJzoAqR0oDPThIKRoXDjTRTzoKfoKyKWEdO0UIIUXmdK4947t0hYhWMSDvoIGoT8gk3Fm"""
    """Gf9wJ+cq6jI7A3O6WAx6iqfaEXzvRfEfwhBdQVY92v/aRq3JtoNTNraLrq6ULEf+N+bIezAirLNXAAAA"""
    """AElFTkSuQmCC"""
)

right_icon = (
    """iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAAAXNSR0IArs4c6QAAAaVJREFUSEuVVW1W"""
    """hDAMTN6K54CbyEl8HkNcqbCKx9ijrEfhHrivvrYp9CPBwi94lMxMMhMQggsBQNvn7S5/8q8REDSdpyLx"""
    """Z2ulECM+yXzAHGYo+VOGxEqRqgksnLLj+A6g+JLRJbERgDQDjz9+Tjdzf1/wZRheZ35+NEFbzLeoUAEB"""
    """PAHAfF+wtSBroWTgQTdlCyTASk31qQKjojYgvwu2Y6ik3EXyMJX6rk+P+gbagTgl3ZxZNlcQ0yXlNhgp"""
    """MdUTCCkxIEZJ5LQIgOljpjZBNCAPlb5phBp0MJMopggoMc0clcUEoVeTBQHAGkHbmXDuWntgagzOisYp"""
    """hy7Cn/tz14TqsyQPl68PRFRZ9b2h+Ckh/PRvXRvujiBozPL6R4N1VKWvpHju37vGhctngttFhY2hTGzF"""
    """z13DfeqGzPdDhHI2hSuAtqk2fZcOC7tIDnfWFiouET0EwBZPLZMEcx8gWZfjZbXwblvSLZvvguSn6UmS"""
    """hZ9dz/f+ANtUyxWk5pYWT/LrC4LGLLtw1uGqkMgTiXCzFigorswyzBXssmMkZazjjfAHBmvcH57jlvcA"""
    """AAAASUVORK5CYII="""
)

top_left_icon = (
    "R0lGODlhZABkAHAAACH5BAEAAPwALAAAAABkAGQAhwAAAAAAMwAAZgAAmQAAzAAA/wArA"
    "AArMwArZgArmQArzAAr/wBVAABVMwBVZgBVmQBVzABV/wCAAACAMwCAZgCAmQCAzACA/w"
    "CqAACqMwCqZgCqmQCqzACq/wDVAADVMwDVZgDVmQDVzADV/wD/AAD/MwD/ZgD/mQD/zAD"
    "//zMAADMAMzMAZjMAmTMAzDMA/zMrADMrMzMrZjMrmTMrzDMr/zNVADNVMzNVZjNVmTNV"
    "zDNV/zOAADOAMzOAZjOAmTOAzDOA/zOqADOqMzOqZjOqmTOqzDOq/zPVADPVMzPVZjPVm"
    "TPVzDPV/zP/ADP/MzP/ZjP/mTP/zDP//2YAAGYAM2YAZmYAmWYAzGYA/2YrAGYrM2YrZm"
    "YrmWYrzGYr/2ZVAGZVM2ZVZmZVmWZVzGZV/2aAAGaAM2aAZmaAmWaAzGaA/2aqAGaqM2a"
    "qZmaqmWaqzGaq/2bVAGbVM2bVZmbVmWbVzGbV/2b/AGb/M2b/Zmb/mWb/zGb//5kAAJkA"
    "M5kAZpkAmZkAzJkA/5krAJkrM5krZpkrmZkrzJkr/5lVAJlVM5lVZplVmZlVzJlV/5mAA"
    "JmAM5mAZpmAmZmAzJmA/5mqAJmqM5mqZpmqmZmqzJmq/5nVAJnVM5nVZpnVmZnVzJnV/5"
    "n/AJn/M5n/Zpn/mZn/zJn//8wAAMwAM8wAZswAmcwAzMwA/8wrAMwrM8wrZswrmcwrzMw"
    "r/8xVAMxVM8xVZsxVmcxVzMxV/8yAAMyAM8yAZsyAmcyAzMyA/8yqAMyqM8yqZsyqmcyq"
    "zMyq/8zVAMzVM8zVZszVmczVzMzV/8z/AMz/M8z/Zsz/mcz/zMz///8AAP8AM/8AZv8Am"
    "f8AzP8A//8rAP8rM/8rZv8rmf8rzP8r//9VAP9VM/9VZv9Vmf9VzP9V//+AAP+AM/+AZv"
    "+Amf+AzP+A//+qAP+qM/+qZv+qmf+qzP+q///VAP/VM//VZv/Vmf/VzP/V////AP//M//"
    "/Zv//mf//zP///wAAAAAAAAAAAAAAAAj/APcJHEiwoMGDCBMOVHbjRjKFECNKnEixosBo"
    "mGzYuMHjzA2NFkOKHBkxWaYb+2yEOcPj4w0wDR2SnEkzYqZkygZ+DJPMzUYwPYDa6HjGB"
    "sGHNZPOzLjxDJqPG9Gg6WGjKNWiLsNk2jfMAJp90JSKnYgmDJqNNqjyCNPSjEq0LlvKta"
    "HsKhoYY/Mm9Jg2LUwzH632xYq2KOGibiQx0AeWYM6cepMqC3O17w3AgTe2NFy1cEuhVZV"
    "ptJFJ2TAGGsNEnqnMzQ2PgDfGdqlyrVyOH6nmDgyVdkuVPRqecWMA8mqJymDYgDn0ckO0"
    "uuHC/OscbecbuoWGae62o0ZMxyn2/2DLcbvu33Blc+xxO/3zn1SvclTJG+Vx4wlFWxXaG"
    "y5g70G9VJ11tGH3FlXA8SDUGWfsowwa+M10RiaZhPHaVgrpU55LMHEYV4GaObfdcy7FVh"
    "lfBr7GV1VJubGddSAVxJhAJ6VlW38bBQjYbOkRqNFLAf4G4opYlUYTDG3gtpZuAz11gwH"
    "gDeMTf5alpxuIuaUHA21YCfljYSFqRCGGIRWlHg9v2YCGmVVdiOKON7wYQ2ppuYVlenNq"
    "Fl2BlwXoV5CjlekRGpgIaRVWnKU1YWfnbXTAjxu69yVUIQooqaRmBuiRRSe9pOZQ5KX5G"
    "m6e/hbUqb0ppyRtksIQQwPbEf+206SSGqjWRxYxKBdltYGIoFW4QaeelcxthGUDo/VH6Z"
    "3uwWQmRWeBuWeYhs3HFnlC/VedgKzasOV1xvKG5keP2jDnh+kV1RJSEzF0qmbk9WabbGl"
    "5uNGI/FFp7r5xXYtmbC2NeMO3UAUIlJmsWoTJWZyBaOFn2V1KW3ydSVrZhps5d6NG5wJV"
    "oG5ukSkRw5lIN3Ff8bbka4G2mdfQjQHzBipUHpm6EWrB/Yimn69VhMkZL1JcIFD5DqssX"
    "NMCey+axmLWHJiEnQvXtRqtOWEmYUGkjBg1Oycr0hr1gGrYPZYHLqt3bmvlcgHeSZUbDl"
    "onskGiaVYxiPRSBgaBaKP/Feq0adf6I3nolV1xiDBlqFmHxn459bheHt2v5D7ibflaaEm"
    "NZXmbfZXfDUkyCheJqMW5ZOUWN26dYDiy/DCvUH1LcNkwMQ0RTqtS+a3Uvzbeul+SA0ll"
    "ziVy1CVbcz5arq+Ej3oDeAdNxiC2F/+oamxff5TnwL7LFh9z4zke+Eaqeospo39p5FT0T"
    "7dlmcowGODSu9mr7B+6LodIGIJcdu+Sr2EiFWagN5CSvcxSpwKDYMgnLLS9xDZdMhOvWE"
    "Ib6tXqM777jayIhqXPwI0gJWuD6FBUwYEhqzZUy0x8LASXzRgOXUOqyp5M1CZ90awz6yM"
    "INCRhOhX6iGIwIJjQ/7DCowoKrnhvsZSsggUiiMFFiBsxiEYgZsNaMSBLsQKVl/gnF8I0"
    "bk8nex+OCqMx2higaZ0hoECMdZvA+cpj6XIgYDrkJYmNrirT4dOk4JTBj2DiBvjJxM/Yl"
    "kQ7mo1X0BEK5t6DHcq96H9lawsYJ2W/ILHFWT0wyFnWA5qhVal5zCkYcO5mtJ/M5znMmV"
    "1m+uY/9VVqMGFQ4z58gC1IPumOGnsYn2pHKuvAbjtX/Jb/OmhLyRHukVRSzUCScZ7X1Es"
    "5yisfEtMWmzsSqHyaS5ei9MWWF96LfFcMUyYFogwGORNNYZCf/GDQgPiByXJk7J7snvQR"
    "dsKgdO6RmRsNx/+A7alqVzpRk508woAtqRMGBxiYklAkLOD1yAaP4l4/vaVQ5cgPdTsh5"
    "ispOtF9kUeNokEDehoy0XVetIF1JNGXJlpPZMUgnBdtp0IZmKfyjVQ6cLnonNgJLzUVJB"
    "Ow6UsM1GnC+M1Jb1ARmCtn1RCYnHGnA9tdP8MJzZvF4AZTVakpraOcqj6nS840jiR+xDM"
    "bHHRgBihXCj/CuHqhJzobmVMDyEU+hbJ0S1XdnuEgOlNKUUVvuiGgGxjWQBNC1FwGQGiY"
    "bkZR5Rw1fPPTiEW1hxpVqUqdMj3jAQjmWNSc657ncuaLGoqfaBUqjxtRZz8TKz+XTDRPy"
    "zvjj/jCU7n/OtZc7bRBbnNrz3Ze8YzrNBdqTipKejLgkh+J0GvOwp4NGVW4Mw0SXQdWUL"
    "oqJ3MMPGxBhSu7dsrUnp/laUW5K7W6upUuBckIQz/C2jmtE1lbWuQTtectl070AIkNprF"
    "m99LP+jZPDVCBUb1KvnLN6aWpUdBH5rYPN8AGrQAYqj2flLwgfstPVNGpZKOqAgP0F68N"
    "AIANSkq+xD6xujDoMF9FvNm0ctesMalXyAzyYB4MFaLtZS0MAGBhC2MniA0BgGwbsFMLJ"
    "zbF8QviZpN8Yw/XV1UHEHD8bqBiDycWqvt1pW7GSRBzqkfCLd4sfm8QgyCqYLM2OHOJh1"
    "rmASM0/8lKTmxDPBxEA6wgsSpgQIcPwGMf/698v+Wub8wjyzXyB00S1nFiW9JjH+9Ox1L"
    "eQp3LrIJEo2XJcDZqfoMYAxVjZap4RcvsDjUMhDSlTVVJK0KftCXtGUALFhawSwxgZSNb"
    "uM3xo/NzeNzmMuNX020OroBanVpim01FCskEhPQTJ1rrkczmgoGkzwjZK8PgztKO9Ztd/"
    "KMjG5nNiYV1klUN1/+1JTdOiVFEWtOcF7GJo9yDgRYiatY33zqI4k5yi1NE5zrnOttv3r"
    "f3RBerX0FDGRFSyKxkGFfh7mtLkg4iZVz16jq/OuJlVvUZqxNEjHv4BvlOMbK4BxcGfec"
    "T5P+8S6FLMhxBupa+PFVOxKFpbYp3VdKVTmyUW0u+HkeV0xG9rULR2YNyGoRdI9mkU+sL"
    "bxiEwcz1hnWHOZ3vWstPxC8BuJm3rtsfBbqHNkD6WJi4ORto4emu0vqjYiDujMdvs/XUA"
    "tvxLeUUu8RVZLyVSlajj9EgbDZ7a4jWOX1bV+X711GVe49xnU2/Ma0o2/lgZPSxjGRQ5q"
    "M2cEMmyidpxd862m2X85HFrQUqI3lOPslSD1DUg0is/gYzCo9A+NKggWjCAGJ4yRbMDnC"
    "sWjgMUn8q1U3P6cGUs5D6+ETgBSl7ilBIEmjXQpplW2bP41XeMQADklUAAK2UjExTQcv/"
    "QLLWfIVlfwU2kLQkMlFk4G995mY2ADEUcgZMkL/8I8kErM8OAzEoI7GSBgaV1mMzBwBig"
    "H/4FwaS9nQC9gnD0HG8J23Sd20phoD4NwwDA3xUFg3KIGBPx3hSBwAJZ4Gr4X5otwX7QA"
    "/yhn2lV3UkiH+SIGlboH1OlxxaB2udBgCS8IIJqH1zh2Sw9nTWpwIjyIN6wQgWpnhgwHY"
    "DI3fSJ2AGYIT4h28gh28dl4SdpgJSWH4PeIW7N3cDs3sCBgBbiH+dlm2kt4Kxdn9leBz6"
    "IGDmonhltgWmR4ZtWH6bpwURt38r2GHzd4eyRwxAhoNoZ3oqsHKAOBYPOHfi1nlB/6SDi"
    "RgeeZh9lAhrhshgkSgWgggDSyhtvKd4AOA5magXXTiD2NdoMDAJoxgZyfF0W9CCS4iDkL"
    "iKerGCrph2dIeJtEgTKgB8T+d+OBg/B7iLY/GB2od2VihgxKiJZ3d2M+iIbaaFy5gUySB"
    "v6ed0qLhj05gUFeJ0QaiHFEhporiNIwF8W3CCVthjdkiOIxGEYBCAX3iKKRYD7CgS0OeE"
    "cXiF9wYAuliPCQF8S/iOsLaHuCZgKOiPE6F/GSh36ectimd9IoiQErEI5yhv7veDISdgl"
    "SaRWgMG7keHKyh3GHeKPMaRCsEI/CeE2Chp0YZksTaMJlkQrYiDsVhm2td2hDj3iKUWkw"
    "QRhALGh1LXY1KmkQbAff1Yj5KACZOACesnCZOwfk+ZCU4pCVOJBpOABlSZBpIAk5EREAA"
    "7"
)

unchecked_icon = (
    "R0lGODlhHQATAP8AAAAAAAAAMwAAZgAAmQAAzAAA/wAzAAAzMwAzZgAzmQAzzAAz/wBmA"
    "ABmMwBmZgBmmQBmzABm/wCZAACZMwCZZgCZmQCZzACZ/wDMAADMMwDMZgDMmQDMzADM/w"
    "D/AAD/MwD/ZgD/mQD/zAD//zMAADMAMzMAZjMAmTMAzDMA/zMzADMzMzMzZjMzmTMzzDM"
    "z/zNmADNmMzNmZjNmmTNmzDNm/zOZADOZMzOZZjOZmTOZzDOZ/zPMADPMMzPMZjPMmTPM"
    "zDPM/zP/ADP/MzP/ZjP/mTP/zDP//2YAAGYAM2YAZmYAmWYAzGYA/2YzAGYzM2YzZmYzm"
    "WYzzGYz/2ZmAGZmM2ZmZmZmmWZmzGZm/2aZAGaZM2aZZmaZmWaZzGaZ/2bMAGbMM2bMZm"
    "bMmWbMzGbM/2b/AGb/M2b/Zmb/mWb/zGb//5kAAJkAM5kAZpkAmZkAzJkA/5kzAJkzM5k"
    "zZpkzmZkzzJkz/5lmAJlmM5lmZplmmZlmzJlm/5mZAJmZM5mZZpmZmZmZzJmZ/5nMAJnM"
    "M5nMZpnMmZnMzJnM/5n/AJn/M5n/Zpn/mZn/zJn//8wAAMwAM8wAZswAmcwAzMwA/8wzA"
    "MwzM8wzZswzmcwzzMwz/8xmAMxmM8xmZsxmmcxmzMxm/8yZAMyZM8yZZsyZmcyZzMyZ/8"
    "zMAMzMM8zMZszMmczMzMzM/8z/AMz/M8z/Zsz/mcz/zMz///8AAP8AM/8AZv8Amf8AzP8"
    "A//8zAP8zM/8zZv8zmf8zzP8z//9mAP9mM/9mZv9mmf9mzP9m//+ZAP+ZM/+ZZv+Zmf+Z"
    "zP+Z///MAP/MM//MZv/Mmf/MzP/M////AP//M///Zv//mf//zP///wAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAACwAAAAAHQATAIcAAAAAADMAAGYAAJkAAMwAAP8AMwAAMzMAM2YAM5kAM8wAM/8A"
    "ZgAAZjMAZmYAZpkAZswAZv8AmQAAmTMAmWYAmZkAmcwAmf8AzAAAzDMAzGYAzJkAzMwAz"
    "P8A/wAA/zMA/2YA/5kA/8wA//8zAAAzADMzAGYzAJkzAMwzAP8zMwAzMzMzM2YzM5kzM8"
    "wzM/8zZgAzZjMzZmYzZpkzZswzZv8zmQAzmTMzmWYzmZkzmcwzmf8zzAAzzDMzzGYzzJk"
    "zzMwzzP8z/wAz/zMz/2Yz/5kz/8wz//9mAABmADNmAGZmAJlmAMxmAP9mMwBmMzNmM2Zm"
    "M5lmM8xmM/9mZgBmZjNmZmZmZplmZsxmZv9mmQBmmTNmmWZmmZlmmcxmmf9mzABmzDNmz"
    "GZmzJlmzMxmzP9m/wBm/zNm/2Zm/5lm/8xm//+ZAACZADOZAGaZAJmZAMyZAP+ZMwCZMz"
    "OZM2aZM5mZM8yZM/+ZZgCZZjOZZmaZZpmZZsyZZv+ZmQCZmTOZmWaZmZmZmcyZmf+ZzAC"
    "ZzDOZzGaZzJmZzMyZzP+Z/wCZ/zOZ/2aZ/5mZ/8yZ///MAADMADPMAGbMAJnMAMzMAP/M"
    "MwDMMzPMM2bMM5nMM8zMM//MZgDMZjPMZmbMZpnMZszMZv/MmQDMmTPMmWbMmZnMmczMm"
    "f/MzADMzDPMzGbMzJnMzMzMzP/M/wDM/zPM/2bM/5nM/8zM////AAD/ADP/AGb/AJn/AM"
    "z/AP//MwD/MzP/M2b/M5n/M8z/M///ZgD/ZjP/Zmb/Zpn/Zsz/Zv//mQD/mTP/mWb/mZn"
    "/mcz/mf//zAD/zDP/zGb/zJn/zMz/zP///wD//zP//2b//5n//8z///8AAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAIRAADCRxIsKDBgwgTKlzIsOGKhxBXNEQYseJEgxUjXiyYEeJGgh0ffhwYUuL"
    "IQCVPogypMuVJlyNhfpQ5M6PKmzhzngwIADs="
)

checked_icon = (
    "R0lGODlhHQATAP8AAAAAAAAAMwAAZgAAmQAAzAAA/wAzAAAzMwAzZgAzmQAzzAAz/wBmA"
    "ABmMwBmZgBmmQBmzABm/wCZAACZMwCZZgCZmQCZzACZ/wDMAADMMwDMZgDMmQDMzADM/w"
    "D/AAD/MwD/ZgD/mQD/zAD//zMAADMAMzMAZjMAmTMAzDMA/zMzADMzMzMzZjMzmTMzzDM"
    "z/zNmADNmMzNmZjNmmTNmzDNm/zOZADOZMzOZZjOZmTOZzDOZ/zPMADPMMzPMZjPMmTPM"
    "zDPM/zP/ADP/MzP/ZjP/mTP/zDP//2YAAGYAM2YAZmYAmWYAzGYA/2YzAGYzM2YzZmYzm"
    "WYzzGYz/2ZmAGZmM2ZmZmZmmWZmzGZm/2aZAGaZM2aZZmaZmWaZzGaZ/2bMAGbMM2bMZm"
    "bMmWbMzGbM/2b/AGb/M2b/Zmb/mWb/zGb//5kAAJkAM5kAZpkAmZkAzJkA/5kzAJkzM5k"
    "zZpkzmZkzzJkz/5lmAJlmM5lmZplmmZlmzJlm/5mZAJmZM5mZZpmZmZmZzJmZ/5nMAJnM"
    "M5nMZpnMmZnMzJnM/5n/AJn/M5n/Zpn/mZn/zJn//8wAAMwAM8wAZswAmcwAzMwA/8wzA"
    "MwzM8wzZswzmcwzzMwz/8xmAMxmM8xmZsxmmcxmzMxm/8yZAMyZM8yZZsyZmcyZzMyZ/8"
    "zMAMzMM8zMZszMmczMzMzM/8z/AMz/M8z/Zsz/mcz/zMz///8AAP8AM/8AZv8Amf8AzP8"
    "A//8zAP8zM/8zZv8zmf8zzP8z//9mAP9mM/9mZv9mmf9mzP9m//+ZAP+ZM/+ZZv+Zmf+Z"
    "zP+Z///MAP/MM//MZv/Mmf/MzP/M////AP//M///Zv//mf//zP///wAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAACwAAAAAHQATAIcAAAAAADMAAGYAAJkAAMwAAP8AMwAAMzMAM2YAM5kAM8wAM/8A"
    "ZgAAZjMAZmYAZpkAZswAZv8AmQAAmTMAmWYAmZkAmcwAmf8AzAAAzDMAzGYAzJkAzMwAz"
    "P8A/wAA/zMA/2YA/5kA/8wA//8zAAAzADMzAGYzAJkzAMwzAP8zMwAzMzMzM2YzM5kzM8"
    "wzM/8zZgAzZjMzZmYzZpkzZswzZv8zmQAzmTMzmWYzmZkzmcwzmf8zzAAzzDMzzGYzzJk"
    "zzMwzzP8z/wAz/zMz/2Yz/5kz/8wz//9mAABmADNmAGZmAJlmAMxmAP9mMwBmMzNmM2Zm"
    "M5lmM8xmM/9mZgBmZjNmZmZmZplmZsxmZv9mmQBmmTNmmWZmmZlmmcxmmf9mzABmzDNmz"
    "GZmzJlmzMxmzP9m/wBm/zNm/2Zm/5lm/8xm//+ZAACZADOZAGaZAJmZAMyZAP+ZMwCZMz"
    "OZM2aZM5mZM8yZM/+ZZgCZZjOZZmaZZpmZZsyZZv+ZmQCZmTOZmWaZmZmZmcyZmf+ZzAC"
    "ZzDOZzGaZzJmZzMyZzP+Z/wCZ/zOZ/2aZ/5mZ/8yZ///MAADMADPMAGbMAJnMAMzMAP/M"
    "MwDMMzPMM2bMM5nMM8zMM//MZgDMZjPMZmbMZpnMZszMZv/MmQDMmTPMmWbMmZnMmczMm"
    "f/MzADMzDPMzGbMzJnMzMzMzP/M/wDM/zPM/2bM/5nM/8zM////AAD/ADP/AGb/AJn/AM"
    "z/AP//MwD/MzP/M2b/M5n/M8z/M///ZgD/ZjP/Zmb/Zpn/Zsz/Zv//mQD/mTP/mWb/mZn"
    "/mcz/mf//zAD/zDP/zGb/zJn/zMz/zP///wD//zP//2b//5n//8z///8AAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAIRAANCRxIsKDBgwgTKlzIsKHDFRAjrnB4UKJFigUtSsRIUGNEjgM9QgQpUOR"
    "EkiZJGkqJUqRKliBhcpQ5U6PKmzhzFgwIADs="
)
