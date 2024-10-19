# SPDX-License-Identifier: GPL-3.0-only
# Copyright 2019 R. A. Gardner

import os
import re
import tkinter as tk
from collections import defaultdict
from itertools import repeat
from tkinter import filedialog, ttk

from openpyxl import load_workbook
from tksheet import Sheet

from .classes import (
    TreeBuilder,
)
from .constants import (
    EF,
    sheet_header_font,
)
from .functions import (
    b32_x_dict,
    bytes_io_wb,
    csv_str_x_data,
    get_json_format,
    get_json_from_file,
    json_to_sheet,
    ws_x_data,
    ws_x_program_data_str,
)
from .toplevels import (
    Ask_Confirm,
    Compare_Report_Popup,
    Error,
)
from .widgets import (
    Button,
    Frame,
    Id_Parent_Column_Selector,
    Readonly_Entry,
)


class Tree_Compare(tk.Frame):
    def __init__(self, parent, C):
        tk.Frame.__init__(self, parent)
        self.C = C
        self.heads1 = []
        self.heads2 = []
        self.data1 = []
        self.data2 = []
        self.sheet1 = []
        self.sheet2 = []
        self.nodes1 = {}
        self.nodes2 = {}
        self.rns1 = {}
        self.rns2 = {}
        self.report = []
        self.ic1 = 0
        self.ic2 = 0
        self.parent_cols1 = []
        self.parent_cols2 = []
        self.row_len1 = 0
        self.row_len2 = 0
        self.shkeys = []

        self.l_frame = Frame(self)
        self.l_frame.config(highlightthickness=0, highlightbackground="black")
        self.l_frame_btns = Frame(self.l_frame)

        self.r_frame = Frame(self)
        self.r_frame.config(highlightthickness=0, highlightbackground="black")
        self.r_frame_btns = Frame(self.r_frame)

        self.l_frame.grid(row=0, column=0, sticky="nswe")
        self.r_frame.grid(row=0, column=1, sticky="nswe")
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1, uniform="x")
        self.grid_columnconfigure(1, weight=1, uniform="x")

        self.l_frame_btns.pack(side="top", fill="both")
        self.r_frame_btns.pack(side="top", fill="both")

        self.sheet_filename1 = Readonly_Entry(self.l_frame_btns, font=EF, theme=self.C.theme)
        self.sheet_filename1.grid(row=1, column=0, padx=2, pady=5, sticky="nswe")

        self.open_sheet1_button = Button(
            self.l_frame_btns,
            text="⯇ Open file",
            style="EFW.Std.TButton",
            command=self.open_sheet1,
        )
        self.open_sheet1_button.grid(row=1, column=1, padx=10, pady=5, sticky="nsw")
        self.open_sheet1_button.config(width=15)

        self.sheet_dropdown_displayed1 = tk.StringVar(self.l_frame_btns)
        self.sheet_dropdown1 = ttk.Combobox(
            self.l_frame_btns,
            textvariable=self.sheet_dropdown_displayed1,
            font=EF,
            state="readonly",
            background="white",
        )
        self.sheet_dropdown1.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
        self.sheet_dropdown1.grid(row=2, padx=2, pady=5, column=0, sticky="nswe")

        self.load_sheet1 = Button(
            self.l_frame_btns,
            text="⯇ Load sheet",
            style="EFW.Std.TButton",
            command=self.load_sheet1,
        )
        self.load_sheet1.config(state="disabled")
        self.load_sheet1.config(width=15)
        self.load_sheet1_STATE = "disabled"
        self.load_sheet1.grid(row=2, column=1, padx=10, pady=5, sticky="nsw")

        self.selector_1 = Id_Parent_Column_Selector(self.l_frame_btns, expand=True)
        self.selector_1.config(width=400, height=330)
        self.selector_1.grid(row=3, column=0, sticky="nswe")

        self.run_compare_button = Button(
            self.l_frame_btns,
            text="Create Report",
            style="TF.Std.TButton",
            command=self.run_comparison,
        )

        self.run_compare_button.grid(row=3, column=1, padx=10, pady=20, sticky="ews")

        self.sheetdisplay1 = Sheet(
            self.l_frame,
            theme=self.C.theme,
            header_font=sheet_header_font,
            outline_thickness=1,
            outline_color="black",
        )
        self.selector_1.link_sheet(self.sheetdisplay1)
        self.sheetdisplay1.enable_bindings(
            (
                "single",
                "drag_select",
                "select_all",
                "copy",
                "rc_popup_menu",
                "column_width_resize",
                "double_click_column_resize",
                "row_height_resize",
                "double_click_row_resize",
                "column_select",
                "row_select",
                "arrowkeys",
                "ctrl_select",
            )
        )
        self.sheetdisplay1.pack(side="top", fill="both", expand=True)

        # __________________ R FRAME ______________________________________________________________

        self.sheet_filename2 = Readonly_Entry(self.r_frame_btns, font=EF, theme=self.C.theme)
        self.sheet_filename2.grid(row=1, column=0, padx=2, pady=5, sticky="nswe")

        self.open_sheet2_button = Button(
            self.r_frame_btns,
            text="⯇ Open file",
            style="EFW.Std.TButton",
            command=self.open_sheet2,
        )
        self.open_sheet2_button.config(width=15)
        self.open_sheet2_button.grid(row=1, column=1, padx=10, pady=5, sticky="nsw")

        self.sheet_dropdown_displayed2 = tk.StringVar(self.r_frame_btns)
        self.sheet_dropdown2 = ttk.Combobox(
            self.r_frame_btns,
            textvariable=self.sheet_dropdown_displayed2,
            font=EF,
            state="readonly",
            background="white",
        )
        self.sheet_dropdown2.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
        self.sheet_dropdown2.grid(row=2, column=0, padx=2, pady=5, sticky="nswe")

        self.load_sheet2 = Button(
            self.r_frame_btns,
            text="⯇ Load sheet",
            style="EFW.Std.TButton",
            command=self.load_sheet2,
        )
        self.load_sheet2.config(state="disabled")
        self.load_sheet2.config(width=15)
        self.load_sheet2_STATE = "disabled"
        self.load_sheet2.grid(row=2, column=1, padx=10, pady=5, sticky="nsw")

        self.selector_2 = Id_Parent_Column_Selector(self.r_frame_btns, expand=True)
        self.selector_2.config(width=400, height=330)
        self.selector_2.grid(row=3, column=0, sticky="nswe")

        self.sheetdisplay2 = Sheet(
            self.r_frame,
            theme=self.C.theme,
            header_font=sheet_header_font,
            outline_thickness=1,
            outline_color="black",
        )
        self.selector_2.link_sheet(self.sheetdisplay2)
        self.sheetdisplay2.enable_bindings(
            (
                "single",
                "drag_select",
                "select_all",
                "copy",
                "rc_popup_menu",
                "column_width_resize",
                "double_click_column_resize",
                "row_height_resize",
                "double_click_row_resize",
                "column_select",
                "row_select",
                "arrowkeys",
                "ctrl_select",
            )
        )
        self.sheetdisplay2.pack(side="top", fill="both", expand=True)

    def enable_widgets(self):
        self.C.menubar_state("normal", start=True)
        self.sheet_filename1.config(state="readonly")
        self.open_sheet1_button.config(state="normal")
        self.sheet_dropdown1.config(state="readonly")
        self.sheet_dropdown1.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
        self.load_sheet1.config(state=self.load_sheet1_STATE)
        self.selector_1.enable_me()
        self.sheetdisplay1.enable_bindings()
        self.sheetdisplay1.basic_bindings(True)

        self.sheet_filename2.config(state="readonly")
        self.open_sheet2_button.config(state="normal")
        self.sheet_dropdown2.config(state="readonly")
        self.sheet_dropdown2.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
        self.load_sheet2.config(state=self.load_sheet2_STATE)
        self.selector_2.enable_me()
        self.sheetdisplay2.enable_bindings()
        self.run_compare_button.config(state="normal")
        self.sheetdisplay2.basic_bindings(True)

    def disable_widgets(self):
        self.C.menubar_state("disabled")
        self.sheet_filename1.config(state="disabled")
        self.open_sheet1_button.config(state="disabled")
        self.sheet_dropdown1.config(state="disabled")
        self.sheet_dropdown1.unbind("<<ComboboxSelected>>")
        self.load_sheet1_STATE = str(self.load_sheet1["state"])
        self.load_sheet1.config(state="disabled")
        self.selector_1.disable_me()
        self.sheetdisplay1.disable_bindings()
        self.sheetdisplay1.basic_bindings(False)

        self.sheet_filename2.config(state="disabled")
        self.open_sheet2_button.config(state="disabled")
        self.sheet_dropdown2.config(state="disabled")
        self.sheet_dropdown2.unbind("<<ComboboxSelected>>")
        self.load_sheet2_STATE = str(self.load_sheet2["state"])
        self.load_sheet2.config(state="disabled")
        self.selector_2.disable_me()
        self.sheetdisplay2.disable_bindings()
        self.run_compare_button.config(state="disabled")
        self.sheetdisplay2.basic_bindings(False)

    def start_work(self, msg=""):
        self.C.status_bar.change_text(msg)
        self.disable_widgets()

    def stop_work(self, msg=""):
        self.C.status_bar.change_text(msg)
        self.enable_widgets()

    def populate(self):
        self.C.change_app_title(title="Comparing sheets")
        self.sheetdisplay1.change_theme(self.C.theme)
        self.sheetdisplay2.change_theme(self.C.theme)
        self.C.show_frame("tree_compare")

    def reset(self):
        try:
            self.C.wb.close()
        except Exception:
            pass
        self.reset_1(staying_on_compare=False)
        self.reset_2(staying_on_compare=False)

    def reset_1(self, staying_on_compare=True):
        try:
            self.C.wb.close()
        except Exception:
            pass
        self.C.wb = None
        self.heads1 = []
        self.data1 = []
        self.sheet1 = []
        self.sheet2 = []
        self.nodes1 = {}
        self.nodes2 = {}
        self.rns1 = {}
        self.rns2 = {}
        self.report = []
        self.ic1 = 0
        self.parent_cols1 = []
        self.row_len1 = 0
        self.sheet_filename1.set_my_value("")
        self.sheet_dropdown1["values"] = []
        self.sheet_dropdown_displayed1.set("")
        self.load_sheet1.config(state="disabled")
        self.selector_1.clear_displays()
        self.sheetdisplay1.dehighlight_cells(all_=True, redraw=False)
        self.sheetdisplay1.dehighlight_cells(canvas="row_index", all_=True, redraw=False)
        self.sheetdisplay1.dehighlight_cells(canvas="header", all_=True, redraw=False)
        self.sheetdisplay2.dehighlight_cells(all_=True, redraw=False)
        self.sheetdisplay2.dehighlight_cells(canvas="row_index", all_=True, redraw=False)
        self.sheetdisplay2.dehighlight_cells(canvas="header", all_=True, redraw=False)
        self.sheetdisplay1.data_reference(newdataref=[], redraw=True)

    def reset_2(self, staying_on_compare=True):
        try:
            self.C.wb.close()
        except Exception:
            pass
        self.C.wb = None
        self.heads2 = []
        self.data2 = []
        self.sheet1 = []
        self.sheet2 = []
        self.nodes1 = {}
        self.nodes2 = {}
        self.rns1 = {}
        self.rns2 = {}
        self.report = []
        self.ic2 = 0
        self.parent_cols2 = []
        self.row_len2 = 0
        self.sheet_filename2.set_my_value("")
        self.sheet_dropdown2["values"] = []
        self.sheet_dropdown_displayed2.set("")
        self.load_sheet2.config(state="disabled")
        self.selector_2.clear_displays()
        self.sheetdisplay1.dehighlight_cells(all_=True, redraw=False)
        self.sheetdisplay1.dehighlight_cells(canvas="row_index", all_=True, redraw=False)
        self.sheetdisplay1.dehighlight_cells(canvas="header", all_=True, redraw=False)
        self.sheetdisplay2.dehighlight_cells(all_=True, redraw=False)
        self.sheetdisplay2.dehighlight_cells(canvas="row_index", all_=True, redraw=False)
        self.sheetdisplay2.dehighlight_cells(canvas="header", all_=True, redraw=False)
        self.sheetdisplay2.data_reference(newdataref=[], redraw=True)

    def open_sheet1(self):
        if self.data1:
            confirm = Ask_Confirm(self, "Note: Opening resets", theme=self.C.theme)
            if not confirm.boolean:
                return
        self.start_work("Loading...   ")
        self.reset_1(False)
        filepath = filedialog.askopenfilename(parent=self.C, title="Select file")
        if not filepath:
            self.stop_work("Program ready")
            return
        try:
            filepath = os.path.normpath(filepath)
        except Exception:
            Error(self, "Filepath invalid   ", theme=self.C.theme)
            self.stop_work("Program ready")
            return
        if not filepath.lower().endswith((".json", ".xlsx", ".xls", ".xlsm", ".csv", ".tsv")):
            Error(self, "Please select excel/csv/json   ", theme=self.C.theme)
            self.stop_work("Program ready")
            return
        check = os.path.isfile(filepath)
        if not check:
            Error(self, "Filepath invalid   ", theme=self.C.theme)
            self.stop_work("Program ready")
            return
        try:
            if filepath.lower().endswith((".csv", ".tsv")):
                with open(filepath, "r") as fh:
                    temp_data = fh.read()
                self.data1 = csv_str_x_data(temp_data)
                if not self.data1:
                    Error(self, "File contains no data   ", theme=self.C.theme)
                    self.stop_work("Program ready")
                    return
                self.load_display1()
                self.stop_work("Program ready")
            elif filepath.lower().endswith(".json"):
                j = get_json_from_file(filepath)
                json_format = get_json_format(j)
                if not json_format:
                    Error(self, "File contains no data   ", theme=self.C.theme)
                    self.stop_work("Program ready")
                    return
                self.data1 = json_to_sheet(
                    j,
                    format_=json_format[0],
                    key=json_format[1],
                    get_format=False,
                )
                if not self.data1:
                    Error(self, "File contains no data   ", theme=self.C.theme)
                    self.stop_work("Program ready")
                    return
                self.load_display1()
                self.stop_work("Program ready")
            else:
                in_mem = bytes_io_wb(filepath)
                self.C.wb = load_workbook(in_mem, read_only=True, data_only=True)
                if len(self.C.wb.sheetnames) < 1:
                    Error(self, "File contains no data   ", theme=self.C.theme)
                    self.stop_work("Program ready")
                    return
                sheetnames = set(self.C.wb.sheetnames)
                if "program_data" in sheetnames:
                    ws = self.C.wb["program_data"]
                    ws.reset_dimensions()
                    try:
                        d = b32_x_dict(ws_x_program_data_str(ws))
                        self.data1 = [[h["name"] for h in d["headers"]]] + d["records"]
                        self.C.wb.close()
                        self.load_display1(
                            idcol=next(c for c, h in enumerate(d["headers"]) if h["type"] == "ID")
                            if d["headers"]
                            else None,
                            parcols=[c for c, h in enumerate(d["headers"]) if h["type"] == "Parent"]
                            if d["headers"]
                            else None,
                        )
                        self.stop_work("Program ready")
                    except Exception:
                        self.data1 = []
                        self.C.wb.close()
                        self.C.wb = load_workbook(in_mem, read_only=True, data_only=True)
                        Error(self, "Error opening program data, select a sheet   ", theme=self.C.theme)
                        self.sheet_dropdown1["values"] = self.C.wb.sheetnames
                        self.sheet_dropdown_displayed1.set(self.C.wb.sheetnames[0])
                        self.stop_work("Program ready")
                        self.open_sheet2_button.config(state="disabled")
                        self.load_sheet1.config(state="normal")
                else:
                    self.sheet_dropdown1["values"] = self.C.wb.sheetnames
                    self.sheet_dropdown_displayed1.set(self.C.wb.sheetnames[0])
                    self.stop_work("Program ready")
                    self.open_sheet2_button.config(state="disabled")
                    self.load_sheet1.config(state="normal")
            self.sheet_filename1.set_my_value(filepath)
        except Exception as error_msg:
            Error(self, f"Error: {error_msg}", theme=self.C.theme)
            self.stop_work("Program ready")

    def load_sheet1(self):
        self.start_work("Loading...   ")
        ws = self.C.wb[self.sheet_dropdown_displayed1.get()]
        ws.reset_dimensions()
        self.data1 = ws_x_data(ws)
        self.C.wb.close()
        self.stop_work("Program ready")
        if not self.data1:
            Error(self, "Sheet contains no data   ", theme=self.C.theme)
            self.load_sheet1.config(state="disabled")
            return
        self.load_sheet1.config(state="disabled")
        self.load_display1()

    def load_display1(self, idcol=None, parcols=None):
        self.row_len1 = max(map(len, self.data1), default=0)
        self.sheetdisplay1.data_reference(newdataref=self.data1, redraw=True)
        self.selector_1.set_columns([h for h in self.data1[0]])
        if idcol is None and parcols is None:
            self.selector_1.detect_id_col()
            self.selector_1.detect_par_cols()
        else:
            self.selector_1.set_id_col(idcol)
            self.selector_1.set_par_cols(parcols)

    def open_sheet2(self):
        if self.data2:
            confirm = Ask_Confirm(self, "Note: Opening resets", theme=self.C.theme)
            if not confirm.boolean:
                return
        self.start_work("Loading...   ")
        self.reset_2(False)
        filepath = filedialog.askopenfilename(parent=self.C, title="Select file")
        if not filepath:
            self.stop_work("Program ready")
            return
        try:
            filepath = os.path.normpath(filepath)
        except Exception:
            Error(self, "Filepath invalid   ", theme=self.C.theme)
            self.stop_work("Program ready")
            return
        if not filepath.lower().endswith((".json", ".xlsx", ".xls", ".xlsm", ".csv", ".tsv")):
            Error(self, "Please select excel/csv/json   ", theme=self.C.theme)
            self.stop_work("Program ready")
            return
        check = os.path.isfile(filepath)
        if not check:
            Error(self, "Filepath invalid   ", theme=self.C.theme)
            self.stop_work("Program ready")
            return
        try:
            if filepath.lower().endswith((".csv", ".tsv")):
                with open(filepath, "r") as fh:
                    temp_data = fh.read()
                self.data2 = csv_str_x_data(temp_data)
                if not self.data2:
                    Error(self, "File contains no data   ", theme=self.C.theme)
                    self.stop_work("Program ready")
                    return
                self.load_display2()
                self.stop_work("Program ready")
            elif filepath.lower().endswith(".json"):
                j = get_json_from_file(filepath)
                json_format = get_json_format(j)
                if not json_format:
                    Error(self, "File contains no data   ", theme=self.C.theme)
                    self.stop_work("Program ready")
                    return
                self.data2 = json_to_sheet(j, format_=json_format[0], key=json_format[1], get_format=False)
                if not self.data2:
                    Error(self, "File contains no data   ", theme=self.C.theme)
                    self.stop_work("Program ready")
                    return
                self.load_display2()
                self.stop_work("Program ready")
            else:
                in_mem = bytes_io_wb(filepath)
                try:
                    self.C.wb = load_workbook(in_mem, read_only=True, data_only=True)
                except Exception:
                    Error(self, "Error opening file   ", theme=self.C.theme)
                    self.stop_work("Program ready")
                    return
                if len(self.C.wb.sheetnames) < 1:
                    Error(self, "File contains no data   ", theme=self.C.theme)
                    self.stop_work("Program ready")
                    return
                sheetnames = set(self.C.wb.sheetnames)
                if "program_data" in sheetnames:
                    ws = self.C.wb["program_data"]
                    ws.reset_dimensions()
                    try:
                        d = b32_x_dict(ws_x_program_data_str(ws))
                        self.data2 = [[h["name"] for h in d["headers"]]] + d["records"]
                        self.C.wb.close()
                        self.load_display2(
                            idcol=next(c for c, h in enumerate(d["headers"]) if h["type"] == "ID")
                            if d["headers"]
                            else None,
                            parcols=[c for c, h in enumerate(d["headers"]) if h["type"] == "Parent"]
                            if d["headers"]
                            else None,
                        )
                        self.stop_work("Program ready")
                    except Exception:
                        self.data2 = []
                        self.C.wb.close()
                        self.C.wb = load_workbook(in_mem, read_only=True, data_only=True)
                        Error(self, "Error opening program data, select a sheet   ", theme=self.C.theme)
                        self.sheet_dropdown2["values"] = self.C.wb.sheetnames
                        self.sheet_dropdown_displayed2.set(self.C.wb.sheetnames[0])
                        self.stop_work("Program ready")
                        self.open_sheet1_button.config(state="disabled")
                        self.load_sheet2.config(state="normal")
                else:
                    self.sheet_dropdown2["values"] = self.C.wb.sheetnames
                    self.sheet_dropdown_displayed2.set(self.C.wb.sheetnames[0])
                    self.stop_work("Program ready")
                    self.open_sheet1_button.config(state="disabled")
                    self.load_sheet2.config(state="normal")
            self.sheet_filename2.set_my_value(filepath)
        except Exception as error_msg:
            Error(self, f"Error: {error_msg}", theme=self.C.theme)
            self.stop_work("Program ready")

    def load_sheet2(self):
        self.start_work("Loading...   ")
        ws = self.C.wb[self.sheet_dropdown_displayed2.get()]
        ws.reset_dimensions()
        self.data2 = ws_x_data(ws)
        self.C.wb.close()
        self.stop_work("Program ready")
        if not self.data2:
            Error(self, "Sheet contains no data   ", theme=self.C.theme)
            self.load_sheet2.config(state="disabled")
            return
        self.load_sheet2.config(state="disabled")
        self.load_display2()

    def load_display2(self, idcol=None, parcols=None):
        self.row_len2 = max(map(len, self.data2), default=0)
        self.sheetdisplay2.data_reference(newdataref=self.data2, redraw=True)
        self.selector_2.set_columns([h for h in self.data2[0]])
        if idcol is None and parcols is None:
            self.selector_2.detect_id_col()
            self.selector_2.detect_par_cols()
        else:
            self.selector_2.set_id_col(idcol)
            self.selector_2.set_par_cols(parcols)

    def set_row_lens(self):
        self.row_len1 = max(map(len, self.sheetdisplay1.data), default=0)
        self.row_len2 = max(map(len, self.sheetdisplay2.data), default=0)

    def heads_comparison(self, heads, datavar, addition):
        if datavar == 1:
            row_len = self.row_len1
        elif datavar == 2:
            row_len = self.row_len2
        if len(heads) < row_len:
            heads += list(repeat("", row_len - len(heads)))
        tally_of_heads = defaultdict(lambda: -1)
        for coln in range(len(heads)):
            cell = heads[coln]
            if not cell:
                cell = f"_MISSING_{coln + 1}"
                addition.append([f" - Missing header in column #{coln + 1}"])
            hk = cell.lower()
            tally_of_heads[hk] += 1
            if tally_of_heads[hk] > 0:
                orig = cell
                x = 1
                while hk in tally_of_heads:
                    cell = f"{orig}_DUPLICATED_{x}"
                    hk = cell.lower()
                    x += 1
                tally_of_heads[hk] += 1
                addition.append([f" - Duplicate header in column #{coln + 1}"])
            heads[coln] = cell
        return heads, addition

    def run_comparison(self):
        self.ic1 = self.selector_1.get_id_col()
        self.parent_cols1 = list(self.selector_1.get_par_cols())
        self.ic2 = self.selector_2.get_id_col()
        self.parent_cols2 = list(self.selector_2.get_par_cols())
        if self.ic1 is None or self.ic2 is None:
            Error(self, "Select an ID column for both sheets", theme=self.C.theme)
            return
        if not self.parent_cols1 or not self.parent_cols2:
            Error(self, "Select parent columns for both sheets", theme=self.C.theme)
            return
        if self.ic1 in self.parent_cols1 or self.ic2 in self.parent_cols2:
            Error(self, "An ID column cannot be the same as a parent column", theme=self.C.theme)
            return
        self.start_work("Creating report...   ")
        sheetname_1 = os.path.basename(self.sheet_filename1.get())
        sheetname_2 = os.path.basename(self.sheet_filename2.get())
        if sheetname_1 == sheetname_2:
            sheetname_1 = "Sheet 1 - Left Panel"
            sheetname_2 = "Sheet 2 - Right Panel"
        self.sheet1 = []
        self.sheet2 = []
        self.nodes1 = {}
        self.nodes2 = {}
        self.rns1 = {}
        self.rns2 = {}
        self.set_row_lens()
        self.report = defaultdict(list)
        self.heads1, addition1 = self.heads_comparison(self.sheetdisplay1.data[0].copy(), 1, [])
        self.sheet1, self.nodes1, addition1, self.rns1 = TreeBuilder().build(
            input_sheet=self.sheetdisplay1.data,
            output_sheet=self.sheet1,
            row_len=self.row_len1,
            ic=self.ic1,
            hiers=self.parent_cols1,
            nodes=self.nodes1,
            warnings=addition1,
            rns=self.rns1,
            add_warnings=True,
            skip_1st=True,
            compare=True,
            fix_associate=True,
            strip=False,
        )
        self.heads2, addition2 = self.heads_comparison(self.sheetdisplay2.data[0].copy(), 2, [])
        self.sheet2, self.nodes2, addition2, self.rns2 = TreeBuilder().build(
            input_sheet=self.sheetdisplay2.data,
            output_sheet=self.sheet2,
            row_len=self.row_len2,
            ic=self.ic2,
            hiers=self.parent_cols2,
            nodes=self.nodes2,
            warnings=addition2,
            rns=self.rns2,
            add_warnings=True,
            skip_1st=True,
            compare=True,
            fix_associate=True,
            strip=False,
        )

        if addition1:
            self.report[f"WARNINGS - {sheetname_1} - "].extend([[warning] for warning in addition1])
        if addition2:
            self.report[f"WARNINGS - {sheetname_2} - "].extend([[warning] for warning in addition2])

        qhst1 = set(self.parent_cols1)
        qhst2 = set(self.parent_cols2)
        pcold = defaultdict(list)
        for i, h in enumerate(self.heads1):
            if i in qhst1:
                pcold[h].append(i)
        for i, h in enumerate(self.heads2):
            if i in qhst2:
                pcold[h].append(i)
        detcold = defaultdict(list)
        qhst1.add(self.ic1)
        qhst2.add(self.ic2)
        for i, h in enumerate(self.heads1):
            if i not in qhst1:
                detcold[h].append(i)
        for i, h in enumerate(self.heads2):
            if i not in qhst2:
                detcold[h].append(i)
        matching_hrs_names = sorted((k for k, v in pcold.items() if len(v) > 1), key=self.srtkey)
        matching_details_names = sorted((k for k, v in detcold.items() if len(v) > 1), key=self.srtkey)

        if self.ic1 == self.ic2 and self.heads1[self.ic1] == self.heads2[self.ic2]:
            self.report["COLUMNS"].append([" - Sheets have the same ID column names and indexes"])
        else:
            self.report["COLUMNS"].append([f" - {sheetname_1} has ID column: {self.ic1 + 1} - {self.heads1[self.ic1]}"])
            self.report["COLUMNS"].append([f" - {sheetname_2} has ID column: {self.ic2 + 1} - {self.heads2[self.ic2]}"])

        if self.parent_cols1 == self.parent_cols2 and [self.heads1[pcol_1] for pcol_1 in self.parent_cols1] == [
            self.heads2[pcol_2] for pcol_2 in self.parent_cols2
        ]:
            self.report["COLUMNS"].append([" - Sheets have the same Parent column names and indexes"])
        else:
            self.report["COLUMNS"].append([f" - {sheetname_1} has parent columns: "])
            for pcol in self.parent_cols1:
                self.report["COLUMNS"].append([f"       Column: {pcol+1} - {self.heads1[pcol]}"])
            self.report["COLUMNS"].append([f" - {sheetname_2} has parent columns: "])
            for pcol in self.parent_cols2:
                self.report["COLUMNS"].append([f"       Column: {pcol+1} - {self.heads2[pcol]}"])

        if len(matching_details_names) > 0:
            hdset1 = {h for i, h in enumerate(self.heads1) if i not in qhst1}
            hdset2 = {h for i, h in enumerate(self.heads2) if i not in qhst2}
            if all(detcold[n][0] == detcold[n][1] for n in matching_details_names):
                self.report["COLUMNS"].append([" - Sheets have the same detail column names and indexes"])
            else:
                self.report["COLUMNS"].append([" - Sheets have following matching detail column names:"])
                for n in matching_details_names:
                    self.report["COLUMNS"].append([f"   - {n}"])
                    self.report["COLUMNS"].append([f"       Column {sheetname_1}: {detcold[n][0]}"])
                    self.report["COLUMNS"].append([f"       Column {sheetname_2}: {detcold[n][1]}"])
            if any(h not in hdset2 for h in hdset1):
                self.report["COLUMNS"].append(
                    [f" - {sheetname_1} has following detail columns that {sheetname_2} doesn't:"]
                )
                for h in hdset1:
                    if h not in hdset2:
                        self.report["COLUMNS"].append([f"       {h}"])
            if any(h not in hdset1 for h in hdset2):
                self.report["COLUMNS"].append(
                    [f" - {sheetname_2} has following detail columns that {sheetname_1} doesn't:"]
                )
                for h in hdset2:
                    if h not in hdset1:
                        self.report["COLUMNS"].append([f"       {h}"])
        else:
            self.report["COLUMNS"].append([" - Sheets have no matching detail column names"])

        shared_ids = any(node in self.nodes2 for node in self.nodes1) or any(
            node in self.nodes1 for node in self.nodes2
        )
        if not shared_ids:
            self.report["MATCHING IDS"].append(["- Sheets have no matching IDs"])

        elif shared_ids:
            missids1 = any(ik not in self.nodes2 for ik in self.nodes1)
            missids2 = any(ik not in self.nodes1 for ik in self.nodes2)
            if missids1:
                self.report["MISSING IDS"].append(
                    [f" - {sheetname_1} has the following IDs that {sheetname_2} doesn't:"]
                )
                self.report["MISSING IDS"].extend(
                    [[f"{self.nodes1[ik].name}"] for ik in self.nodes1 if ik not in self.nodes2]
                )
            if missids2:
                self.report["MISSING IDS"].append(
                    [f" - {sheetname_2} has the following IDs that {sheetname_1} doesn't:"]
                )
                self.report["MISSING IDS"].extend(
                    [[f"{self.nodes2[ik].name}"] for ik in self.nodes2 if ik not in self.nodes1]
                )

            self.report["MATCHING IDS"].append(["ID", "DIFFERENCE", sheetname_1, sheetname_2])
            if matching_hrs_names:
                if self.row_len1 >= self.row_len2:
                    for row in self.sheet2:
                        ID = row[self.ic2]
                        if (ik := ID.lower()) in self.nodes1:
                            for nx in matching_hrs_names:
                                h1 = pcold[nx][0]
                                h2 = pcold[nx][1]
                                if self.nodes1[ik].ps[h1]:
                                    p1 = self.nodes1[ik].ps[h1].k
                                else:
                                    p1 = self.nodes1[ik].ps[h1]
                                if self.nodes2[ik].ps[h2]:
                                    p2 = self.nodes2[ik].ps[h2].k
                                else:
                                    p2 = self.nodes2[ik].ps[h2]
                                if p1 != p2 and p1 is None:
                                    if p2 == "":
                                        self.report["MATCHING IDS"].append(
                                            [
                                                f"{ID}",
                                                f"Present in hierarchy: {nx} in {sheetname_2} and not {sheetname_1}",
                                                "Not present",
                                                "Appears as top ID",
                                            ]
                                        )
                                    elif p2:
                                        self.report["MATCHING IDS"].append(
                                            [
                                                f"{ID}",
                                                f"Present in hierarchy: {nx} in {sheetname_2} and not {sheetname_1}",
                                                "Not present",
                                                f"{self.nodes2[ik].ps[h2].name}",
                                            ]
                                        )
                                elif p1 != p2 and p2 is None:
                                    if p1 == "":
                                        self.report["MATCHING IDS"].append(
                                            [
                                                f"{ID}",
                                                f"Present in hierarchy: {nx} in {sheetname_1} and not {sheetname_2}",
                                                "Appears as top ID",
                                                "Not present",
                                            ]
                                        )
                                    elif p1:
                                        self.report["MATCHING IDS"].append(
                                            [
                                                f"{ID}",
                                                f"Present in hierarchy: {nx} in {sheetname_1} and not {sheetname_2}",
                                                f"{self.nodes1[ik].ps[h1].name}",
                                                "Not present",
                                            ]
                                        )
                                elif p1 != p2 and p1 == "":
                                    self.report["MATCHING IDS"].append(
                                        [
                                            f"{ID}",
                                            f"Parents in hierarchy: {nx}",
                                            "Appears as top ID",
                                            f"{self.nodes2[ik].ps[h2].name}",
                                        ]
                                    )
                                elif p1 != p2 and p2 == "":
                                    self.report["MATCHING IDS"].append(
                                        [
                                            f"{ID}",
                                            f"Parents in hierarchy: {nx}",
                                            f"{self.nodes1[ik].ps[h1].name}",
                                            "Appears as top ID",
                                        ]
                                    )
                                elif p1 != p2:
                                    self.report["MATCHING IDS"].append(
                                        [
                                            f"{ID}",
                                            f"Parents in hierarchy: {nx}",
                                            f"{self.nodes1[ik].ps[h1].name}",
                                            f"{self.nodes2[ik].ps[h2].name}",
                                        ]
                                    )
                            for nx in matching_details_names:
                                c1 = self.sheet1[self.rns1[ik]][detcold[nx][0]]
                                c2 = row[detcold[nx][1]]
                                if c1.lower() != c2.lower():
                                    self.report["MATCHING IDS"].append(
                                        [f"{ID}", f"Details in column: {nx}", f"{c1}", f"{c2}"]
                                    )
                elif self.row_len1 < self.row_len2:
                    for row in self.sheet1:
                        ID = row[self.ic1]
                        if (ik := ID.lower()) in self.nodes2:
                            for nx in matching_hrs_names:
                                h1 = pcold[nx][0]
                                h2 = pcold[nx][1]
                                if self.nodes1[ik].ps[h1]:
                                    p1 = self.nodes1[ik].ps[h1].k
                                else:
                                    p1 = self.nodes1[ik].ps[h1]
                                if self.nodes2[ik].ps[h2]:
                                    p2 = self.nodes2[ik].ps[h2].k
                                else:
                                    p2 = self.nodes2[ik].ps[h2]
                                if p1 != p2 and p1 is None:
                                    if p2 == "":
                                        self.report["MATCHING IDS"].append(
                                            [
                                                f"{ID}",
                                                f"Present in hierarchy: {nx} in {sheetname_2} and not {sheetname_1}",
                                                "Not present",
                                                "Appears as top ID",
                                            ]
                                        )
                                    elif p2:
                                        self.report["MATCHING IDS"].append(
                                            [
                                                f"{ID}",
                                                f"Present in hierarchy: {nx} in {sheetname_2} and not {sheetname_1}",
                                                "Not present",
                                                f"{self.nodes2[ik].ps[h2].name}",
                                            ]
                                        )
                                elif p1 != p2 and p2 is None:
                                    if p1 == "":
                                        self.report["MATCHING IDS"].append(
                                            [
                                                f"{ID}",
                                                f"Present in hierarchy: {nx} in {sheetname_1} and not {sheetname_2}",
                                                "Appears as top ID",
                                                "Not present",
                                            ]
                                        )
                                    elif p1:
                                        self.report["MATCHING IDS"].append(
                                            [
                                                f"{ID}",
                                                f"Present in hierarchy: {nx} in {sheetname_1} and not {sheetname_2}",
                                                f"{self.nodes1[ik].ps[h1].name}",
                                                "Not present",
                                            ]
                                        )
                                elif p1 != p2 and p1 == "":
                                    self.report["MATCHING IDS"].append(
                                        [
                                            f"{ID}",
                                            f"Parents in hierarchy: {nx}",
                                            "Appears as top ID",
                                            f"{self.nodes2[ik].ps[h2].name}",
                                        ]
                                    )
                                elif p1 != p2 and p2 == "":
                                    self.report["MATCHING IDS"].append(
                                        [
                                            f"{ID}",
                                            f"Parents in hierarchy: {nx}",
                                            f"{self.nodes1[ik].ps[h1].name}",
                                            "Appears as top ID",
                                        ]
                                    )
                                elif p1 != p2:
                                    self.report["MATCHING IDS"].append(
                                        [
                                            f"{ID}",
                                            f"Parents in hierarchy: {nx}",
                                            f"{self.nodes1[ik].ps[h1].name}",
                                            f"{self.nodes2[ik].ps[h2].name}",
                                        ]
                                    )
                            for nx in matching_details_names:
                                c1 = row[detcold[nx][0]]
                                c2 = self.sheet2[self.rns2[ik]][detcold[nx][1]]
                                if c1.lower() != c2.lower():
                                    self.report["MATCHING IDS"].append(
                                        [f"{ID}", f"Details in column: {nx}", f"{c1}", f"{c2}"]
                                    )

            elif not matching_hrs_names:
                if self.row_len1 >= self.row_len2:
                    for row in self.sheet2:
                        ID = row[self.ic2]
                        if (ik := ID.lower()) in self.nodes1:
                            for nx in matching_details_names:
                                c1 = self.sheet1[self.rns1[ik]][detcold[nx][0]]
                                c2 = row[detcold[nx][1]]
                                if c1.lower() != c2.lower():
                                    self.report["MATCHING IDS"].append(
                                        [f"{ID}", f"Details in column: {nx}", f"{c1}", f"{c2}"]
                                    )
                elif self.row_len1 < self.row_len2:
                    for row in self.sheet1:
                        ID = row[self.ic1]
                        if (ik := ID.lower()) in self.nodes2:
                            for nx in matching_details_names:
                                c1 = row[detcold[nx][0]]
                                c2 = self.sheet2[self.rns2[ik]][detcold[nx][1]]
                                if c1.lower() != c2.lower():
                                    self.report["MATCHING IDS"].append(
                                        [f"{ID}", f"Details in column: {nx}", f"{c1}", f"{c2}"]
                                    )
        self.sheetname_1 = sheetname_1
        self.sheetname_2 = sheetname_2

        self.stop_work("Program ready")
        Compare_Report_Popup(self, theme=self.C.theme)

    def srtkey(self, e):
        return [int(c) if c.isdigit() else c.lower() for c in re.split("([0-9]+)", e)]
